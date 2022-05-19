"""
Add mechanism to extend VocExcel with more commands

The new commands be may be run either before or after VocExcel.

Copyright (c) 2022 David Linke (ORCID: 0000-0002-5898-1820)
"""
import argparse
import datetime
import glob
import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from rdflib import URIRef

from vocexcel.convert import main
from vocexcel.utils import EXCEL_FILE_ENDINGS, RDF_FILE_ENDINGS, KNOWN_FILE_ENDINGS
from vocexcel.models import ORGANISATIONS, ORGANISATIONS_INVERSE

from django.utils.text import slugify

__version__ = "0.2.0-dev"

ORGANISATIONS["NFDI4Cat"] = URIRef("http://example.org/nfdi4cat/")
ORGANISATIONS["LIKAT"] = URIRef("https://www.catalysis.de/")
ORGANISATIONS_INVERSE.update({v: k for k, v in ORGANISATIONS.items()})
NOW = datetime.datetime.now().strftime("%Y%m%dT%H%M%S")


def is_file_available(fname, ftype):
    if not type(ftype) is list:
        ftype = [ftype]
    if fname is None or not os.path.exists(fname):
        print(f"File not found: {fname}s")
        return False
    if "excel" in ftype and fname.suffix.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
        return True
    if "rdf" in ftype and fname.suffix.lower().endswith(tuple(RDF_FILE_ENDINGS)):
        return True
    return False


def has_file_in_more_than_one_format(dir_):
    files = [
        os.path.normcase(f)
        for f in glob.glob(os.path.join(dir_, "*.*"))
        if f.endswith(tuple(KNOWN_FILE_ENDINGS))
    ]
    file_names = [os.path.splitext(f)[0] for f in files]
    unique_file_names = set(file_names)
    if len(file_names) == len(unique_file_names):
        return False
    else:
        seen = set()
        duplicates = [x for x in file_names if x in seen or seen.add(x)]
        return duplicates


def is_supported_template(wb):
    # TODO add check for Excel template version
    return True


def add_IRI(fpath, outfile):
    """
    Add IRIs from preferred label in col A of sheets Concepts & Collections

    Safe and valid IRIs are created using django's slugify function.
    Column A is only updated for the rows where the cell is empty.
    """
    print(f"\nRunning add_IRI for file {fpath}")
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)
    VOC_BASE_IRI = wb["Concept Scheme"].cell(row=2, column=2).value
    if not VOC_BASE_IRI.endswith("/"):
        VOC_BASE_IRI += "/"

    # process both worksheets
    subsequent_empty_rows = 0
    for sheet in ["Concepts", "Collections"]:
        ws = wb[sheet]
        # iterate over first two columns; skip header and start from row 3
        for row in ws.iter_rows(min_row=3, max_col=2):
            if not row[0].value and row[1].value:
                concept_iri = VOC_BASE_IRI + slugify(row[1].value)
                concept_iri += "-coll" if sheet == "Collections" else ""
                ws.cell(row[0].row, column=1).value = concept_iri
                subsequent_empty_rows = 0
            elif row[0].value is None and row[1].value is None:
                # stop processing a sheet after 3 empty rows
                if subsequent_empty_rows < 2:
                    subsequent_empty_rows += 1
                else:
                    break

    wb.save(outfile)
    print(f"Saved updated file as {outfile}")


def add_related(fpath, outfile):
    """
    Add related Children URI and Members URI by preferred label if none is present.

    In detail the folowing sheets are modified:
    "Concepts": update col. G "Children URI" from col. J "Children by Pref. Label"
    "Collections": update col. D "Members URI" from col. F "Members by Pref. Label"
    """
    print(f"\nRunning add_related for file {fpath}")
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)
    # process both worksheets
    subsequent_empty_rows = 0
    for sheet in ["Concepts", "Collections"]:
        print(f"Processing sheet '{sheet}'")
        ws = wb[sheet]
        pref_label_lookup = {}
        # read all preferred labels from the sheet
        for row in ws.iter_rows(min_row=3, max_col=2):
            if row[0].value and row[1].value:
                pref_label_lookup[slugify(row[1].value)] = row[0].value
                subsequent_empty_rows = 0
            elif row[0].value is None and row[1].value is None:
                # stop processing a sheet after 3 empty rows
                if subsequent_empty_rows < 2:
                    subsequent_empty_rows += 1
                else:
                    subsequent_empty_rows = 0
                    break

        # update read all peferred labels from the sheet
        col_to_fill = 6 if sheet == "Concepts" else 3
        col_preflabel = 9 if sheet == "Concepts" else 5
        for row in ws.iter_rows(min_row=3, max_col=col_preflabel + 1):
            if not row[col_to_fill].value and row[col_preflabel].value:
                # update cell in col_to_fill
                pref_label_slugs = [
                    slugify(pl.strip()) for pl in row[col_preflabel].value.split(",")
                ]
                found = [
                    pref_label_lookup[pl]
                    for pl in pref_label_slugs
                    if pl in pref_label_lookup
                ]
                ws.cell(row[0].row, column=col_to_fill + 1).value = ", ".join(found)
                subsequent_empty_rows = 0
            elif row[0].value is None and row[1].value is None:
                # stop processing a sheet after 3 empty rows
                if subsequent_empty_rows < 2:
                    subsequent_empty_rows += 1
                else:
                    break

    wb.save(outfile)
    print(f"Saved updated file as {outfile}")


def check(fpath, outfile):
    """
    Check vocabulary in Excel sheet

    In detail the following checks are performed:
    Concepts:
    - The language-specific preferred label of a concept must be unique.
      The comparison ignores case.
    - The "Concept IRI" must be unique; this is the case when no language
      is used more than once per concept.
    """
    print(f"\nRunning check of Concepts sheet for file {fpath}")
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)
    ws = wb["Concepts"]
    color = PatternFill("solid", start_color="00FFCC00")  # orange

    subsequent_empty_rows = 0
    seen_preflabels = []
    # seen_conceptIRIs = defaultdict(list)
    seen_conceptIRIs = []
    failed_check = False
    for row_no, row in enumerate(ws.iter_rows(min_row=3, max_col=3), 3):
        if row[0].value and row[1].value:
            conceptIRI, preflabel, lang = [c.value.strip() for c in row]

            new_preflabel = f'"{preflabel}"@{lang}'.lower()
            if new_preflabel in seen_preflabels:
                failed_check = True
                print(f"Duplicate of Preferred Label found: {new_preflabel}")
                # colorise problematic cells
                row[1].fill = row[2].fill = color
                seen_in_row = 3 + seen_preflabels.index(new_preflabel)
                ws[f"B{seen_in_row}"].fill = ws[f"C{seen_in_row}"].fill = color
            else:
                seen_preflabels.append(new_preflabel)

            new_conceptIRI = f'"{conceptIRI}"@{lang.lower()}'
            if new_conceptIRI in seen_conceptIRIs:
                failed_check = True
                print(
                    f'Same Concept IRI "{conceptIRI}" used more than once for language "{lang}"'
                )
                # colorise problematic cells
                row[0].fill = row[2].fill = color
                seen_in_row = 3 + seen_conceptIRIs.index(new_conceptIRI)
                ws[f"A{seen_in_row}"].fill = ws[f"C{seen_in_row}"].fill = color
            else:
                seen_conceptIRIs.append(new_conceptIRI)

            subsequent_empty_rows = 0
        elif row[0].value is None and row[1].value is None:
            # stop processing a sheet after 3 empty rows
            if subsequent_empty_rows < 2:
                subsequent_empty_rows += 1
            else:
                subsequent_empty_rows = 0
                break

    if failed_check:
        wb.save(outfile)
        print(f"Saved file with highlighted errors as {outfile}")
    else:
        print(f"All checks passed succesfully.")


def run_vocexcel(args=None):
    main(args)


def wrapper(args=None):
    print("Running VocExcel wrapper")

    if args is None:  # vocexcel4cat run via entrypoint
        args = sys.argv[1:]

    has_args = True if args else False

    parser = argparse.ArgumentParser(
        prog="vocexcel4cat", formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    parser.add_argument(
        "-v",
        "--version",
        help="The version of this wrapper of VocExcel.",
        action="store_true",
    )

    parser.add_argument(
        "-i",
        "--add_IRI",
        help=(
            "Add IRI (http://example.org/...) for concepts and collections "
            "if none is present."
        ),
        action="store_true",
    )

    parser.add_argument(
        "-od",
        "--output_directory",
        help=(
            "Specify directory where files should be written to. "
            "The directory is created if required."
        ),
        type=Path,
        required=False,
    )

    parser.add_argument(
        "-r",
        "--add_related",
        help=(
            "Add related Children URI and Members URI by preferred label "
            "if none is present."
        ),
        action="store_true",
    )

    parser.add_argument(
        "-c",
        "--check",
        help=(
            "Perform various checks on vocabulary in Excel file (detect duplicates,...)"
        ),
        action="store_true",
    )

    parser.add_argument(
        "-f",
        "--forward",
        help=(
            "Forward file resulting from other running other options to vocexcel."
        ),
        action="store_true",
    )

    parser.add_argument(
        "file_to_preprocess",
        nargs="?",  # allow 0 or 1 file name as argument
        type=Path,
        help="Either the file to process or a directory with files to process.",
    )

    # args_wrapper = parser.parse_args(args)
    args_wrapper, vocexcel_args = parser.parse_known_args(args)

    if not has_args:
        # show help if no args are given
        parser.print_help()
        parser.exit()

    outdir = args_wrapper.output_directory
    if outdir is not None and not os.path.isdir(outdir):
        os.makedirs(outdir, exist_ok=True)

    if args_wrapper.version:
        print(f"{__version__}")
    elif args_wrapper.add_IRI or args_wrapper.add_related or args_wrapper.check:
        funcs = [
            m
            for m, to_run in zip(
                [add_IRI, add_related, check],
                [args_wrapper.add_IRI, args_wrapper.add_related, args_wrapper.check],
            )
            if to_run
        ]
        if os.path.isdir(args_wrapper.file_to_preprocess):
            for xlf in glob.glob(
                os.path.join(args_wrapper.file_to_preprocess, "*.xlsx")
            ):
                fprefix, fsuffix = xlf.rsplit(".", 1)
                fprefix = os.path.split(fprefix)[1]  # split off leading dirs
                if outdir is None:
                    outfile = Path(f"{fprefix}_{NOW}.{fsuffix}")
                else:
                    outfile = Path(outdir) / Path(f"{fprefix}_{NOW}.{fsuffix}")
                infile = xlf
                for func in funcs:
                    func(infile, outfile)
                    infile = outfile
                if args_wrapper.forward:
                    locargs = list(vocexcel_args)
                    locargs.append(str(infile))
                    main(locargs)
        elif is_file_available(args_wrapper.file_to_preprocess, ftype="excel"):
            fprefix, fsuffix = str(args_wrapper.file_to_preprocess).rsplit(".", 1)
            if outdir is None:
                outfile = Path(f"{fprefix}_{NOW}.{fsuffix}")
            else:
                outfile = Path(outdir) / Path(f"{fprefix}_{NOW}.{fsuffix}")
            infile = args_wrapper.file_to_preprocess
            for func in funcs:
                func(infile, outfile)
                infile = outfile
            if args_wrapper.forward:
                locargs = list(vocexcel_args)
                locargs.append(str(infile))
                main(locargs)
        else:
            parser.exit()
    elif args_wrapper and args_wrapper.file_to_preprocess:
        if os.path.isdir(args_wrapper.file_to_preprocess):
            dir_ = args_wrapper.file_to_preprocess
            if duplicates := has_file_in_more_than_one_format(dir_):
                print(
                    "Files may only be present in one format. Found more than one "
                    "format for:\n  " + "\n  ".join(duplicates)
                )
                parser.exit()
            for xlf in glob.glob(os.path.join(dir_, "*.xlsx")):
                print(f"Calling VocExcel for Excel file {xlf}")
                locargs = list(vocexcel_args)
                locargs.append(xlf)
                if outdir is not None:
                    fprefix, fsuffix = str(xlf).rsplit(".", 1)

                    outfile = Path(f"{fprefix}_{NOW}.{fsuffix}")

                    output_fname = Path(outdir) / os.path.split(xlf)[1]
                    locargs = ["--outputfile", str(output_fname)] + locargs
                main(locargs)
            for ttlf in glob.glob(os.path.join(dir_, "*.ttl")) + glob.glob(
                os.path.join(dir_, "*.turtle")
            ):
                print(f"Calling VocExcel for RDF/turtle file {ttlf}")
                locargs = list(vocexcel_args)
                locargs.append(ttlf)
                if outdir is not None:
                    output_fname = Path(outdir) / os.path.split(ttlf)[1]
                    locargs = ["--outputfile", str(output_fname)] + locargs
                main(locargs)
        elif is_file_available(args_wrapper.file_to_preprocess, ftype=["excel", "rdf"]):
            print(f"Calling VocExcel for file {args_wrapper.file_to_preprocess}")
            main(args)
        else:
            if os.path.exists(args_wrapper.file_to_preprocess):
                print(f"Cannot convert file {args_wrapper.file_to_preprocess}")
                endings = ", ".join(
                    [f".{e}" for e in EXCEL_FILE_ENDINGS]
                    + list(RDF_FILE_ENDINGS.keys())
                )
                print(f"Files for processing must end with one of {endings}.")
            parser.exit()
    else:
        print("\nThis part should not be reached!")

    return args_wrapper


if __name__ == "__main__":
    w = wrapper(sys.argv[1:])
