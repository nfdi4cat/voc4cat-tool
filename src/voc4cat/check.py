import glob
import logging
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from voc4cat import config
from voc4cat.checks import (
    Voc4catError,
    check_for_removed_iris,
    check_hierarchical_redundancy,
    check_number_of_files_in_inbox,
    validate_vocabulary_files_for_ci_workflow,
)
from voc4cat.convert import (
    DEFAULT_PROFILE,
    get_bundled_profiles,
    resolve_profile,
    validate_with_profile,
)
from voc4cat.models_v1 import CONCEPTS_READ_CONFIG, CONCEPTS_SHEET_NAME, ConceptV1
from voc4cat.transform import join_split_turtle
from voc4cat.utils import (
    EXCEL_FILE_ENDINGS,
    RDF_FILE_ENDINGS,
)
from voc4cat.xlsx_common import (
    XLSXFieldAnalyzer,
    XLSXRowCalculator,
    adjust_all_tables_length,
)

logger = logging.getLogger(__name__)


def check_xlsx(fpath: Path, outfile: Path) -> int:
    """
    Complex checks of the xlsx file not handled by pydantic model validation

    Checks/validation implemented here need/use information from several rows
    and can therefore not be easily done in models.py with pydantic.
    However, SHACL validation does also catch the problem, at least in pre-
    liminary testing. So this function may get removed later. Doing the check
    in Excel has the advantage that the cell position can be added to the
    validation message which would not work with SHACL validation.

    For concepts:
    - The "Concept IRI" must be unique. However, it can be present in several
      rows as long as the languages in the rows with the same IRI are
      different. This condition is fulfilled when no language is used more
      than once per concept.
    """
    logger.debug("Running check of Concepts sheet for file %s", fpath)
    wb = openpyxl.load_workbook(fpath)
    ws = wb["Concepts"]
    color = PatternFill("solid", start_color="00FFCC00")  # orange

    # Calculate data start row dynamically using xlsx-pydantic infrastructure
    field_analyses = XLSXFieldAnalyzer.analyze_model(ConceptV1)
    fields = list(field_analyses.values())
    row_calculator = XLSXRowCalculator(CONCEPTS_READ_CONFIG)
    data_start_row = row_calculator.get_data_start_row(fields)

    subsequent_empty_rows = 0
    seen_concept_iris = []
    failed_checks = 0
    # v1.0 template: data starts after header row, columns are IRI(A), Language(B)
    for row in ws.iter_rows(min_row=data_start_row, max_col=2):  # pragma: no branch
        if row[0].value and row[1].value:
            concept_iri, lang = (
                c.value.strip() if c.value is not None else "" for c in row
            )
            # Check that IRI is valid.
            #            config.IDRANGES
            # Check that IRI is used for exactly one concept.
            new_concept_iri = f'"{concept_iri}"@{lang.lower()}'
            if new_concept_iri in seen_concept_iris:
                failed_checks += 1
                msg = (
                    f'Same Concept IRI "{concept_iri}" used more than once for '
                    f'language "{lang}"'
                )
                logger.error(msg)
                # colorize problematic cells (columns A and B for IRI and Language)
                row[0].fill = color
                row[1].fill = color
                previously_seen_in_row = data_start_row + seen_concept_iris.index(
                    new_concept_iri
                )
                ws[f"A{previously_seen_in_row}"].fill = color
                ws[f"B{previously_seen_in_row}"].fill = color
            else:
                seen_concept_iris.append(new_concept_iri)

            subsequent_empty_rows = 0

        # stop processing a sheet after 3 empty rows
        elif subsequent_empty_rows < 2:  # noqa: PLR2004
            subsequent_empty_rows += 1
        else:
            subsequent_empty_rows = 0
            break

    if failed_checks:
        wb.save(outfile)
        logger.info("-> Saved file with highlighted errors as %s", outfile)
        # Extend size (length) of tables in all sheets
        adjust_all_tables_length(
            outfile,
            rows_pre_allocated=config.xlsx_rows_pre_allocated,
            active_sheet=CONCEPTS_SHEET_NAME,
        )
        return

    logger.info("-> xlsx check passed for file: %s", fpath)


def ci_post(args):
    prev_vocab_dir, vocab_new = args.ci_post, args.VOCAB
    for vocfile in glob.glob(str(vocab_new.resolve() / "*.ttl")):
        new = Path(vocfile)
        prev_split_voc = prev_vocab_dir / new.stem
        # The prev version could be in split form in a vocabulary directory
        if (
            prev_split_voc.exists()
            and prev_split_voc.is_dir()
            and any(prev_split_voc.glob("*.ttl"))
        ):
            # Create a single vocab out of the directory
            logger.debug("-> previous version is a split vocabulary, joining...")
            join_split_turtle(prev_split_voc)

        prev = prev_vocab_dir / new.name
        if not prev.exists():
            logger.debug(
                '-> previous version of vocabulary "%s" does not exist.',
                new.name,
            )
            continue
        check_for_removed_iris(prev, new)
        logger.info("-> Check ci-post passed.")


# ===== check command & helpers to validate cmd options =====


def _check_ci_args(args):
    msg = ""
    if args.ci_pre:
        if args.VOCAB and args.VOCAB.is_dir() and args.ci_pre.is_dir():
            return
        msg = "Need two dirs for ci_pre!"
    if args.ci_post:
        if args.VOCAB and args.VOCAB.is_dir() and args.ci_post.is_dir():
            return
        msg = "Need two dirs for ci_post!"
    if msg:
        logging.error(msg)
        raise Voc4catError(msg)


def check(args):
    logger.debug("Check subcommand started!")

    _check_ci_args(args)

    if not args.VOCAB and not args.listprofiles:
        args._parser.print_help()
        sys.exit(2)

    if args.listprofiles:
        cwd = Path.cwd()
        s = "\nKnown profiles:\n\nSource\tToken\tPath\n------\t-----\t----\n"
        # Bundled profiles (dynamically scanned)
        bundled = get_bundled_profiles()
        for token in sorted(bundled.keys()):
            try:
                rel_path = bundled[token].relative_to(cwd)
            except ValueError:
                rel_path = bundled[token]
            s += f"bundled\t{token}\t{rel_path}\n"
        # Custom profiles from config (if config is loaded)
        if not config.IDRANGES.default_config and config.IDRANGES_PATH:
            for vocab_name, vocab_config in config.IDRANGES.vocabs.items():
                if vocab_config.profile_local_path:
                    profile_path = (
                        config.IDRANGES_PATH.parent / vocab_config.profile_local_path
                    ).resolve()
                    try:
                        rel_path = profile_path.relative_to(cwd)
                    except ValueError:
                        rel_path = profile_path
                    s += f"config\t{vocab_name}\t{rel_path}\n"
        print(s.rstrip())
        return

    if args.ci_pre:
        inbox_dir, vocab_dir = args.ci_pre, args.VOCAB
        check_number_of_files_in_inbox(inbox_dir)
        validate_vocabulary_files_for_ci_workflow(vocab_dir, inbox_dir)
        logger.info("-> Check ci-pre passed.")
        return

    if args.ci_post:
        ci_post(args)
        return

    files = [args.VOCAB] if args.VOCAB.is_file() else [*Path(args.VOCAB).iterdir()]
    xlsx_files = [f for f in files if f.suffix.lower() in EXCEL_FILE_ENDINGS]
    rdf_files = [f for f in files if f.suffix.lower() in RDF_FILE_ENDINGS]

    # check xlsx files
    for file in xlsx_files:
        outfile = file if args.outdir is None else args.outdir / file.name
        if outfile == file and not args.inplace:
            logger.warning(
                'This command will overwrite the existing file "%s". '
                'Use the flag "--inplace" to enforce replacement or '
                'supply an output directory with flag "--outdir".',
                file,
            )
            return
        check_xlsx(file, outfile)

    # validate rdf files with profile/pyshacl
    all_redundancies = {}  # file -> list of redundancies
    for file in rdf_files:
        logger.debug("Running SHACL validation for file %s", file)
        # Priority: CLI --profile (highest) > config profile_local_path > default
        # If user explicitly set --profile (not default), use it
        user_set_profile = args.profile != DEFAULT_PROFILE
        if user_set_profile:
            effective_profile = args.profile
        else:
            # Check config for vocab-specific profile
            vocab_name = file.stem.lower()
            vocab_config = config.IDRANGES.vocabs.get(vocab_name)
            if (
                vocab_config
                and vocab_config.profile_local_path
                and config.IDRANGES_PATH
            ):
                effective_profile = str(
                    (
                        config.IDRANGES_PATH.parent / vocab_config.profile_local_path
                    ).resolve()
                )
            else:
                effective_profile = args.profile  # default
        validate_with_profile(
            str(file),
            profile=effective_profile,
            error_level=args.fail_at_level,
        )
        # Get profile name for log message
        _, profile_name = resolve_profile(effective_profile)
        logger.info("-> The file is valid according to the %s profile.", profile_name)

        # Check for redundant hierarchical relationships if requested
        if args.redundant_hierarchies:
            redundancies = check_hierarchical_redundancy(file)
            if redundancies:
                all_redundancies[file] = redundancies

    # Report all redundant hierarchical relationships at the end
    if args.redundant_hierarchies:
        logger.info("Checking for redundant hierarchical relationships.")
    if all_redundancies:
        for file, redundancies in all_redundancies.items():
            logger.error("File: %s", file.name)
            for concept, ancestor, via_parent in redundancies:
                logger.error(
                    "  From %s remove Parent IRI (skos:broader) %s",
                    concept,
                    ancestor,
                )
                logger.error(
                    "       (already reachable via %s)",
                    via_parent,
                )
        total = sum(len(r) for r in all_redundancies.values())
        logger.error("Total: %d redundant relationship(s) to remove", total)
    else:
        logger.info("-> No redundant hierarchical relationships detected.")
