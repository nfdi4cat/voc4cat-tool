"""Born as a wrapper to extend VocExcel with more commands."""

import glob
import logging
import os
from collections import defaultdict
from itertools import count, zip_longest
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, PatternFill

from voc4cat import config
from voc4cat.util import (
    dag_from_indented_text,
    dag_from_narrower,
    dag_to_narrower,
    dag_to_node_levels,
    get_concept_and_level_from_indented_line,
)

logger = logging.getLogger(__name__)


def is_supported_template(wb):
    # TODO add check for xlsx template version
    return True


def make_ids(
    fpath: Path, outfile: Path, search_prefix: str, start_id: int, base_iri=None
):
    """
    Replace all prefix:suffix CURIEs using IDs counting up from start_id

    If base_iri is None the new IRI is a concatenation of VOC_BASE_IRI and ID.
    If base_iri is given the new IRI is the concatenation of base_iri and ID.
    """
    logger.info("\nReplacing '%s' IRIs.", search_prefix)
    # Load in data_only mode to get cell values not formulas.
    wb = openpyxl.load_workbook(fpath, data_only=True)
    is_supported_template(wb)
    if base_iri is None:
        base_iri = wb["Concept Scheme"].cell(row=2, column=2).value
        if base_iri is None:
            base_iri = "https://example.org/"
            wb["Concept Scheme"].cell(row=2, column=2).value = base_iri
            logger.warning("No concept scheme IRI found, using https://example.org/")

    # TODO if config is set: Check that file name is in config
    voc_config = config.IDRANGES.vocabs.get(fpath.stem, {})
    id_length = voc_config.get("id_length", 7)

    id_gen = count(start_id)
    replaced_iris = {}
    # process primary iri column in both worksheets
    for sheet in ["Concepts", "Collections"]:
        ws = wb[sheet]
        # iterate over first two columns; skip header and start from row 3
        for row in ws.iter_rows(min_row=3, max_col=2):  # pragma: no branch
            if row[0].value:
                iri = str(row[0].value)
                if not iri.startswith(search_prefix):
                    continue
                if iri in replaced_iris:
                    iri_new = replaced_iris[iri]
                else:
                    iri_new = base_iri + f"{next(id_gen):0{id_length}d}"
                    msg = f"[{sheet}] Replaced CURIE {iri} by {iri_new}"
                    logger.debug(msg)
                    replaced_iris[iri] = iri_new
                row[0].value = iri_new

    # replace iri by new_iri in all columns where it might be referenced
    cols_to_update = {
        "Concepts": [6],
        "Collections": [3],
        "Additional Concept Features": [0, 1, 2, 3, 4, 5],
    }
    for sheet, cols in cols_to_update.items():
        ws = wb[sheet]
        # iterate over first two columns; skip header and start from row 3
        for row in ws.iter_rows(min_row=3, max_col=1 + max(cols)):  # pragma: no branch
            # stop processing a sheet after 3 empty rows
            for col in cols:
                if row[col].value:
                    iris = [iri.strip() for iri in str(row[col].value).split(",")]
                    count_new_iris = [1 for iri in iris if iri in replaced_iris].count(
                        1
                    )
                    if count_new_iris:
                        new_iris = [replaced_iris.get(iri, iri) for iri in iris]
                        msg = (
                            f"[{sheet}] Replaced {count_new_iris} CURIs "
                            f"in cell {row[col].coordinate}"
                        )
                        logger.debug(msg)
                        row[col].value = ", ".join(new_iris)
                    # subsequent_empty_rows = 0

    wb.save(outfile)
    logger.info("Saved updated file as %s", outfile)
    return 0


def hierarchy_from_indent(fpath, outfile, sep):
    """
    Convert indentation hierarchy of concepts to children-URI form.

    If separator character(s) are given they will be replaced with
    the xlsx default indent. This is also the default if sep is None.
    """
    logger.info("Reading concepts from file %s", fpath)
    # Load in data_only mode to get cell values not formulas.
    wb = openpyxl.load_workbook(fpath, data_only=True)
    is_supported_template(wb)
    # process both worksheets
    subsequent_empty_rows = 0
    ws = wb["Concepts"]
    concepts_indented = []
    row_by_iri = defaultdict(dict)
    # read concepts, determine their indentation level, then clear indentation
    col_last = 9
    max_row = 0
    for row in ws.iter_rows(min_row=3, max_col=col_last):  # pragma: no branch
        iri = row[0].value
        if iri and row[1].value:
            row_no = row[0].row
            lang = row[2].value

            if sep is None:  # xlsx indentation
                level = int(row[1].alignment.indent)
                ws.cell(row_no, column=2).alignment = Alignment(indent=0)
            else:
                descr, level = get_concept_and_level_from_indented_line(
                    row[1].value, sep=sep
                )
                ws.cell(row_no, column=2).value = descr
                ws.cell(row_no, column=2).alignment = Alignment(indent=0)
            concepts_indented.append(level * " " + iri)

            if iri in row_by_iri:  # merge needed
                # compare fields, merge if one is empty, error if different values
                new_data = [
                    ws.cell(row_no, col_no).value for col_no in range(2, col_last)
                ]
                old_data = row_by_iri[iri].get(lang, [None] * (len(new_data)))
                merged = []
                for old, new in zip(old_data, new_data, strict=True):
                    if (old and new) and (old != new):
                        msg = f"Cannot merge rows for {iri}. Resolve differences manually."
                        raise ValueError(msg)
                    merged.append(old if old else new)
                row_by_iri[iri][lang] = merged
            else:
                row_by_iri[iri][lang] = [
                    ws.cell(row_no, col_no).value for col_no in range(2, col_last)
                ]
            max_row = row_no

        # stop processing a sheet after 3 empty rows
        elif subsequent_empty_rows < 2:  # noqa: PLR2004
            subsequent_empty_rows += 1
        else:
            subsequent_empty_rows = 0
            break

    term_dag = dag_from_indented_text("\n".join(concepts_indented))
    children_by_iri = dag_to_narrower(term_dag)

    # Write rows to xlsx-file, write translation directly after each concept.
    col_children_iri = 7
    row = 3
    for iri in children_by_iri:
        for lang in row_by_iri[iri]:
            ws.cell(row, column=1).value = iri
            for col, stored_value in zip_longest(
                range(2, col_last), row_by_iri[iri][lang]
            ):
                ws.cell(row, column=col).value = stored_value
            ws.cell(row, column=col_children_iri).value = ", ".join(
                children_by_iri[iri]
            )
            row += 1

    while row <= max_row:
        # Clear remaining rows.
        for col in range(1, col_last):
            ws.cell(row, column=col).value = None
        row += 1

    wb.save(outfile)
    logger.info("Saved file with children-IRI hierarchy as %s", outfile)
    return 0


def hierarchy_to_indent(fpath, outfile, sep):
    """
    Convert concept hierarchy in children-URI form to indentation.

    If separator character(s) are given they will be used.
    If sep is None, xlsx default indent will be used.
    """
    logger.info("Reading concepts from file %s", fpath)
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)
    ws = wb["Concepts"]

    concept_children_dict = {}
    subsequent_empty_rows = 0
    row_by_iri = defaultdict(dict)
    col_last = 9
    # read all IRI, preferred labels, children_uris from the sheet
    for rows_total, row in enumerate(  # pragma: no branch
        ws.iter_rows(min_row=3, max_col=col_last, values_only=True)
    ):
        if row[0] and row[1]:
            iri = row[0]
            lang = row[2]
            children_uris = [] if not row[6] else [c.strip() for c in row[6].split(",")]
            # We need to check if ChildrenIRI, Provenance & Source Vocab URL
            # are consistent across languages since SKOS has no support for
            # per language statements. (SKOS-XL would add this)
            if iri in row_by_iri:  # merge needed
                # compare fields, merge if one is empty, error if different values
                new_data = [row[col_no] for col_no in range(6, col_last)]
                old_data = row_by_iri[iri][next(iter(row_by_iri[iri].keys()))][5:]
                merged = []
                for old, new in zip(old_data, new_data):
                    if (old and new) and (old != new):
                        msg = f'Merge conflict for concept {iri}. New: "{new}" - Before: "{old}"'
                        raise ValueError(msg)
                    merged.append(old if old else new)
                row_by_iri[iri][lang] = [row[col] for col in range(1, 6)] + merged
            else:
                row_by_iri[iri][lang] = [row[col] for col in range(1, col_last)]
                concept_children_dict[iri] = children_uris

        # stop processing a sheet after 3 empty rows
        elif subsequent_empty_rows < 2:  # noqa: PLR2004
            subsequent_empty_rows += 1
        else:
            subsequent_empty_rows = 0
            rows_total -= 2  # noqa: PLW2901
            break
    else:
        pass

    term_dag = dag_from_narrower(concept_children_dict)
    concept_levels = dag_to_node_levels(term_dag)

    # Write indented rows to xlsx-file but
    # 1 - each concept translation only once
    # 2 - details only once per concept and language
    iri_written = []
    row = 3
    for iri, level in concept_levels:
        for transl_no, lang in enumerate(row_by_iri[iri]):
            if transl_no and (iri, lang) in iri_written:  # case 1
                continue
            ws.cell(row, 1).value = iri
            concept_text = row_by_iri[iri][lang][0]
            if sep is None:
                ws.cell(row, 2).value = concept_text
                ws.cell(row, 2).alignment = Alignment(indent=level)
            else:
                ws.cell(row, 2).value = sep * level + concept_text
                ws.cell(row, 2).alignment = Alignment(indent=0)
            ws.cell(row, 3).value = lang

            if (iri, lang) in iri_written:  # case 2
                for col, _ in zip_longest(
                    range(4, col_last + 1), row_by_iri[iri][lang][2:]
                ):
                    ws.cell(row, column=col).value = None
                row += 1
                continue

            for col, stored_value in zip_longest(
                range(4, col_last + 1), row_by_iri[iri][lang][2:]
            ):
                ws.cell(row, column=col).value = stored_value
            # clear children IRI column G
            ws.cell(row, column=7).value = None
            row += 1
            iri_written.append((iri, lang))

    wb.save(outfile)
    logger.info("Saved file with indentation hierarchy as %s", outfile)
    return 0


def run_pylode(file_path, output_path):
    """
    Generate pyLODE documentation.
    """
    import pylode

    turtle_files = find_files_to_document(file_path)
    if not turtle_files:
        return 1

    for turtle_file in turtle_files:
        logger.info("Building pyLODE documentation for %s", turtle_file)
        filename = Path(turtle_file)  # .resolve())
        outdir = output_path / filename.stem
        outdir.mkdir(exist_ok=True)
        outfile = outdir / "index.html"
        # breakpoint()
        html = pylode.MakeDocco(
            # we use str(abs path) because pyLODE 2.x does not find the file otherwise
            input_data_file=str(filename.resolve()),
            outputformat="html",
            profile="vocpub",
        )
        html.document(destination=outfile)
        # Fix html-doc: Do not show overview section with black div for image.
        with open(outfile) as html_file:
            content = html_file.read()
        content = content.replace(
            '<section id="overview">',
            '<section id="overview" style="display: none;">',
        )
        with open(outfile, "w") as html_file:
            html_file.write(content)
        logger.info("-> %s", outfile.resolve().as_uri())
    return 0


def run_ontospy(file_path, output_path):
    """
    Generate ontospy documentation (single-page html & dendrogram).
    """
    import ontospy
    from ontospy.gendocs.viz.viz_d3dendogram import Dataviz
    from ontospy.gendocs.viz.viz_html_single import HTMLVisualizer

    turtle_files = find_files_to_document(file_path)
    if not turtle_files:
        return 1

    for turtle_file in turtle_files:
        title = config.VOCAB_TITLE

        logger.info("Building ontospy documentation for %s", turtle_file)
        specific_output_path = (Path(output_path) / Path(turtle_file).stem).resolve()

        g = ontospy.Ontospy(Path(turtle_file).resolve().as_uri())

        docs = HTMLVisualizer(g, title=title)
        docs_path = os.path.join(specific_output_path, "docs")
        docs.build(docs_path)  # => build and save docs/visualization.

        viz = Dataviz(g, title=title)
        viz_path = os.path.join(specific_output_path, "dendro")
        viz.build(viz_path)  # => build and save docs/visualization.

    return 0


def find_files_to_document(file_path):
    if Path(file_path).is_dir():
        turtle_files = glob.glob(f"{file_path}/*.ttl")
        if not turtle_files:
            logger.warning("No turtle file(s) found to document in %s", file_path)
            turtle_files = []
    elif Path(file_path).exists():
        turtle_files = [file_path]
    else:
        logger.warning("File/dir not found (for docs): %s", file_path)
        turtle_files = []
    return turtle_files


def build_docs(file_path, output_path, doc_builder):
    """
    Generate documentation for a file or directory of files.
    """
    if doc_builder == "ontospy":
        errcode = run_ontospy(file_path, output_path)
    elif doc_builder == "pylode":
        errcode = run_pylode(file_path, output_path)
    else:
        logger.error("Unsupported document builder '%s'.", doc_builder)
        errcode = 1
    return errcode


def check(fpath: Path, outfile: Path) -> int:
    """
    Complex checks of the xlsx file not handled by pydantic model validation

    Checks/validation implemented here need/use information from several rows
    and can therefore not be easily done in models.py with pydantic.
    However, SHACL validation does also catch the problem, at least in pre-
    liminary testing. So this function may get removed later. Doing the check
    in Excel has the advantage that the cell position can be added to the
    validation message which would not work with SHACL validation.

    For concepts:
    - The "Concept IRI" must be unique. However, it can be present in several
      rows as long as the languages in the rows with the same IRI are
      different. This condition is fulfilled when no language is used more
      than once per concept.
    """
    logger.info("Running check of Concepts sheet for file %s", fpath)
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)
    ws = wb["Concepts"]
    color = PatternFill("solid", start_color="00FFCC00")  # orange

    subsequent_empty_rows = 0
    seen_concept_iris = []
    failed_checks = 0
    for row in ws.iter_rows(min_row=3, max_col=3):  # pragma: no branch
        if row[0].value and row[1].value:
            concept_iri, _, lang = (
                c.value.strip() if c.value is not None else "" for c in row
            )
            # Check that IRI is valid.
            #            config.IDRANGES
            # Check that IRI is used for exactly one concept.
            new_concept_iri = f'"{concept_iri}"@{lang.lower()}'
            if new_concept_iri in seen_concept_iris:
                failed_checks += 1
                msg = (
                    f'Same Concept IRI "{concept_iri}" used more than once for '
                    f'language "{lang}"'
                )
                logger.error(msg)
                # colorize problematic cells
                row[0].fill = color
                row[2].fill = color
                previously_seen_in_row = 3 + seen_concept_iris.index(new_concept_iri)
                ws[f"A{previously_seen_in_row}"].fill = color
                ws[f"C{previously_seen_in_row}"].fill = color
            else:
                seen_concept_iris.append(new_concept_iri)

            subsequent_empty_rows = 0

        # stop processing a sheet after 3 empty rows
        elif subsequent_empty_rows < 2:  # noqa: PLR2004
            subsequent_empty_rows += 1
        else:
            subsequent_empty_rows = 0
            break

    if failed_checks:
        wb.save(outfile)
        logger.info("-> Saved file with highlighted errors as %s", outfile)
        return 1

    logger.info("-> Extended xlsx checks passed successfully.")

    if fpath != outfile:  # Save to new directory (important for --forward option)
        wb.save(outfile)
        msg = f"Saved checked file as {outfile}"
        logger.info(msg)

    return 0
