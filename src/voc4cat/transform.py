import logging
import os
import shutil
from collections import defaultdict
from itertools import count, zip_longest
from pathlib import Path
from urllib.parse import urlsplit

import openpyxl
from openpyxl.styles import Alignment
from rdflib import DCTERMS, OWL, RDF, SDO, SKOS, XSD, Graph, Literal, URIRef

from voc4cat import config
from voc4cat.checks import Voc4catError
from voc4cat.models import ORGANISATIONS
from voc4cat.utils import (
    EXCEL_FILE_ENDINGS,
    RDF_FILE_ENDINGS,
    adjust_length_of_tables,
    is_supported_template,
)

logger = logging.getLogger(__name__)


def extract_numeric_id_from_iri(iri):
    iri_path = urlsplit(iri).path
    reverse_id = []
    for char in reversed(iri_path):
        if char.isdigit():
            reverse_id.append(char)
        elif char == "/":
            continue
        else:
            break
    return "".join(reversed(reverse_id))


def write_split_turtle(vocab_graph: Graph, outdir: Path) -> None:
    """
    Write each concept, collection and concept scheme to a separate turtle file.

    The ids are used as filenames.
    """
    outdir.mkdir(exist_ok=True)
    query = "SELECT ?iri WHERE {?iri a %s.}"

    for skos_class in ["skos:Concept", "skos:Collection", "skos:ConceptScheme"]:
        qresults = vocab_graph.query(query % skos_class, initNs={"skos": SKOS})
        # Iterate over search results and write each concept, collection and
        # concept scheme to a separate turtle file using id as filename.
        for qresult in qresults:
            iri = qresult["iri"]
            tmp_graph = Graph()
            tmp_graph += vocab_graph.triples((iri, None, None))
            id_part = extract_numeric_id_from_iri(iri)
            if skos_class == "skos:ConceptScheme":
                outfile = outdir / "concept_scheme.ttl"
            else:
                outfile = outdir / f"{id_part}.ttl"
            tmp_graph.serialize(destination=outfile, format="longturtle")
        logger.debug("-> wrote %i %ss-file(s).", len(qresults), skos_class)


def autoversion_cs(graph: Graph) -> Graph:
    """Set modified date and version if "requested" via environment variables."""
    if any(graph.triples((None, RDF.type, SKOS.ConceptScheme))):
        cs, _, _ = next(graph.triples((None, RDF.type, SKOS.ConceptScheme)))
    if os.getenv("VOC4CAT_MODIFIED") is not None:
        graph.remove((None, DCTERMS.modified, None))
        date_modified = os.getenv("VOC4CAT_MODIFIED")
        graph.add((cs, DCTERMS.modified, Literal(date_modified, datatype=XSD.date)))
    if os.getenv("VOC4CAT_VERSION") is not None:
        version = os.getenv("VOC4CAT_VERSION")
        if version is not None and not version.startswith("v"):
            msg = 'Invalid environment variable VOC4CAT_VERSION "%s". Version must start with letter "v".'
            logger.error(msg, version)
            raise Voc4catError(msg % version)
        graph.remove((None, OWL.versionInfo, None))
        graph.add(
            (cs, OWL.versionInfo, Literal(version)),
        )
    return graph


def join_split_turtle(vocab_dir: Path) -> Graph:
    # Search recursively all turtle files belonging to the concept scheme
    turtle_files = vocab_dir.rglob("*.ttl")
    # Create an empty RDF graph to hold the concept scheme
    cs_graph = Graph()
    # Load each turtle file into a separate graph and merge it into the concept scheme graph
    for file in turtle_files:
        graph = Graph().parse(file, format="turtle")
        # Set modified date if "requested" via environment variable.
        if file.name == "concept_scheme.ttl" or any(
            graph.triples((None, RDF.type, SKOS.ConceptScheme))
        ):
            graph = autoversion_cs(graph)
        cs_graph += graph

    # TODO The metadata should also be present in a separate ttl-file and
    #    joined just like the other.
    # Newer for vocpub profile require creator_/publisher for vocabulary:
    # Requirement 2.1.6 Each vocabulary MUST have at least one creator,
    #   indicated using sdo:creator or dcterms:creator predicate and exactly
    #   one publisher, indicated using sdo:publisher or dcterms:publisher,
    #   all of which MUST be IRIs indicating instances of sdo:Person,
    #   or sdo:Organization.
    # Get the publisher from the first concept scheme
    publisher = next(cs_graph.triples((None, DCTERMS.publisher, None)))[2]
    tg = Graph()
    org = ORGANISATIONS.get(publisher, URIRef(publisher))
    tg.add((org, RDF.type, SDO.Organization))
    tg.add(
        (
            org,
            SDO.name,
            # should be name but there is no field in the template 0.43
            Literal(
                ORGANISATIONS.get(publisher, publisher),
            ),
        )
    )
    tg.add(
        (
            org,
            SDO.url,
            Literal(
                ORGANISATIONS.get(publisher, publisher),
                datatype=URIRef(XSD.anyURI),
            ),
        )
    )
    cs_graph += tg

    return cs_graph


def make_ids(
    fpath: Path, outfile: Path, search_prefix: str, start_id: int, base_iri=None
):
    """
    Replace all prefix:suffix CURIEs using IDs counting up from start_id

    If base_iri is None the new IRI is a concatenation of VOC_BASE_IRI and ID.
    If base_iri is given the new IRI is the concatenation of base_iri and ID.
    """
    logger.info("Replacing '%s' IRIs.", search_prefix)
    # Load in data_only mode to get cell values not formulas.
    wb = openpyxl.load_workbook(fpath, data_only=True)
    is_supported_template(wb)
    if base_iri is None:
        base_iri = wb["Concept Scheme"].cell(row=2, column=2).value
        if base_iri is None:
            base_iri = "https://example.org/"
            wb["Concept Scheme"].cell(row=2, column=2).value = base_iri
            logger.warning("No concept scheme IRI found, using https://example.org/")

    # TODO if config is set: Check that file name is in config
    voc_config = config.IDRANGES.vocabs.get(fpath.stem, {})
    id_length = voc_config.get("id_length", 7)

    id_gen = count(start_id)
    replaced_iris = {}
    # process primary iri column in both worksheets
    for sheet in ["Concepts", "Collections"]:
        ws = wb[sheet]
        # iterate over first two columns; skip header and start from row 3
        for row in ws.iter_rows(min_row=3, max_col=2):  # pragma: no branch
            if row[0].value:
                iri = str(row[0].value)
                if not iri.startswith(search_prefix):
                    continue
                if iri in replaced_iris:
                    iri_new = replaced_iris[iri]
                else:
                    iri_new = base_iri + f"{next(id_gen):0{id_length}d}"
                    msg = f"[{sheet}] Replaced CURIE {iri} by {iri_new}"
                    logger.debug(msg)
                    replaced_iris[iri] = iri_new
                row[0].value = iri_new

    # replace iri by new_iri in all columns where it might be referenced
    cols_to_update = {
        "Concepts": [6],
        "Collections": [3],
        "Additional Concept Features": [0, 1, 2, 3, 4, 5],
    }
    for sheet, cols in cols_to_update.items():
        ws = wb[sheet]
        # iterate over first two columns; skip header and start from row 3
        for row in ws.iter_rows(min_row=3, max_col=1 + max(cols)):  # pragma: no branch
            # stop processing a sheet after 3 empty rows
            for col in cols:
                if row[col].value:
                    iris = [iri.strip() for iri in str(row[col].value).split(",")]
                    count_new_iris = [1 for iri in iris if iri in replaced_iris].count(
                        1
                    )
                    if count_new_iris:
                        new_iris = [replaced_iris.get(iri, iri) for iri in iris]
                        msg = (
                            f"[{sheet}] Replaced {count_new_iris} CURIs "
                            f"in cell {row[col].coordinate}"
                        )
                        logger.debug(msg)
                        row[col].value = ", ".join(new_iris)
                    # subsequent_empty_rows = 0

    wb.save(outfile)
    logger.info("Saved updated file as %s", outfile)






# ===== transform command & helpers to validate cmd options =====


def _check_make_ids_args(args):
    """Validate make_ids arguments"""
    try:
        start_id = int(args.make_ids[1])
        msg = "" if start_id > 0 else "Start ID must be greater than zero."
    except ValueError:
        msg = "Start ID must be an integer number."
    if msg:
        logger.error(msg)
        raise Voc4catError(msg)
    if ":" in args.make_ids[0]:
        prefix, base_iri = args.make_ids[0].split(":", 1)
        if not base_iri.strip().startswith("http"):
            msg = 'The base_iri must be in IRI-form and start with "http".'
            logging.error(msg)
            raise Voc4catError(msg)
        base_iri = base_iri.strip()
    else:
        prefix, base_iri = args.make_ids[0], None
    return prefix.strip(), base_iri, start_id




def _transform_xlsx(file, args):
    if args.make_ids:
        prefix, base_iri, start_id = _check_make_ids_args(args)

    logger.debug('Processing "%s"', file)

    outfile = file if args.outdir is None else args.outdir / file.name
    any_action = args.make_ids
    if outfile == file and not args.inplace and any_action:
        logger.warning(
            'This command will overwrite the existing file "%s".'
            'Use the flag "--inplace" to enforce replacement or '
            'supply an output directory with flag "--outdir".',
            file,
        )
        return
    if args.make_ids:
        make_ids(file, outfile, prefix, start_id, base_iri)
    else:
        logger.debug("-> nothing to do for xlsx files!")

    # Extend size (length) of tables in all sheets
    adjust_length_of_tables(outfile, rows_pre_allocated=config.xlsx_rows_pre_allocated)


def _transform_rdf(file, args):
    if args.split:
        vocab_graph = Graph().parse(str(file), format=RDF_FILE_ENDINGS[file.suffix])
        vocab_dir = (
            args.outdir / file.with_suffix("").name
            if args.outdir
            else file.with_suffix("")
        )
        vocab_dir.mkdir(exist_ok=True)
        write_split_turtle(vocab_graph, vocab_dir)
        logger.info("-> wrote split vocabulary to: %s", vocab_dir)
        if args.inplace:
            logger.debug("-> going to remove %s", file)
            file.unlink()
    else:
        logger.debug("-> nothing to do for rdf files!")


def transform(args):
    logger.debug("Transform subcommand started!")

    files = [args.VOCAB] if args.VOCAB.is_file() else [*Path(args.VOCAB).iterdir()]
    xlsx_files = [f for f in files if f.suffix.lower() in EXCEL_FILE_ENDINGS]

    rdf_files = [f for f in files if f.suffix.lower() in RDF_FILE_ENDINGS]

    if args.VOCAB.is_file() and (len(xlsx_files) + len(rdf_files)) == 0:
        logger.warning("Unsupported filetype: %s", args.VOCAB)

    if args.join:
        rdf_dirs = [d for d in Path(args.VOCAB).iterdir() if any(d.glob("*.ttl"))]
    else:
        rdf_dirs = []

    # transform xlsx files (could be a separate function)
    for file in xlsx_files:
        _transform_xlsx(file, args)

    for file in rdf_files:
        logger.debug('Processing "%s"', file)
        _transform_rdf(file, args)

    for rdf_dir in rdf_dirs:
        logger.debug('Processing rdf files in "%s"', rdf_dir)
        # The if..else is not required now. It is a frame for future additions.
        if args.join:
            vocab_graph = join_split_turtle(rdf_dir)
            dest = (
                (args.outdir / rdf_dir.name).with_suffix(".ttl")
                if args.outdir
                else rdf_dir.with_suffix(".ttl")
            )
            vocab_graph.serialize(destination=str(dest), format="turtle")
            logger.info("-> joined vocabulary into: %s", dest)
            if args.inplace:
                logger.debug("-> going to remove %s", rdf_dir)
                shutil.rmtree(rdf_dir, ignore_errors=True)
        else:  # pragma: no cover
            logger.debug("-> nothing to do!")
