import logging
from datetime import datetime
from typing import Any

from curies import Converter
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from voc4cat import config, models
from voc4cat.utils import ConversionError, split_and_tidy

logger = logging.getLogger(__name__)


header_names_to_skip = [
    "Concepts",
    "Concepts*",
    "Concept IRI",
    "Concept IRI*",
    "Concept Relations",
    "Concept Extras",
]


def write_prefix_sheet(wb: Workbook, prefix_map):
    """
    Write prefix_dict to prefix sheet
    """
    ws = wb["Prefix Sheet"]

    # Clear the sheet except for the header.
    ws.delete_rows(2, ws.max_row - 1)

    for prefix, iri in prefix_map.items():
        ws.append([prefix, iri])


def split_multi_iri(cell_value: str | None, prefix_converter: Converter) -> list[str]:
    """
    Split a string of IRIs separated by a comma into a list of IRIs
    """
    if cell_value is None:
        return []
    iris_nomalised = []
    for line in cell_value.split(","):
        if not line.strip():
            continue
        iri = line.split()[0].strip()
        iris_nomalised.append(prefix_converter.expand(iri) or iri)
    return iris_nomalised


def clean(cell_value: str | datetime | None) -> Any:
    """
    Clean a string cell by removing leading and trailing whitespace
    """
    if isinstance(cell_value, str):
        return cell_value.strip()
    return cell_value


def extract_concepts_and_collections(
    q: Worksheet,
    r: Worksheet,
    s: Worksheet,
    vocab_name: str = "",
) -> tuple[list[models.Concept], list[models.Collection]]:
    concepts = []
    collections = []
    concept_data = {}

    prefix_converter = config.CURIES_CONVERTER_MAP.get(vocab_name, Converter({}))

    # Iterating over the concept page
    for col in q.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if cell.value is None or cell.value in header_names_to_skip:
                continue
            uri = clean(q[f"A{row}"].value.split()[0])
            uri = prefix_converter.expand(uri) or uri
            concept_data[uri] = {
                "uri": uri,
                "curie": config.curies_converter.compress(uri),
                "pref_label": clean(q[f"B{row}"].value),
                "pl_language_code": split_and_tidy(q[f"C{row}"].value),
                "definition": clean(q[f"D{row}"].value),
                "def_language_code": split_and_tidy(q[f"E{row}"].value),
                "alt_labels": split_and_tidy(q[f"F{row}"].value),
                "children": split_multi_iri(q[f"G{row}"].value, prefix_converter),
                "provenance": clean(q[f"H{row}"].value),
                # Note in the new template, source_vocab is synonymous with source vocab uri
                "source_vocab": clean(q[f"I{row}"].value),
            }

    # Iterating over the additional concept page
    additional_concept_iris = []
    for col in r.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if cell.value is None or cell.value in header_names_to_skip:
                continue
            uri = clean(r[f"A{row}"].value.split()[0])
            uri = prefix_converter.expand(uri) or uri
            if uri in additional_concept_iris:
                msg = f"Concept IRI {uri} used a second time in sheet {r} at row {row} but must be unique."
                raise ConversionError(msg)
            if uri not in concept_data:
                msg = f'Concept "{uri}" from sheet {r} at row {row} not present in {q} sheet.'
                raise ConversionError(msg)
            data = {
                # additional concept features sheets
                "related_match": split_multi_iri(r[f"B{row}"].value, prefix_converter),
                "close_match": split_multi_iri(r[f"C{row}"].value, prefix_converter),
                "exact_match": split_multi_iri(r[f"D{row}"].value, prefix_converter),
                "narrow_match": split_multi_iri(r[f"E{row}"].value, prefix_converter),
                "broad_match": split_multi_iri(r[f"F{row}"].value, prefix_converter),
                "vocab_name": vocab_name,
            }
            concept_data[uri].update(**data)
            additional_concept_iris.append(uri)

    # validate all concept data (unfortunately, we can't point to sheet and row of error)
    for uri, concept_data_item in concept_data.items():
        try:
            c = models.Concept(**concept_data_item)
        except ValidationError as exc:
            msg = f"Concept processing error for Concept IRI* = {uri}: {exc}"
            raise ConversionError(msg) from exc
        concepts.append(c)

    # iterating over the collections page
    for col in s.iter_cols(max_col=1):
        for cell in col:
            row = cell.row
            if cell.value is None or cell.value in [
                "Collections",
                "Collection URI",
                "Collection IRI",
            ]:
                continue

            data_collection = {
                "uri": clean(s[f"A{row}"].value.split()[0]),
                "pref_label": clean(s[f"B{row}"].value),
                "definition": clean(s[f"C{row}"].value),
                "members": split_multi_iri(s[f"D{row}"].value, prefix_converter),
                "provenance": clean(s[f"E{row}"].value),
                "vocab_name": vocab_name,
            }

            try:
                c = models.Collection(**data_collection)
                collections.append(c)
            except ValidationError as exc:
                msg = f"Collection processing error, likely at sheet {s}, row {row}, and has error: {exc}"
                raise ConversionError(msg) from exc

    return concepts, collections


def extract_concept_scheme(sheet: Worksheet, vocab_name: str = ""):
    prefix_converter = config.CURIES_CONVERTER_MAP.get(vocab_name, Converter({}))
    uri = sheet["B2"].value
    cs = models.ConceptScheme(
        uri=(prefix_converter.expand(uri) or uri) if uri else None,
        title=clean(sheet["B3"].value),
        description=clean(sheet["B4"].value),
        created=clean(sheet["B5"].value),
        modified=(
            clean(sheet["B6"].value)
            if sheet["B6"].value is not None
            else clean(sheet["B5"].value)
        ),
        creator=clean(sheet["B7"].value),
        publisher=clean(sheet["B8"].value),
        version=clean(sheet["B9"].value) if sheet["B9"].value is not None else "",
        provenance=clean(sheet["B10"].value),
        custodian=clean(sheet["B11"].value),
        pid=clean(sheet["B12"].value),
        vocab_name=vocab_name,
    )
    return cs  # noqa: RET504
