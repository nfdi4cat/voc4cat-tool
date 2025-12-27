import glob
import logging
import os
from pathlib import Path

from openpyxl import load_workbook

from voc4cat.checks import Voc4catError
from voc4cat.models_v1 import (
    COLLECTIONS_SHEET_NAME,
    CONCEPT_SCHEME_SHEET_NAME,
    CONCEPTS_SHEET_NAME,
    ID_RANGES_SHEET_NAME,
    MAPPINGS_SHEET_NAME,
    PREFIXES_SHEET_NAME,
    RESERVED_SHEET_NAMES,
)

logger = logging.getLogger(__name__)

EXCEL_FILE_ENDINGS = [".xlsx"]
RDF_FILE_ENDINGS = {
    ".ttl": "ttl",
    ".rdf": "xml",
    ".xml": "xml",
    ".json-ld": "json-ld",
    ".json": "json-ld",
    ".nt": "nt",
    ".n3": "n3",
}
KNOWN_FILE_ENDINGS = [str(x) for x in RDF_FILE_ENDINGS] + EXCEL_FILE_ENDINGS


class ConversionError(Exception):
    pass


def split_and_tidy(cell_value: str):
    # note this may not work in list of things that contain commas. Need to consider revising
    # to allow comma-separated values where it'll split in commas but not in things enclosed in quotes.
    if cell_value == "" or cell_value is None:
        return []
    entries = [x.strip() for x in cell_value.strip().split(",")]
    return [x for x in entries if x]


def has_file_in_multiple_formats(dir_):
    files = [
        os.path.normcase(f)
        for f in glob.glob(os.path.join(dir_, "*.*"))
        if f.endswith(tuple(KNOWN_FILE_ENDINGS))
    ]
    file_names = [os.path.splitext(f)[0] for f in files]
    unique_file_names = set(file_names)
    if len(file_names) == len(unique_file_names):
        return False
    seen = set()
    return [x for x in file_names if x in seen or seen.add(x)]


# =============================================================================
# Shared Template Utilities
# =============================================================================


def validate_template_sheets(template_path: Path) -> None:
    """Validate that template doesn't contain reserved sheet names.

    Args:
        template_path: Path to the template xlsx file.

    Raises:
        Voc4catError: If template contains any reserved sheet names.
    """
    wb = load_workbook(template_path, read_only=True)
    template_sheets = set(wb.sheetnames)
    wb.close()

    conflicting = template_sheets & RESERVED_SHEET_NAMES
    if conflicting:
        msg = (
            f"Template contains reserved sheet names that will be auto-created: "
            f'"{", ".join(sorted(conflicting))}". '
            f"Please remove or rename these sheets in the template."
        )
        logger.error(msg)
        raise Voc4catError(msg)


def get_template_sheet_names(template_path: Path) -> list[str]:
    """Get the list of sheet names from a template file.

    Args:
        template_path: Path to the template xlsx file.

    Returns:
        List of sheet names in the template, in their original order.
    """
    wb = load_workbook(template_path, read_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()
    return sheet_names


def reorder_sheets_with_template(
    wb, template_sheet_names: list[str] | None = None
) -> None:
    """Reorder sheets: template sheets first (if any), then auto-created sheets.

    Args:
        wb: Workbook to reorder.
        template_sheet_names: List of sheet names from template (in order).
                             If provided, these sheets are placed first,
                             followed by auto-created sheets.
    """
    auto_created_order = [
        CONCEPT_SCHEME_SHEET_NAME,
        CONCEPTS_SHEET_NAME,
        COLLECTIONS_SHEET_NAME,
        MAPPINGS_SHEET_NAME,
        ID_RANGES_SHEET_NAME,
        PREFIXES_SHEET_NAME,
    ]

    current_sheets = wb.sheetnames

    if template_sheet_names is not None:
        # Template sheets first (in original order), then auto-created
        new_order = []
        # First: template sheets in their original order
        for sheet_name in template_sheet_names:
            if sheet_name in current_sheets:
                new_order.append(sheet_name)
        # Second: auto-created sheets in expected order
        for sheet_name in auto_created_order:
            if sheet_name in current_sheets and sheet_name not in new_order:
                new_order.append(sheet_name)
        # Third: any remaining sheets (shouldn't happen, but defensive)
        for sheet_name in current_sheets:
            if sheet_name not in new_order:
                new_order.append(sheet_name)
    else:
        # No template: auto-created first, then others
        new_order = []
        for sheet_name in auto_created_order:
            if sheet_name in current_sheets:
                new_order.append(sheet_name)
        for sheet_name in current_sheets:
            if sheet_name not in new_order:
                new_order.append(sheet_name)

    for idx, sheet_name in enumerate(new_order):
        wb.move_sheet(sheet_name, offset=idx - wb.sheetnames.index(sheet_name))
