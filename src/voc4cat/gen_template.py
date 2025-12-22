"""Template generator for voc4cat v1.0 vocabulary templates.

This module generates Excel templates for v1.0 vocabularies using the
xlsx infrastructure (xlsx_api, xlsx_table, xlsx_keyvalue).
"""

from __future__ import annotations

import logging
import shutil
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook

from voc4cat.checks import Voc4catError
from voc4cat.config import load_config
from voc4cat.convert_v1_helpers import extract_creator_names, generate_history_note
from voc4cat.models_v1 import (
    COLLECTIONS_EXPORT_CONFIG,
    COLLECTIONS_SHEET_NAME,
    CONCEPT_SCHEME_SHEET_NAME,
    CONCEPT_SCHEME_SHEET_TITLE,
    CONCEPTS_EXPORT_CONFIG,
    CONCEPTS_SHEET_NAME,
    DEFAULT_PREFIXES,
    EXAMPLE_COLLECTIONS,
    EXAMPLE_CONCEPT_SCHEME,
    EXAMPLE_CONCEPTS,
    EXAMPLE_MAPPINGS,
    ID_RANGES_EXPORT_CONFIG,
    ID_RANGES_SHEET_NAME,
    MAPPINGS_EXPORT_CONFIG,
    MAPPINGS_SHEET_NAME,
    PREFIXES_EXPORT_CONFIG,
    PREFIXES_SHEET_NAME,
    ConceptSchemeV1,
    ConceptV1,
    IDRangeInfoV1,
    PrefixV1,
)
from voc4cat.utils import (
    EXCEL_FILE_ENDINGS,
    get_template_sheet_names,
    reorder_sheets_with_template,
    validate_template_sheets,
)
from voc4cat.xlsx_api import export_to_xlsx
from voc4cat.xlsx_common import XLSXFieldAnalyzer, XLSXRowCalculator
from voc4cat.xlsx_keyvalue import XLSXKeyValueConfig

if TYPE_CHECKING:
    from voc4cat.config import Vocab

logger = logging.getLogger(__name__)


def generate_template_v1(
    output_path: Path | None = None,
    vocab_config: Vocab | None = None,
    template_path: Path | None = None,
) -> Workbook:
    """Generate a complete v1.0 vocabulary template workbook.

    Creates an Excel workbook with the following sheets:
    - Concept Scheme (key-value format)
    - Concepts (table format with example data)
    - Collections (table format with example data)
    - Mappings (table format with example data)
    - Prefixes (table format with default prefixes)

    Args:
        output_path: Optional path to save the workbook. If provided, the
                    workbook is saved to this path.
        vocab_config: Optional vocabulary configuration from idranges.toml.
                     If provided, the template will be populated with
                     vocabulary-specific metadata, prefixes, and ID ranges.
        template_path: Optional path to an xlsx template file. If provided,
                      the template's sheets are preserved and placed before
                      the auto-generated vocabulary sheets.

    Returns:
        The generated Workbook object.
    """

    # We need to build the workbook sheet by sheet
    # The xlsx_api functions work with files, so we use a temp approach

    if output_path is None:
        # Create a temporary file path for building
        temp_dir = Path(tempfile.mkdtemp())
        temp_path = temp_dir / "template_v1.0.xlsx"
    else:
        temp_path = output_path

    # If template provided, copy it to output location first
    if template_path is not None:
        shutil.copy(template_path, temp_path)
        logger.debug("Copied template from %s to %s", template_path, temp_path)

    # Create the workbook by exporting each sheet
    # 1. Concept Scheme (key-value format)
    _export_concept_scheme(temp_path, vocab_config)

    # 2. Concepts (table format)
    _export_concepts(temp_path)

    # 3. Collections (table format)
    _export_collections(temp_path)

    # 4. Mappings (table format)
    _export_mappings(temp_path)

    # 5. ID Ranges (table format, read-only)
    _export_id_ranges(temp_path, vocab_config)

    # 6. Prefixes (table format)
    _export_prefixes(temp_path, vocab_config)

    # Load the workbook to return and possibly adjust
    wb = load_workbook(temp_path)

    # Get template sheet names for ordering (if template was used)
    template_sheet_names = None
    if template_path is not None:
        template_sheet_names = get_template_sheet_names(template_path)

    # Reorder sheets to match expected order
    reorder_sheets_with_template(wb, template_sheet_names)

    # Set freeze panes for Concepts sheet (dynamically calculated)
    if CONCEPTS_SHEET_NAME in wb.sheetnames:
        field_analyses = XLSXFieldAnalyzer.analyze_model(ConceptV1)
        fields = list(field_analyses.values())
        row_calculator = XLSXRowCalculator(CONCEPTS_EXPORT_CONFIG)
        data_start_row = row_calculator.get_data_start_row(fields)
        wb[CONCEPTS_SHEET_NAME].freeze_panes = f"A{data_start_row}"

    # Save if output_path was provided
    if output_path:
        wb.save(output_path)

    return wb


def _export_concept_scheme(filepath: Path, vocab_config: Vocab | None = None) -> None:
    """Export Concept Scheme sheet in key-value format (read-only).

    Args:
        filepath: Path to the Excel file.
        vocab_config: Optional vocabulary configuration. If provided, uses
                     config values for scheme metadata instead of example data.
    """
    if vocab_config is not None:
        # Determine history_note: config value or auto-generate
        if vocab_config.history_note and vocab_config.history_note.strip():
            history_note = vocab_config.history_note.strip()
        else:
            # Auto-generate from created_date and creator names
            creator_names = extract_creator_names(vocab_config.creator or "")
            history_note = generate_history_note(
                vocab_config.created_date or "", creator_names
            )

        # Use vocabulary-specific values from config
        concept_scheme = ConceptSchemeV1(
            vocabulary_iri=vocab_config.vocabulary_iri,
            prefix=vocab_config.prefix,
            title=vocab_config.title,
            description=vocab_config.description,
            created_date=vocab_config.created_date,
            creator=vocab_config.creator,
            publisher=vocab_config.publisher,
            custodian=vocab_config.custodian,
            history_note=history_note,
            catalogue_pid=vocab_config.catalogue_pid,
            documentation=vocab_config.documentation,
            issue_tracker=vocab_config.issue_tracker,
            helpdesk=vocab_config.helpdesk,
            repository=vocab_config.repository,
            homepage=vocab_config.homepage,
            conforms_to=vocab_config.conforms_to,
        )
    else:
        concept_scheme = EXAMPLE_CONCEPT_SCHEME

    config = XLSXKeyValueConfig(
        title=CONCEPT_SCHEME_SHEET_TITLE,
        table_style="TableStyleMedium16",
    )
    export_to_xlsx(
        concept_scheme,
        filepath,
        format_type="keyvalue",
        config=config,
        sheet_name=CONCEPT_SCHEME_SHEET_NAME,
    )


def _export_concepts(filepath: Path) -> None:
    """Export Concepts sheet in table format."""
    export_to_xlsx(
        EXAMPLE_CONCEPTS,
        filepath,
        format_type="table",
        config=CONCEPTS_EXPORT_CONFIG,
        sheet_name=CONCEPTS_SHEET_NAME,
    )


def _export_collections(filepath: Path) -> None:
    """Export Collections sheet in table format."""
    export_to_xlsx(
        EXAMPLE_COLLECTIONS,
        filepath,
        format_type="table",
        config=COLLECTIONS_EXPORT_CONFIG,
        sheet_name=COLLECTIONS_SHEET_NAME,
    )


def _export_mappings(filepath: Path) -> None:
    """Export Mappings sheet in table format."""
    export_to_xlsx(
        EXAMPLE_MAPPINGS,
        filepath,
        format_type="table",
        config=MAPPINGS_EXPORT_CONFIG,
        sheet_name=MAPPINGS_SHEET_NAME,
    )


def _export_id_ranges(filepath: Path, vocab_config: Vocab | None = None) -> None:
    """Export ID Ranges sheet in table format (read-only).

    Args:
        filepath: Path to the Excel file.
        vocab_config: Optional vocabulary configuration. If provided, populates
                     the ID ranges from the config's id_range entries.
    """
    if vocab_config is not None and vocab_config.id_range:
        # Build ID range entries from config
        id_ranges = []
        for idr in vocab_config.id_range:
            # Format the range as "first_id - last_id"
            range_str = f"{idr.first_id} - {idr.last_id}"
            # Use gh_name if available, otherwise use ORCID
            name = idr.gh_name if idr.gh_name else str(idr.orcid or "")
            id_ranges.append(
                IDRangeInfoV1(
                    gh_name=name,
                    id_range=range_str,
                    unused_ids="",  # Will be computed during conversion
                )
            )
    else:
        # Export with empty row to get headers only
        id_ranges = [IDRangeInfoV1()]

    export_to_xlsx(
        id_ranges,
        filepath,
        format_type="table",
        config=ID_RANGES_EXPORT_CONFIG,
        sheet_name=ID_RANGES_SHEET_NAME,
    )


def _export_prefixes(filepath: Path, vocab_config: Vocab | None = None) -> None:
    """Export Prefixes sheet in table format (read-only).

    Args:
        filepath: Path to the Excel file.
        vocab_config: Optional vocabulary configuration. If provided, includes
                     prefix_map entries from the config in addition to defaults.
    """
    # Start with default prefixes
    prefixes = list(DEFAULT_PREFIXES)

    if vocab_config is not None and vocab_config.prefix_map:
        # Add prefixes from config's prefix_map
        existing_prefixes = {p.prefix for p in prefixes}
        for prefix, namespace in vocab_config.prefix_map.items():
            if prefix not in existing_prefixes:
                prefixes.append(PrefixV1(prefix=prefix, namespace=str(namespace)))

    export_to_xlsx(
        prefixes,
        filepath,
        format_type="table",
        config=PREFIXES_EXPORT_CONFIG,
        sheet_name=PREFIXES_SHEET_NAME,
    )


def template_cmd(args) -> None:
    """CLI command handler for template generation.

    Args:
        args: Parsed command-line arguments containing:
            - outdir: Output directory for the template (from common options)
            - template_version: Template version to generate
            - template: Optional path to base template xlsx file
            - VOCAB: Vocabulary name for the output filename
            - config: Optional path to config file (idranges.toml)
    """
    version = getattr(args, "template_version", "v1.0")
    outdir = getattr(args, "outdir", None)
    template_path = getattr(args, "template", None)
    vocab = args.VOCAB
    config_path = getattr(args, "config", None)

    if version != "v1.0":
        msg = f"Unsupported template version: {version}"
        raise ValueError(msg)

    # Validate template if provided
    if template_path is not None:
        if not template_path.exists():
            msg = f"Template file not found: {template_path}"
            logger.error(msg)
            raise Voc4catError(msg)
        if template_path.suffix.lower() not in EXCEL_FILE_ENDINGS:
            msg = 'Template file must be of type ".xlsx".'
            logger.error(msg)
            raise Voc4catError(msg)
        # Validate template doesn't contain reserved sheet names
        validate_template_sheets(template_path)

    # Determine output path
    if outdir is None:
        outdir = Path.cwd()
    else:
        outdir = Path(outdir)
        outdir.mkdir(parents=True, exist_ok=True)

    # Use vocab name for filename, strip any extension if provided
    vocab_name = Path(vocab).stem if "." in vocab else vocab
    output_path = outdir / f"{vocab_name}.xlsx"

    # Check if file already exists
    if output_path.exists():
        logger.error("File already exists: %s", output_path)
        return

    # Load config if provided and extract vocabulary-specific settings
    vocab_config = None
    if config_path is not None:
        config_file = Path(config_path)
        if config_file.exists():
            load_config(config_file)
            # Import after load_config to get updated globals
            from voc4cat.config import IDRANGES  # noqa: PLC0415

            # Look up vocabulary in config (case-insensitive)
            vocab_key = vocab_name.lower()
            if vocab_key in IDRANGES.vocabs:
                vocab_config = IDRANGES.vocabs[vocab_key]
                logger.info(
                    "Using config for vocabulary '%s' from %s", vocab_name, config_path
                )
            else:
                logger.warning(
                    "Vocabulary '%s' not found in config file %s. "
                    "Generating template with example data.",
                    vocab_name,
                    config_path,
                )
        else:
            logger.warning(
                "Config file not found: %s. Generating template with example data.",
                config_path,
            )

    logger.info("Generating %s template at: %s", version, output_path)

    generate_template_v1(output_path, vocab_config, template_path=template_path)

    logger.info("Template generated successfully: %s", output_path)
    print(f"Generated template: {output_path}")
