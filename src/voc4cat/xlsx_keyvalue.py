"""
Key-value format implementation for single records with field/value pairs.

This module contains all key-value-specific functionality including:
- Key-value configuration with column customization
- Key-value formatter for field/value layout
- Key-value processor for import/export operations
"""

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError

from .xlsx_common import (
    MAX_SHEETNAME_LENGTH,
    FieldAnalysis,
    XLSXConfig,
    XLSXDeserializationError,
    XLSXFieldAnalyzer,
    XLSXFormatter,
    XLSXProcessor,
    XLSXSerializationError,
)


# Key-value specific configuration
@dataclass
class XLSXKeyValueConfig(XLSXConfig):
    """Configuration for key-value format (single record)."""

    field_column_header: str = "Field"
    value_column_header: str = "Value"
    unit_column_header: str = "Unit"
    requiredness_column_header: str = "Required"
    description_column_header: str = "Description"
    meaning_column_header: str = "Meaning"
    field_column_width: int = 25
    value_column_width: int = 30
    unit_column_width: int = 10
    requiredness_column_width: int = 20
    description_column_width: int = 50
    meaning_column_width: int = 40
    table_style: str = "TableStyleMedium9"


# Key-value formatter
class XLSXKeyValueFormatter(XLSXFormatter):
    """Handles key-value format with field names and values."""

    def format_export(
        self, worksheet: Worksheet, data: BaseModel, fields: list[FieldAnalysis]
    ) -> None:
        """Format data as key-value pairs."""
        config = self.config
        if not isinstance(config, XLSXKeyValueConfig):
            msg = "XLSXKeyValueFormatter requires XLSXKeyValueConfig"
            raise TypeError(msg)

        # Add title at row 1 (if present)
        title = config.title
        if title == "Data Export":  # Default title
            title = f"{data.__class__.__name__} Instance"

        if title:
            self._add_kv_title(worksheet, title, config)

        # Add headers starting at row 3 (leaving row 2 empty)
        self._add_kv_headers(worksheet, config, fields)

        # Write field data
        self._write_field_data(worksheet, data, fields, config)

        # Create and add Excel table
        table = self._create_kv_table(worksheet, len(fields), config, fields)
        worksheet.add_table(table)

        # Add validation (always enabled)
        self._add_data_validation(worksheet, fields, config)

        # Auto-adjust columns - dynamic based on which optional columns are shown
        col_layout = self._get_column_layout(fields)
        self._auto_adjust_columns(worksheet, len(col_layout))

    def _get_column_layout(self, fields: list[FieldAnalysis]) -> dict[str, str]:
        """Get column layout based on which optional columns are present.

        Order: Field, Value, Unit, Requiredness, Description, Meaning
        """
        # Use row_calculator's visibility methods which respect config toggles
        show_units = self.row_calculator._should_show_units(fields)
        show_requiredness = self.row_calculator._should_show_requiredness(fields)
        show_descriptions = self.row_calculator._should_show_descriptions(fields)
        show_meanings = self.row_calculator._should_show_meanings(fields)

        # Build column mapping dynamically
        # Required: Field (A), Value (B)
        # Optional: Unit, Requiredness, Description, Meaning
        columns = {"field": "A", "value": "B"}
        next_col = "C"

        if show_units:
            columns["unit"] = next_col
            next_col = chr(ord(next_col) + 1)

        if show_requiredness:
            columns["requiredness"] = next_col
            next_col = chr(ord(next_col) + 1)

        if show_descriptions:
            columns["description"] = next_col
            next_col = chr(ord(next_col) + 1)

        if show_meanings:
            columns["meaning"] = next_col

        return columns

    def _add_kv_title(
        self, worksheet: Worksheet, title: str, config: XLSXKeyValueConfig
    ) -> None:
        """Add title to worksheet at configured start row."""
        if title:
            title_row = self.row_calculator.get_title_row()
            cell = worksheet.cell(row=title_row, column=1)
            cell.value = title
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="left")

    def _add_kv_headers(
        self,
        worksheet: Worksheet,
        config: XLSXKeyValueConfig,
        fields: list[FieldAnalysis],
    ) -> None:
        """Add field/value headers with optional unit/description/meaning columns."""
        header_row = self.row_calculator.get_first_content_row()
        col_layout = self._get_column_layout(fields)

        # Required columns
        worksheet[f"{col_layout['field']}{header_row}"] = config.field_column_header
        worksheet[f"{col_layout['value']}{header_row}"] = config.value_column_header

        # Optional columns - only write if present in layout
        if "unit" in col_layout:
            worksheet[f"{col_layout['unit']}{header_row}"] = config.unit_column_header
        if "requiredness" in col_layout:
            worksheet[f"{col_layout['requiredness']}{header_row}"] = (
                config.requiredness_column_header
            )
        if "description" in col_layout:
            worksheet[f"{col_layout['description']}{header_row}"] = (
                config.description_column_header
            )
        if "meaning" in col_layout:
            worksheet[f"{col_layout['meaning']}{header_row}"] = (
                config.meaning_column_header
            )

    def _write_field_data(
        self,
        worksheet: Worksheet,
        data: BaseModel,
        fields: list[FieldAnalysis],
        config: XLSXKeyValueConfig,
    ) -> None:
        """Write field data in key-value format."""
        row = self.row_calculator.get_first_content_row() + 1
        col_layout = self._get_column_layout(fields)

        for field_analysis in fields:
            value = getattr(data, field_analysis.name, None)

            # Write field name
            display_name = self._get_field_display_name(field_analysis)
            field_name_cell = worksheet[f"{col_layout['field']}{row}"]
            field_name_cell.value = display_name
            # Apply formatting to field name cell (treat as text field)
            text_field_analysis = FieldAnalysis(name="field_name", field_type=str)
            self._apply_data_cell_formatting(field_name_cell, text_field_analysis)

            # Write field value
            try:
                formatted_value = self.serialization_engine.serialize_value(
                    value, field_analysis
                )
                value_cell = worksheet[f"{col_layout['value']}{row}"]
                value_cell.value = formatted_value
                # Apply special formatting for Value column (always left-aligned)
                # Skip formatting if disabled in configuration
                if not (
                    hasattr(self.config, "enable_cell_formatting")
                    and not self.config.enable_cell_formatting
                ):
                    # Value column gets left alignment regardless of data type
                    wrap_text = self._should_wrap_text(field_analysis.field_type)

                    value_cell.alignment = Alignment(
                        horizontal="left",  # Always left-align Value column
                        vertical="center",
                        wrap_text=wrap_text,
                    )
            except Exception as e:
                raise XLSXSerializationError(field_analysis.name, value, e) from e

            # Write optional columns - only if present in layout

            # Unit column (optional)
            if "unit" in col_layout:
                unit_value = (
                    field_analysis.xlsx_metadata.unit
                    if field_analysis.xlsx_metadata
                    and field_analysis.xlsx_metadata.unit
                    else ""
                )
                unit_cell = worksheet[f"{col_layout['unit']}{row}"]
                unit_cell.value = unit_value
                # Apply formatting to unit cell (treat as text field)
                text_field_analysis = FieldAnalysis(name="unit", field_type=str)
                self._apply_data_cell_formatting(unit_cell, text_field_analysis)

            # Requiredness column (optional)
            if "requiredness" in col_layout:
                field_info = data.__class__.model_fields.get(field_analysis.name)
                if field_info:
                    is_required, default_value = (
                        XLSXFieldAnalyzer.get_requiredness_info(
                            field_info, field_analysis
                        )
                    )
                    req_text = XLSXFieldAnalyzer.format_requiredness_text(
                        is_required, default_value
                    )
                else:
                    req_text = ""
                req_cell = worksheet[f"{col_layout['requiredness']}{row}"]
                req_cell.value = req_text
                # Apply formatting to requiredness cell (treat as text field)
                text_field_analysis = FieldAnalysis(name="requiredness", field_type=str)
                self._apply_data_cell_formatting(req_cell, text_field_analysis)

            # Description column (optional)
            if "description" in col_layout:
                description = (
                    field_analysis.xlsx_metadata.description
                    if field_analysis.xlsx_metadata
                    and field_analysis.xlsx_metadata.description
                    else ""
                )
                description_cell = worksheet[f"{col_layout['description']}{row}"]
                description_cell.value = description
                # Apply formatting to description cell (treat as text field)
                text_field_analysis = FieldAnalysis(name="description", field_type=str)
                self._apply_data_cell_formatting(description_cell, text_field_analysis)

            # Meaning column (optional)
            if "meaning" in col_layout:
                meaning_value = (
                    field_analysis.xlsx_metadata.meaning
                    if field_analysis.xlsx_metadata
                    and field_analysis.xlsx_metadata.meaning
                    else ""
                )
                meaning_cell = worksheet[f"{col_layout['meaning']}{row}"]
                meaning_cell.value = meaning_value
                # Apply formatting to meaning cell (treat as text field)
                text_field_analysis = FieldAnalysis(name="meaning", field_type=str)
                self._apply_data_cell_formatting(meaning_cell, text_field_analysis)

            row += 1

    def _create_kv_table(
        self,
        worksheet: Worksheet,
        num_fields: int,
        config: XLSXKeyValueConfig,
        fields: list[FieldAnalysis],
    ) -> Table:
        """Create Excel table for key-value data."""
        # Calculate number of columns from layout
        col_layout = self._get_column_layout(fields)
        num_columns = len(col_layout)

        # Calculate table range - headers start at dynamic position
        start_row = self.row_calculator.get_first_content_row()
        end_row = start_row + num_fields  # header row + data rows
        end_col_letter = get_column_letter(num_columns)

        table_range = f"A{start_row}:{end_col_letter}{end_row}"

        # Create table name (clean and unique)
        table_name = f"KeyValue_{worksheet.title.replace(' ', '_')}"
        table_name = re.sub(r"[^a-zA-Z0-9_]", "_", table_name)
        if len(table_name) > MAX_SHEETNAME_LENGTH:
            table_name = table_name[:MAX_SHEETNAME_LENGTH]

        # Create table object
        table = Table(displayName=table_name, ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name=config.table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        return table

    def _add_data_validation(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        config: XLSXKeyValueConfig,
    ) -> None:
        """Add data validation for enum fields in key-value format.

        For enum lists that exceed Excel's 255 character limit for inline
        formulas, creates a hidden sheet with the values and uses a named range.
        """
        # Calculate data start row - where the values column starts
        data_start_row = self.row_calculator.get_first_content_row() + 1

        # Go through each field and add validation if it's an enum
        for i, field_analysis in enumerate(fields):
            if field_analysis.enum_values:
                # Calculate the row for this specific field
                field_row = data_start_row + i
                # Value column is always column B
                validation_range = f"B{field_row}"
                self._add_enum_validation(worksheet, field_analysis, validation_range)

    def parse_import(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        model_class: type[BaseModel],
    ) -> BaseModel:
        """Parse key-value data from worksheet."""
        config = self.config
        if not isinstance(config, XLSXKeyValueConfig):
            msg = "XLSXKeyValueFormatter requires XLSXKeyValueConfig"
            raise TypeError(msg)

        field_dict = {field.name: field for field in fields}
        field_data = self._read_field_data(worksheet, field_dict, model_class, config)

        try:
            return model_class(**field_data)
        except ValidationError as e:
            msg = f"Failed to create {model_class.__name__} instance: {e}"
            raise ValueError(msg) from e

    def _detect_column_layout_from_headers(
        self,
        worksheet: Worksheet,
        config: XLSXKeyValueConfig,
    ) -> dict[str, str]:
        """Detect column layout by reading the header row."""
        header_row = self.row_calculator.get_first_content_row()
        col_layout = {}

        # Scan columns A through H to find headers
        for col_idx in range(1, 9):  # A=1, B=2, ..., H=8
            col_letter = get_column_letter(col_idx)
            header_value = worksheet[f"{col_letter}{header_row}"].value
            if header_value:
                header_text = str(header_value).strip()
                if header_text == config.field_column_header:
                    col_layout["field"] = col_letter
                elif header_text == config.value_column_header:
                    col_layout["value"] = col_letter
                elif header_text == config.unit_column_header:
                    col_layout["unit"] = col_letter
                elif header_text == config.requiredness_column_header:
                    col_layout["requiredness"] = col_letter
                elif header_text == config.description_column_header:
                    col_layout["description"] = col_letter
                elif header_text == config.meaning_column_header:
                    col_layout["meaning"] = col_letter

        return col_layout

    def _read_field_data(
        self,
        worksheet: Worksheet,
        field_dict: dict[str, FieldAnalysis],
        model_class: type[BaseModel],
        config: XLSXKeyValueConfig,
    ) -> dict[str, Any]:
        """Read field data from worksheet."""
        field_data = {}

        # Detect column layout from headers (supports both old and new formats)
        col_layout = self._detect_column_layout_from_headers(worksheet, config)

        # Get column letters (with fallback for required columns)
        field_col = col_layout.get("field", "A")
        value_col = col_layout.get("value", "B")
        unit_col = col_layout.get("unit")  # Optional
        meaning_col = col_layout.get("meaning")  # Optional

        start_row = self.row_calculator.get_first_content_row() + 1
        for row in range(start_row, worksheet.max_row + 1):
            field_name_cell = worksheet[f"{field_col}{row}"]
            value_cell = worksheet[f"{value_col}{row}"]

            # Read unit (if column exists)
            worksheet_unit = ""
            if unit_col:
                unit_cell = worksheet[f"{unit_col}{row}"]
                worksheet_unit = unit_cell.value
                worksheet_unit = worksheet_unit.strip() if worksheet_unit else ""

            # Read meaning (if column exists)
            worksheet_meaning = ""
            if meaning_col:
                meaning_cell = worksheet[f"{meaning_col}{row}"]
                worksheet_meaning = meaning_cell.value
                worksheet_meaning = (
                    worksheet_meaning.strip() if worksheet_meaning else ""
                )

            if not field_name_cell.value:
                continue

            # Convert display name back to field name
            display_name = str(field_name_cell.value)
            field_name = self._find_field_name_from_display(
                display_name, field_dict, config
            )

            if field_name and field_name in field_dict:
                field_analysis = field_dict[field_name]
                raw_value = value_cell.value

                # Validate unit consistency (if field has unit defined and column exists)
                if unit_col:
                    expected_unit = (
                        field_analysis.xlsx_metadata.unit
                        if field_analysis.xlsx_metadata
                        else None
                    )
                    if (
                        expected_unit
                        and worksheet_unit
                        and worksheet_unit != expected_unit
                    ):
                        msg = (
                            f"Unit mismatch for field '{field_name}': "
                            f"expected '{expected_unit}', found '{worksheet_unit}'"
                        )
                        raise ValueError(msg)

                # Validate meaning consistency (if field has meaning defined and column exists)
                if meaning_col:
                    expected_meaning = (
                        field_analysis.xlsx_metadata.meaning
                        if field_analysis.xlsx_metadata
                        else None
                    )
                    if (
                        expected_meaning
                        and worksheet_meaning
                        and worksheet_meaning != expected_meaning
                    ):
                        msg = (
                            f"Meaning mismatch for field '{field_name}': "
                            f"expected '{expected_meaning}', found '{worksheet_meaning}'"
                        )
                        raise ValueError(msg)

                try:
                    converted_value = self.serialization_engine.deserialize_value(
                        raw_value, field_analysis, model_class
                    )
                    # If value is None, only add it if the field accepts None (is optional)
                    # Otherwise, skip it so Pydantic uses the model's default value
                    if converted_value is not None:
                        field_data[field_name] = converted_value
                    elif field_analysis.is_optional:
                        # Field accepts None, so include it
                        field_data[field_name] = None
                    # else: skip None for non-optional fields with defaults
                except Exception as e:
                    raise XLSXDeserializationError(field_name, raw_value, e) from e

        return field_data

    def _find_field_name_from_display(
        self,
        display_name: str,
        field_dict: dict[str, FieldAnalysis],
        config: XLSXKeyValueConfig,
    ) -> str | None:
        """Find the actual field name from the display name."""
        # Check against field display names (either custom or auto-generated)
        for field_name, field_analysis in field_dict.items():
            expected_display = self._get_field_display_name(field_analysis)
            if expected_display == display_name:
                return field_name

        return None


# Key-value processor
class XLSXKeyValueProcessor(XLSXProcessor):
    """Processor for key-value format."""

    def export(
        self, data: BaseModel, filepath: Path, sheet_name: str | None = None
    ) -> None:
        """Export single model to XLSX file."""
        model_class = data.__class__
        sheet_name = sheet_name or model_class.__name__

        # Analyze fields
        field_analyses = self.field_analyzer.analyze_model(model_class)
        fields = list(field_analyses.values())
        filtered_fields = self._filter_and_order_fields(fields)

        workbook, worksheet = self._prepare_workbook(filepath, sheet_name)
        self.formatter.format_export(worksheet, data, filtered_fields)
        workbook.save(filepath)

    def import_data(
        self,
        filepath: Path,
        model_class: type[BaseModel],
        sheet_name: str | None = None,
    ) -> BaseModel:
        """Import single model from XLSX file."""
        sheet_name = sheet_name or model_class.__name__

        workbook = load_workbook(filepath, data_only=True)

        if sheet_name not in workbook.sheetnames:
            msg = f"Sheet '{sheet_name}' not found in workbook"
            raise ValueError(msg)

        worksheet = workbook[sheet_name]

        # Analyze fields
        field_analyses = self.field_analyzer.analyze_model(model_class)
        fields = list(field_analyses.values())
        filtered_fields = self._filter_and_order_fields(fields)

        # Parse import
        return self.formatter.parse_import(worksheet, filtered_fields, model_class)
