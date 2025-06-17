"""
Public API and factory for creating appropriate processors.

This module provides the main public API for XLSX processing, including:
- Factory for creating processors
- Public API functions for export/import
- Utility functions for model enhancement
- Auto-detection logic for format selection
"""

from pathlib import Path
from typing import Annotated, Any

from openpyxl import load_workbook
from pydantic import BaseModel

from .xlsx_common import XLSXConfig, XLSXMetadata
from .xlsx_keyvalue import (
    XLSXKeyValueConfig,
    XLSXKeyValueFormatter,
    XLSXKeyValueProcessor,
)
from .xlsx_table import (
    JoinConfiguration,
    XLSXJoinedTableFormatter,
    XLSXJoinedTableProcessor,
    XLSXTableConfig,
    XLSXTableFormatter,
    XLSXTableProcessor,
)


class XLSXProcessorFactory:
    """Factory to create appropriate processors."""

    @staticmethod
    def create_table_processor(
        config: XLSXTableConfig | None = None,
    ) -> XLSXTableProcessor:
        """Create processor for tabular format (multiple records)."""
        config = config or XLSXTableConfig()
        formatter = XLSXTableFormatter(config)
        return XLSXTableProcessor(config, formatter)

    @staticmethod
    def create_keyvalue_processor(
        config: XLSXKeyValueConfig | None = None,
    ) -> XLSXKeyValueProcessor:
        """Create processor for key-value format (single record)."""
        config = config or XLSXKeyValueConfig()
        formatter = XLSXKeyValueFormatter(config)
        return XLSXKeyValueProcessor(config, formatter)

    @staticmethod
    def create_joined_table_processor(
        join_config: JoinConfiguration,
        config: XLSXTableConfig | None = None,
    ) -> XLSXJoinedTableProcessor:
        """Create processor for joined models in tabular format.

        Args:
            join_config: Configuration for joining models
            config: Optional table configuration

        Returns:
            XLSXJoinedTableProcessor instance
        """
        config = config or XLSXTableConfig()

        # Set default title if not specified
        if config.title is None:
            primary_name = join_config.primary_model.__name__
            related_names = ", ".join(join_config.related_models.keys())
            config.title = f"{primary_name} with {related_names}"

        formatter = XLSXJoinedTableFormatter(config, join_config)
        return XLSXJoinedTableProcessor(config, formatter)


def create_xlsx_wrapper(
    original_model: type[BaseModel],
    metadata_map: dict[str, XLSXMetadata],
    base_wrapper: type[BaseModel] | None = None,
) -> type[BaseModel]:
    """Create wrapper model with XLSX metadata support and proper inheritance.

    Args:
        original_model: The original Pydantic model to enhance
        metadata_map: Dictionary mapping field names to XLSXMetadata
        base_wrapper: Optional base wrapper class to inherit from (for inheritance chains)

    Returns:
        Enhanced model class with XLSX metadata

    Example:
        ```python
        # Simple usage
        XLSXReaction = create_xlsx_wrapper(Reaction, {
            "temperature": XLSXMetadata(unit="°C", meaning="Reaction temperature"),
            "catalysts": XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
        })

        # With inheritance
        XLSXBaseReaction = create_xlsx_wrapper(BaseReaction, base_metadata)
        XLSXAdvancedReaction = create_xlsx_wrapper(
            AdvancedReaction,
            advanced_metadata,
            base_wrapper=XLSXBaseReaction
        )
        ```
    """
    # Build complete annotations including all fields and metadata
    all_annotations = {}

    # Get all field types from the original model (including inherited ones)
    for field_name, field_info in original_model.model_fields.items():
        field_type = field_info.annotation

        # Check if we have metadata for this field
        if field_name in metadata_map:
            all_annotations[field_name] = Annotated[
                field_type, metadata_map[field_name]
            ]
        # Check if base wrapper has metadata for this field
        elif (
            base_wrapper
            and hasattr(base_wrapper, "__annotations__")
            and field_name in base_wrapper.__annotations__
        ):
            # Use the annotated type from base wrapper (preserves metadata)
            all_annotations[field_name] = base_wrapper.__annotations__[field_name]
        else:
            # No metadata, use original type
            all_annotations[field_name] = field_type

    # Determine the base class for inheritance
    base_class = base_wrapper if base_wrapper else original_model

    # Create a completely new class with proper annotations
    # This is necessary because Pydantic doesn't support modifying annotations after class creation
    class_name = f"XLSX{original_model.__name__}"

    # Create the class dynamically with proper annotations
    return type(
        class_name,
        (base_class,),
        {
            "__annotations__": all_annotations,
            "__module__": original_model.__module__,
            "__qualname__": f"XLSX{original_model.__qualname__}",
        },
    )


def export_to_xlsx(
    data: Any,
    filepath: Path | str,
    format_type: str = "table",
    config: XLSXConfig | None = None,
    sheet_name: str | None = None,
) -> None:
    """Universal export function.

    Args:
        data: Data to export (single model or sequence of models)
        filepath: Path to save the Excel file
        format_type: Format type ("table", "keyvalue").
                    Defaults to "table" format.
        config: Optional configuration object
        sheet_name: Optional sheet name
    """
    if isinstance(filepath, str):
        filepath = Path(filepath)

    if format_type == "table":
        table_config = config if isinstance(config, XLSXTableConfig) else None
        table_processor = XLSXProcessorFactory.create_table_processor(table_config)
        table_processor.export(data, filepath, sheet_name)
    elif format_type == "keyvalue":
        kv_config = config if isinstance(config, XLSXKeyValueConfig) else None
        kv_processor = XLSXProcessorFactory.create_keyvalue_processor(kv_config)
        kv_processor.export(data, filepath, sheet_name)
    else:
        msg = f"Unsupported format type: {format_type}"
        raise ValueError(msg)


def import_from_xlsx(
    filepath: Path | str,
    model_class: type[BaseModel],
    format_type: str = "auto",
    config: XLSXConfig | None = None,
    sheet_name: str | None = None,
) -> Any:
    """Universal import function.

    Args:
        filepath: Path to the Excel file
        model_class: Pydantic model class to import into
        format_type: Format type ("auto", "table", "keyvalue")
        config: Optional configuration object
        sheet_name: Optional sheet name

    Returns:
        Single model instance or list of model instances
    """
    if isinstance(filepath, str):
        filepath = Path(filepath)

    if format_type == "auto":
        # Try to detect format from file structure by checking sheet names and layout
        workbook = load_workbook(filepath, data_only=True)
        detect_sheet = sheet_name or model_class.__name__

        if detect_sheet in workbook.sheetnames:
            worksheet = workbook[detect_sheet]
            # Simple heuristic: detect key-value format by looking for "Field" and "Value" headers
            try:
                # Check different possible row positions for headers
                for header_row in [1, 2, 3]:
                    field_header = worksheet[f"A{header_row}"].value
                    value_header = worksheet[f"B{header_row}"].value

                    # Check if this looks like a key-value format
                    if field_header == "Field" and value_header == "Value":
                        # Check remaining columns to see if they contain key-value metadata columns
                        remaining_headers = []
                        for col_letter in ["C", "D", "E", "F"]:
                            header_value = worksheet[f"{col_letter}{header_row}"].value
                            if header_value:
                                remaining_headers.append(header_value)

                        # Key-value format should have combinations of Unit, Meaning, Description
                        valid_kv_headers = {"Unit", "Meaning", "Description"}
                        if all(
                            header in valid_kv_headers for header in remaining_headers
                        ):
                            format_type = "keyvalue"
                            break
                        format_type = "table"
                        break
                else:
                    format_type = "table"
            except Exception:
                format_type = "table"  # Default fallback
        else:
            format_type = "table"  # Default fallback

    if format_type == "table":
        table_config = config if isinstance(config, XLSXTableConfig) else None
        table_processor = XLSXProcessorFactory.create_table_processor(table_config)
        return table_processor.import_data(filepath, model_class, sheet_name)
    if format_type == "keyvalue":
        kv_config = config if isinstance(config, XLSXKeyValueConfig) else None
        kv_processor = XLSXProcessorFactory.create_keyvalue_processor(kv_config)
        return kv_processor.import_data(filepath, model_class, sheet_name)
    msg = f"Unsupported format type: {format_type}"
    raise ValueError(msg)
