"""Template generator for voc4cat v1.0 vocabulary templates.

This module generates Excel templates for v1.0 vocabularies using the
xlsx infrastructure (xlsx_api, xlsx_table, xlsx_keyvalue).
"""

from pathlib import Path

from openpyxl import Workbook

from voc4cat.models_v1 import (
    DEFAULT_PREFIXES,
    EXAMPLE_COLLECTIONS,
    EXAMPLE_CONCEPT_SCHEME,
    EXAMPLE_CONCEPTS,
    EXAMPLE_MAPPINGS,
)
from voc4cat.xlsx_api import export_to_xlsx
from voc4cat.xlsx_keyvalue import XLSXKeyValueConfig
from voc4cat.xlsx_table import XLSXTableConfig


def generate_template_v1(output_path: Path | None = None) -> Workbook:
    """Generate a complete v1.0 vocabulary template workbook.

    Creates an Excel workbook with the following sheets:
    - Concept Scheme (key-value format)
    - Concepts (table format with example data)
    - Collections (table format with example data)
    - Mappings (table format with example data)
    - Prefixes (table format with default prefixes)

    Args:
        output_path: Optional path to save the workbook. If provided, the
                    workbook is saved to this path.

    Returns:
        The generated Workbook object.
    """
    from openpyxl import load_workbook

    # We need to build the workbook sheet by sheet
    # The xlsx_api functions work with files, so we use a temp approach

    if output_path is None:
        # Create a temporary file path for building
        import tempfile

        temp_dir = Path(tempfile.mkdtemp())
        temp_path = temp_dir / "template_v1.xlsx"
    else:
        temp_path = output_path

    # Create the workbook by exporting each sheet
    # 1. Concept Scheme (key-value format)
    _export_concept_scheme(temp_path)

    # 2. Concepts (table format)
    _export_concepts(temp_path)

    # 3. Collections (table format)
    _export_collections(temp_path)

    # 4. Mappings (table format)
    _export_mappings(temp_path)

    # 5. Prefixes (table format)
    _export_prefixes(temp_path)

    # Load the workbook to return and possibly adjust
    wb = load_workbook(temp_path)

    # Reorder sheets to match expected order
    _reorder_sheets(wb)

    # Set freeze panes for Concepts sheet
    if "Concepts" in wb.sheetnames:
        wb["Concepts"].freeze_panes = "A5"

    # Save if output_path was provided
    if output_path:
        wb.save(output_path)

    return wb


def _export_concept_scheme(filepath: Path) -> None:
    """Export Concept Scheme sheet in key-value format."""
    config = XLSXKeyValueConfig(
        title="Concept Scheme",
    )
    export_to_xlsx(
        EXAMPLE_CONCEPT_SCHEME,
        filepath,
        format_type="keyvalue",
        config=config,
        sheet_name="Concept Scheme",
    )


def _export_concepts(filepath: Path) -> None:
    """Export Concepts sheet in table format."""
    config = XLSXTableConfig(
        title="Concepts",
        freeze_panes=True,
    )
    export_to_xlsx(
        EXAMPLE_CONCEPTS,
        filepath,
        format_type="table",
        config=config,
        sheet_name="Concepts",
    )


def _export_collections(filepath: Path) -> None:
    """Export Collections sheet in table format."""
    config = XLSXTableConfig(
        title="Collections",
    )
    export_to_xlsx(
        EXAMPLE_COLLECTIONS,
        filepath,
        format_type="table",
        config=config,
        sheet_name="Collections",
    )


def _export_mappings(filepath: Path) -> None:
    """Export Mappings sheet in table format."""
    config = XLSXTableConfig(
        title="Mappings",
    )
    export_to_xlsx(
        EXAMPLE_MAPPINGS,
        filepath,
        format_type="table",
        config=config,
        sheet_name="Mappings",
    )


def _export_prefixes(filepath: Path) -> None:
    """Export Prefixes sheet in table format with default prefixes."""
    config = XLSXTableConfig(
        title="Prefix mappings",
    )
    export_to_xlsx(
        DEFAULT_PREFIXES,
        filepath,
        format_type="table",
        config=config,
        sheet_name="Prefixes",
    )


def _reorder_sheets(wb: Workbook) -> None:
    """Reorder sheets to match expected template order."""
    expected_order = [
        "Concept Scheme",
        "Concepts",
        "Collections",
        "Mappings",
        "Prefixes",
    ]

    # Get current sheet order
    current_sheets = wb.sheetnames

    # Build new order based on expected, then any remaining
    new_order = []
    for sheet_name in expected_order:
        if sheet_name in current_sheets:
            new_order.append(sheet_name)
    for sheet_name in current_sheets:
        if sheet_name not in new_order:
            new_order.append(sheet_name)

    # Reorder
    for idx, sheet_name in enumerate(new_order):
        wb.move_sheet(sheet_name, offset=idx - wb.sheetnames.index(sheet_name))


def template_cmd(args) -> None:
    """CLI command handler for template generation.

    Args:
        args: Parsed command-line arguments containing:
            - outdir: Output directory for the template (from common options)
            - template_version: Template version to generate (currently only "1.0")
    """
    import logging

    logger = logging.getLogger(__name__)

    version = getattr(args, "template_version", "1.0")
    outdir = getattr(args, "outdir", None)

    if version != "1.0":
        msg = f"Unsupported template version: {version}"
        raise ValueError(msg)

    # Determine output path
    if outdir is None:
        outdir = Path.cwd()
    else:
        outdir = Path(outdir)
        outdir.mkdir(parents=True, exist_ok=True)

    output_path = outdir / f"blank_{version.replace('.', '-')}.xlsx"

    logger.info("Generating v%s template at: %s", version, output_path)

    generate_template_v1(output_path)

    logger.info("Template generated successfully: %s", output_path)
    print(f"Generated template: {output_path}")
