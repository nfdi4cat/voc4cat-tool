"""
Table format implementation for multiple records with headers and rows.

This module contains all table-specific functionality including:
- Table configuration with validation support
- Table formatter for rows and columns layout
- Table processor for import/export operations
- Joined model support for complex relationships
"""

import re
from collections.abc import Sequence
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError
from pydantic_core import PydanticUndefined

from .xlsx_common import (
    MAX_SHEETNAME_LENGTH,
    FieldAnalysis,
    XLSXConfig,
    XLSXDeserializationError,
    XLSXFieldAnalyzer,
    XLSXFormatter,
    XLSXProcessor,
    XLSXSerializationError,
)


# Table-specific configuration
@dataclass
class XLSXTableConfig(XLSXConfig):
    """Configuration for table format (multiple records)."""

    auto_filter: bool = False
    table_style: str = "TableStyleMedium9"
    header_row_color: str | None = None
    freeze_panes: bool = True
    bold_fields: set[str] = field(default_factory=set)
    rows_pre_allocated: int = 0


# Join configuration for complex relationships
@dataclass
class JoinConfiguration:
    """Configuration for joining models in a full outer join."""

    primary_model: type[BaseModel]
    related_models: dict[str, type[BaseModel]]  # field_name -> model_class
    join_keys: dict[str, str]  # related_field -> primary_field
    flattened_fields: list[str]  # ordered list of output fields
    field_mappings: dict[
        str, tuple[str, str]
    ]  # output_field -> (model_name, field_name)
    list_fields: set[str] = field(
        default_factory=set
    )  # fields that should be comma-separated


# Joined model processor
class JoinedModelProcessor:
    """Handles processing of joined models for XLSX export/import."""

    @staticmethod
    def _get_join_key_fields(join_config: JoinConfiguration) -> set[str]:
        """Get all fields that are used as join keys."""
        return set(join_config.join_keys.values())

    @staticmethod
    def flatten_joined_data(
        data: Sequence[BaseModel], join_config: JoinConfiguration
    ) -> list[dict[str, Any]]:
        """Convert hierarchical model data to flattened rows for Excel."""
        flattened_rows = []

        for primary_item in data:
            # Get related data (e.g., translations)
            related_data = {}
            for field_name, _model_class in join_config.related_models.items():
                field_value = getattr(primary_item, field_name, [])
                if isinstance(field_value, list):
                    related_data[field_name] = field_value
                else:
                    related_data[field_name] = [field_value] if field_value else []

            # If there are related items, create a row for each
            has_related = any(related_data.values())
            if has_related:
                # Find the related field with the most items to drive the join
                max_related_items = 0
                primary_related_field = None
                for field_name, items in related_data.items():
                    if len(items) > max_related_items:
                        max_related_items = len(items)
                        primary_related_field = field_name

                if primary_related_field:
                    primary_related_items = related_data[primary_related_field]

                    for i, related_item in enumerate(primary_related_items):
                        row = JoinedModelProcessor._create_flattened_row(
                            primary_item,
                            related_item,
                            join_config,
                            is_primary_row=(i == 0),
                        )
                        flattened_rows.append(row)
                else:
                    # No related items, create a single row with primary data only
                    row = JoinedModelProcessor._create_flattened_row(
                        primary_item, None, join_config, is_primary_row=True
                    )
                    flattened_rows.append(row)
            else:
                # No related items, create a single row with primary data only
                row = JoinedModelProcessor._create_flattened_row(
                    primary_item, None, join_config, is_primary_row=True
                )
                flattened_rows.append(row)

        return flattened_rows

    @staticmethod
    def _create_flattened_row(
        primary_item: BaseModel,
        related_item: BaseModel | None,
        join_config: JoinConfiguration,
        is_primary_row: bool,
    ) -> dict[str, Any]:
        """Create a single flattened row from primary and related items."""
        row = {}
        join_key_fields = JoinedModelProcessor._get_join_key_fields(join_config)

        for field_name in join_config.flattened_fields:
            if field_name in join_config.field_mappings:
                model_name, model_field = join_config.field_mappings[field_name]

                if model_name == "primary":
                    # Always include join key fields in every row (for table readability after sorting)
                    if field_name in join_key_fields:
                        value = getattr(primary_item, model_field, None)
                        row[field_name] = value if value is not None else ""
                    # Only include other primary fields in the first row
                    elif is_primary_row:
                        value = getattr(primary_item, model_field, None)
                        if field_name in join_config.list_fields and isinstance(
                            value, list
                        ):
                            row[field_name] = (
                                ", ".join(str(v) for v in value) if value else ""
                            )
                        else:
                            row[field_name] = value if value is not None else ""
                    else:
                        row[field_name] = ""
                elif model_name == "related" and related_item:
                    value = getattr(related_item, model_field, None)
                    if field_name in join_config.list_fields and isinstance(
                        value, list
                    ):
                        row[field_name] = (
                            ", ".join(str(v) for v in value) if value else ""
                        )
                    else:
                        row[field_name] = value if value is not None else ""
                else:
                    row[field_name] = ""
            else:
                row[field_name] = ""

        return row

    @staticmethod
    def reconstruct_joined_data(
        flattened_data: list[dict[str, Any]], join_config: JoinConfiguration
    ) -> list[BaseModel]:
        """Convert flattened Excel data back to hierarchical models."""
        # Group rows by primary key (typically concept_uri)
        primary_key = None
        for field_name in join_config.flattened_fields:
            if field_name in join_config.field_mappings:
                model_name, model_field = join_config.field_mappings[field_name]
                if model_name == "primary":
                    primary_key = field_name
                    break

        if not primary_key:
            msg = "No primary key found in field mappings"
            raise ValueError(msg)

        grouped_data: dict[str, list[dict[str, Any]]] = {}
        for row in flattened_data:
            key_value = row.get(primary_key)
            if key_value and str(key_value).strip():
                if key_value not in grouped_data:
                    grouped_data[key_value] = []
                grouped_data[key_value].append(row)

        # Reconstruct models
        reconstructed_models = []
        for _key_value, rows in grouped_data.items():
            primary_data = {}
            related_items = []

            # Process each row
            for row_idx, row in enumerate(rows):
                # Extract primary model data (only from first row)
                if row_idx == 0:
                    for field_name in join_config.flattened_fields:
                        if field_name in join_config.field_mappings:
                            model_name, model_field = join_config.field_mappings[
                                field_name
                            ]
                            if model_name == "primary":
                                value = row.get(field_name)
                                if field_name in join_config.list_fields:
                                    if isinstance(value, str) and value.strip():
                                        primary_data[model_field] = [
                                            item.strip()
                                            for item in value.split(",")
                                            if item.strip()
                                        ]
                                    else:
                                        primary_data[model_field] = []
                                else:
                                    primary_data[model_field] = (
                                        value if value is not None else ""
                                    )  # type: ignore[assignment]

                # Extract related model data
                related_data = {}
                has_related_data = False
                for field_name in join_config.flattened_fields:
                    if field_name in join_config.field_mappings:
                        model_name, model_field = join_config.field_mappings[field_name]
                        if model_name == "related":
                            value = row.get(field_name)
                            if value is not None and str(value).strip():
                                has_related_data = True
                                if field_name in join_config.list_fields:
                                    if isinstance(value, str) and value.strip():
                                        related_data[model_field] = [
                                            item.strip()
                                            for item in value.split(",")
                                            if item.strip()
                                        ]
                                    else:
                                        related_data[model_field] = []
                                else:
                                    related_data[model_field] = value
                            elif field_name in join_config.list_fields:
                                related_data[model_field] = []
                            else:
                                related_data[model_field] = None

                # Ensure join key is included in related data
                join_key_fields = JoinedModelProcessor._get_join_key_fields(join_config)
                for join_key_field in join_key_fields:
                    # Get the primary field name that corresponds to this join key
                    primary_field_name = None
                    for field_name, (
                        model_name,
                        model_field,
                    ) in join_config.field_mappings.items():
                        if model_name == "primary" and field_name == join_key_field:
                            primary_field_name = model_field
                            break

                    if primary_field_name and primary_field_name in primary_data:
                        related_data[primary_field_name] = primary_data[
                            primary_field_name
                        ]

                if has_related_data:
                    # Create related model instance
                    related_model_class = next(
                        iter(join_config.related_models.values())
                    )
                    try:
                        related_instance = related_model_class(**related_data)
                        related_items.append(related_instance)
                    except Exception as e:
                        msg = f"Error creating related model instance: {e}"
                        raise ValueError(msg) from e

            # Add related items to primary data
            related_field_name = next(iter(join_config.related_models.keys()))
            primary_data[related_field_name] = related_items  # type: ignore[assignment]

            # Create primary model instance
            try:
                primary_instance = join_config.primary_model(**primary_data)
                reconstructed_models.append(primary_instance)
            except Exception as e:
                msg = f"Error creating primary model instance: {e}"
                raise ValueError(msg) from e

        return reconstructed_models


# Table formatter
class XLSXTableFormatter(XLSXFormatter):
    """Handles tabular format with rows and columns."""

    def format_export(
        self,
        worksheet: Worksheet,
        data: Sequence[BaseModel],
        fields: list[FieldAnalysis],
    ) -> None:
        """Format data as a table with headers and rows."""
        config = self.config

        # Add title
        self._add_title(worksheet, config.title)

        # Add field meanings (if enabled and any field has meanings)
        self._add_field_meanings(worksheet, fields)

        # Add field descriptions (if enabled)
        self._add_field_descriptions(worksheet, fields, data)

        # Add field units (if enabled and any field has units)
        self._add_field_units(worksheet, fields)

        # Add field requiredness (if enabled)
        if data:
            model_class = data[0].__class__
            self._add_field_requiredness(worksheet, fields, model_class)

        # Add headers
        self._add_headers(worksheet, fields)

        # Write data rows
        self._write_data_rows(worksheet, data, fields)

        # Create table
        table = self._create_table(worksheet, fields, len(data))
        worksheet.add_table(table)

        # Add validation (always enabled)
        if isinstance(config, XLSXTableConfig):
            self._add_data_validation(worksheet, fields, len(data))

        # Auto-adjust columns
        self._auto_adjust_columns(worksheet, len(fields))

    def _add_field_descriptions(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        data: Sequence[BaseModel],
    ) -> None:
        """Add field descriptions above headers."""
        if not isinstance(self.config, XLSXTableConfig):
            return

        if not data:
            return

        # Check if descriptions should be shown
        if not self.row_calculator._should_show_descriptions(fields):
            return

        description_row = self.row_calculator.get_description_row(fields)
        start_col_idx = self.config.start_column

        for i, field_analysis in enumerate(fields):
            col_letter = get_column_letter(start_col_idx + i)
            description_cell = f"{col_letter}{description_row}"

            # Get field description from XLSX metadata
            description = ""
            if (
                field_analysis.xlsx_metadata
                and field_analysis.xlsx_metadata.description
            ):
                description = field_analysis.xlsx_metadata.description

            worksheet[description_cell] = description

            # Style description cell
            worksheet[description_cell].font = Font(
                italic=True,
                size=getattr(self.config, "description_font_size", 9),
                color=getattr(self.config, "description_color", "666666"),
            )
            worksheet[description_cell].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    def _add_field_meanings(
        self, worksheet: Worksheet, fields: list[FieldAnalysis]
    ) -> None:
        """Add field meanings above headers."""
        if not isinstance(self.config, XLSXTableConfig):
            return

        # Check if meanings should be shown
        if not self.row_calculator._should_show_meanings(fields):
            return

        # Calculate meaning row position (account for title + empty row)
        meaning_row = self.row_calculator.get_meaning_row(fields)

        start_col_idx = self.config.start_column

        for i, field_analysis in enumerate(fields):
            col_letter = get_column_letter(start_col_idx + i)
            meaning_cell = f"{col_letter}{meaning_row}"

            # Write meaning or empty string
            meaning_text = (
                field_analysis.xlsx_metadata.meaning
                if field_analysis.xlsx_metadata and field_analysis.xlsx_metadata.meaning
                else ""
            )
            worksheet[meaning_cell] = meaning_text

            # Style meaning cell
            if meaning_text:  # Only style cells with meanings
                worksheet[meaning_cell].font = Font(
                    italic=getattr(self.config, "meaning_style_italic", True),
                    size=getattr(self.config, "meaning_font_size", 9),
                    color=getattr(self.config, "meaning_color", "666666"),
                )
                worksheet[meaning_cell].alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    def _add_field_units(
        self, worksheet: Worksheet, fields: list[FieldAnalysis]
    ) -> None:
        """Add field units above headers."""
        if not isinstance(self.config, XLSXTableConfig):
            return

        # Check if units should be shown
        if not self.row_calculator._should_show_units(fields):
            return

        # Calculate unit row position
        unit_row = self.row_calculator.get_unit_row(fields)

        start_col_idx = self.config.start_column

        for i, field_analysis in enumerate(fields):
            col_letter = get_column_letter(start_col_idx + i)
            unit_cell = f"{col_letter}{unit_row}"

            # Write unit or empty string
            unit_text = (
                field_analysis.xlsx_metadata.unit
                if field_analysis.xlsx_metadata and field_analysis.xlsx_metadata.unit
                else ""
            )
            worksheet[unit_cell] = unit_text

            # Style unit cell
            if unit_text:  # Only style cells with units
                worksheet[unit_cell].font = Font(
                    italic=getattr(self.config, "unit_style_italic", True),
                    size=getattr(self.config, "unit_font_size", 9),
                    color=getattr(self.config, "unit_color", "666666"),
                )
                worksheet[unit_cell].alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    def _add_field_requiredness(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        model_class: type[BaseModel],
    ) -> None:
        """Add field requiredness row above headers.

        Shows "Yes" for required fields, "No" for optional fields,
        and "No (default: value)" for fields with non-trivial defaults.
        """
        if not isinstance(self.config, XLSXTableConfig):
            return

        # Check if requiredness should be shown
        if not self.row_calculator._should_show_requiredness(fields):
            return

        requiredness_row = self.row_calculator.get_requiredness_row(fields)
        start_col_idx = self.config.start_column

        for i, field_analysis in enumerate(fields):
            col_letter = get_column_letter(start_col_idx + i)
            req_cell = f"{col_letter}{requiredness_row}"

            # Get requiredness info from the model
            field_info = model_class.model_fields.get(field_analysis.name)
            if field_info:
                is_required, default_value = XLSXFieldAnalyzer.get_requiredness_info(
                    field_info, field_analysis
                )
                # Use labels ("Required"/"Optional") for table format since there's no header
                req_text = XLSXFieldAnalyzer.format_requiredness_text(
                    is_required, default_value, use_labels=True
                )
            else:
                req_text = ""

            worksheet[req_cell] = req_text

            # Style requiredness cell
            worksheet[req_cell].font = Font(
                italic=True,
                size=9,
                color="666666",
            )
            worksheet[req_cell].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )

    def _add_headers(self, worksheet: Worksheet, fields: list[FieldAnalysis]) -> None:
        """Add column headers."""
        # Determine header row based on whether meanings, descriptions, and units are shown
        header_row = self.row_calculator.get_header_row(fields)

        start_col_idx = self.config.start_column

        for i, field_analysis in enumerate(fields):
            col_letter = get_column_letter(start_col_idx + i)
            header_cell = f"{col_letter}{header_row}"

            header_text = self._format_header_text(field_analysis)
            worksheet[header_cell] = header_text
            worksheet[header_cell].alignment = Alignment(
                wrap_text=True, vertical="center"
            )

            # Only apply explicit header styling if configured
            # Otherwise, let the table style control header appearance (font, color, fill)
            if (
                hasattr(self.config, "header_row_color")
                and self.config.header_row_color
            ):
                worksheet[header_cell].font = Font(bold=True)
                header_fill = PatternFill(
                    start_color=self.config.header_row_color,
                    end_color=self.config.header_row_color,
                    fill_type="solid",
                )
                worksheet[header_cell].fill = header_fill

    def _write_data_rows(
        self,
        worksheet: Worksheet,
        data: Sequence[BaseModel],
        fields: list[FieldAnalysis],
    ) -> None:
        """Write data rows to worksheet."""
        start_col_idx = self.config.start_column

        # Calculate data start row - account for title, meanings, descriptions, and units
        data_start_row = self.row_calculator.get_data_start_row(fields)

        for row_idx, item in enumerate(data, start=data_start_row):
            for col_idx, field_analysis in enumerate(fields):
                col_letter = get_column_letter(start_col_idx + col_idx)
                cell = f"{col_letter}{row_idx}"

                value = getattr(item, field_analysis.name, None)

                try:
                    formatted_value = self.serialization_engine.serialize_value(
                        value, field_analysis
                    )
                    data_cell = worksheet[cell]
                    data_cell.value = formatted_value
                    # Apply data cell formatting (vertical center, left align, text wrap for strings)
                    self._apply_data_cell_formatting(data_cell, field_analysis)
                    # Apply bold formatting for specified fields
                    if field_analysis.name in self.config.bold_fields:
                        data_cell.font = Font(bold=True)
                except Exception as e:
                    raise XLSXSerializationError(field_analysis.name, value, e) from e

    def _create_table(
        self, worksheet: Worksheet, fields: list[FieldAnalysis], data_rows: int
    ) -> Table:
        """Create Excel table."""
        start_col_idx = self.config.start_column
        end_col_idx = start_col_idx + len(fields) - 1
        end_col_letter = get_column_letter(end_col_idx)

        # Calculate table start row (header row, not title/description/meaning/unit rows)
        table_start_row = self.row_calculator.get_header_row(fields)

        # Calculate table end row
        table_end_row = table_start_row + data_rows

        start_col_letter = get_column_letter(self.config.start_column)
        table_range = (
            f"{start_col_letter}{table_start_row}:{end_col_letter}{table_end_row}"
        )

        table_name = f"Table_{worksheet.title.replace(' ', '_')}"
        table_name = re.sub(r"[^a-zA-Z0-9_]", "_", table_name)
        if len(table_name) > MAX_SHEETNAME_LENGTH:
            table_name = table_name[:MAX_SHEETNAME_LENGTH]

        table = Table(displayName=table_name, ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name=self.config.table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        return table

    def _add_data_validation(
        self, worksheet: Worksheet, fields: list[FieldAnalysis], data_rows: int
    ) -> None:
        """Add data validation for enum fields.

        For enum lists that exceed Excel's 255 character limit for inline
        formulas, creates a hidden sheet with the values and uses a named range.
        """
        start_col_idx = self.config.start_column

        # Calculate data start row - account for title, meanings, descriptions, and units
        data_start_row = self.row_calculator.get_data_start_row(fields)

        for i, field_analysis in enumerate(fields):
            if field_analysis.enum_values:
                col_letter = get_column_letter(start_col_idx + i)
                validation_range = (
                    f"{col_letter}{data_start_row}:"
                    f"{col_letter}{data_start_row + data_rows - 1}"
                )
                self._add_enum_validation(worksheet, field_analysis, validation_range)

    def parse_import(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        model_class: type[BaseModel],
    ) -> list[BaseModel]:
        """Parse tabular data from worksheet."""
        # Read headers
        headers = self._read_headers(worksheet, fields)

        # Read data rows
        data_rows = self._read_data_rows(worksheet, headers, fields)

        # Convert to models
        models = []
        errors = []

        for row_idx, row_data in enumerate(data_rows, start=self.config.start_row + 1):
            try:
                model_data = self._convert_row_to_model_data(
                    row_data, fields, model_class
                )
                model = model_class(**model_data)
                models.append(model)
            except ValidationError as e:
                errors.append(f"Row {row_idx}: {e}")
            except Exception as e:
                errors.append(f"Row {row_idx}: {e}")

        if errors:
            error_msg = "Import errors found:\n" + "\n".join(errors)
            raise ValueError(error_msg)

        return models

    def _read_headers(
        self, worksheet: Worksheet, fields: list[FieldAnalysis]
    ) -> dict[str, int]:
        """Read and map headers to column indices."""
        headers = {}
        start_col_idx = self.config.start_column

        # Determine header row based on whether title, meanings, descriptions, and units are shown
        header_row = self.row_calculator.get_header_row(fields)

        for i, field_analysis in enumerate(fields):
            col_letter = get_column_letter(start_col_idx + i)
            header_cell = f"{col_letter}{header_row}"
            header_value = worksheet[header_cell].value

            if header_value:
                expected_header = self._format_header_text(field_analysis)
                if str(header_value).strip() == expected_header:
                    headers[field_analysis.name] = i

        return headers

    def _read_data_rows(
        self,
        worksheet: Worksheet,
        headers: dict[str, int],
        fields: list[FieldAnalysis] | None = None,
    ) -> list[dict[str, Any]]:
        """Read data rows from worksheet."""
        data_rows = []
        start_col_idx = self.config.start_column
        max_row = worksheet.max_row

        # Calculate data start row - account for title, meanings, descriptions, and units
        data_start_row = self.row_calculator.get_data_start_row(fields)

        for row_idx in range(data_start_row, max_row + 1):
            row_data = {}
            has_data = False

            for field_name, col_offset in headers.items():
                col_letter = get_column_letter(start_col_idx + col_offset)
                cell = f"{col_letter}{row_idx}"
                value = worksheet[cell].value

                # Consider a value as "data" only if it's not None and not empty string
                if value is not None and value != "":
                    has_data = True

                row_data[field_name] = value

            if has_data:
                data_rows.append(row_data)
            else:
                break

        return data_rows

    def _convert_row_to_model_data(
        self,
        row_data: dict[str, Any],
        fields: list[FieldAnalysis],
        model_class: type[BaseModel],
    ) -> dict[str, Any]:
        """Convert row data to model-compatible format."""

        model_data: dict[str, Any] = {}

        for field_analysis in fields:
            value = row_data.get(field_analysis.name)

            # Handle empty/null values
            if value is None or (isinstance(value, str) and value.strip() == ""):
                # Check if field is optional by looking at the model field definition
                field_def = model_class.model_fields.get(field_analysis.name)
                if field_def:
                    # Check if field is optional (has None in Union) or has a default value
                    field_type = field_def.annotation
                    has_default = (
                        hasattr(field_def, "default")
                        and field_def.default is not PydanticUndefined
                    )
                    is_optional_type = (
                        field_analysis.is_optional
                        or XLSXFieldAnalyzer.is_optional_type(field_type)
                    )

                    if not is_optional_type and not has_default:
                        msg = f"Required field '{field_analysis.name}' is empty"
                        raise ValueError(msg)

                    # If field accepts None (is_optional_type), include None in data
                    # If field has non-None default, skip it so Pydantic uses default
                    if is_optional_type:
                        model_data[field_analysis.name] = None
                    # else: skip - let Pydantic use the model's default
                    continue

            try:
                model_data[field_analysis.name] = (
                    self.serialization_engine.deserialize_value(
                        value, field_analysis, model_class
                    )
                )
            except Exception as e:
                raise XLSXDeserializationError(field_analysis.name, value, e) from e

        return model_data


# Joined table formatter
class XLSXJoinedTableFormatter(XLSXTableFormatter):
    """Handles joined models with tabular format."""

    def __init__(self, config: XLSXConfig, join_config: JoinConfiguration):
        super().__init__(config)
        self.join_config = join_config

    def format_export(
        self,
        worksheet: Worksheet,
        data: Sequence[BaseModel],
        fields: list[FieldAnalysis],
    ) -> None:
        """Format joined model data as a table with headers and rows."""
        # Flatten the joined data
        flattened_data = JoinedModelProcessor.flatten_joined_data(
            data, self.join_config
        )

        # Create field analyses for the flattened structure
        flattened_fields = self._create_flattened_field_analyses(fields)

        # Add title
        self._add_title(worksheet, self.config.title)

        # Add field meanings, descriptions, and units if enabled
        self._add_field_meanings(worksheet, flattened_fields)
        self._add_field_descriptions(worksheet, flattened_fields, data)
        self._add_field_units(worksheet, flattened_fields)

        # Add headers
        self._add_headers(worksheet, flattened_fields)

        # Write flattened data rows
        self._write_flattened_data_rows(worksheet, flattened_data, flattened_fields)

        # Create table
        table = self._create_table(worksheet, flattened_fields, len(flattened_data))
        worksheet.add_table(table)

        # Add validation (always enabled)
        if isinstance(self.config, XLSXTableConfig):
            self._add_data_validation(worksheet, flattened_fields, len(flattened_data))

        # Auto-adjust columns
        self._auto_adjust_columns(worksheet, len(flattened_fields))

    def _create_flattened_field_analyses(
        self, original_fields: list[FieldAnalysis]
    ) -> list[FieldAnalysis]:
        """Create field analyses for the flattened structure."""
        flattened_fields = []

        for field_name in self.join_config.flattened_fields:
            field_analysis = None

            if field_name in self.join_config.field_mappings:
                model_name, model_field = self.join_config.field_mappings[field_name]

                # Try to get the field analysis from the appropriate model
                if model_name == "primary":
                    source_model = self.join_config.primary_model
                else:
                    source_model = next(iter(self.join_config.related_models.values()))

                # Analyze the source model to get field information
                source_field_analyses = XLSXFieldAnalyzer.analyze_model(source_model)
                source_fields = list(source_field_analyses.values())
                source_field_map = {field.name: field for field in source_fields}

                if model_field in source_field_map:
                    # Use the original field analysis but with the flattened field name
                    original_field = source_field_map[model_field]
                    field_analysis = FieldAnalysis(
                        name=field_name,
                        field_type=original_field.field_type,
                        is_optional=True,  # Flattened fields are typically optional
                        enum_values=original_field.enum_values,
                        xlsx_metadata=original_field.xlsx_metadata,
                    )

            # Fallback to generic string field if we can't find the original
            if field_analysis is None:
                field_analysis = FieldAnalysis(
                    name=field_name,
                    field_type=str,
                    is_optional=True,
                )

            flattened_fields.append(field_analysis)

        return flattened_fields

    def _write_flattened_data_rows(
        self,
        worksheet: Worksheet,
        flattened_data: list[dict[str, Any]],
        fields: list[FieldAnalysis],
    ) -> None:
        """Write flattened data rows to worksheet."""
        start_col_idx = self.config.start_column

        # Calculate data start row - account for title, meanings, descriptions, and units
        data_start_row = self.row_calculator.get_data_start_row(fields)

        for row_idx, row_data in enumerate(flattened_data, start=data_start_row):
            for col_idx, field_analysis in enumerate(fields):
                col_letter = get_column_letter(start_col_idx + col_idx)
                cell = f"{col_letter}{row_idx}"

                value = row_data.get(field_analysis.name)

                try:
                    # Use the serialization engine to format the value
                    formatted_value = self.serialization_engine.serialize_value(
                        value, field_analysis
                    )
                    data_cell = worksheet[cell]
                    data_cell.value = formatted_value
                    # Apply data cell formatting
                    self._apply_data_cell_formatting(data_cell, field_analysis)
                except Exception as e:
                    raise XLSXSerializationError(field_analysis.name, value, e) from e

    def parse_import(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        model_class: type[BaseModel],
    ) -> list[BaseModel]:
        """Parse joined model data from worksheet."""
        # Create flattened field analyses
        flattened_fields = self._create_flattened_field_analyses(fields)

        # Read headers
        headers = self._read_headers(worksheet, flattened_fields)

        # Read data rows
        data_rows = self._read_data_rows(worksheet, headers, flattened_fields)

        # Convert to flattened format expected by JoinedModelProcessor
        flattened_data = []
        for row_data in data_rows:
            flattened_row = {}
            for field_name in self.join_config.flattened_fields:
                # Get the corresponding field analysis
                field_analysis = next(
                    (f for f in flattened_fields if f.name == field_name), None
                )
                if field_analysis:
                    raw_value = row_data.get(field_name)
                    try:
                        # Use the serialization engine to deserialize the value
                        converted_value = self.serialization_engine.deserialize_value(
                            raw_value, field_analysis, model_class
                        )
                        # If value is None, only add it if the field accepts None
                        # Otherwise, skip it so Pydantic uses the model's default
                        if converted_value is not None:
                            flattened_row[field_name] = converted_value
                        elif field_analysis.is_optional:
                            flattened_row[field_name] = None
                    except Exception as e:
                        raise XLSXDeserializationError(field_name, raw_value, e) from e
                else:
                    value = row_data.get(field_name)
                    if value is not None:
                        flattened_row[field_name] = value
            flattened_data.append(flattened_row)

        # Use JoinedModelProcessor to reconstruct the models
        return JoinedModelProcessor.reconstruct_joined_data(
            flattened_data, self.join_config
        )


# Table processor
class XLSXTableProcessor(XLSXProcessor):
    """Processor for tabular format."""

    def export(
        self, data: Sequence[BaseModel], filepath: Path, sheet_name: str | None = None
    ) -> None:
        """Export sequence of models to XLSX file."""
        if not data:
            msg = "No data provided for export"
            raise ValueError(msg)

        model_class = data[0].__class__
        sheet_name = sheet_name or model_class.__name__

        # Analyze fields
        field_analyses = self.field_analyzer.analyze_model(model_class)
        fields = list(field_analyses.values())
        filtered_fields = self._filter_and_order_fields(fields)

        workbook, worksheet = self._prepare_workbook(filepath, sheet_name)
        self.formatter.format_export(worksheet, data, filtered_fields)
        workbook.save(filepath)

    def import_data(
        self,
        filepath: Path,
        model_class: type[BaseModel],
        sheet_name: str | None = None,
    ) -> list[BaseModel]:
        """Import sequence of models from XLSX file."""
        sheet_name = sheet_name or model_class.__name__

        workbook = load_workbook(filepath, data_only=True)

        if sheet_name not in workbook.sheetnames:
            msg = f"Sheet '{sheet_name}' not found in workbook"
            raise ValueError(msg)

        worksheet = workbook[sheet_name]

        # Analyze fields
        field_analyses = self.field_analyzer.analyze_model(model_class)
        fields = list(field_analyses.values())
        filtered_fields = self._filter_and_order_fields(fields)

        # Parse import
        return self.formatter.parse_import(worksheet, filtered_fields, model_class)


# Joined table processor
class XLSXJoinedTableProcessor(XLSXTableProcessor):
    """Processor for joined models in tabular format."""

    def __init__(self, config: XLSXConfig, formatter: XLSXJoinedTableFormatter):
        super().__init__(config, formatter)

    def export(
        self, data: Sequence[BaseModel], filepath: Path, sheet_name: str | None = None
    ) -> None:
        """Export sequence of joined models to XLSX file."""
        if not data:
            msg = "No data provided for export"
            raise ValueError(msg)

        # Get the primary model class from the join configuration
        primary_model_class = self.formatter.join_config.primary_model
        sheet_name = sheet_name or primary_model_class.__name__

        # Analyze fields from the primary model
        field_analyses = self.field_analyzer.analyze_model(primary_model_class)
        fields = list(field_analyses.values())
        filtered_fields = self._filter_and_order_fields(fields)

        workbook, worksheet = self._prepare_workbook(filepath, sheet_name)
        self.formatter.format_export(worksheet, data, filtered_fields)
        workbook.save(filepath)

    def import_data(
        self,
        filepath: Path,
        model_class: type[BaseModel] | None = None,
        sheet_name: str | None = None,
    ) -> list[BaseModel]:
        """Import sequence of joined models from XLSX file."""
        # Use the primary model class from the join configuration if not specified
        if model_class is None:
            model_class = self.formatter.join_config.primary_model

        sheet_name = sheet_name or model_class.__name__

        workbook = load_workbook(filepath, data_only=True)

        if sheet_name not in workbook.sheetnames:
            msg = f"Sheet '{sheet_name}' not found in workbook"
            raise ValueError(msg)

        worksheet = workbook[sheet_name]

        # Analyze fields from the primary model
        field_analyses = self.field_analyzer.analyze_model(model_class)
        fields = list(field_analyses.values())
        filtered_fields = self._filter_and_order_fields(fields)

        # Parse import using the joined formatter
        return self.formatter.parse_import(worksheet, filtered_fields, model_class)
