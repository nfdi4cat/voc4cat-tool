"""Converter for RDF vocabularies to/from v1.0 Excel template format.

This module provides functions to convert between RDF vocabularies and
the v1.0 Excel template structure, supporting bidirectional conversion:
- RDF -> XLSX: Extract data from RDF graphs into v1.0 template
- XLSX -> RDF: Read v1.0 template and generate RDF graph

The two-way conversion is designed to be lossless (isomorphic graphs).
"""

import datetime
import logging
import os
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal as TypingLiteral

import curies
from openpyxl import load_workbook
from openpyxl.styles import Font
from pydantic import BaseModel
from rdflib import (
    DCAT,
    DCTERMS,
    FOAF,
    OWL,
    PROV,
    RDF,
    RDFS,
    SKOS,
    XSD,
    BNode,
    Graph,
    Literal,
    Namespace,
    URIRef,
)
from rdflib.collection import Collection as RDFCollection

from voc4cat import config
from voc4cat.checks import Voc4catError
from voc4cat.convert_v1_helpers import (
    add_provenance_triples_to_graph,
    build_id_range_info,
    derive_contributors,
    expand_curie,
    extract_creator_names,
    extract_used_ids,
    format_iri_with_label,
    generate_history_note,
    validate_deprecation,
    validate_entity_deprecation,
)
from voc4cat.models_v1 import (
    COLLECTIONS_EXPORT_CONFIG,
    COLLECTIONS_READ_CONFIG,
    COLLECTIONS_SHEET_NAME,
    CONCEPT_SCHEME_SHEET_NAME,
    CONCEPT_SCHEME_SHEET_TITLE,
    CONCEPTS_EXPORT_CONFIG,
    CONCEPTS_READ_CONFIG,
    CONCEPTS_SHEET_NAME,
    ID_RANGES_EXPORT_CONFIG,
    ID_RANGES_SHEET_NAME,
    MAPPINGS_EXPORT_CONFIG,
    MAPPINGS_READ_CONFIG,
    MAPPINGS_SHEET_NAME,
    PREFIXES_EXPORT_CONFIG,
    PREFIXES_READ_CONFIG,
    PREFIXES_SHEET_NAME,
    TEMPLATE_VERSION,
    CollectionObsoletionReason,
    CollectionV1,
    ConceptObsoletionReason,
    ConceptSchemeV1,
    ConceptV1,
    IDRangeInfoV1,
    MappingV1,
    OrderedChoice,
    PrefixV1,
)
from voc4cat.utils import (
    EXCEL_FILE_ENDINGS,
    RDF_FILE_ENDINGS,
    get_template_sheet_names,
    reorder_sheets_with_template,
)
from voc4cat.xlsx_api import export_to_xlsx, import_from_xlsx
from voc4cat.xlsx_common import (
    MetadataToggleConfig,
    MetadataVisibility,
    XLSXFieldAnalyzer,
    XLSXRowCalculator,
)
from voc4cat.xlsx_keyvalue import XLSXKeyValueConfig
from voc4cat.xlsx_table import XLSXTableConfig

# schema.org namespace (not in rdflib by default)
SDO = Namespace("https://schema.org/")

logger = logging.getLogger(__name__)


# =============================================================================
# Enum Conversion Helpers
# =============================================================================


def string_to_concept_obsoletion_enum(value: str) -> ConceptObsoletionReason | None:
    """Convert string to ConceptObsoletionReason enum or None.

    Args:
        value: The obsoletion reason string from RDF.

    Returns:
        The matching enum value, or None if empty or no match found.
    """
    if not value:
        return None
    for reason in ConceptObsoletionReason:
        if reason.value == value:
            return reason
    # Non-standard reason - log warning and return None
    logger.warning("Non-standard obsoletion reason for concept: %s", value)
    return None


def string_to_collection_obsoletion_enum(
    value: str,
) -> CollectionObsoletionReason | None:
    """Convert string to CollectionObsoletionReason enum or None.

    Args:
        value: The obsoletion reason string from RDF.

    Returns:
        The matching enum value, or None if empty or no match found.
    """
    if not value:
        return None
    for reason in CollectionObsoletionReason:
        if reason.value == value:
            return reason
    # Non-standard reason - log warning and return None
    logger.warning("Non-standard obsoletion reason for collection: %s", value)
    return None


def string_to_ordered_enum(value: str | bool) -> OrderedChoice | None:
    """Convert string/bool to OrderedChoice enum or None.

    Args:
        value: The ordered flag value (string or bool) from RDF.

    Returns:
        OrderedChoice.YES if true/yes, OrderedChoice.NO if no, None otherwise.
    """
    if isinstance(value, bool):
        return OrderedChoice.YES if value else None
    if not value:
        return None
    if value.strip().lower() == "yes":
        return OrderedChoice.YES
    if value.strip().lower() == "no":
        return OrderedChoice.NO
    return None


# =============================================================================
# RDF Extraction Functions
# =============================================================================


def lookup_entity_name(graph: Graph, entity_iri: str) -> str:
    """Look up schema:name for an entity and format as 'name <url>'.

    Looks for schema:name on the entity IRI. If found and not a URL,
    returns 'name url'. Otherwise returns just the URL.

    Args:
        graph: The RDF graph to query.
        entity_iri: The IRI of the entity (Person or Organization).

    Returns:
        Formatted string: 'name url' if name found, else just 'url'.
    """
    entity_ref = URIRef(entity_iri)
    name = None

    # Look up schema:name
    for obj in graph.objects(entity_ref, SDO.name):
        name = str(obj)
        break

    # If name exists and is not a URL, format as "name url"
    if name and not name.startswith("http"):
        return f"{name} {entity_iri}"
    return entity_iri


def lookup_entity_name_safe(graph: Graph, value: str) -> str:
    """Look up entity name, handling both URIs and literal strings.

    For URI values (starting with http), looks up schema:name and formats
    as 'name url'. For literal strings, returns them as-is.

    This is useful for fields like dcat:contactPoint that may be either
    a URI reference or a literal string in different RDF sources.

    Args:
        graph: The RDF graph to query.
        value: Either a URI or a literal string value.

    Returns:
        Formatted string: 'name url' for URIs, or the literal value as-is.
    """
    if value.startswith("http"):
        return lookup_entity_name(graph, value)
    # For literal values, return as-is
    return value


def extract_concept_scheme_from_rdf(graph: Graph) -> dict:
    """Extract ConceptScheme data from an RDF graph.

    Args:
        graph: The RDF graph to extract from.

    Returns:
        Dictionary with concept scheme data fields.
    """
    holder = {
        "vocabulary_iri": "",
        "title": "",
        "description": "",
        "created_date": "",
        "modified_date": "",
        "creator": "",
        "contributor": "",
        "publisher": "",
        "version": "",
        "history_note": "",
        "custodian": "",
        "catalogue_pid": "",
    }

    # Collect multiple values for fields that can have multiple entries
    creators = []
    contributors = []
    publishers = []
    custodians = []

    for s in graph.subjects(RDF.type, SKOS.ConceptScheme):
        holder["vocabulary_iri"] = str(s)

        for p, o in graph.predicate_objects(s):
            if p == SKOS.prefLabel:
                holder["title"] = str(o)
            elif p == SKOS.definition:
                holder["description"] = str(o)
            elif p == DCTERMS.created:
                holder["created_date"] = str(o)
            elif p == DCTERMS.modified:
                holder["modified_date"] = str(o)
            elif p == DCTERMS.creator:
                creators.append(str(o))
            elif p == DCTERMS.contributor:
                contributors.append(str(o))
            elif p == DCTERMS.publisher:
                publishers.append(str(o))
            elif p == OWL.versionInfo:
                holder["version"] = str(o)
            elif p in [SKOS.historyNote, DCTERMS.provenance, PROV.wasDerivedFrom]:
                holder["history_note"] = str(o)
            elif p == DCAT.contactPoint:
                custodians.append(str(o))
            elif p == RDFS.seeAlso:
                holder["catalogue_pid"] = str(o)

        # Look up names from schema:name and format as "name url"
        holder["creator"] = "\n".join(
            lookup_entity_name(graph, iri) for iri in creators
        )
        holder["contributor"] = "\n".join(
            lookup_entity_name(graph, iri) for iri in contributors
        )
        holder["publisher"] = "\n".join(
            lookup_entity_name(graph, iri) for iri in publishers
        )
        holder["custodian"] = "\n".join(
            lookup_entity_name_safe(graph, value) for value in custodians
        )

        # Only process the first ConceptScheme found
        # TODO: log warning or error if multiple found
        break

    return holder


def extract_concepts_from_rdf(graph: Graph) -> dict[str, dict[str, dict]]:
    """Extract Concepts from an RDF graph, grouped by IRI and language.

    Args:
        graph: The RDF graph to extract from.

    Returns:
        Nested dict: {concept_iri: {language: concept_data_dict}}
        Each concept_data_dict contains: preferred_label, definition,
        alternate_labels, parent_iris, source_vocab_iri, change_note,
        editorial_note, obsolete_reason, influenced_by_iris,
        source_vocab_license, source_vocab_rights_holder.
    """
    # Structure: {iri: {lang: {field: value}}}
    concepts_by_iri_lang: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(dict)
    )

    # First pass: collect all data per concept
    concept_data: dict[str, dict] = defaultdict(
        lambda: {
            "pref_labels": {},  # {lang: label}
            "definitions": {},  # {lang: definition}
            "alt_labels": {},  # {lang: [labels]}
            "editorial_notes": {},  # {lang: note}
            "parent_iris": [],
            "source_vocab_iri": "",
            "source_vocab_license": "",
            "source_vocab_rights_holder": "",
            "change_note": "",
            "obsolete_reason": "",
            "is_deprecated": False,
            "influenced_by_iris": [],
            "replaced_by_iri": "",
        }
    )

    for s in graph.subjects(RDF.type, SKOS.Concept):
        iri = str(s)
        data = concept_data[iri]

        for p, o in graph.predicate_objects(s):
            if p == SKOS.prefLabel:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                data["pref_labels"][lang] = str(o)
            elif p == SKOS.definition:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                data["definitions"][lang] = str(o)
            elif p == SKOS.altLabel:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                if lang not in data["alt_labels"]:
                    data["alt_labels"][lang] = []
                data["alt_labels"][lang].append(str(o))
            elif p == SKOS.editorialNote:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                data["editorial_notes"][lang] = str(o)
            elif p == SKOS.broader:
                data["parent_iris"].append(str(o))
            elif p == PROV.hadPrimarySource:
                data["source_vocab_iri"] = str(o)
            elif p == DCTERMS.license:
                data["source_vocab_license"] = str(o)
            elif p == DCTERMS.rightsHolder:
                data["source_vocab_rights_holder"] = str(o)
            elif p == SKOS.changeNote:
                data["change_note"] = str(o)
            elif p == OWL.deprecated:
                data["is_deprecated"] = str(o).lower() == "true"
            elif p == SKOS.historyNote:
                # historyNote is used for obsoletion reason when deprecated
                data["obsolete_reason"] = str(o)
            elif p == PROV.wasInfluencedBy:
                data["influenced_by_iris"].append(str(o))
            elif p == DCTERMS.isReplacedBy:
                data["replaced_by_iri"] = str(o)

    # Second pass: organize by IRI and language
    for iri, data in concept_data.items():
        # Determine all languages present for this concept
        all_langs = set(data["pref_labels"].keys()) | set(data["definitions"].keys())

        # Put "en" first if available
        langs_ordered = sorted(all_langs)
        if "en" in langs_ordered:
            langs_ordered.remove("en")
            langs_ordered.insert(0, "en")

        for lang in langs_ordered:
            # Only include obsolete_reason if concept is deprecated
            obsolete_reason = data["obsolete_reason"] if data["is_deprecated"] else ""

            concepts_by_iri_lang[iri][lang] = {
                "preferred_label": data["pref_labels"].get(lang, ""),
                "definition": data["definitions"].get(lang, ""),
                "alternate_labels": data["alt_labels"].get(lang, []),
                "editorial_note": data["editorial_notes"].get(lang, ""),
                "parent_iris": data["parent_iris"],
                "source_vocab_iri": data["source_vocab_iri"],
                "source_vocab_license": data["source_vocab_license"],
                "source_vocab_rights_holder": data["source_vocab_rights_holder"],
                "change_note": data["change_note"],
                "obsolete_reason": obsolete_reason,
                "influenced_by_iris": data["influenced_by_iris"],
                "replaced_by_iri": data["replaced_by_iri"],
                "is_deprecated": data["is_deprecated"],
            }

    return dict(concepts_by_iri_lang)


def extract_collections_from_rdf(graph: Graph) -> dict[str, dict[str, dict]]:
    """Extract Collections from an RDF graph, grouped by IRI and language.

    Args:
        graph: The RDF graph to extract from.

    Returns:
        Nested dict: {collection_iri: {language: collection_data_dict}}
        Each collection_data_dict contains: preferred_label, definition, change_note,
        editorial_note, obsolete_reason, ordered, members.
    """
    collections_by_iri_lang: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(dict)
    )

    # First pass: collect all data per collection
    collection_data: dict[str, dict] = defaultdict(
        lambda: {
            "pref_labels": {},  # {lang: label}
            "definitions": {},  # {lang: definition}
            "editorial_notes": {},  # {lang: note}
            "change_note": "",
            "obsolete_reason": "",
            "is_deprecated": False,
            "ordered": False,
            "members": [],  # member IRIs (concepts or collections) - unordered
            "ordered_members": [],  # member IRIs in order - for OrderedCollection
            "replaced_by_iri": "",
        }
    )

    # Process both Collection and OrderedCollection types
    collection_iris = set()
    for s in graph.subjects(RDF.type, SKOS.Collection):
        collection_iris.add(s)
    for s in graph.subjects(RDF.type, SKOS.OrderedCollection):
        collection_iris.add(s)

    for s in collection_iris:
        iri = str(s)
        data = collection_data[iri]

        # Check if it's an OrderedCollection
        if (s, RDF.type, SKOS.OrderedCollection) in graph:
            data["ordered"] = True

        for p, o in graph.predicate_objects(s):
            if p == SKOS.prefLabel:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                data["pref_labels"][lang] = str(o)
            elif p == SKOS.definition:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                data["definitions"][lang] = str(o)
            elif p == SKOS.editorialNote:
                lang = o.language if isinstance(o, Literal) and o.language else "en"
                data["editorial_notes"][lang] = str(o)
            elif p == SKOS.changeNote:
                data["change_note"] = str(o)
            elif p == OWL.deprecated:
                data["is_deprecated"] = str(o).lower() == "true"
            elif p == SKOS.historyNote:
                # historyNote is used for obsoletion reason when deprecated
                data["obsolete_reason"] = str(o)
            elif p == SKOS.member:
                data["members"].append(str(o))
            elif p == SKOS.memberList:
                # Parse RDF List for ordered members
                try:
                    rdf_list = RDFCollection(graph, o)
                    data["ordered_members"] = [str(m) for m in rdf_list]
                except Exception:
                    # If parsing fails, fall back to empty list
                    data["ordered_members"] = []
            elif p == DCTERMS.isReplacedBy:
                data["replaced_by_iri"] = str(o)

    # Second pass: organize by IRI and language
    for iri, data in collection_data.items():
        all_langs = set(data["pref_labels"].keys()) | set(data["definitions"].keys())

        langs_ordered = sorted(all_langs)
        if "en" in langs_ordered:
            langs_ordered.remove("en")
            langs_ordered.insert(0, "en")

        # Only include obsolete_reason if collection is deprecated
        obsolete_reason = data["obsolete_reason"] if data["is_deprecated"] else ""

        for lang in langs_ordered:
            collections_by_iri_lang[iri][lang] = {
                "preferred_label": data["pref_labels"].get(lang, ""),
                "definition": data["definitions"].get(lang, ""),
                "editorial_note": data["editorial_notes"].get(lang, ""),
                "change_note": data["change_note"],
                "obsolete_reason": obsolete_reason,
                "ordered": data["ordered"],
                "members": data["members"],
                "ordered_members": data["ordered_members"],
                "replaced_by_iri": data["replaced_by_iri"],
                "is_deprecated": data["is_deprecated"],
            }

    return dict(collections_by_iri_lang)


def extract_mappings_from_rdf(graph: Graph) -> dict[str, dict]:
    """Extract mapping relations from an RDF graph.

    Args:
        graph: The RDF graph to extract from.

    Returns:
        Dict: {concept_iri: {related_matches: [], close_matches: [], ...}}
    """
    mappings: dict[str, dict] = defaultdict(
        lambda: {
            "related_matches": [],
            "close_matches": [],
            "exact_matches": [],
            "narrower_matches": [],
            "broader_matches": [],
        }
    )

    for s in graph.subjects(RDF.type, SKOS.Concept):
        iri = str(s)
        has_mappings = False

        for p, o in graph.predicate_objects(s):
            if p == SKOS.relatedMatch:
                mappings[iri]["related_matches"].append(str(o))
                has_mappings = True
            elif p == SKOS.closeMatch:
                mappings[iri]["close_matches"].append(str(o))
                has_mappings = True
            elif p == SKOS.exactMatch:
                mappings[iri]["exact_matches"].append(str(o))
                has_mappings = True
            elif p == SKOS.narrowMatch:
                mappings[iri]["narrower_matches"].append(str(o))
                has_mappings = True
            elif p == SKOS.broadMatch:
                mappings[iri]["broader_matches"].append(str(o))
                has_mappings = True

        # Only keep entries that have at least one mapping
        if not has_mappings and iri in mappings:
            del mappings[iri]

    return dict(mappings)


def build_concept_to_collections_map(graph: Graph) -> dict[str, list[str]]:
    """Build a mapping from concept IRIs to the collections they belong to.

    This inverts the skos:member relationship.

    Args:
        graph: The RDF graph to analyze.

    Returns:
        Dict: {concept_iri: [collection_iris]}
    """
    concept_to_collections: dict[str, list[str]] = defaultdict(list)

    for collection_iri in graph.subjects(RDF.type, SKOS.Collection):
        for member in graph.objects(collection_iri, SKOS.member):
            member_iri = str(member)
            # Only map if member is a concept (not another collection)
            if (member, RDF.type, SKOS.Concept) in graph:
                concept_to_collections[member_iri].append(str(collection_iri))

    return dict(concept_to_collections)


def build_collection_hierarchy_map(graph: Graph) -> dict[str, list[str]]:
    """Build a mapping from collection IRIs to their parent collections.

    Collections can be members of other collections (hierarchy).

    Args:
        graph: The RDF graph to analyze.

    Returns:
        Dict: {child_collection_iri: [parent_collection_iris]}
    """
    collection_to_parents: dict[str, list[str]] = defaultdict(list)

    for parent_iri in graph.subjects(RDF.type, SKOS.Collection):
        for member in graph.objects(parent_iri, SKOS.member):
            member_iri = str(member)
            # Only map if member is a collection (not a concept)
            if (member, RDF.type, SKOS.Collection) in graph:
                collection_to_parents[member_iri].append(str(parent_iri))

    return dict(collection_to_parents)


def build_concept_to_ordered_collections_map(
    graph: Graph,
) -> dict[str, dict[str, int]]:
    """Build mapping from concept IRIs to ordered collections with positions.

    Args:
        graph: The RDF graph to analyze.

    Returns:
        Dict: {concept_iri: {collection_iri: position}}
        Position is 1-indexed (first member is position 1).
    """

    concept_to_ordered: dict[str, dict[str, int]] = defaultdict(dict)

    for collection_iri in graph.subjects(RDF.type, SKOS.OrderedCollection):
        # Get the memberList
        member_list_node = graph.value(collection_iri, SKOS.memberList)
        if member_list_node:
            try:
                rdf_list = RDFCollection(graph, member_list_node)
                for position, member in enumerate(rdf_list, start=1):
                    member_iri = str(member)
                    # Only map if member is a concept
                    if (member, RDF.type, SKOS.Concept) in graph:
                        concept_to_ordered[member_iri][str(collection_iri)] = position
            except Exception as e:
                # If parsing fails, warn and skip this collection
                logger.warning(
                    "Failed to parse ordered collection %s: %s", collection_iri, e
                )

    return dict(concept_to_ordered)


# =============================================================================
# Model Conversion Functions
# =============================================================================


def rdf_concept_scheme_to_v1(data: dict) -> ConceptSchemeV1:
    """Convert extracted concept scheme data to ConceptSchemeV1 model.

    Args:
        data: Dictionary with concept scheme fields.

    Returns:
        ConceptSchemeV1 model instance.
    """
    return ConceptSchemeV1(
        template_version=TEMPLATE_VERSION,
        vocabulary_iri=data.get("vocabulary_iri", ""),
        title=data.get("title", ""),
        description=data.get("description", ""),
        created_date=data.get("created_date", ""),
        modified_date=data.get("modified_date", ""),
        creator=data.get("creator", ""),
        contributor=data.get("contributor", ""),
        publisher=data.get("publisher", ""),
        version=data.get("version", ""),
        history_note=data.get("history_note", ""),
        custodian=data.get("custodian", ""),
        catalogue_pid=data.get("catalogue_pid", ""),
    )


def config_to_concept_scheme_v1(
    vocab_config: "config.Vocab",
    rdf_scheme: ConceptSchemeV1 | None = None,
    derived_contributors: str = "",
) -> ConceptSchemeV1:
    """Build ConceptSchemeV1 from idranges.toml config with optional RDF fallback.

    Config values override RDF values. RDF fills gaps for empty config fields.
    Warnings are logged for required fields (vocabulary_iri, title) that are
    empty in both sources.

    Args:
        vocab_config: Vocab configuration from idranges.toml.
        rdf_scheme: Optional ConceptSchemeV1 extracted from RDF (for fallback).
        derived_contributors: Auto-derived contributors from ID range usage.

    Returns:
        ConceptSchemeV1 model instance with merged data.
    """

    def get_field(
        config_val: str, rdf_val: str, field_name: str, required: bool = False
    ) -> str:
        """Get field value, preferring config over RDF."""
        result = config_val.strip() if config_val else ""
        if not result and rdf_val:
            result = rdf_val.strip() if rdf_val else ""
        if not result and required:
            logger.warning(
                "ConceptScheme field '%s' is empty in both config and RDF.", field_name
            )
        return result

    # Get RDF values (or empty defaults)
    rdf = rdf_scheme or ConceptSchemeV1()

    # Determine history_note: config > RDF > auto-generate
    history_note = ""
    if vocab_config.history_note and vocab_config.history_note.strip():
        history_note = vocab_config.history_note.strip()
    elif rdf.history_note:
        history_note = rdf.history_note
    else:
        # Auto-generate from created_date and creator names
        created_date = get_field(
            vocab_config.created_date, rdf.created_date, "created_date"
        )
        creator = get_field(vocab_config.creator, rdf.creator, "creator")
        creator_names = extract_creator_names(creator)
        history_note = generate_history_note(created_date, creator_names)

    return ConceptSchemeV1(
        template_version=TEMPLATE_VERSION,
        vocabulary_iri=get_field(
            vocab_config.vocabulary_iri,
            rdf.vocabulary_iri,
            "vocabulary_iri",
            required=True,
        ),
        prefix=get_field(vocab_config.prefix, rdf.prefix, "prefix"),
        title=get_field(vocab_config.title, rdf.title, "title", required=True),
        description=get_field(vocab_config.description, rdf.description, "description"),
        created_date=get_field(
            vocab_config.created_date, rdf.created_date, "created_date"
        ),
        # Auto-generated fields come from RDF only (not in config)
        modified_date=rdf.modified_date,
        version=rdf.version,
        # Contributor: derived from ID range usage if provided, else from RDF
        contributor=derived_contributors if derived_contributors else rdf.contributor,
        # Multi-line fields from config
        creator=get_field(vocab_config.creator, rdf.creator, "creator"),
        publisher=get_field(vocab_config.publisher, rdf.publisher, "publisher"),
        custodian=get_field(vocab_config.custodian, rdf.custodian, "custodian"),
        # URL fields
        catalogue_pid=get_field(
            vocab_config.catalogue_pid, rdf.catalogue_pid, "catalogue_pid"
        ),
        documentation=get_field(
            vocab_config.documentation, rdf.documentation, "documentation"
        ),
        issue_tracker=get_field(
            vocab_config.issue_tracker, rdf.issue_tracker, "issue_tracker"
        ),
        helpdesk=get_field(vocab_config.helpdesk, rdf.helpdesk, "helpdesk"),
        repository=get_field(vocab_config.repository, rdf.repository, "repository"),
        homepage=get_field(vocab_config.homepage, rdf.homepage, "homepage"),
        conforms_to=get_field(vocab_config.conforms_to, rdf.conforms_to, "conforms_to"),
        # History note: config > RDF > auto-generated
        history_note=history_note,
    )


def rdf_concepts_to_v1(
    concepts_data: dict[str, dict[str, dict]],
    concept_to_collections: dict[str, list[str]],
    concept_to_ordered_collections: dict[str, dict[str, int]] | None = None,
    collections_data: dict[str, dict[str, dict]] | None = None,
    vocab_name: str = "",
    provenance_template: str = "",
    repository_url: str = "",
) -> list[ConceptV1]:
    """Convert extracted concepts to ConceptV1 models.

    Creates one ConceptV1 row per (concept_iri, language) combination.
    The first row for each concept includes: parent_iris, member_of_collections,
    member_of_ordered_collection, source_vocab_iri/license/rights_holder,
    change_note, obsolete_reason, influenced_by_iris, provenance.
    Subsequent rows (other languages) have structural fields empty but include
    editorial_note per language.

    IRI columns (parent_iris, member_of_collections, etc.) are formatted with
    English preferred labels when available: "curie (label)".

    Args:
        concepts_data: Nested dict from extract_concepts_from_rdf.
        concept_to_collections: Mapping from concept IRI to collection IRIs.
        concept_to_ordered_collections: Mapping from concept IRI to
            {ordered_collection_iri: position}.
        collections_data: Optional nested dict from extract_collections_from_rdf,
            used for looking up collection labels.
        vocab_name: Vocabulary name for provenance URL generation.
        provenance_template: Jinja template for provenance URLs.
        repository_url: Repository URL from config for GitHub auto-detection.

    Returns:
        List of ConceptV1 model instances.
    """
    concepts_v1 = []
    concept_to_ordered_collections = concept_to_ordered_collections or {}
    collections_data = collections_data or {}

    # Use curies converter to compress IRIs
    converter = config.curies_converter

    for concept_iri, lang_data in concepts_data.items():
        is_first_row = True

        # Validate deprecation and get provenance URL
        provenance_url = validate_entity_deprecation(
            entity_iri=concept_iri,
            lang_data=lang_data,
            vocab_name=vocab_name,
            provenance_template=provenance_template,
            repository_url=repository_url,
            obsoletion_reason_enum=ConceptObsoletionReason,
            entity_type="concept",
        )

        # Get deprecation info from any language (it's the same for all)
        first_lang_data = next(iter(lang_data.values()), {})
        replaced_by_iri = first_lang_data.get("replaced_by_iri", "")

        for lang, data in lang_data.items():
            # Compress the concept IRI
            concept_iri_display = converter.compress(concept_iri, passthrough=True)

            # Format alternate labels (join with | separator, no spaces around |)
            alt_labels = data.get("alternate_labels", [])
            alt_labels_str = " | ".join(alt_labels) if alt_labels else ""

            # Editorial note is per-language
            editorial_note = data.get("editorial_note", "")

            # Only include structural data in first row
            if is_first_row:
                # Format parent IRIs with labels
                parent_iris = data.get("parent_iris", [])
                parent_iris_strs = [
                    format_iri_with_label(p, concepts_data) for p in parent_iris
                ]
                parent_iris_str = "\n".join(parent_iris_strs)

                # Format member_of_collections with labels (regular collections only)
                collections = concept_to_collections.get(concept_iri, [])
                collections_strs = [
                    format_iri_with_label(c, concepts_data, collections_data)
                    for c in collections
                ]
                member_of_collections_str = "\n".join(collections_strs)

                # Format member_of_ordered_collection with labels (format: collIRI (label) # pos)
                ordered_colls = concept_to_ordered_collections.get(concept_iri, {})
                ordered_parts = []
                for coll_iri, position in ordered_colls.items():
                    coll_display = format_iri_with_label(
                        coll_iri, concepts_data, collections_data
                    )
                    ordered_parts.append(f"{coll_display} # {position}")
                member_of_ordered_collection_str = "\n".join(ordered_parts)

                # Source vocab attribution
                source_vocab_iri = data.get("source_vocab_iri", "")
                source_vocab_license = data.get("source_vocab_license", "")
                source_vocab_rights_holder = data.get("source_vocab_rights_holder", "")

                # Notes
                change_note = data.get("change_note", "")
                replaced_by = (
                    converter.compress(replaced_by_iri, passthrough=True)
                    if replaced_by_iri
                    else ""
                )
                obsolete_reason = string_to_concept_obsoletion_enum(
                    data.get("obsolete_reason", "")
                )

                # Influenced by IRIs with labels
                influenced_iris = data.get("influenced_by_iris", [])
                influenced_iris_strs = [
                    format_iri_with_label(i, concepts_data) for i in influenced_iris
                ]
                influenced_by_iris_str = "\n".join(influenced_iris_strs)

                provenance = provenance_url
                is_first_row = False
            else:
                parent_iris_str = ""
                member_of_collections_str = ""
                member_of_ordered_collection_str = ""
                source_vocab_iri = ""
                source_vocab_license = ""
                source_vocab_rights_holder = ""
                change_note = ""
                replaced_by = ""
                obsolete_reason = None
                influenced_by_iris_str = ""
                provenance = ""

            concepts_v1.append(
                ConceptV1(
                    concept_iri=concept_iri_display,
                    language_code=lang,
                    preferred_label=data.get("preferred_label", ""),
                    definition=data.get("definition", ""),
                    alternate_labels=alt_labels_str,
                    parent_iris=parent_iris_str,
                    member_of_collections=member_of_collections_str,
                    member_of_ordered_collection=member_of_ordered_collection_str,
                    provenance=provenance,
                    change_note=change_note,
                    editorial_note=editorial_note,
                    influenced_by_iris=influenced_by_iris_str,
                    source_vocab_iri=source_vocab_iri,
                    source_vocab_license=source_vocab_license,
                    source_vocab_rights_holder=source_vocab_rights_holder,
                    obsolete_reason=obsolete_reason,
                    replaced_by=replaced_by,
                )
            )

    return concepts_v1


def rdf_collections_to_v1(
    collections_data: dict[str, dict[str, dict]],
    collection_to_parents: dict[str, list[str]],
    concepts_data: dict[str, dict[str, dict]] | None = None,
    vocab_name: str = "",
    provenance_template: str = "",
    repository_url: str = "",
) -> list[CollectionV1]:
    """Convert extracted collections to CollectionV1 models.

    Creates one CollectionV1 row per (collection_iri, language) combination.
    The first row includes: parent_collection_iris, ordered, change_note,
    obsolete_reason, provenance.
    Subsequent rows (other languages) have structural fields empty but include
    editorial_note per language.

    Parent collection IRIs are formatted with English preferred labels when
    available: "curie (label)".

    Args:
        collections_data: Nested dict from extract_collections_from_rdf.
        collection_to_parents: Mapping from collection IRI to parent collection IRIs.
        concepts_data: Optional nested dict for label lookups (unused, for consistency).
        vocab_name: Vocabulary name for provenance URL generation.
        provenance_template: Jinja template for provenance URLs.
        repository_url: Repository URL from config for GitHub auto-detection.

    Returns:
        List of CollectionV1 model instances.
    """
    collections_v1 = []

    converter = config.curies_converter

    for collection_iri, lang_data in collections_data.items():
        is_first_row = True

        # Validate deprecation and get provenance URL
        provenance_url = validate_entity_deprecation(
            entity_iri=collection_iri,
            lang_data=lang_data,
            vocab_name=vocab_name,
            provenance_template=provenance_template,
            repository_url=repository_url,
            obsoletion_reason_enum=CollectionObsoletionReason,
            entity_type="collection",
        )

        # Get deprecation info from any language (it's the same for all)
        first_lang_data = next(iter(lang_data.values()), {})
        replaced_by_iri = first_lang_data.get("replaced_by_iri", "")

        for lang, data in lang_data.items():
            collection_iri_display = converter.compress(
                collection_iri, passthrough=True
            )

            # Editorial note is per-language
            editorial_note = data.get("editorial_note", "")

            if is_first_row:
                # Format parent collection IRIs with labels
                # Parent collections are looked up in collections_data itself
                parents = collection_to_parents.get(collection_iri, [])
                parents_strs = [
                    format_iri_with_label(p, {}, collections_data) for p in parents
                ]
                parent_iris_str = "\n".join(parents_strs)

                # Ordered flag
                ordered = string_to_ordered_enum(data.get("ordered", False))

                # Notes
                change_note = data.get("change_note", "")
                replaced_by = (
                    converter.compress(replaced_by_iri, passthrough=True)
                    if replaced_by_iri
                    else ""
                )
                obsolete_reason = string_to_collection_obsoletion_enum(
                    data.get("obsolete_reason", "")
                )

                provenance = provenance_url
                is_first_row = False
            else:
                parent_iris_str = ""
                ordered = None
                change_note = ""
                replaced_by = ""
                obsolete_reason = None
                provenance = ""

            collections_v1.append(
                CollectionV1(
                    collection_iri=collection_iri_display,
                    language_code=lang,
                    preferred_label=data.get("preferred_label", ""),
                    definition=data.get("definition", ""),
                    parent_collection_iris=parent_iris_str,
                    ordered=ordered,
                    provenance=provenance,
                    change_note=change_note,
                    editorial_note=editorial_note,
                    obsolete_reason=obsolete_reason,
                    replaced_by=replaced_by,
                )
            )

    return collections_v1


def rdf_mappings_to_v1(
    mappings_data: dict[str, dict],
    concepts_data: dict[str, dict[str, dict]] | None = None,
) -> list[MappingV1]:
    """Convert extracted mappings to MappingV1 models.

    The concept_iri column is formatted with the English preferred label
    when available: "curie (label)".

    Args:
        mappings_data: Dict from extract_mappings_from_rdf.
        concepts_data: Optional nested dict for looking up concept labels.

    Returns:
        List of MappingV1 model instances.
    """
    mappings_v1 = []
    concepts_data = concepts_data or {}

    converter = config.curies_converter

    for concept_iri, data in mappings_data.items():
        concept_iri_display = format_iri_with_label(concept_iri, concepts_data)

        # Format each mapping type as space-separated IRIs
        related = " ".join(
            converter.compress(m, passthrough=True)
            for m in data.get("related_matches", [])
        )
        close = " ".join(
            converter.compress(m, passthrough=True)
            for m in data.get("close_matches", [])
        )
        exact = " ".join(
            converter.compress(m, passthrough=True)
            for m in data.get("exact_matches", [])
        )
        narrower = " ".join(
            converter.compress(m, passthrough=True)
            for m in data.get("narrower_matches", [])
        )
        broader = " ".join(
            converter.compress(m, passthrough=True)
            for m in data.get("broader_matches", [])
        )

        mappings_v1.append(
            MappingV1(
                concept_iri=concept_iri_display,
                related_matches=related,
                close_matches=close,
                exact_matches=exact,
                narrower_matches=narrower,
                broader_matches=broader,
            )
        )

    return mappings_v1


def build_prefixes_v1() -> list[PrefixV1]:
    """Build prefix list from the current curies converter.

    Returns:
        List of PrefixV1 model instances.
    """
    prefixes_v1 = []

    for prefix, namespace in config.curies_converter.prefix_map.items():
        prefixes_v1.append(PrefixV1(prefix=prefix, namespace=namespace))

    return prefixes_v1


# =============================================================================
# Excel Export Function
# =============================================================================


def _get_v1_table_row_info(
    model_class: type[BaseModel], title: str = ""
) -> tuple[int, int]:
    """Get header and data start rows for v1.0 vocabulary tables.

    Uses the xlsx-pydantic infrastructure to calculate row positions
    based on the standard v1.0 config (with requiredness shown).

    Args:
        model_class: The Pydantic model class (e.g., ConceptV1).
        title: Optional title for the table (affects row positions).

    Returns:
        Tuple of (header_row, data_start_row).
    """
    config = XLSXTableConfig(
        title=title,
        metadata_visibility=MetadataToggleConfig(requiredness=MetadataVisibility.SHOW),
    )
    field_analyses = XLSXFieldAnalyzer.analyze_model(model_class)
    fields = list(field_analyses.values())
    row_calculator = XLSXRowCalculator(config)

    return row_calculator.get_header_row(fields), row_calculator.get_data_start_row(
        fields
    )


def _find_column_by_header(worksheet, header_name: str, header_row: int) -> int | None:
    """Find column index by header name in a worksheet.

    Args:
        worksheet: openpyxl worksheet object.
        header_name: The header text to search for.
        header_row: Row number where headers are located.

    Returns:
        Column index (1-based) if found, None otherwise.
    """
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value == header_name:
            return col_idx
    return None


def _extract_entity_id_from_provenance_url(url: str) -> str:
    """Extract entity ID from a provenance URL.

    The URL typically ends with "<entity_id>.ttl" or similar.

    Args:
        url: Provenance URL like "https://github.com/.../blame/.../0000001.ttl"

    Returns:
        The entity ID (e.g., "0000001") or empty string if not found.
    """
    # Remove trailing slashes and get the last path segment
    path = url.rstrip("/").split("/")[-1]
    # Remove file extension if present
    if "." in path:
        path = path.rsplit(".", 1)[0]
    return path


def _add_provenance_hyperlinks(
    workbook, sheet_name: str, model_class: type[BaseModel]
) -> None:
    """Add hyperlinks to provenance cells in a sheet.

    Finds the "Provenance (read-only)" column and sets cell.hyperlink for any
    cell value that looks like a URL (starts with "http"). The cell value is
    replaced with a friendly display text "git blame for <entity_id>".

    Args:
        workbook: openpyxl Workbook object.
        sheet_name: Name of the sheet to process.
        model_class: The Pydantic model class for row calculation.
    """
    if sheet_name not in workbook.sheetnames:
        return

    header_row, data_start_row = _get_v1_table_row_info(model_class, title=sheet_name)

    ws = workbook[sheet_name]
    prov_col = _find_column_by_header(ws, "Provenance (read-only)", header_row)

    if prov_col is None:
        return

    for row_idx in range(data_start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=prov_col)
        value = cell.value
        if value and isinstance(value, str) and value.startswith("http"):
            entity_id = _extract_entity_id_from_provenance_url(value)
            cell.hyperlink = value
            cell.value = f"git blame for {entity_id}" if entity_id else "git blame"
            _apply_hyperlink_style(cell)


def _add_concept_iri_hyperlinks(
    workbook, sheet_name: str, model_class: type[BaseModel]
) -> None:
    """Add hyperlinks to concept IRI cells in a sheet.

    Finds the "Concept IRI*" column and sets cell.hyperlink by expanding
    the CURIE to a full IRI.

    Args:
        workbook: openpyxl Workbook object.
        sheet_name: Name of the sheet to process.
        model_class: The Pydantic model class for row calculation.
    """
    if sheet_name not in workbook.sheetnames:
        return

    header_row, data_start_row = _get_v1_table_row_info(model_class, title=sheet_name)

    ws = workbook[sheet_name]
    iri_col = _find_column_by_header(ws, "Concept IRI*", header_row)

    if iri_col is None:
        return

    converter = config.curies_converter

    for row_idx in range(data_start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=iri_col)
        value = cell.value
        if value and isinstance(value, str):
            # Value might be "curie (label)" format - extract just the curie
            curie = value.split(" (")[0].strip()
            # Try to expand the CURIE to full IRI
            full_iri = converter.expand(curie)
            if full_iri and full_iri != curie:  # expand returns original if no match
                cell.hyperlink = full_iri
                _apply_hyperlink_style(cell)


def _apply_hyperlink_style(cell) -> None:
    """Apply standard hyperlink styling to a cell (blue, underlined).

    Args:
        cell: openpyxl cell object.
    """
    cell.font = Font(
        name=cell.font.name,
        size=cell.font.size,
        bold=cell.font.bold,
        italic=cell.font.italic,
        color="0563C1",
        underline="single",
    )


def _add_entity_iri_hyperlinks(
    workbook, sheet_name: str, column_header: str, model_class: type[BaseModel]
) -> None:
    """Add hyperlinks to entity IRI cells (Concepts or Collections).

    Finds the specified column and sets cell.hyperlink by expanding
    the CURIE to a full IRI. The display text (CURIE) is preserved.

    Args:
        workbook: openpyxl Workbook object.
        sheet_name: Name of the sheet to process.
        column_header: Header text of the IRI column (e.g., "Concept IRI*").
        model_class: The Pydantic model class for row calculation.
    """
    if sheet_name not in workbook.sheetnames:
        return

    header_row, data_start_row = _get_v1_table_row_info(model_class, title=sheet_name)

    ws = workbook[sheet_name]
    iri_col = _find_column_by_header(ws, column_header, header_row)

    if iri_col is None:
        return

    converter = config.curies_converter

    for row_idx in range(data_start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=iri_col)
        value = cell.value
        if value and isinstance(value, str):
            curie = value.strip()
            # Try to expand the CURIE to full IRI
            full_iri = converter.expand(curie)
            if full_iri and full_iri != curie:
                cell.hyperlink = full_iri
                _apply_hyperlink_style(cell)


def _add_vocabulary_iri_hyperlink(workbook, sheet_name: str = "Concept Scheme") -> None:
    """Add hyperlink to the Vocabulary IRI field in Concept Scheme sheet.

    The Concept Scheme sheet uses key-value format where the Vocabulary IRI
    is in a specific cell. This function finds and hyperlinks it.

    Args:
        workbook: openpyxl Workbook object.
        sheet_name: Name of the Concept Scheme sheet.
    """
    if sheet_name not in workbook.sheetnames:
        return

    ws = workbook[sheet_name]

    # Find the row with "Vocabulary IRI" key (column A) and get value from column B
    for row_idx in range(1, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=1)
        if key_cell.value == "Vocabulary IRI":
            value_cell = ws.cell(row=row_idx, column=2)
            iri_value = value_cell.value
            if (
                iri_value
                and isinstance(iri_value, str)
                and iri_value.startswith("http")
            ):
                # Make the IRI clickable
                value_cell.hyperlink = iri_value.strip()
                _apply_hyperlink_style(value_cell)
            break


def export_vocabulary_v1(
    concept_scheme: ConceptSchemeV1,
    concepts: list[ConceptV1],
    collections: list[CollectionV1],
    mappings: list[MappingV1],
    prefixes: list[PrefixV1],
    output_path: Path,
    id_ranges: list[IDRangeInfoV1] | None = None,
    template_path: Path | None = None,
) -> None:
    """Export v1.0 vocabulary data to Excel.

    Uses export_to_xlsx() for each sheet.

    Args:
        concept_scheme: ConceptSchemeV1 model instance.
        concepts: List of ConceptV1 model instances.
        collections: List of CollectionV1 model instances.
        mappings: List of MappingV1 model instances.
        prefixes: List of PrefixV1 model instances.
        output_path: Path to save the Excel file.
        id_ranges: Optional list of IDRangeInfoV1 model instances. If provided,
                  an "ID Ranges" sheet is added showing contributor allocations.
        template_path: Optional path to an xlsx template file. If provided,
                      the template's sheets are preserved and placed before
                      the auto-generated vocabulary sheets.
    """
    # If template provided, copy it to output location first
    if template_path is not None:
        shutil.copy(template_path, output_path)
        logger.debug("Copied template from %s to %s", template_path, output_path)

    # 1. Concept Scheme (key-value format, read-only)
    kv_config = XLSXKeyValueConfig(
        title=CONCEPT_SCHEME_SHEET_TITLE,
        table_style="TableStyleMedium16",
    )
    export_to_xlsx(
        concept_scheme,
        output_path,
        format_type="keyvalue",
        config=kv_config,
        sheet_name=CONCEPT_SCHEME_SHEET_NAME,
    )

    # 2. Concepts (table format)
    if not concepts:
        # Export empty concepts sheet with just headers
        concepts = [
            ConceptV1(
                concept_iri="", language_code="", preferred_label="", definition=""
            )
        ]

    export_to_xlsx(
        concepts,
        output_path,
        format_type="table",
        config=CONCEPTS_EXPORT_CONFIG,
        sheet_name=CONCEPTS_SHEET_NAME,
    )

    # 3. Collections (table format)
    if not collections:
        # Empty placeholder row - all values empty/None so it will be skipped on import
        collections = [
            CollectionV1(
                collection_iri="",
                language_code="",
                preferred_label="",
                definition="",
                ordered=None,  # Override default to ensure row is empty
            )
        ]

    export_to_xlsx(
        collections,
        output_path,
        format_type="table",
        config=COLLECTIONS_EXPORT_CONFIG,
        sheet_name=COLLECTIONS_SHEET_NAME,
    )

    # 4. Mappings (table format)
    if not mappings:
        mappings = [MappingV1(concept_iri="")]

    export_to_xlsx(
        mappings,
        output_path,
        format_type="table",
        config=MAPPINGS_EXPORT_CONFIG,
        sheet_name=MAPPINGS_SHEET_NAME,
    )
    # 5. ID Ranges (table format, read-only)
    if not id_ranges:
        id_ranges = [IDRangeInfoV1()]

    export_to_xlsx(
        id_ranges,
        output_path,
        format_type="table",
        config=ID_RANGES_EXPORT_CONFIG,
        sheet_name=ID_RANGES_SHEET_NAME,
    )

    # 6. Prefixes (table format, read-only)
    export_to_xlsx(
        prefixes,
        output_path,
        format_type="table",
        config=PREFIXES_EXPORT_CONFIG,
        sheet_name=PREFIXES_SHEET_NAME,
    )

    # Post-processing: reorder sheets, add hyperlinks, set freeze panes
    wb = load_workbook(output_path)

    # Get template sheet names for ordering (if template was used)
    template_sheet_names = None
    if template_path is not None:
        template_sheet_names = get_template_sheet_names(template_path)

    reorder_sheets_with_template(wb, template_sheet_names)

    # Add hyperlinks to entity IRI columns (Concepts and Collections sheets)
    _add_entity_iri_hyperlinks(wb, CONCEPTS_SHEET_NAME, "Concept IRI*", ConceptV1)
    _add_entity_iri_hyperlinks(
        wb, COLLECTIONS_SHEET_NAME, "Collection IRI", CollectionV1
    )

    # Add hyperlinks to provenance columns (Concepts and Collections sheets)
    _add_provenance_hyperlinks(wb, CONCEPTS_SHEET_NAME, ConceptV1)
    _add_provenance_hyperlinks(wb, COLLECTIONS_SHEET_NAME, CollectionV1)

    # Add hyperlinks to concept IRI column in Mappings sheet (with labels)
    _add_concept_iri_hyperlinks(wb, MAPPINGS_SHEET_NAME, MappingV1)

    # Set freeze panes for Concepts sheet (dynamically calculated)
    if CONCEPTS_SHEET_NAME in wb.sheetnames:
        _, data_start_row = _get_v1_table_row_info(ConceptV1, title=CONCEPTS_SHEET_NAME)
        wb[CONCEPTS_SHEET_NAME].freeze_panes = f"A{data_start_row}"

    # Add hyperlink to Vocabulary IRI in Concept Scheme
    _add_vocabulary_iri_hyperlink(wb, CONCEPT_SCHEME_SHEET_NAME)

    wb.save(output_path)
    wb.close()


# =============================================================================
# Main Converter Function
# =============================================================================


def rdf_to_excel_v1(
    file_to_convert_path: Path,
    output_file_path: Path | None = None,
    vocab_config: "config.Vocab | None" = None,
    template_path: Path | None = None,
) -> Path:
    """Convert an RDF vocabulary to v1.0 Excel template.

    If vocab_config is provided, ConceptScheme metadata is merged from config
    (config values override RDF, RDF fills gaps). The "Concept Scheme (read-only)"
    sheet is populated from this merged data.

    Args:
        file_to_convert_path: Path to the RDF file (.ttl, .rdf, etc.).
        output_file_path: Optional path for output Excel file.
                         Defaults to same name with .xlsx extension.
        vocab_config: Optional Vocab config from idranges.toml. If provided,
                     scheme metadata is merged (config overrides RDF).
        template_path: Optional path to an xlsx template file. If provided,
                      the template's sheets are preserved and placed before
                      the auto-generated vocabulary sheets.

    Returns:
        Path to the generated Excel file.

    Raises:
        ValueError: If the input file is not a supported RDF format.
    """
    if file_to_convert_path.suffix.lower() not in RDF_FILE_ENDINGS:
        msg = (
            "Files for conversion must end with one of the RDF file formats: "
            f"'{', '.join(RDF_FILE_ENDINGS.keys())}'"
        )
        raise ValueError(msg)

    vocab_name = file_to_convert_path.stem.lower()

    # Set up curies converter for this vocabulary
    config.curies_converter = config.CURIES_CONVERTER_MAP.get(
        vocab_name, config.curies_converter
    )

    # Parse RDF file
    logger.info("Parsing RDF file: %s", file_to_convert_path)
    graph = Graph().parse(
        str(file_to_convert_path),
        format=RDF_FILE_ENDINGS[file_to_convert_path.suffix.lower()],
    )

    # Extract data from RDF
    logger.debug("Extracting concept scheme...")
    cs_data = extract_concept_scheme_from_rdf(graph)

    logger.debug("Extracting concepts...")
    concepts_data = extract_concepts_from_rdf(graph)

    logger.debug("Extracting collections...")
    collections_data = extract_collections_from_rdf(graph)

    logger.debug("Extracting mappings...")
    mappings_data = extract_mappings_from_rdf(graph)

    logger.debug("Building concept-to-collections map...")
    concept_to_collections = build_concept_to_collections_map(graph)

    logger.debug("Building concept-to-ordered-collections map...")
    concept_to_ordered_collections = build_concept_to_ordered_collections_map(graph)

    logger.debug("Building collection hierarchy map...")
    collection_to_parents = build_collection_hierarchy_map(graph)

    # Convert to v1.0 models
    logger.debug("Converting to v1.0 models...")

    # Build ConceptScheme - either from config (merged with RDF) or RDF only
    if vocab_config is not None:
        logger.debug(
            "Using vocab config for ConceptScheme metadata (config overrides RDF)."
        )
        rdf_scheme = rdf_concept_scheme_to_v1(cs_data)
        concept_scheme_v1 = config_to_concept_scheme_v1(vocab_config, rdf_scheme)
    else:
        concept_scheme_v1 = rdf_concept_scheme_to_v1(cs_data)

    # Get provenance URL config
    provenance_template = vocab_config.provenance_url_template if vocab_config else ""
    repository_url = vocab_config.repository if vocab_config else ""

    concepts_v1 = rdf_concepts_to_v1(
        concepts_data,
        concept_to_collections,
        concept_to_ordered_collections,
        collections_data,  # For looking up collection labels
        vocab_name,
        provenance_template,
        repository_url,
    )
    collections_v1 = rdf_collections_to_v1(
        collections_data,
        collection_to_parents,
        concepts_data,  # For consistency (unused, collections reference other collections)
        vocab_name,
        provenance_template,
        repository_url,
    )
    mappings_v1 = rdf_mappings_to_v1(mappings_data, concepts_data)
    prefixes_v1 = build_prefixes_v1()

    # Build ID range info and derive contributors if vocab_config is provided
    id_ranges_v1: list[IDRangeInfoV1] | None = None
    if vocab_config is not None and vocab_config.id_range:
        logger.debug("Building ID range info...")
        used_ids = extract_used_ids(concepts_v1, collections_v1, vocab_config)
        id_ranges_v1 = build_id_range_info(vocab_config, used_ids)
        # Derive contributors from ID range usage
        logger.debug("Deriving contributors from ID range usage...")
        derived_contributors = derive_contributors(vocab_config, used_ids)
        if derived_contributors:
            concept_scheme_v1 = concept_scheme_v1.model_copy(
                update={"contributor": derived_contributors}
            )

    # Determine output path
    if output_file_path is None:
        output_file_path = file_to_convert_path.with_suffix(".xlsx")

    # Export to Excel
    logger.info("Exporting to Excel: %s", output_file_path)
    export_vocabulary_v1(
        concept_scheme_v1,
        concepts_v1,
        collections_v1,
        mappings_v1,
        prefixes_v1,
        output_file_path,
        id_ranges=id_ranges_v1,
        template_path=template_path,
    )

    logger.info("Conversion complete: %s", output_file_path)
    return output_file_path


# =============================================================================
# XLSX -> RDF Conversion (Step 3)
# =============================================================================

# --- Aggregated Data Structures ---

# TODO Do we need these dataclasses, or can we just use pydantic models?


@dataclass
class AggregatedConcept:
    """Aggregated concept data from multiple XLSX rows (one per language)."""

    iri: str
    pref_labels: dict[str, str] = field(default_factory=dict)  # {lang: label}
    definitions: dict[str, str] = field(default_factory=dict)  # {lang: definition}
    alt_labels: dict[str, list[str]] = field(default_factory=dict)  # {lang: [labels]}
    editorial_notes: dict[str, str] = field(default_factory=dict)  # {lang: note}
    parent_iris: list[str] = field(default_factory=list)
    member_of_collections: list[str] = field(default_factory=list)
    # {ordered_collection_iri: position}
    ordered_collection_positions: dict[str, int] = field(default_factory=dict)
    source_vocab_iri: str = ""
    source_vocab_license: str = ""
    source_vocab_rights_holder: str = ""
    change_note: str = ""
    obsolete_reason: str = ""
    influenced_by_iris: list[str] = field(default_factory=list)
    replaced_by_iri: str = ""


@dataclass
class AggregatedCollection:
    """Aggregated collection data from multiple XLSX rows (one per language)."""

    iri: str
    pref_labels: dict[str, str] = field(default_factory=dict)  # {lang: label}
    definitions: dict[str, str] = field(default_factory=dict)  # {lang: definition}
    editorial_notes: dict[str, str] = field(default_factory=dict)  # {lang: note}
    parent_collection_iris: list[str] = field(default_factory=list)
    ordered: bool = False
    change_note: str = ""
    obsolete_reason: str = ""
    replaced_by_iri: str = ""


# --- XLSX Reading Functions ---


def read_concept_scheme_v1(filepath: Path) -> ConceptSchemeV1:
    """Read ConceptScheme data from v1.0 XLSX file.

    Args:
        filepath: Path to the XLSX file.

    Returns:
        ConceptSchemeV1 model instance.
    """
    return import_from_xlsx(
        filepath,
        ConceptSchemeV1,
        format_type="keyvalue",
        sheet_name=CONCEPT_SCHEME_SHEET_NAME,
    )


def read_concepts_v1(filepath: Path) -> list[ConceptV1]:
    """Read Concepts from v1.0 XLSX file.

    Args:
        filepath: Path to the XLSX file.

    Returns:
        List of ConceptV1 model instances (one per row).
    """
    return import_from_xlsx(
        filepath,
        ConceptV1,
        format_type="table",
        config=CONCEPTS_READ_CONFIG,
        sheet_name=CONCEPTS_SHEET_NAME,
    )


def read_collections_v1(filepath: Path) -> list[CollectionV1]:
    """Read Collections from v1.0 XLSX file.

    Args:
        filepath: Path to the XLSX file.

    Returns:
        List of CollectionV1 model instances (one per row).
    """
    return import_from_xlsx(
        filepath,
        CollectionV1,
        format_type="table",
        config=COLLECTIONS_READ_CONFIG,
        sheet_name=COLLECTIONS_SHEET_NAME,
    )


def read_mappings_v1(filepath: Path) -> list[MappingV1]:
    """Read Mappings from v1.0 XLSX file.

    Args:
        filepath: Path to the XLSX file.

    Returns:
        List of MappingV1 model instances (one per row).
    """
    return import_from_xlsx(
        filepath,
        MappingV1,
        format_type="table",
        config=MAPPINGS_READ_CONFIG,
        sheet_name=MAPPINGS_SHEET_NAME,
    )


def read_prefixes_v1(filepath: Path) -> list[PrefixV1]:
    """Read Prefixes from v1.0 XLSX file.

    Args:
        filepath: Path to the XLSX file.

    Returns:
        List of PrefixV1 model instances.
    """
    return import_from_xlsx(
        filepath,
        PrefixV1,
        format_type="table",
        config=PREFIXES_READ_CONFIG,
        sheet_name=PREFIXES_SHEET_NAME,
    )


# --- CURIE Handling ---


def build_curies_converter_from_prefixes(prefixes: list[PrefixV1]) -> curies.Converter:
    """Build a curies converter from prefix list.

    Args:
        prefixes: List of PrefixV1 with prefix and namespace.

    Returns:
        Configured curies.Converter.
    """
    records = [
        curies.Record(prefix=p.prefix, uri_prefix=p.namespace)
        for p in prefixes
        if p.prefix and p.namespace
    ]
    return curies.Converter(records)


def strip_label_from_iri(iri_with_label: str) -> str:
    """Strip the optional label suffix from an IRI string.

    The label format is "curie (label)" or "iri (label)".
    Returns just the curie/iri portion.

    Args:
        iri_with_label: String like "voc4cat:0000016 (catalyst form)" or just "voc4cat:0000016".

    Returns:
        The IRI/CURIE portion without the label.
    """
    # Strip whitespace first
    iri_with_label = iri_with_label.strip()
    # Check if there's a label suffix " (something)"
    if " (" in iri_with_label and iri_with_label.endswith(")"):
        return iri_with_label.split(" (")[0].strip()
    return iri_with_label


def expand_iri_list(iri_string: str, converter: curies.Converter) -> list[str]:
    """Expand a newline-separated string of CURIEs/IRIs to list of full IRIs.

    IRIs may include optional labels in format "curie (label)" which are
    stripped before expansion.

    Args:
        iri_string: Newline-separated CURIEs or IRIs, possibly with labels.
        converter: Curies converter to use.

    Returns:
        List of full IRI strings.
    """
    if not iri_string or not iri_string.strip():
        return []
    results = []
    for ipart in iri_string.split("\n"):
        part = ipart.strip()
        if not part:
            continue
        # Strip label if present
        iri_only = strip_label_from_iri(part)
        results.append(expand_curie(iri_only, converter))
    return results


# --- Aggregation Functions ---


def parse_ordered_collection_positions(
    position_string: str, converter: curies.Converter
) -> dict[str, int]:
    """Parse ordered collection position string to dict.

    Format: "collIRI (label) # pos" or "collIRI # pos", newline-separated.
    Labels are optional and stripped before IRI expansion.

    Args:
        position_string: String like "ex:coll1 (Label 1) # 1\\nex:coll2 (Label 2) # 2"
        converter: Curies converter for IRI expansion.

    Returns:
        Dict mapping collection IRI to position (1-indexed).
    """
    if not position_string or not position_string.strip():
        return {}

    result: dict[str, int] = {}

    # Split by newline - each entry is one line
    for iline in position_string.split("\n"):
        line = iline.strip()
        if not line or "#" not in line:
            continue

        # Split by "#" to separate IRI (with optional label) from position
        parts = line.split("#")
        if len(parts) < 2:  # noqa: PLR2004
            continue

        # Left part is "collIRI (label)" or just "collIRI"
        coll_with_label = parts[0].strip()
        coll_curie = strip_label_from_iri(coll_with_label)

        # Right part is the position
        try:
            position = int(parts[1].strip())
        except ValueError:
            continue

        coll_iri = expand_curie(coll_curie, converter)
        result[coll_iri] = position

    return result


def aggregate_concepts(
    concept_rows: list[ConceptV1], converter: curies.Converter
) -> dict[str, AggregatedConcept]:
    """Aggregate multi-row concept data into single AggregatedConcept per IRI.

    The v1.0 template has one row per (concept_iri, language). This function
    merges them back:
    - First row for each IRI contains structural data (parent_iris, etc.)
    - All rows contribute language-specific data (pref_label, definition,
      alt_labels, editorial_note)

    Also validates deprecation consistency and auto-adds "OBSOLETE " prefix
    to English prefLabel if needed.

    Args:
        concept_rows: List of ConceptV1 from XLSX.
        converter: Curies converter for IRI expansion.

    Returns:
        Dict mapping full IRI to AggregatedConcept.
    """
    concepts: dict[str, AggregatedConcept] = {}

    for row in concept_rows:
        # Skip empty rows
        if not row.concept_iri:
            continue

        # Strip label if present (format: "curie (label)")
        concept_iri_only = strip_label_from_iri(row.concept_iri)
        iri = expand_curie(concept_iri_only, converter)

        if iri not in concepts:
            # First row for this concept - create new entry with structural data
            change_note = row.change_note or ""
            replaced_by_iri = (
                expand_curie(row.replaced_by, converter) if row.replaced_by else ""
            )

            concepts[iri] = AggregatedConcept(
                iri=iri,
                parent_iris=expand_iri_list(row.parent_iris, converter),
                member_of_collections=expand_iri_list(
                    row.member_of_collections, converter
                ),
                ordered_collection_positions=parse_ordered_collection_positions(
                    row.member_of_ordered_collection, converter
                ),
                source_vocab_iri=expand_curie(row.source_vocab_iri, converter)
                if row.source_vocab_iri
                else "",
                source_vocab_license=row.source_vocab_license or "",
                source_vocab_rights_holder=row.source_vocab_rights_holder or "",
                change_note=change_note,
                obsolete_reason=row.obsolete_reason.value
                if row.obsolete_reason
                else "",
                influenced_by_iris=expand_iri_list(row.influenced_by_iris, converter),
                replaced_by_iri=replaced_by_iri,
            )

        concept = concepts[iri]

        # Add language-specific data
        lang = row.language_code or "en"
        if row.preferred_label:
            concept.pref_labels[lang] = row.preferred_label
        if row.definition:
            concept.definitions[lang] = row.definition
        if row.alternate_labels:
            # Split by "|" separator (with or without spaces), trim each label
            labels = [lbl.strip() for lbl in row.alternate_labels.split("|")]
            labels = [lbl for lbl in labels if lbl]  # Remove empty strings
            concept.alt_labels[lang] = labels
        if row.editorial_note:
            concept.editorial_notes[lang] = row.editorial_note

    # Second pass: validate deprecation for English prefLabels
    for iri, concept in concepts.items():
        en_pref_label = concept.pref_labels.get("en", "")
        if en_pref_label:
            is_deprecated = bool(concept.obsolete_reason)
            corrected_label, errors = validate_deprecation(
                pref_label=en_pref_label,
                is_deprecated=is_deprecated,
                history_note=concept.obsolete_reason,  # Already a string
                valid_reasons=[e.value for e in ConceptObsoletionReason],
                entity_iri=iri,
                entity_type="concept",
            )
            for error in errors:
                logger.error(error)
            # Update the English prefLabel
            concept.pref_labels["en"] = corrected_label

    return concepts


def aggregate_collections(
    collection_rows: list[CollectionV1], converter: curies.Converter
) -> dict[str, AggregatedCollection]:
    """Aggregate multi-row collection data into single AggregatedCollection per IRI.

    The v1.0 template has one row per (collection_iri, language). This function
    merges them back:
    - First row for each IRI contains structural data (parent_iris, ordered, etc.)
    - All rows contribute language-specific data (pref_label, definition,
      editorial_note)

    Also validates deprecation consistency and auto-adds "OBSOLETE " prefix
    to English prefLabel if needed.

    Args:
        collection_rows: List of CollectionV1 from XLSX.
        converter: Curies converter for IRI expansion.

    Returns:
        Dict mapping full IRI to AggregatedCollection.
    """
    collections: dict[str, AggregatedCollection] = {}

    for row in collection_rows:
        # Skip empty rows
        if not row.collection_iri:
            continue

        iri = expand_curie(row.collection_iri, converter)

        if iri not in collections:
            # First row for this collection - create new entry with structural data
            # Parse ordered flag (now an enum)
            is_ordered = row.ordered == OrderedChoice.YES if row.ordered else False

            change_note = row.change_note or ""
            replaced_by_iri = (
                expand_curie(row.replaced_by, converter) if row.replaced_by else ""
            )

            collections[iri] = AggregatedCollection(
                iri=iri,
                parent_collection_iris=expand_iri_list(
                    row.parent_collection_iris, converter
                ),
                ordered=is_ordered,
                change_note=change_note,
                obsolete_reason=row.obsolete_reason.value
                if row.obsolete_reason
                else "",
                replaced_by_iri=replaced_by_iri,
            )

        collection = collections[iri]

        # Add language-specific data
        lang = row.language_code or "en"
        if row.preferred_label:
            collection.pref_labels[lang] = row.preferred_label
        if row.definition:
            collection.definitions[lang] = row.definition
        if row.editorial_note:
            collection.editorial_notes[lang] = row.editorial_note

    # Second pass: validate deprecation for English prefLabels
    for iri, collection in collections.items():
        en_pref_label = collection.pref_labels.get("en", "")
        if en_pref_label:
            is_deprecated = bool(collection.obsolete_reason)
            corrected_label, errors = validate_deprecation(
                pref_label=en_pref_label,
                is_deprecated=is_deprecated,
                history_note=collection.obsolete_reason,  # Already a string
                valid_reasons=[e.value for e in CollectionObsoletionReason],
                entity_iri=iri,
                entity_type="collection",
            )
            for error in errors:
                logger.error(error)
            # Update the English prefLabel
            collection.pref_labels["en"] = corrected_label

    return collections


# --- Inverse Relationship Builders ---


def build_collection_members_from_concepts(
    concepts: dict[str, AggregatedConcept],
) -> dict[str, list[str]]:
    """Build collection -> members map from concept membership data.

    Inverts the concept.member_of_collections relationship.

    Args:
        concepts: Dict of aggregated concepts.

    Returns:
        Dict: {collection_iri: [member_concept_iris]}
    """
    collection_members: dict[str, list[str]] = defaultdict(list)

    for concept_iri, concept in concepts.items():
        for collection_iri in concept.member_of_collections:
            collection_members[collection_iri].append(concept_iri)

    return dict(collection_members)


def build_narrower_map(
    concepts: dict[str, AggregatedConcept],
) -> dict[str, list[str]]:
    """Build parent -> children (narrower) map from broader relationships.

    Inverts the concept.parent_iris (broader) relationship.

    Args:
        concepts: Dict of aggregated concepts.

    Returns:
        Dict: {parent_iri: [child_iris]}
    """
    narrower: dict[str, list[str]] = defaultdict(list)

    for concept_iri, concept in concepts.items():
        for parent_iri in concept.parent_iris:
            narrower[parent_iri].append(concept_iri)

    return dict(narrower)


def build_ordered_collection_members(
    concepts: dict[str, AggregatedConcept],
) -> dict[str, list[str]]:
    """Build ordered collection -> ordered members list from concept positions.

    Uses the concept.ordered_collection_positions to build ordered member lists.

    Args:
        concepts: Dict of aggregated concepts.

    Returns:
        Dict: {ordered_collection_iri: [member_iris_in_order]}
        Members are sorted by their position values.
    """
    # Collect members with positions
    collection_members: dict[str, list[tuple[int, str]]] = defaultdict(list)

    for concept_iri, concept in concepts.items():
        for coll_iri, position in concept.ordered_collection_positions.items():
            collection_members[coll_iri].append((position, concept_iri))

    # Sort by position and return just the IRIs
    result: dict[str, list[str]] = {}
    for coll_iri, members in collection_members.items():
        members.sort(key=lambda x: x[0])  # Sort by position
        result[coll_iri] = [iri for _, iri in members]

    return result


def build_collection_hierarchy_members(
    collections: dict[str, "AggregatedCollection"],
) -> dict[str, list[str]]:
    """Build parent collection -> child collections map from hierarchy data.

    Inverts the collection.parent_collection_iris relationship to get
    which collections are members of other collections.

    Args:
        collections: Dict of aggregated collections.

    Returns:
        Dict: {parent_collection_iri: [child_collection_iris]}
    """
    hierarchy_members: dict[str, list[str]] = defaultdict(list)

    for child_iri, collection in collections.items():
        for parent_iri in collection.parent_collection_iris:
            hierarchy_members[parent_iri].append(child_iri)

    return dict(hierarchy_members)


# --- Identifier Extraction ---


def extract_identifier(iri: str, id_pattern: "re.Pattern[str]") -> str:
    """Extract local ID from IRI for dcterms:identifier.

    Uses the vocabulary's ID pattern to extract digits from the end of the IRI.

    For 'https://example.org/vocab_0000004' with 7-digit pattern -> '0000004'
    For 'https://example.org/0000004' with 7-digit pattern -> '0000004'

    Args:
        iri: Full IRI string.
        id_pattern: Compiled regex pattern with 'identifier' named group.

    Returns:
        Local ID string (digits only) suitable for dcterms:identifier,
        or empty string if extraction fails.
    """
    match = id_pattern.search(iri)
    if match:
        return match.group("identifier")

    logger.error(
        "Could not extract ID from IRI using pattern %s: %s", id_pattern.pattern, iri
    )
    return ""


# --- RDF Graph Building Functions ---


def parse_name_url(line: str) -> tuple[str, str]:
    """Parse a string in format 'Name with spaces https://url' into (name, url).

    Splits from the right at the first space before the URL.

    Args:
        line: String in format "<name> <URL>" or just "<URL>".

    Returns:
        Tuple of (name, url). Name is empty string if only URL provided.
    """
    line = line.strip()
    if not line:
        return "", ""

    # Split from right: "Name with spaces https://url" -> ["Name with spaces", "https://url"]
    parts = line.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].startswith("http"):  # noqa: PLR2004
        return parts[0], parts[1]
    if line.startswith("http"):
        # URL only, no name
        return "", line
    # No URL found, treat entire line as name
    return line, ""


def is_orcid_url(url: str) -> bool:
    """Check if URL is an ORCID identifier."""
    return "orcid.org" in url.lower()


def is_ror_url(url: str) -> bool:
    """Check if URL is a ROR identifier."""
    return "ror.org" in url.lower()


def build_organization_graph(org_iri: str, name: str = "") -> Graph:
    """Build RDF graph for an Organization.

    Args:
        org_iri: IRI of the organization (creator or publisher).
        name: Name of the organization. If not provided, extracts from IRI.

    Returns:
        Graph with Organization triples.
    """
    g = Graph()
    org = URIRef(org_iri)

    g.add((org, RDF.type, SDO.Organization))

    # Use provided name or fall back to extracting from IRI
    org_name = name if name else org_iri
    g.add((org, SDO.name, Literal(org_name)))
    g.add((org, SDO.url, Literal(org_iri, datatype=XSD.anyURI)))

    return g


def build_person_graph(person_iri: str, name: str = "") -> Graph:
    """Build RDF graph for a Person.

    Args:
        person_iri: IRI of the person (e.g., ORCID URL).
        name: Name of the person. If not provided, uses IRI as fallback.

    Returns:
        Graph with Person triples.
    """
    g = Graph()
    person = URIRef(person_iri)

    g.add((person, RDF.type, SDO.Person))

    # Use provided name or fall back to IRI
    person_name = name if name else person_iri
    g.add((person, SDO.name, Literal(person_name)))
    g.add((person, SDO.url, Literal(person_iri, datatype=XSD.anyURI)))

    return g


def build_entity_graph(url: str, name: str, field_type: str) -> Graph:
    """Build Person or Organization graph based on field type and URL pattern.

    Type determination rules:
    - ORCID URL -> always Person
    - ROR URL -> always Organization
    - publisher field (no pattern match) -> Organization
    - creator/contributor/custodian (no pattern match) -> Person

    Args:
        url: The entity IRI.
        name: Name of the entity.
        field_type: One of "publisher", "creator", "contributor", "custodian".

    Returns:
        Graph with Person or Organization triples.
    """
    if is_orcid_url(url):
        return build_person_graph(url, name)
    if is_ror_url(url) or field_type == "publisher":
        return build_organization_graph(url, name)
    # creator, contributor, custodian default to Person
    return build_person_graph(url, name)


def build_concept_scheme_graph(
    cs: ConceptSchemeV1,
    concepts: dict[str, AggregatedConcept],
    id_pattern: "re.Pattern[str] | None",
) -> Graph:
    """Build RDF graph for ConceptScheme.

    Args:
        cs: ConceptSchemeV1 data.
        concepts: Aggregated concepts (to compute hasTopConcept).
        id_pattern: Compiled regex pattern for extracting IDs.

    Returns:
        Graph with ConceptScheme triples.
    """
    g = Graph()
    scheme_iri = URIRef(cs.vocabulary_iri)

    # Type
    g.add((scheme_iri, RDF.type, SKOS.ConceptScheme))

    # Identifier - use catalogue_pid if provided, otherwise extract from vocabulary IRI
    if not cs.catalogue_pid and id_pattern:
        identifier = extract_identifier(cs.vocabulary_iri, id_pattern)
        if identifier:
            g.add(
                (
                    scheme_iri,
                    DCTERMS.identifier,
                    Literal(identifier, datatype=XSD.token),
                )
            )

    # Basic metadata
    if cs.title:
        g.add((scheme_iri, SKOS.prefLabel, Literal(cs.title, lang="en")))
    if cs.description:
        g.add((scheme_iri, SKOS.definition, Literal(cs.description, lang="en")))

    # Dates
    if cs.created_date:
        g.add(
            (
                scheme_iri,
                DCTERMS.created,
                Literal(cs.created_date, datatype=XSD.date),
            )
        )
    if cs.modified_date:
        g.add(
            (
                scheme_iri,
                DCTERMS.modified,
                Literal(cs.modified_date, datatype=XSD.date),
            )
        )

    # Creator and Publisher (with Person/Organization triples)
    # Format is multi-line "<name> <URL>" - use parse_name_url for extraction
    creator_urls: list[str] = []
    if cs.creator:
        for iline in cs.creator.strip().split("\n"):
            name, url = parse_name_url(iline)
            if url:
                creator_urls.append(url)
                g.add((scheme_iri, DCTERMS.creator, URIRef(url)))
                g += build_entity_graph(url, name, "creator")

    if cs.publisher:
        for iline in cs.publisher.strip().split("\n"):
            name, url = parse_name_url(iline)
            if url:
                g.add((scheme_iri, DCTERMS.publisher, URIRef(url)))
                # Only add entity graph if not already added as creator
                if url not in creator_urls:
                    g += build_entity_graph(url, name, "publisher")

    # Contributors (with Person/Organization triples when URL present)
    if cs.contributor:
        for iline in cs.contributor.strip().split("\n"):
            name, url = parse_name_url(iline)
            if url:
                g.add((scheme_iri, DCTERMS.contributor, URIRef(url)))
                # Only add entity graph if not already added as creator
                if url not in creator_urls:
                    g += build_entity_graph(url, name, "contributor")
            elif iline.strip():
                # Fallback to literal if no URL
                g.add((scheme_iri, DCTERMS.contributor, Literal(iline.strip())))

    # Version
    if cs.version:
        g.add((scheme_iri, OWL.versionInfo, Literal(cs.version)))

    # History note (satisfies vocpub requirement 2.1.7 for ConceptScheme origins)
    if cs.history_note:
        g.add((scheme_iri, SKOS.historyNote, Literal(cs.history_note, lang="en")))

    # Custodian (with Person/Organization triples when URL present)
    if cs.custodian:
        for iline in cs.custodian.strip().split("\n"):
            name, url = parse_name_url(iline)
            if url:
                g.add((scheme_iri, DCAT.contactPoint, URIRef(url)))
                # Only add entity graph if not already added as creator
                if url not in creator_urls:
                    g += build_entity_graph(url, name, "custodian")
            elif iline.strip():
                # Fallback to literal if no URL
                g.add((scheme_iri, DCAT.contactPoint, Literal(iline.strip())))

    # Catalogue PID - both dct:identifier and rdfs:seeAlso
    if cs.catalogue_pid:
        g.add((scheme_iri, DCTERMS.identifier, Literal(cs.catalogue_pid)))
        if cs.catalogue_pid.startswith("http"):
            g.add((scheme_iri, RDFS.seeAlso, URIRef(cs.catalogue_pid)))
        else:
            g.add((scheme_iri, RDFS.seeAlso, Literal(cs.catalogue_pid)))

    # Homepage
    if cs.homepage:
        if cs.homepage.startswith("http"):
            g.add((scheme_iri, FOAF.homepage, URIRef(cs.homepage)))
        else:
            g.add((scheme_iri, FOAF.homepage, Literal(cs.homepage)))

    # Conforms to (SHACL profile)
    if cs.conforms_to:
        for iline in cs.conforms_to.strip().split("\n"):
            line = iline.strip()
            if not line:
                continue
            if line.startswith("http"):
                g.add((scheme_iri, DCTERMS.conformsTo, URIRef(line)))
            else:
                g.add((scheme_iri, DCTERMS.conformsTo, Literal(line)))

    # hasTopConcept - concepts with no broader
    for concept_iri, concept in concepts.items():
        if not concept.parent_iris:
            g.add((scheme_iri, SKOS.hasTopConcept, URIRef(concept_iri)))

    # Note: dcterms:hasPart (ConceptScheme -> Collection) is not used in v1.0
    # Collections link to scheme via skos:inScheme instead

    return g


def build_concept_graph(
    concept: AggregatedConcept,
    scheme_iri: URIRef,
    narrower_map: dict[str, list[str]],
    id_pattern: "re.Pattern[str] | None",
    vocab_name: str = "",
    provenance_template: str = "",
    repository_url: str = "",
) -> Graph:
    """Build RDF graph for a single Concept.

    Args:
        concept: Aggregated concept data.
        scheme_iri: URIRef of the ConceptScheme.
        narrower_map: Map of parent -> children for narrower relationships.
        id_pattern: Compiled regex pattern for extracting IDs.
        vocab_name: Vocabulary name for provenance URL generation.
        provenance_template: Jinja template for provenance URLs.
        repository_url: Repository URL from config for GitHub auto-detection.

    Returns:
        Graph with Concept triples.
    """
    g = Graph()
    c = URIRef(concept.iri)

    # Type
    g.add((c, RDF.type, SKOS.Concept))

    # Identifier
    if id_pattern:
        identifier = extract_identifier(concept.iri, id_pattern)
        if identifier:
            g.add((c, DCTERMS.identifier, Literal(identifier, datatype=XSD.token)))

    # Labels per language
    for lang, label in concept.pref_labels.items():
        g.add((c, SKOS.prefLabel, Literal(label, lang=lang)))

    for lang, definition in concept.definitions.items():
        g.add((c, SKOS.definition, Literal(definition, lang=lang)))

    for lang, labels in concept.alt_labels.items():
        for label in labels:
            g.add((c, SKOS.altLabel, Literal(label, lang=lang)))

    # Editorial notes per language
    for lang, note in concept.editorial_notes.items():
        g.add((c, SKOS.editorialNote, Literal(note, lang=lang)))

    # Broader (parent)
    for parent_iri in concept.parent_iris:
        g.add((c, SKOS.broader, URIRef(parent_iri)))

    # Narrower (computed inverse)
    for child_iri in narrower_map.get(concept.iri, []):
        g.add((c, SKOS.narrower, URIRef(child_iri)))

    # In scheme
    g.add((c, SKOS.inScheme, scheme_iri))

    # Source vocabulary attribution (verbatim copy)
    if concept.source_vocab_iri:
        g.add((c, PROV.hadPrimarySource, URIRef(concept.source_vocab_iri)))
    if concept.source_vocab_license:
        g.add((c, DCTERMS.license, URIRef(concept.source_vocab_license)))
    if concept.source_vocab_rights_holder:
        g.add((c, DCTERMS.rightsHolder, Literal(concept.source_vocab_rights_holder)))

    # Influenced by IRIs (prov:wasInfluencedBy)
    for influenced_iri in concept.influenced_by_iris:
        g.add((c, PROV.wasInfluencedBy, URIRef(influenced_iri)))

    # Top concept of (if no broader)
    if not concept.parent_iris:
        g.add((c, SKOS.topConceptOf, scheme_iri))

    # Change note
    if concept.change_note:
        g.add((c, SKOS.changeNote, Literal(concept.change_note, lang="en")))

    # Obsoletion (deprecated concept)
    if concept.obsolete_reason:
        g.add((c, OWL.deprecated, Literal(True)))
        g.add((c, SKOS.historyNote, Literal(concept.obsolete_reason, lang="en")))

    # Replaced by (dct:isReplacedBy)
    if concept.replaced_by_iri:
        g.add((c, DCTERMS.isReplacedBy, URIRef(concept.replaced_by_iri)))

    # Provenance (git blame URL)
    add_provenance_triples_to_graph(
        g, c, vocab_name, provenance_template, repository_url
    )

    return g


def build_collection_graph(
    collection: AggregatedCollection,
    scheme_iri: URIRef,
    collection_members: dict[str, list[str]],
    ordered_collection_members: dict[str, list[str]] | None = None,
    id_pattern: "re.Pattern[str] | None" = None,
    vocab_name: str = "",
    provenance_template: str = "",
    repository_url: str = "",
) -> Graph:
    """Build RDF graph for a single Collection.

    Args:
        collection: Aggregated collection data.
        scheme_iri: URIRef of the ConceptScheme.
        collection_members: Map of collection -> member IRIs (unordered).
        ordered_collection_members: Map of ordered collection -> member IRIs
            in order. Used for building skos:memberList.
        id_pattern: Compiled regex pattern for extracting IDs.
        vocab_name: Vocabulary name for provenance URL generation.
        provenance_template: Jinja template for provenance URLs.
        repository_url: Repository URL from config for GitHub auto-detection.

    Returns:
        Graph with Collection triples.
    """
    ordered_collection_members = ordered_collection_members or {}
    g = Graph()
    c = URIRef(collection.iri)

    # Type - either OrderedCollection or Collection
    if collection.ordered:
        g.add((c, RDF.type, SKOS.OrderedCollection))
    else:
        g.add((c, RDF.type, SKOS.Collection))

    # Identifier
    if id_pattern:
        identifier = extract_identifier(collection.iri, id_pattern)
        if identifier:
            g.add((c, DCTERMS.identifier, Literal(identifier, datatype=XSD.token)))

    # Labels per language
    for lang, label in collection.pref_labels.items():
        g.add((c, SKOS.prefLabel, Literal(label, lang=lang)))

    for lang, definition in collection.definitions.items():
        g.add((c, SKOS.definition, Literal(definition, lang=lang)))

    # Editorial notes per language
    for lang, note in collection.editorial_notes.items():
        g.add((c, SKOS.editorialNote, Literal(note, lang=lang)))

    # Members - different handling for ordered vs unordered collections
    if collection.ordered:
        # OrderedCollection uses skos:memberList with RDF List
        ordered_members = ordered_collection_members.get(collection.iri, [])
        if ordered_members:
            list_node = BNode()
            member_refs = [URIRef(m) for m in ordered_members]
            RDFCollection(g, list_node, member_refs)
            g.add((c, SKOS.memberList, list_node))
    else:
        # Regular Collection uses skos:member
        for member_iri in collection_members.get(collection.iri, []):
            g.add((c, SKOS.member, URIRef(member_iri)))

    # In scheme
    g.add((c, SKOS.inScheme, scheme_iri))

    # rdfs:isDefinedBy - points to ConceptScheme (convention)
    g.add((c, RDFS.isDefinedBy, scheme_iri))

    # Change note
    if collection.change_note:
        g.add((c, SKOS.changeNote, Literal(collection.change_note, lang="en")))

    # Obsoletion (deprecated collection)
    if collection.obsolete_reason:
        g.add((c, OWL.deprecated, Literal(True)))
        g.add((c, SKOS.historyNote, Literal(collection.obsolete_reason, lang="en")))

    # Replaced by (dct:isReplacedBy)
    if collection.replaced_by_iri:
        g.add((c, DCTERMS.isReplacedBy, URIRef(collection.replaced_by_iri)))

    # Provenance (git blame URL)
    add_provenance_triples_to_graph(
        g, c, vocab_name, provenance_template, repository_url
    )

    return g


def build_mappings_graph(
    mappings: list[MappingV1], converter: curies.Converter
) -> Graph:
    """Build RDF graph for all mappings.

    Args:
        mappings: List of MappingV1 from XLSX.
        converter: Curies converter for IRI expansion.

    Returns:
        Graph with mapping triples.
    """
    g = Graph()

    for mapping in mappings:
        if not mapping.concept_iri:
            continue

        # Strip label if present (format: "curie (label)")
        concept_iri_only = strip_label_from_iri(mapping.concept_iri)
        concept_iri = URIRef(expand_curie(concept_iri_only, converter))

        # Related matches
        for match_iri in expand_iri_list(mapping.related_matches, converter):
            g.add((concept_iri, SKOS.relatedMatch, URIRef(match_iri)))

        # Close matches
        for match_iri in expand_iri_list(mapping.close_matches, converter):
            g.add((concept_iri, SKOS.closeMatch, URIRef(match_iri)))

        # Exact matches
        for match_iri in expand_iri_list(mapping.exact_matches, converter):
            g.add((concept_iri, SKOS.exactMatch, URIRef(match_iri)))

        # Narrower matches
        for match_iri in expand_iri_list(mapping.narrower_matches, converter):
            g.add((concept_iri, SKOS.narrowMatch, URIRef(match_iri)))

        # Broader matches
        for match_iri in expand_iri_list(mapping.broader_matches, converter):
            g.add((concept_iri, SKOS.broadMatch, URIRef(match_iri)))

    return g


# --- Main XLSX -> RDF Converter ---


def excel_to_rdf_v1(
    file_to_convert_path: Path,
    output_file_path: Path | None = None,
    output_format: TypingLiteral["turtle", "xml", "json-ld"] = "turtle",
    output_type: TypingLiteral["file", "graph"] = "file",
    *,
    vocab_config: "config.Vocab",
) -> Path | Graph:
    """Convert a v1.0 Excel template to RDF vocabulary.

    ConceptScheme metadata is read from vocab_config (idranges.toml).
    The "Concept Scheme" sheet in Excel is read-only and never read.

    Args:
        file_to_convert_path: Path to the XLSX file.
        output_file_path: Optional path for output RDF file.
                         Defaults to same name with .ttl extension.
        output_format: RDF serialization format ("turtle", "xml", "json-ld").
        output_type: "file" to serialize to file, "graph" to return Graph object.
        vocab_config: Vocab config from idranges.toml with scheme metadata.

    Returns:
        Path to the generated RDF file, or Graph object if output_type="graph".

    Raises:
        ValueError: If the input file is not a supported Excel format.
    """
    if file_to_convert_path.suffix.lower() not in EXCEL_FILE_ENDINGS:
        msg = (
            "Files for conversion must end with one of the Excel file formats: "
            f"'{', '.join(EXCEL_FILE_ENDINGS)}'"
        )
        raise ValueError(msg)

    logger.info("Reading XLSX file: %s", file_to_convert_path)

    # Get vocabulary name from filename (used for provenance URLs)
    vocab_name = file_to_convert_path.stem.lower()

    logger.debug("Reading Prefixes...")
    prefixes = read_prefixes_v1(file_to_convert_path)

    logger.debug("Reading Concepts...")
    concept_rows = read_concepts_v1(file_to_convert_path)

    logger.debug("Reading Collections...")
    collection_rows = read_collections_v1(file_to_convert_path)

    logger.debug("Reading Mappings...")
    mapping_rows = read_mappings_v1(file_to_convert_path)

    # Build curies converter from prefixes
    converter = build_curies_converter_from_prefixes(prefixes)

    # Aggregate multi-row data
    logger.debug("Aggregating concepts...")
    concepts = aggregate_concepts(concept_rows, converter)

    logger.debug("Aggregating collections...")
    collections = aggregate_collections(collection_rows, converter)

    # Derive contributors from ID range usage
    derived_contributors = ""
    if vocab_config.id_range:
        logger.debug("Deriving contributors from ID range usage...")
        used_ids = extract_used_ids(concept_rows, collection_rows, vocab_config)
        derived_contributors = derive_contributors(vocab_config, used_ids)

    # Build ConceptScheme from config (Excel sheet is read-only, never read)
    concept_scheme = config_to_concept_scheme_v1(
        vocab_config, derived_contributors=derived_contributors
    )

    # Auto-generate version and modified_date from environment variables (if set)
    version = os.getenv("VOC4CAT_VERSION", "")
    if version:
        if not version.startswith("v"):
            msg = (
                f'Invalid environment variable VOC4CAT_VERSION "{version}". '
                'Version must start with letter "v".'
            )
            raise Voc4catError(msg)
        concept_scheme.version = version

    modified_date = (
        os.getenv("VOC4CAT_MODIFIED")
        or datetime.datetime.now(tz=datetime.timezone.utc).date().isoformat()
    )
    concept_scheme.modified_date = modified_date

    # Build inverse relationships
    logger.debug("Building inverse relationships...")
    collection_members = build_collection_members_from_concepts(concepts)
    # Add child collections as members of parent collections
    collection_hierarchy_members = build_collection_hierarchy_members(collections)
    for parent_iri, child_iris in collection_hierarchy_members.items():
        if parent_iri not in collection_members:
            collection_members[parent_iri] = []
        collection_members[parent_iri].extend(child_iris)
    ordered_collection_members = build_ordered_collection_members(concepts)
    narrower_map = build_narrower_map(concepts)

    # Get provenance URL config
    provenance_template = vocab_config.provenance_url_template if vocab_config else ""
    repository_url = vocab_config.repository if vocab_config else ""

    # Get ID pattern for identifier extraction
    id_pattern = config.ID_PATTERNS.get(vocab_name)
    if not id_pattern:
        logger.warning(
            "No ID pattern found for vocab '%s', identifiers will be skipped",
            vocab_name,
        )

    # Build the complete graph
    logger.debug("Building RDF graph...")
    scheme_iri = URIRef(concept_scheme.vocabulary_iri)

    graph = build_concept_scheme_graph(concept_scheme, concepts, id_pattern)

    for concept in concepts.values():
        graph += build_concept_graph(
            concept,
            scheme_iri,
            narrower_map,
            id_pattern,
            vocab_name,
            provenance_template,
            repository_url,
        )

    for collection in collections.values():
        graph += build_collection_graph(
            collection,
            scheme_iri,
            collection_members,
            ordered_collection_members,
            id_pattern,
            vocab_name,
            provenance_template,
            repository_url,
        )

    graph += build_mappings_graph(mapping_rows, converter)

    # Bind prefixes for nice serialization
    for prefix_model in prefixes:
        if prefix_model.prefix and prefix_model.namespace:
            graph.bind(prefix_model.prefix, Namespace(prefix_model.namespace))

    logger.info("Built graph with %d triples", len(graph))

    if output_type == "graph":
        return graph

    # Serialize to file
    if output_file_path is None:
        if output_format == "xml":
            suffix = ".rdf"
        elif output_format == "json-ld":
            suffix = ".json-ld"
        else:
            suffix = ".ttl"
        output_file_path = file_to_convert_path.with_suffix(suffix)

    logger.info("Serializing to: %s", output_file_path)
    graph.serialize(destination=str(output_file_path), format=output_format)

    logger.info("Conversion complete: %s", output_file_path)
    return output_file_path


# --- Debugging Utilities ---


def compare_graphs(g1: Graph, g2: Graph) -> dict:
    """Compare two graphs and return differences.

    Useful for debugging round-trip conversion issues.

    Args:
        g1: First graph (e.g., original).
        g2: Second graph (e.g., round-tripped).

    Returns:
        Dict with "only_in_g1" and "only_in_g2" triple lists.
    """
    only_in_g1 = list(g1 - g2)
    only_in_g2 = list(g2 - g1)

    return {
        "only_in_g1": only_in_g1,
        "only_in_g2": only_in_g2,
        "g1_count": len(g1),
        "g2_count": len(g2),
        "common_count": len(g1) - len(only_in_g1),
    }


# =============================================================================
# Re-exports for backwards compatibility
# =============================================================================

# Re-export helper functions that were moved to convert_v1_helpers.py
# These are re-exported via the import at the top of this file:
# - DEFAULT_PROVENANCE_TEMPLATE
# - OBSOLETE_PREFIX
# - build_id_range_info
# - build_provenance_url
# - derive_contributors
# - expand_curie
# - extract_entity_id_from_iri
# - extract_github_repo_from_url
# - extract_used_ids
# - format_change_note_with_replaced_by
# - format_contributor_string
# - format_iri_with_label
# - parse_replaced_by_from_change_note
# - validate_deprecation

# Note: convert_rdf_043_to_v1 is NOT re-exported here to avoid circular imports.
# Import it directly from voc4cat.convert_043 instead.
