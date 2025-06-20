import datetime
import logging
import os
from itertools import chain
from typing import Annotated

from openpyxl import Workbook
from pydantic import (
    AnyHttpUrl,
    BaseModel,
    BeforeValidator,
    Field,
    field_validator,
    model_validator,
)
from rdflib import (
    DCAT,
    DCTERMS,
    OWL,
    RDF,
    RDFS,
    SDO,
    SKOS,
    XSD,
    Graph,
    Literal,
    Namespace,
    URIRef,
)
from rdflib.namespace import NamespaceManager
from typing_extensions import Self

from voc4cat import config
from voc4cat.fields import ORCID_PATTERN, ORCIDIdentifier, RORIdentifier

logger = logging.getLogger(__name__)

# If defined here the key can be used in xlsx instead of an URI.
ORGANISATIONS = {
    "NFDI4Cat": URIRef("http://example.org/nfdi4cat/"),
    "LIKAT": URIRef("https://ror.org/029hg0311"),
    # from rdflib.vocexcel (used in old tests)
    "CGI": URIRef("https://linked.data.gov.au/org/cgi-gtwg"),
    "GSQ": URIRef("https://linked.data.gov.au/org/gsq"),
}

ORGANISATIONS_INVERSE = {uref: name for name, uref in ORGANISATIONS.items()}


def make_iri_qualifier_listing(item, concepts_by_iri):
    """Return listing of item with one "uri (pref.label)" per row."""
    child_lines = []
    for uri in item:
        uri = str(uri)
        uri_str = config.curies_converter.compress(uri, passthrough=True)
        if "en" not in concepts_by_iri[uri]:
            child_lines.append(f"{uri_str}")
            continue
        # we must be careful: not all concepts have all languages
        pref_label_in_lang = concepts_by_iri[uri]["en"].pref_label[0]
        child_lines.append(f"{uri_str} ({pref_label_in_lang})")
    return ",\n".join(child_lines)


# === Pydantic validators used by more than one model ===


def split_curie_list(v):
    """Pydantic before validator to split a comma-separated list of CURIEs."""
    if v is None:
        return []
    if isinstance(v, list):
        return v
    return (p.strip() for p in v.split(","))


def normalise_curie_to_uri(v):
    """Pydantic after validator to convert CURIEs to URIs."""
    v = str(v)  # Ensure v is a string for processing with curies_converter
    try:
        v = config.curies_converter.standardize_curie(v) or v
        v_as_uri = config.curies_converter.expand(v) or v
    except ValueError:
        # curies fails for invalid URLs, e.g. "www.wikipedia.de"
        # We keep the problematic value and let pydantic validate it.
        v_as_uri = v
    if not v_as_uri.startswith("http"):
        msg = f'"{v}" is not a valid IRI (missing URL scheme).'
        raise ValueError(msg)
    return v_as_uri


def normalise_curies_to_uris(values: list) -> list:
    """Pydantic after validator to convert a list of CURIEs to URIs."""
    return [normalise_curie_to_uri(v) for v in values]


def check_uri_vs_config(values):
    """Root validator to check if uri starts with permanent_iri_part.

    Note that it will pass validation if permanent_iri_part is empty.
    """
    voc_conf = config.IDRANGES.vocabs.get(values.vocab_name, {})
    if not voc_conf:
        return values

    perm_iri_part = str(getattr(voc_conf, "permanent_iri_part", ""))
    iri, *_fragment = str(getattr(values, "uri", "")).split("#", 1)
    if not iri.startswith(perm_iri_part):
        msg = "Invalid IRI %s - It must start with %s"
        raise ValueError(msg % (iri, perm_iri_part))
    id_part_of_iri = iri.split(perm_iri_part)[-1]
    if any(not c.isdigit() for c in id_part_of_iri):
        msg = 'Invalid ID part "%s" in IRI %s. The ID part may only contain digits.'
        raise ValueError(msg % (id_part_of_iri, iri))

    id_pattern = config.ID_PATTERNS.get(values.vocab_name)
    match = id_pattern.search(iri)
    if not match:
        msg = "ID part of %s is not matching the configured pattern of %s digits."
        raise ValueError(msg % (str(getattr(values, "uri", "")), voc_conf.id_length))

    return values


def check_used_id(values):
    """
    Root validator to check if the ID part of the IRI is allowed for the actor.

    The first entry in provenance is used as relevant actor. For this actor the
    allowed ID range(s) are read from the config.
    """
    voc_conf = config.IDRANGES.vocabs.get(values.vocab_name, {})
    if not voc_conf:
        return values

    # provenance example value: "0000-0001-2345-6789 Provenance description, ..."
    if hasattr(values, "provenance"):
        actor = values.provenance.split(",", 1)[0].split(" ", 1)[0].strip()
    else:
        actor = ""
    iri, *_fragment = str(getattr(values, "uri", "")).split("#", 1)
    id_pattern = config.ID_PATTERNS.get(values.vocab_name, None)
    if id_pattern is not None:
        match = id_pattern.search(iri)
        if match:
            id_ = int(match["identifier"])
            if actor in config.ID_RANGES_BY_ACTOR:
                actors_ids = config.ID_RANGES_BY_ACTOR[actor]
            else:
                # Because the use of provenance is not well defined and it is not validated
                # a GitHub name may be in incorrect case in the provenance field. (#122)
                # So we look up for lower-cased actor as well:
                actors_ids = config.ID_RANGES_BY_ACTOR[actor.lower()]
            allowed = any(first <= id_ <= last for first, last in actors_ids)
            if not allowed:
                msg = 'ID of IRI %s is not in allowed ID range(s) of actor "%s" (from provenance).'
                raise ValueError(msg % (str(getattr(values, "uri", "")), actor))
    return values


# === Pydantic models ===


class ConceptScheme(BaseModel):
    uri: AnyHttpUrl
    title: str
    description: str
    created: datetime.date
    modified: datetime.date
    creator: RORIdentifier | ORCIDIdentifier | str = Field(union_mode="left_to_right")
    publisher: RORIdentifier | str = Field(union_mode="left_to_right")
    provenance: str
    version: str = "automatic"
    custodian: str | None = None
    pid: str | None = None
    vocab_name: str = Field("", exclude=True)

    # @validator("creator")
    @field_validator("creator", mode="after")
    @classmethod
    def creator_must_be_from_list(cls, v):
        if isinstance(v, str):
            if v.startswith("http") or ORCID_PATTERN.match(v):
                print(f"Creator: {v}")
                return v
            if v not in ORGANISATIONS:
                msg = f"Creator must be an ORCID or ROR ID or a string from the organisations list: {', '.join(ORGANISATIONS)}"
                raise ValueError(msg)
        return v

    # @validator("modified")
    @field_validator("modified", mode="after")
    @classmethod
    def set_modified_date_from_env(cls, v):
        if os.getenv("VOC4CAT_MODIFIED") is not None:
            v = datetime.date.fromisoformat(os.getenv("VOC4CAT_MODIFIED", ""))
        return v

    # @validator("publisher")
    @field_validator("publisher", mode="after")
    @classmethod
    def publisher_must_be_from_list(cls, v):
        if isinstance(v, str):
            if v.startswith("http"):
                return v
            if v not in ORGANISATIONS:
                msg = f"Publisher must be an ROR ID or a string from the organisations list: {', '.join(ORGANISATIONS)}"
                raise ValueError(msg)
        return v

    # @validator("version")
    @field_validator("version", mode="after")
    @classmethod
    def version_from_env(cls, v):
        # Don't track version in GitHub.
        v = "automatic" if os.getenv("CI") is not None else str(v)
        version_from_env = os.getenv("VOC4CAT_VERSION")
        if version_from_env is not None:
            if not version_from_env.startswith("v"):
                msg = f'Invalid environment variable VOC4CAT_VERSION "{version_from_env}". Version must start with letter "v".'
                raise ValueError(msg)
            v = version_from_env
        return v

    def to_graph(self):
        g = Graph()

        # <http://example.org/nfdi4cat/> a sdo:Organization ;
        # sdo:name "NFDI4Cat" ;
        # sdo:url "https://nfdi4cat.org"^^xsd:anyURI .
        tg = Graph()
        org = ORGANISATIONS.get(self.publisher, URIRef(self.publisher))
        tg.add((org, RDF.type, SDO.Organization))
        tg.add(
            (
                org,
                SDO.name,
                # should be name but there is no field in the template 0.43
                Literal(
                    ORGANISATIONS.get(self.publisher, self.publisher),
                ),
            )
        )
        tg.add(
            (
                org,
                SDO.url,
                Literal(
                    ORGANISATIONS.get(self.publisher, self.publisher),
                    datatype=URIRef(XSD.anyURI),
                ),
            )
        )
        g += tg

        v = URIRef(str(self.uri))
        # For dcterms:identifier
        identifier = v.split("#")[-1] if "#" in v else v.split("/")[-1]
        g.add((v, DCTERMS.identifier, Literal(identifier, datatype=XSD.token)))

        g.add((v, RDF.type, SKOS.ConceptScheme))
        g.add((v, SKOS.prefLabel, Literal(self.title, lang="en")))
        g.add((v, SKOS.definition, Literal(self.description, lang="en")))
        g.add((v, DCTERMS.created, Literal(self.created, datatype=XSD.date)))
        g.add((v, DCTERMS.modified, Literal(self.modified, datatype=XSD.date)))
        g.add(
            (v, DCTERMS.creator, ORGANISATIONS.get(self.creator, URIRef(self.creator)))
        )
        g.add(
            (
                v,
                DCTERMS.publisher,
                ORGANISATIONS.get(self.publisher, URIRef(self.publisher)),
            )
        )
        g.add((v, OWL.versionInfo, Literal(self.version)))
        g.add((v, SKOS.historyNote, Literal(self.provenance, lang="en")))
        if self.custodian is not None:
            g.add((v, DCAT.contactPoint, Literal(self.custodian)))
        if self.pid is not None:
            if self.pid.startswith("http"):
                g.add((v, RDFS.seeAlso, URIRef(self.pid)))
            else:
                g.add((v, RDFS.seeAlso, Literal(self.pid)))

        g.namespace_manager = NamespaceManager(g)
        converter = config.CURIES_CONVERTER_MAP.get(
            self.vocab_name, config.curies_converter
        )
        for prefix, uri_prefix in converter.bimap.items():
            g.namespace_manager.bind(prefix, Namespace(uri_prefix))
        return g

    def to_excel(self, wb: Workbook):
        ws = wb["Concept Scheme"]
        ws["B2"] = config.curies_converter.expand(str(self.uri), passthrough=True)
        ws["B2"].hyperlink = str(self.uri)
        ws["B3"] = self.title
        ws["B4"] = self.description
        ws["B5"] = self.created.isoformat()
        ws["B6"] = None if self.modified is None else self.modified.isoformat()
        ws["B7"] = self.creator
        ws["B8"] = self.publisher
        ws["B9"] = self.version
        ws["B10"] = self.provenance
        ws["B11"] = self.custodian
        ws["B12"] = self.pid

        # reset row height for all table rows including header to default
        start_row, last_row = 2, 12
        for row in range(start_row, last_row + 1):
            ws.row_dimensions[row].height = None


class Concept(BaseModel):
    uri: Annotated[AnyHttpUrl, BeforeValidator(normalise_curie_to_uri)]
    pref_label: str | list[str]
    alt_labels: list[str] = []
    pl_language_code: list[str] = []
    definition: str | list[str]
    def_language_code: list[str] = []
    children: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    source_vocab: (
        Annotated[AnyHttpUrl, BeforeValidator(normalise_curie_to_uri)] | None
    ) = None
    provenance: str | None = None
    related_match: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    close_match: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    exact_match: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    narrow_match: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    broad_match: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    vocab_name: str = Field("", exclude=True)

    # We validate with a reusable (external) validators. With a pre-validator,
    # which is applied before all others, we split and to convert from CURIE to URI.

    # _uri_list_fields = [
    #     "children",
    #     "related_match",
    #     "close_match",
    #     "exact_match",
    #     "narrow_match",
    #     "broad_match",
    # ]
    # _split_uri_list = validator(*_uri_list_fields, mode="before")(
    #     split_curie_list
    # )

    # _normalize_uri = validator(
    #     "uri",
    #     "source_vocab",
    #     *_uri_list_fields,
    #     each_item=True,
    #     mode="before",
    # )(normalise_curie_to_uri)

    # _check_uri_vs_config = model_validator(mode="after")(check_uri_vs_config)
    @model_validator(mode="after")
    def _check_uri_vs_config(self) -> Self:
        check_uri_vs_config(self)
        return self

    # _check_used_id = model_validator(mode="after")(check_used_id)
    @model_validator(mode="after")
    def _check_used_id(self) -> Self:
        check_used_id(self)
        return self

    def to_graph(self):
        g = Graph()
        c = URIRef(str(self.uri))

        g.add((c, RDF.type, SKOS.Concept))
        # For dcterms:identifier
        identifier = c.split("#")[-1] if "#" in c else c.split("/")[-1]
        g.add((c, DCTERMS.identifier, Literal(identifier, datatype=XSD.token)))

        if not self.pl_language_code:
            g.add((c, SKOS.prefLabel, Literal(self.pref_label, lang="en")))
        else:
            for lang_code in self.pl_language_code:
                g.add((c, SKOS.prefLabel, Literal(self.pref_label, lang=lang_code)))
        if self.alt_labels is not None:
            for alt_label in self.alt_labels:
                g.add((c, SKOS.altLabel, Literal(alt_label, lang="en")))
        if not self.def_language_code:
            g.add((c, SKOS.definition, Literal(self.definition, lang="en")))
        else:
            for lang_code in self.def_language_code:
                g.add((c, SKOS.definition, Literal(self.definition, lang=lang_code)))
        for child in self.children:
            g.add((c, SKOS.narrower, URIRef(str(child))))
            g.add((URIRef(str(child)), SKOS.broader, c))
        if self.source_vocab is not None:
            g.add((c, RDFS.isDefinedBy, URIRef(str(self.source_vocab))))
        if self.provenance is not None:
            g.add((c, SKOS.historyNote, Literal(self.provenance, lang="en")))
        if self.related_match is not None:
            for related_match in self.related_match:
                g.add((c, SKOS.relatedMatch, URIRef(str(related_match))))
        if self.close_match:
            for close_match in self.close_match:
                g.add((c, SKOS.closeMatch, URIRef(str(close_match))))
        if self.exact_match is not None:
            for exact_match in self.exact_match:
                g.add((c, SKOS.exactMatch, URIRef(str(exact_match))))
        if self.narrow_match is not None:
            for narrow_match in self.narrow_match:
                g.add((c, SKOS.narrowMatch, URIRef(str(narrow_match))))
        if self.broad_match is not None:
            for broad_match in self.broad_match:
                g.add((c, SKOS.broadMatch, URIRef(str(broad_match))))

        return g

    def to_excel(
        self,
        wb: Workbook,
        row_no_concepts: int,
        row_no_features: int,
        concepts_by_iri: dict,
    ) -> int:
        """ "
        Export Concept to Excel using one row per language

        Non-labels like Children, Provenance are reported only for the first
        language. If "en" is among the used languages, it is reported first.

        Return the row number of next row after the ones filled.
        """
        ws = wb["Concepts"]

        # determine the languages with full and partial translation
        pref_labels = {
            lang: pl for pl, lang in zip(self.pref_label, self.pl_language_code)
        }
        definitions = {
            lang: d for d, lang in zip(self.definition, self.def_language_code)
        }
        fully_translated = [lbl for lbl in pref_labels if lbl in definitions]
        partially_translated = [
            lbl
            for lbl in chain(pref_labels.keys(), definitions.keys())
            if lbl not in fully_translated
        ]

        # put "en" first if available
        if "en" in fully_translated:
            fully_translated.remove("en")
            fully_translated.insert(0, "en")

        first_row_exported = False
        for lang in chain(fully_translated, partially_translated):
            ws[f"A{row_no_concepts}"].value = config.curies_converter.compress(
                str(self.uri), passthrough=True
            )
            ws[f"A{row_no_concepts}"].hyperlink = str(self.uri)
            ws[f"B{row_no_concepts}"] = pref_labels.get(lang, "")
            ws[f"C{row_no_concepts}"] = lang
            ws[f"D{row_no_concepts}"] = definitions.get(lang, "")
            ws[f"E{row_no_concepts}"] = lang
            ws[f"H{row_no_concepts}"] = self.provenance

            if first_row_exported:
                row_no_concepts += 1
                continue

            first_row_exported = True
            ws[f"F{row_no_concepts}"] = ",\n".join(self.alt_labels)
            ws[f"G{row_no_concepts}"] = make_iri_qualifier_listing(
                self.children, concepts_by_iri
            )
            if self.source_vocab:
                ws[f"I{row_no_concepts}"] = config.curies_converter.compress(
                    str(self.source_vocab), passthrough=True
                )
                ws[f"I{row_no_concepts}"].hyperlink = str(self.source_vocab)
            else:
                ws[f"I{row_no_concepts}"] = None
                ws[f"I{row_no_concepts}"].hyperlink = None
            row_no_concepts += 1

        # Fill Additional Concept Features sheet
        if any(
            [
                self.related_match,
                self.close_match,
                self.exact_match,
                self.narrow_match,
                self.broad_match,
            ]
        ):
            ws = wb["Additional Concept Features"]
            ws[f"A{row_no_features}"].value = (
                config.curies_converter.compress(str(self.uri), passthrough=True)
                + f" ({pref_labels.get('en', '')})"
            )
            ws[f"A{row_no_features}"].hyperlink = str(self.uri)
            ws[f"B{row_no_features}"] = make_iri_qualifier_listing(
                self.related_match, concepts_by_iri
            )
            ws[f"C{row_no_features}"] = make_iri_qualifier_listing(
                self.close_match, concepts_by_iri
            )
            ws[f"D{row_no_features}"] = make_iri_qualifier_listing(
                self.exact_match, concepts_by_iri
            )
            ws[f"E{row_no_features}"] = make_iri_qualifier_listing(
                self.narrow_match, concepts_by_iri
            )
            ws[f"F{row_no_features}"] = make_iri_qualifier_listing(
                self.broad_match, concepts_by_iri
            )

        return row_no_concepts


class Collection(BaseModel):
    uri: Annotated[AnyHttpUrl, BeforeValidator(normalise_curie_to_uri)]
    pref_label: str
    definition: str
    members: Annotated[
        list[AnyHttpUrl],
        BeforeValidator(split_curie_list),
        BeforeValidator(normalise_curies_to_uris),
    ] = []
    provenance: str | None = None
    vocab_name: str = Field("", exclude=True)

    # _split_uri_list = validator("members", pre=True)(split_curie_list)
    # _normalize_uri = validator(
    #     "uri", "members", each_item=True, pre=True
    # )(normalise_curie_to_uri)

    _check_uri_vs_config = model_validator(mode="after")(check_uri_vs_config)
    _check_used_id = model_validator(mode="after")(check_used_id)

    def to_graph(self, cs):
        g = Graph()
        c = URIRef(str(self.uri))
        g.add((c, RDF.type, SKOS.Collection))

        # rdfs:isDefinedBy <https://example.org> ;
        g.add((c, RDFS.isDefinedBy, cs))
        # skos:inScheme <https://example.org> ;
        g.add((c, SKOS.inScheme, cs))

        # for dcterms:identifier
        identifier = c.split("#")[-1] if "#" in c else c.split("/")[-1]
        g.add((c, DCTERMS.identifier, Literal(identifier, datatype=XSD.token)))

        g.add((c, SKOS.prefLabel, Literal(self.pref_label, lang="en")))
        g.add((c, SKOS.definition, Literal(self.definition, lang="en")))
        for member in self.members:
            g.add((c, SKOS.member, URIRef(str(member))))
        if self.provenance is not None:
            g.add((c, SKOS.historyNote, Literal(self.provenance, lang="en")))

        return g

    def to_excel(self, wb: Workbook, row_no: int, concepts_by_iri: dict) -> None:
        ws = wb["Collections"]
        ws[f"A{row_no}"].value = config.curies_converter.compress(
            str(self.uri), passthrough=True
        )  # + f" ({self.pref_label})"
        ws[f"A{row_no}"].hyperlink = str(self.uri)
        ws[f"B{row_no}"] = self.pref_label
        ws[f"C{row_no}"] = self.definition
        ws[f"D{row_no}"] = make_iri_qualifier_listing(self.members, concepts_by_iri)
        ws[f"E{row_no}"] = self.provenance


class Vocabulary(BaseModel):
    concept_scheme: ConceptScheme
    concepts: list[Concept]
    collections: list[Collection]

    def to_graph(self):
        g = self.concept_scheme.to_graph()

        cs = URIRef(str(self.concept_scheme.uri))
        for concept in self.concepts:
            g += concept.to_graph()
            g.add((URIRef(str(concept.uri)), SKOS.inScheme, cs))

        # create as Top Concepts those Concepts that have no skos:narrower properties with them as objects
        for s in g.subjects(SKOS.inScheme, cs):
            is_tc = True
            for _ in g.objects(s, SKOS.broader):
                is_tc = False
            if is_tc:
                g.add((cs, SKOS.hasTopConcept, s))
                g.add((s, SKOS.topConceptOf, cs))

        for collection in self.collections:
            g += collection.to_graph(cs)
            g.add((URIRef(str(collection.uri)), DCTERMS.isPartOf, cs))
            g.add((cs, DCTERMS.hasPart, URIRef(str(collection.uri))))

        return g
