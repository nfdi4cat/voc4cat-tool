"""
Common XLSX functionality shared by both table and key-value formats.

This module contains shared infrastructure including:
- Base classes for configuration, formatting, and processing
- Field analysis and metadata handling
- Serialization engine for converting Python objects to/from Excel values
- Converters for common data patterns
- Exception classes
- Pre-configured type aliases
"""

import json
import logging
from abc import ABC, abstractmethod
from collections.abc import Callable
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Annotated, Any, Union, get_args, get_origin

from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from pydantic_core import PydanticUndefined

logger = logging.getLogger(__name__)

# Excel's limit for data validation formula length
EXCEL_DV_FORMULA_LIMIT = 255
MAX_SHEETNAME_LENGTH = 31


# Exception classes
class XLSXSerializationError(ValueError):
    """Raised when custom serialization fails."""

    def __init__(self, field_name: str, value: Any, original_error: Exception):
        self.field_name = field_name
        self.value = value
        self.original_error = original_error
        super().__init__(
            f"Error serializing field '{field_name}' with value '{value}': {original_error}"
        )


class XLSXDeserializationError(ValueError):
    """Raised when custom deserialization fails."""

    def __init__(self, field_name: str, value: Any, original_error: Exception):
        self.field_name = field_name
        self.value = value
        self.original_error = original_error
        super().__init__(
            f"Error deserializing field '{field_name}' with value '{value}': {original_error}"
        )


# Metadata and field analysis
@dataclass(frozen=True)
class XLSXMetadata:
    """XLSX-specific metadata for Pydantic fields."""

    unit: str | None = None
    meaning: str | None = None
    display_name: str | None = None
    description: str | None = None
    separator_pattern: "SeparatorPattern | None" = None  # Forward reference
    xlsx_serializer: Callable | None = None
    xlsx_deserializer: Callable | None = None


class MetadataVisibility(Enum):
    """Control visibility of metadata rows/columns in XLSX output.

    AUTO: Show if any field has this metadata type (current default behavior)
    SHOW: Force show (empty cells for fields without metadata)
    HIDE: Force hide (never show this metadata type)
    """

    AUTO = "auto"
    SHOW = "show"
    HIDE = "hide"


@dataclass
class MetadataToggleConfig:
    """Configuration for metadata visibility toggles.

    Controls whether unit, requiredness, description, and meaning
    rows (table format) or columns (key-value format) are displayed.
    """

    unit: MetadataVisibility = MetadataVisibility.AUTO
    requiredness: MetadataVisibility = MetadataVisibility.AUTO
    description: MetadataVisibility = MetadataVisibility.AUTO
    meaning: MetadataVisibility = MetadataVisibility.AUTO


@dataclass
class FieldAnalysis:
    """Runtime analysis data for Pydantic model fields."""

    name: str
    field_type: type | None
    is_optional: bool = False
    enum_values: list[str] = field(default_factory=list)
    xlsx_metadata: XLSXMetadata | None = None


def _validate_unit_usage(field_name: str, field_type: type, unit: str | None) -> None:
    """Validate that units are only used with numeric fields."""
    if unit is None:
        return

    # Check if field type is numeric
    numeric_types = {int, float, complex}

    # Handle Optional[numeric] types (Union[numeric, None])
    if get_origin(field_type) is Union:
        args = get_args(field_type)
        if len(args) == 2 and type(None) in args:  # noqa: PLR2004
            non_none_type = next(arg for arg in args if arg is not type(None))
            if non_none_type not in numeric_types:
                msg = (
                    f"Unit '{unit}' specified for non-numeric field '{field_name}' "
                    f"of type {non_none_type}. Units are only valid for numeric fields."
                )
                raise ValueError(msg)
            return

    # Handle Python 3.10+ union syntax (numeric | None)
    elif hasattr(field_type, "__args__"):
        args = field_type.__args__
        if len(args) == 2 and type(None) in args:  # noqa: PLR2004
            non_none_type = next(arg for arg in args if arg is not type(None))
            if non_none_type not in numeric_types:
                msg = (
                    f"Unit '{unit}' specified for non-numeric field '{field_name}' "
                    f"of type {non_none_type}. Units are only valid for numeric fields."
                )
                raise ValueError(msg)
            return

    # Check if field type is numeric
    if field_type not in numeric_types:
        msg = (
            f"Unit '{unit}' specified for non-numeric field '{field_name}' "
            f"of type {field_type}. Units are only valid for numeric fields."
        )
        raise ValueError(msg)


class XLSXFieldAnalyzer:
    """Analyzes Pydantic model fields for XLSX processing."""

    @staticmethod
    def analyze_model(model: type[BaseModel]) -> dict[str, FieldAnalysis]:
        """Analyze all fields in a Pydantic model."""
        if not issubclass(model, BaseModel):
            msg = f"Expected Pydantic BaseModel, got {type(model)}"
            raise TypeError(msg)

        field_analyses = {}
        for field_name, field_info in model.model_fields.items():
            field_analyses[field_name] = XLSXFieldAnalyzer.analyze_field(
                field_name, field_info, model
            )

        return field_analyses

    @staticmethod
    def analyze_field(
        field_name: str, field_info: Any, model: type[BaseModel]
    ) -> FieldAnalysis:
        """Analyze a single field in a Pydantic model."""
        field_type = field_info.annotation
        is_optional = XLSXFieldAnalyzer.is_optional_type(field_type)
        enum_values = XLSXFieldAnalyzer.get_enum_values(field_type)
        xlsx_metadata = XLSXFieldAnalyzer.extract_xlsx_metadata(field_info)

        # Validate unit usage
        if xlsx_metadata and xlsx_metadata.unit:
            _validate_unit_usage(field_name, field_type, xlsx_metadata.unit)

        return FieldAnalysis(
            name=field_name,
            field_type=field_type,
            is_optional=is_optional,
            enum_values=enum_values,
            xlsx_metadata=xlsx_metadata,
        )

    @staticmethod
    def is_optional_type(field_type: type) -> bool:
        """Check if a type is Optional (Union with None)."""
        origin = get_origin(field_type)
        if origin is Union:
            args = get_args(field_type)
            return len(args) == 2 and type(None) in args  # noqa: PLR2004
        # Handle Python 3.10+ union syntax (str | None)
        if hasattr(field_type, "__args__"):
            args = field_type.__args__
            return len(args) == 2 and type(None) in args  # noqa: PLR2004
        return False

    @staticmethod
    def get_enum_values(field_type: type) -> list[str]:
        """Extract enum values if the field type is an enum."""
        origin = get_origin(field_type)

        # Handle Union types (both typing.Union and Python 3.10+ | syntax)
        if origin is Union or (
            hasattr(field_type, "__args__") and len(get_args(field_type) or []) > 1
        ):
            args = get_args(field_type) if origin is Union else field_type.__args__
            non_none_args = [arg for arg in args if arg is not type(None)]
            if len(non_none_args) == 1:
                field_type = non_none_args[0]

        if isinstance(field_type, type) and issubclass(field_type, Enum):
            return [item.value for item in field_type]
        return []

    @staticmethod
    def extract_xlsx_metadata(field_info: Any) -> XLSXMetadata | None:
        """Extract XLSX metadata from field info."""
        # First check if Pydantic v2 has already processed the metadata
        if hasattr(field_info, "metadata") and field_info.metadata:
            for metadata_item in field_info.metadata:
                if isinstance(metadata_item, XLSXMetadata):
                    return metadata_item

        # Fallback: check if the annotation is still Annotated (older behavior)
        if (
            hasattr(field_info, "annotation")
            and get_origin(field_info.annotation) is Annotated
        ):
            args = get_args(field_info.annotation)
            if len(args) > 1:
                for metadata in args[1:]:
                    if isinstance(metadata, XLSXMetadata):
                        return metadata

        return None

    @staticmethod
    def get_field_display_name(
        field_name: str, xlsx_metadata: XLSXMetadata | None
    ) -> str:
        """Get display name for a field, using metadata if available."""
        if xlsx_metadata and xlsx_metadata.display_name:
            return xlsx_metadata.display_name
        return field_name.replace("_", " ").title()

    @staticmethod
    def get_field_description(
        field_info: Any, xlsx_metadata: XLSXMetadata | None
    ) -> str | None:
        """Get field description from metadata or field info."""
        if xlsx_metadata and xlsx_metadata.description:
            return xlsx_metadata.description
        return getattr(field_info, "description", None)

    @staticmethod
    def get_field_unit(xlsx_metadata: XLSXMetadata | None) -> str | None:
        """Get field unit from metadata."""
        if xlsx_metadata:
            return xlsx_metadata.unit
        return None

    @staticmethod
    def get_field_meaning(xlsx_metadata: XLSXMetadata | None) -> str | None:
        """Get field meaning from metadata."""
        if xlsx_metadata:
            return xlsx_metadata.meaning
        return None

    @staticmethod
    def has_custom_serializers(xlsx_metadata: XLSXMetadata | None) -> bool:
        """Check if field has custom serialization functions."""
        return (
            xlsx_metadata is not None
            and xlsx_metadata.xlsx_serializer is not None
            and xlsx_metadata.xlsx_deserializer is not None
        )

    @staticmethod
    def get_custom_serializer(xlsx_metadata: XLSXMetadata | None) -> Callable | None:
        """Get custom serializer function from metadata."""
        if xlsx_metadata:
            return xlsx_metadata.xlsx_serializer
        return None

    @staticmethod
    def get_custom_deserializer(xlsx_metadata: XLSXMetadata | None) -> Callable | None:
        """Get custom deserializer function from metadata."""
        if xlsx_metadata:
            return xlsx_metadata.xlsx_deserializer
        return None

    @staticmethod
    def get_separator_pattern(
        xlsx_metadata: XLSXMetadata | None,
    ) -> "SeparatorPattern | None":
        """Get separator pattern from metadata."""
        if xlsx_metadata:
            return xlsx_metadata.separator_pattern
        return None

    @staticmethod
    def get_requiredness_info(
        field_info: Any, field_analysis: "FieldAnalysis"
    ) -> tuple[bool, Any]:
        """Determine if field is required and get its default value.

        A field is required if:
        1. It is not an optional type (no Union with None)
        2. It has no default value

        Args:
            field_info: Pydantic field info object
            field_analysis: Analyzed field data

        Returns:
            Tuple of (is_required: bool, default_value: Any or PydanticUndefined)
        """
        # Check if field has default value
        has_default = (
            hasattr(field_info, "default")
            and field_info.default is not PydanticUndefined
        )

        # Field is required if not optional and has no default
        is_required = not field_analysis.is_optional and not has_default

        default_value = field_info.default if has_default else PydanticUndefined

        return is_required, default_value

    @staticmethod
    def format_requiredness_text(
        is_required: bool, default_value: Any, use_labels: bool = False
    ) -> str:
        """Format requiredness text for display in Excel.

        Args:
            is_required: Whether the field is required
            default_value: The field's default value (or PydanticUndefined)
            use_labels: If True, use "Required"/"Optional" labels (for table format).
                       If False, use "Yes"/"No" (for key-value format with header).

        Returns:
            - "Yes" or "Required" for required fields
            - "No" or "Optional" for optional fields with no default or trivial default
            - "No/Optional (default: value)" for optional fields with non-trivial default
        """
        required_text = "Required" if use_labels else "Yes"
        optional_text = "Optional" if use_labels else "No"

        if is_required:
            return required_text

        # Trivial defaults that don't need to be shown
        trivial_defaults = (None, "", [], {})

        if default_value is PydanticUndefined or default_value in trivial_defaults:
            return optional_text

        # Non-trivial default - format for display
        if isinstance(default_value, Enum):
            display_default = default_value.value
        elif isinstance(default_value, (list, dict)):
            display_default = json.dumps(default_value, default=str)
        else:
            display_default = str(default_value)

        return f"{optional_text} (default: {display_default})"


# Serialization engine
class XLSXSerializationEngine:
    """Centralized serialization/deserialization logic."""

    def __init__(self):
        self.field_analyzer = XLSXFieldAnalyzer()

    def serialize_value(self, value: Any, field_analysis: FieldAnalysis) -> Any:
        """Common serialization logic for both formats."""
        if value is None:
            return ""

        # Use custom serializer if available
        if (
            field_analysis.xlsx_metadata
            and field_analysis.xlsx_metadata.xlsx_serializer
        ):
            return field_analysis.xlsx_metadata.xlsx_serializer(value)

        # Handle Enum types BEFORE basic types (str, Enum subclasses are also str)
        if isinstance(value, Enum):
            return value.value

        # Handle basic types that Excel supports natively
        if isinstance(value, bool | int | float | str | date | datetime):
            return value

        # For complex types, convert to string representation
        if isinstance(value, list | dict):
            return json.dumps(value, default=str, ensure_ascii=False)

        return str(value)

    def deserialize_value(
        self,
        raw_value: Any,
        field_analysis: FieldAnalysis,
        model_class: type[BaseModel],
    ) -> Any:
        """Common deserialization logic for both formats."""
        if raw_value is None or raw_value == "":
            return None

        # Use custom deserializer if available
        if (
            field_analysis.xlsx_metadata
            and field_analysis.xlsx_metadata.xlsx_deserializer
        ):
            return field_analysis.xlsx_metadata.xlsx_deserializer(raw_value)

        # Handle enum values
        if field_analysis.enum_values:
            original_field_type = model_class.model_fields[
                field_analysis.name
            ].annotation
            enum_type: type | None
            if original_field_type is not None and XLSXFieldAnalyzer.is_optional_type(
                original_field_type
            ):
                args = get_args(original_field_type)
                enum_type = next(arg for arg in args if arg is not type(None))
            else:
                enum_type = original_field_type

            # Find the matching enum value
            if (
                enum_type
                and isinstance(enum_type, type)
                and issubclass(enum_type, Enum)
            ):
                for enum_item in enum_type:
                    if enum_item.value == raw_value:
                        return enum_item
            msg = f"Invalid enum value '{raw_value}'"
            raise ValueError(msg)

        # Handle basic types
        return self._convert_basic_types(raw_value, field_analysis.field_type)

    def _convert_basic_types(self, raw_value: Any, field_type: type | None) -> Any:
        """Convert raw value to basic Python types."""
        if field_type is None:
            return str(raw_value)

        type_converters = {
            str: lambda x: str(x),
            int: lambda x: int(x),
            float: lambda x: float(x),
            bool: lambda x: self._convert_bool(x),
            date: lambda x: self._convert_datetime(x, date),
            datetime: lambda x: self._convert_datetime(x, datetime),
        }

        if field_type in type_converters:
            return type_converters[field_type](raw_value)

        # Handle complex types
        if isinstance(raw_value, str):
            try:
                parsed = json.loads(raw_value)
                if isinstance(parsed, list | dict):
                    return parsed
            except (json.JSONDecodeError, ValueError):
                pass

        return str(raw_value)

    def _convert_bool(self, raw_value: Any) -> bool:
        """Convert raw value to boolean."""
        if isinstance(raw_value, bool):
            return raw_value
        return str(raw_value).lower() in ("true", "1", "yes", "on")

    def _convert_datetime(self, raw_value: Any, field_type: type) -> date | datetime:
        """Convert raw value to date or datetime."""
        if isinstance(raw_value, date | datetime):
            return raw_value
        if field_type is date:
            return date.fromisoformat(str(raw_value))
        if field_type is datetime:
            return datetime.fromisoformat(str(raw_value))
        return datetime.fromisoformat(str(raw_value))


# Separator patterns and converters
class EscapeMode(Enum):
    """Enumeration for handling separators within text elements."""

    NONE = "none"
    BACKSLASH = "backslash"
    QUOTE = "quote"


@dataclass(frozen=True)
class SeparatorPattern:
    """Configuration for separator patterns with escaping support."""

    separator: str
    prefix: str = ""
    suffix: str = ""
    escape_mode: EscapeMode = EscapeMode.NONE

    @property
    def full_pattern(self) -> str:
        """Get the full separator pattern with prefix and suffix."""
        return f"{self.prefix}{self.separator}{self.suffix}"

    def escape_item(self, item: str) -> str:
        """Escape an item according to the configured escape mode."""
        if self.escape_mode == EscapeMode.BACKSLASH:
            return item.replace(self.separator, f"\\{self.separator}")
        if self.escape_mode == EscapeMode.QUOTE:
            if self.separator in item:
                return f'"{item}"'
            return item
        return item

    def split_escaped(self, text: str) -> list[str]:
        """Split text while respecting escape sequences."""
        if self.escape_mode == EscapeMode.BACKSLASH:
            return self._split_backslash_escaped(text)
        if self.escape_mode == EscapeMode.QUOTE:
            return self._split_quote_escaped(text)
        return text.split(self.full_pattern)

    def _split_backslash_escaped(self, text: str) -> list[str]:
        """Split text with backslash escaping."""
        escaped_separator = f"\\{self.separator}"
        placeholder = "\x00ESCAPED_SEP\x00"

        # Replace escaped separators with placeholder
        temp_text = text.replace(escaped_separator, placeholder)

        # Split on unescaped separators
        parts = temp_text.split(self.full_pattern)

        # Restore escaped separators
        return [part.replace(placeholder, self.separator) for part in parts]

    def _split_quote_escaped(self, text: str) -> list[str]:
        """Split text with quote escaping."""
        parts = []
        current = ""
        in_quotes = False
        i = 0

        while i < len(text):
            char = text[i]

            if char == '"' and not in_quotes:
                in_quotes = True
            elif char == '"' and in_quotes:
                in_quotes = False
            elif (
                not in_quotes
                and text[i : i + len(self.full_pattern)] == self.full_pattern
            ):
                parts.append(current)
                current = ""
                i += len(self.full_pattern) - 1
            else:
                current += char

            i += 1

        if current:
            parts.append(current)

        return parts


class XLSXConverters:
    """Flexible converter functions with configurable separator patterns."""

    # Common patterns as class constants
    COMMA = SeparatorPattern(",", "", " ")  # ", "
    PIPE = SeparatorPattern("|", " ", " ")  # " | "
    SEMICOLON = SeparatorPattern(";", "", " ")  # "; "
    NEWLINE = SeparatorPattern("\n")  # "\n"
    TAB = SeparatorPattern("\t")  # "\t"
    ARROW = SeparatorPattern("->", " ", " ")  # " -> "
    DOUBLE_COLON = SeparatorPattern("::", "", "")  # "::"

    # Patterns with escaping
    COMMA_ESCAPED = SeparatorPattern(
        ",", "", " ", EscapeMode.BACKSLASH
    )  # ", " with \, escaping
    COMMA_QUOTED = SeparatorPattern(
        ",", "", " ", EscapeMode.QUOTE
    )  # ", " with "..." quoting
    PIPE_ESCAPED = SeparatorPattern(
        "|", " ", " ", EscapeMode.BACKSLASH
    )  # " | " with \| escaping

    @staticmethod
    def create_separated_converters(pattern: SeparatorPattern | str):
        """Create serializer/deserializer pair for any separator pattern.

        Args:
            pattern: SeparatorPattern object or simple separator string

        Returns:
            Tuple of (serializer_func, deserializer_func)
        """
        if isinstance(pattern, str):
            pattern = SeparatorPattern(pattern)

        full_separator = pattern.full_pattern

        def to_separated(value: list[Any]) -> str:
            """Convert list to separated string."""
            if value is None:
                return ""
            # Apply escaping to each item before joining (handle empty strings properly)
            escaped_items = [
                pattern.escape_item(str(item)) if item != "" else "" for item in value
            ]
            return full_separator.join(escaped_items)

        def from_separated(value: str) -> list[str]:
            """Convert separated string to list."""
            if not value or not value.strip():
                return []
            # Use the pattern's escape-aware splitting
            return pattern.split_escaped(str(value))

        return to_separated, from_separated

    @staticmethod
    def create_int_separated_converters(pattern: SeparatorPattern | str):
        """Create serializer/deserializer pair for integer lists."""
        if isinstance(pattern, str):
            pattern = SeparatorPattern(pattern)

        to_separated, from_separated_str = XLSXConverters.create_separated_converters(
            pattern
        )

        def to_int_separated(value: list[int]) -> str:
            """Convert list of integers to separated string."""
            if value is None:
                return ""
            return to_separated(value)

        def from_int_separated(value: str) -> list[int]:
            """Convert separated string to list of integers."""
            if not value or not value.strip():
                return []
            try:
                return [int(item) for item in from_separated_str(value)]
            except ValueError as e:
                msg = f"Cannot convert '{value}' to list of integers: {e}"
                raise ValueError(msg) from e

        return to_int_separated, from_int_separated

    # Convenience methods for common patterns
    @classmethod
    def comma_converters(cls):
        """Get comma-separated converters (', ')."""
        return cls.create_separated_converters(cls.COMMA)

    @classmethod
    def pipe_converters(cls):
        """Get pipe-separated converters (' | ')."""
        return cls.create_separated_converters(cls.PIPE)

    @classmethod
    def semicolon_converters(cls):
        """Get semicolon-separated converters ('; ')."""
        return cls.create_separated_converters(cls.SEMICOLON)

    @classmethod
    def newline_converters(cls):
        """Get newline-separated converters."""
        return cls.create_separated_converters(cls.NEWLINE)

    @classmethod
    def tab_converters(cls):
        """Get tab-separated converters."""
        return cls.create_separated_converters(cls.TAB)

    @classmethod
    def arrow_converters(cls):
        """Get arrow-separated converters (' -> ')."""
        return cls.create_separated_converters(cls.ARROW)

    @classmethod
    def double_colon_converters(cls):
        """Get double-colon-separated converters ('::')."""
        return cls.create_separated_converters(cls.DOUBLE_COLON)

    # Escaping variants
    @classmethod
    def comma_escaped_converters(cls):
        """Get comma-separated converters with backslash escaping."""
        return cls.create_separated_converters(cls.COMMA_ESCAPED)

    @classmethod
    def comma_quoted_converters(cls):
        """Get comma-separated converters with quote escaping."""
        return cls.create_separated_converters(cls.COMMA_QUOTED)

    @classmethod
    def pipe_escaped_converters(cls):
        """Get pipe-separated converters with backslash escaping."""
        return cls.create_separated_converters(cls.PIPE_ESCAPED)

    # Integer list converters
    @classmethod
    def comma_int_converters(cls):
        """Get comma-separated integer converters."""
        return cls.create_int_separated_converters(cls.COMMA)

    @classmethod
    def pipe_int_converters(cls):
        """Get pipe-separated integer converters."""
        return cls.create_int_separated_converters(cls.PIPE)

    # JSON converters
    @staticmethod
    def dict_to_json_string(value: dict) -> str:
        """Convert dictionary to JSON string."""
        if value is None:
            return ""
        return json.dumps(value)

    @staticmethod
    def json_string_to_dict(value: str) -> dict:
        """Convert JSON string to dictionary."""
        if not value or not value.strip():
            return {}
        try:
            return json.loads(value)
        except json.JSONDecodeError as e:
            msg = f"Cannot parse '{value}' as JSON: {e}"
            raise ValueError(msg) from e

    @staticmethod
    def list_to_json_string(value: list) -> str:
        """Convert list to JSON string."""
        if value is None:
            return ""
        return json.dumps(value)

    @staticmethod
    def json_string_to_list(value: str) -> list:
        """Convert JSON string to list."""
        if not value or not value.strip():
            return []
        try:
            return json.loads(value)
        except json.JSONDecodeError as e:
            msg = f"Cannot parse '{value}' as JSON: {e}"
            raise ValueError(msg) from e

    # Boolean converters
    @staticmethod
    def bool_to_yes_no(value: bool) -> str:
        """Convert boolean to Yes/No string."""
        return "Yes" if value else "No"

    @staticmethod
    def yes_no_to_bool(value: str) -> bool:
        """Convert Yes/No string to boolean."""
        value = value.strip().lower()
        if value in ("yes", "y", "true", "1"):
            return True
        if value in ("no", "n", "false", "0"):
            return False
        msg = (
            f"Cannot convert '{value}' to boolean. Expected Yes/No, True/False, or 1/0."
        )
        raise ValueError(msg)

    # Date converters
    @staticmethod
    def date_to_iso_string(value: date) -> str:
        """Convert date to ISO string (YYYY-MM-DD)."""
        return value.isoformat()

    @staticmethod
    def iso_string_to_date(value: str) -> date:
        """Convert ISO string to date."""
        try:
            return datetime.fromisoformat(value).date()
        except ValueError as e:
            msg = f"Cannot parse '{value}' as ISO date: {e}"
            raise ValueError(msg) from e

    # DateTime converters
    @staticmethod
    def datetime_to_iso_string(value: datetime) -> str:
        """Convert datetime to ISO string without timezone offset."""
        return value.replace(tzinfo=None).isoformat()

    @staticmethod
    def iso_string_to_datetime(value: str) -> datetime:
        """Convert ISO string to datetime, assuming UTC if no timezone."""
        try:
            dt = datetime.fromisoformat(value)
            # If no timezone info, assume UTC
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        except ValueError as e:
            msg = f"Cannot parse '{value}' as ISO datetime: {e}"
            raise ValueError(msg) from e
        else:
            return dt


def _create_separator_metadata(
    pattern: SeparatorPattern | str, description: str | None = None
) -> XLSXMetadata:
    """Helper to create XLSXMetadata with separator converters."""
    to_separated, from_separated = XLSXConverters.create_separated_converters(pattern)
    return XLSXMetadata(
        description=description,
        separator_pattern=pattern
        if isinstance(pattern, SeparatorPattern)
        else SeparatorPattern(pattern),
        xlsx_serializer=to_separated,
        xlsx_deserializer=from_separated,
    )


# Pre-configured type aliases for common use cases
XLSXCommaList = Annotated[
    list[str], _create_separator_metadata(XLSXConverters.COMMA, "Comma-separated list")
]
XLSXPipeList = Annotated[
    list[str], _create_separator_metadata(XLSXConverters.PIPE, "Pipe-separated list")
]
XLSXSemicolonList = Annotated[
    list[str],
    _create_separator_metadata(XLSXConverters.SEMICOLON, "Semicolon-separated list"),
]
XLSXNewlineList = Annotated[
    list[str],
    _create_separator_metadata(XLSXConverters.NEWLINE, "Newline-separated list"),
]
XLSXTabList = Annotated[
    list[str], _create_separator_metadata(XLSXConverters.TAB, "Tab-separated list")
]
XLSXArrowList = Annotated[
    list[str], _create_separator_metadata(XLSXConverters.ARROW, "Arrow-separated list")
]
XLSXDoubleColonList = Annotated[
    list[str],
    _create_separator_metadata(
        XLSXConverters.DOUBLE_COLON, "Double-colon-separated list"
    ),
]

# Escaping variants
XLSXCommaEscapedList = Annotated[
    list[str],
    _create_separator_metadata(
        XLSXConverters.COMMA_ESCAPED, "Comma-separated list with backslash escaping"
    ),
]
XLSXCommaQuotedList = Annotated[
    list[str],
    _create_separator_metadata(
        XLSXConverters.COMMA_QUOTED, "Comma-separated list with quote escaping"
    ),
]
XLSXPipeEscapedList = Annotated[
    list[str],
    _create_separator_metadata(
        XLSXConverters.PIPE_ESCAPED, "Pipe-separated list with backslash escaping"
    ),
]

# JSON types
XLSXJSONDict = Annotated[
    dict,
    XLSXMetadata(
        description="Dictionary as JSON string",
        xlsx_serializer=XLSXConverters.dict_to_json_string,
        xlsx_deserializer=XLSXConverters.json_string_to_dict,
    ),
]
XLSXJSONList = Annotated[
    list,
    XLSXMetadata(
        description="List as JSON string",
        xlsx_serializer=XLSXConverters.list_to_json_string,
        xlsx_deserializer=XLSXConverters.json_string_to_list,
    ),
]

# Boolean types
XLSXYesNoBoolean = Annotated[
    bool,
    XLSXMetadata(
        description="Boolean as Yes/No string",
        xlsx_serializer=XLSXConverters.bool_to_yes_no,
        xlsx_deserializer=XLSXConverters.yes_no_to_bool,
    ),
]

# Date types
XLSXISODate = Annotated[
    date,
    XLSXMetadata(
        description="Date as ISO string (YYYY-MM-DD)",
        xlsx_serializer=XLSXConverters.date_to_iso_string,
        xlsx_deserializer=XLSXConverters.iso_string_to_date,
    ),
]
XLSXISODateTime = Annotated[
    datetime,
    XLSXMetadata(
        description="DateTime as ISO string",
        xlsx_serializer=XLSXConverters.datetime_to_iso_string,
        xlsx_deserializer=XLSXConverters.iso_string_to_datetime,
    ),
]


# Base configuration classes
@dataclass
class XLSXConfig:
    """Base configuration for XLSX operations."""

    title: str | None = None
    start_row: int = 1
    start_column: int = 1
    include_fields: set[str] | None = None
    exclude_fields: set[str] | None = None
    field_order: list[str] | None = None
    metadata_visibility: MetadataToggleConfig | None = None

    def should_include_field(self, field_name: str) -> bool:
        """Check if a field should be included based on configuration."""
        if self.include_fields is not None:
            return field_name in self.include_fields
        if self.exclude_fields is not None:
            return field_name not in self.exclude_fields
        return True

    def get_ordered_fields(self, available_fields: list[str]) -> list[str]:
        """Get fields in the specified order, filtering as needed."""
        if self.field_order is not None:
            # Start with specified order
            ordered = [f for f in self.field_order if f in available_fields]
            # Add remaining fields that aren't in field_order
            remaining = [f for f in available_fields if f not in self.field_order]
            ordered.extend(remaining)
            result = ordered
        else:
            result = available_fields

        # Apply include/exclude filters
        return [f for f in result if self.should_include_field(f)]


# Base formatter class
class XLSXFormatter(ABC):
    """Base formatter interface for different Excel layouts."""

    def __init__(self, config: XLSXConfig):
        self.config = config
        self.serialization_engine = XLSXSerializationEngine()
        self.row_calculator = XLSXRowCalculator(config)

    @abstractmethod
    def format_export(
        self, worksheet: Worksheet, data: Any, fields: list[FieldAnalysis]
    ) -> None:
        """Format data for export to Excel."""

    @abstractmethod
    def parse_import(
        self,
        worksheet: Worksheet,
        fields: list[FieldAnalysis],
        model_class: type[BaseModel],
    ) -> Any:
        """Parse Excel data for import."""

    def _get_field_display_name(self, field_analysis: FieldAnalysis) -> str:
        """Get display name for a field, with fallback to auto-generated title case."""
        if field_analysis.xlsx_metadata and field_analysis.xlsx_metadata.display_name:
            return field_analysis.xlsx_metadata.display_name
        return " ".join(word.capitalize() for word in field_analysis.name.split("_"))

    def _apply_data_cell_formatting(self, cell, field_analysis: FieldAnalysis) -> None:
        """Apply consistent formatting to data cells (not headers)."""
        # Skip formatting if disabled in configuration
        if (
            hasattr(self.config, "enable_cell_formatting")
            and not self.config.enable_cell_formatting
        ):
            return

        # Vertical alignment: center for all data cells
        vertical_align = "center"

        # Text wrapping: enabled for string fields
        wrap_text = self._should_wrap_text(field_analysis.field_type)

        # Apply formatting (preserve Excel's default horizontal alignment)
        cell.alignment = Alignment(vertical=vertical_align, wrap_text=wrap_text)

    def _should_wrap_text(self, field_type: type | None) -> bool:
        """Determine if a field should have text wrapping enabled."""
        if not field_type:
            return False

        # Check for direct string type
        if field_type is str:
            return True

        # Handle Union types (including Optional)
        origin = get_origin(field_type)
        if origin is Union or (
            hasattr(field_type, "__args__") and len(get_args(field_type) or []) > 1
        ):
            args = get_args(field_type) if origin is Union else field_type.__args__
            non_none_args = [arg for arg in args if arg is not type(None)]

            # If there's exactly one non-None type, check that type
            if len(non_none_args) == 1:
                return self._should_wrap_text(non_none_args[0])

            # If multiple non-None types, check if any is string
            for arg in non_none_args:
                if self._should_wrap_text(arg):
                    return True

        # Handle Annotated types (e.g., Annotated[str, XLSXMetadata(...)])
        if get_origin(field_type) is Annotated:
            # For Annotated types, the first argument is the actual type
            args = get_args(field_type)
            if args:
                return self._should_wrap_text(args[0])

        return False

    def _format_header_text(self, field_analysis: FieldAnalysis) -> str:
        """Format field name as readable header text."""
        return self._get_field_display_name(field_analysis)

    def _add_title(self, worksheet: Worksheet, title: str) -> None:
        """Add title to worksheet."""
        if title:
            title_row = self.row_calculator.get_title_row()
            cell = worksheet.cell(row=title_row, column=self.config.start_column)
            cell.value = title
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="left")

    def _auto_adjust_columns(self, worksheet: Worksheet, num_columns: int) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx in range(1, num_columns + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if not isinstance(cell, MergedCell) and cell.value and cell.row > 1:
                        max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _create_validation_list_range(
        self, workbook: Workbook, field_name: str, values: list[str]
    ) -> str:
        """Create a named range for validation values on a hidden sheet.

        Args:
            workbook: The workbook to add the hidden sheet to.
            field_name: Name of the field (used for named range).
            values: List of enum values.

        Returns:
            Named range reference for data validation formula.
        """
        # Hidden sheet name for validation lists
        hidden_sheet_name = "_ValidationLists"

        # Get or create hidden sheet
        if hidden_sheet_name not in workbook.sheetnames:
            hidden_sheet = workbook.create_sheet(hidden_sheet_name)
            hidden_sheet.sheet_state = "hidden"
            # Track which column to use next
            hidden_sheet["A1"] = "Validation Lists (hidden)"
            workbook._validation_list_col = 2  # Start from column B
        else:
            hidden_sheet = workbook[hidden_sheet_name]

        # Get next available column
        if not hasattr(workbook, "_validation_list_col"):
            workbook._validation_list_col = 2
        col_idx = workbook._validation_list_col
        workbook._validation_list_col += 1

        col_letter = get_column_letter(col_idx)

        # Write values to column
        for row_idx, value in enumerate(values, start=1):
            hidden_sheet.cell(row=row_idx, column=col_idx, value=value)

        # Create named range - sanitize field name for Excel
        range_name = f"ValidationList_{field_name.replace(' ', '_')}"
        range_ref = f"'{hidden_sheet_name}'!${col_letter}$1:${col_letter}${len(values)}"

        # Remove existing definition if present (in case of regeneration)
        if range_name in workbook.defined_names:
            del workbook.defined_names[range_name]

        workbook.defined_names[range_name] = DefinedName(
            name=range_name, attr_text=range_ref
        )

        # Named range reference without = prefix for data validation
        return range_name

    def _add_enum_validation(
        self,
        worksheet: Worksheet,
        field_analysis: FieldAnalysis,
        validation_range: str,
    ) -> None:
        """Add dropdown validation for an enum field.

        Handles formula length limits by creating a hidden sheet with named
        ranges for long enum lists.

        Args:
            worksheet: The worksheet to add validation to.
            field_analysis: Field analysis with enum_values.
            validation_range: Cell range to apply validation (e.g., "B5" or "C3:C100").
        """
        if not field_analysis.enum_values:
            return

        # Check if inline formula would exceed Excel's limit
        inline_formula = f'"{",".join(field_analysis.enum_values)}"'

        if len(inline_formula) <= EXCEL_DV_FORMULA_LIMIT:
            # Use inline formula (shorter lists)
            formula1 = inline_formula
        else:
            # Use named range on hidden sheet (longer lists)
            formula1 = self._create_validation_list_range(
                worksheet.parent,
                field_analysis.name,
                field_analysis.enum_values,
            )

        dv = DataValidation(
            type="list",
            formula1=formula1,
            allow_blank=field_analysis.is_optional,
        )
        # Excel limits error message to 255 characters
        error_msg = (
            f"Invalid value. Must be one of: {', '.join(field_analysis.enum_values)}"
        )
        if len(error_msg) > EXCEL_DV_FORMULA_LIMIT:
            error_msg = "Invalid value. Please select from the dropdown list."
        dv.error = error_msg
        dv.errorTitle = "Invalid Input"

        worksheet.add_data_validation(dv)
        dv.add(validation_range)


# Base processor class
class XLSXProcessor(ABC):
    """Base class for all XLSX processors."""

    def __init__(self, config: XLSXConfig, formatter: XLSXFormatter):
        self.config = config
        self.formatter = formatter
        self.field_analyzer = XLSXFieldAnalyzer()

    @abstractmethod
    def export(self, data: Any, filepath: Path, sheet_name: str | None = None) -> None:
        """Export data to XLSX file."""

    @abstractmethod
    def import_data(
        self,
        filepath: Path,
        model_class: type[BaseModel],
        sheet_name: str | None = None,
    ) -> Any:
        """Import data from XLSX file."""

    def _filter_and_order_fields(
        self, fields: list[FieldAnalysis]
    ) -> list[FieldAnalysis]:
        """Filter and order fields based on configuration."""
        # Filter fields
        filtered_fields = [
            field for field in fields if self.config.should_include_field(field.name)
        ]

        # Order fields if custom order is specified
        if self.config.field_order:
            ordered_fields = []
            field_dict = {field.name: field for field in filtered_fields}

            # Add fields in specified order
            for field_name in self.config.field_order:
                if field_name in field_dict:
                    ordered_fields.append(field_dict[field_name])
                    del field_dict[field_name]

            # Add remaining fields not in order list
            ordered_fields.extend(field_dict.values())
            return ordered_fields

        return filtered_fields

    def _prepare_workbook(
        self, filepath: Path, sheet_name: str
    ) -> tuple[Workbook, Worksheet]:
        """Create or load workbook and prepare worksheet for export.

        Handles:
        - Loading existing workbook or creating new one
        - Cleaning up default "Sheet"
        - Removing existing sheet with same name
        - Creating new sheet with specified name

        Args:
            filepath: Path to the XLSX file.
            sheet_name: Name for the worksheet.

        Returns:
            Tuple of (workbook, worksheet).
        """
        if filepath.exists() and filepath.stat().st_size > 0:
            try:
                workbook = load_workbook(filepath)
            except Exception:
                # If file exists but is not a valid Excel file, create new workbook
                workbook = Workbook()
                if "Sheet" in workbook.sheetnames:
                    workbook.remove(workbook["Sheet"])
        else:
            workbook = Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        # Remove existing sheet with same name if it exists
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

        worksheet = workbook.create_sheet(title=sheet_name)
        return workbook, worksheet


# Row calculation utilities
class XLSXRowCalculator:
    """Utility class to calculate row positions for different XLSX elements."""

    def __init__(self, config: XLSXConfig):
        self.config = config

    def get_title_row(self) -> int:
        """Get the row number for the title."""
        return self.config.start_row

    def get_first_content_row(self) -> int:
        """Get the first row where content begins.

        For key-value format: this is the header row (Field|Value|Unit|Description|Meaning).
        For table format: this is the first metadata row (meanings, descriptions, or units).
        """
        row = self.config.start_row
        if self.config.title:
            row += 2  # title + 1 empty row
        return row

    def get_meaning_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the row number for field meanings."""
        row = self.config.start_row
        if self.config.title:
            row += 2  # title + empty row
        return row

    def get_description_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the row number for field descriptions."""
        row = self.get_meaning_row(fields)
        if fields and self._should_show_meanings(fields):
            row += 1
        return row

    def get_unit_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the row number for field units."""
        row = self.get_description_row(fields)
        if fields and self._should_show_descriptions(fields):
            row += 1
        return row

    def get_requiredness_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the row number for field requiredness.

        Position: After unit row, before header row.
        """
        row = self.get_unit_row(fields)
        if fields and self._should_show_units(fields):
            row += 1
        return row

    def get_header_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the row number for headers."""
        row = self.get_requiredness_row(fields)
        if fields and self._should_show_requiredness(fields):
            row += 1
        return row

    def get_data_start_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the starting row number for data (after headers)."""
        return self.get_header_row(fields) + 1

    def get_table_start_row(self, fields: list[FieldAnalysis] | None = None) -> int:
        """Get the starting row for Excel table (same as header row)."""
        return self.get_header_row(fields)

    def get_table_end_row(
        self, fields: list[FieldAnalysis] | None = None, data_rows: int = 0
    ) -> int:
        """Get the ending row for Excel table (header + data rows)."""
        return self.get_table_start_row(fields) + data_rows

    @staticmethod
    def _has_any_meanings(fields: list[FieldAnalysis]) -> bool:
        """Check if any field has meanings defined."""
        return any(
            field_analysis.xlsx_metadata and field_analysis.xlsx_metadata.meaning
            for field_analysis in fields
        )

    @staticmethod
    def _has_any_descriptions(fields: list[FieldAnalysis]) -> bool:
        """Check if any field has descriptions defined."""
        return any(
            field_analysis.xlsx_metadata and field_analysis.xlsx_metadata.description
            for field_analysis in fields
        )

    @staticmethod
    def _has_any_units(fields: list[FieldAnalysis]) -> bool:
        """Check if any field has units defined."""
        return any(
            field_analysis.xlsx_metadata and field_analysis.xlsx_metadata.unit
            for field_analysis in fields
        )

    def _should_show_meanings(self, fields: list[FieldAnalysis]) -> bool:
        """Check if meanings row/column should be shown based on config and data."""
        visibility = (
            self.config.metadata_visibility.meaning
            if self.config.metadata_visibility
            else MetadataVisibility.AUTO
        )
        if visibility == MetadataVisibility.HIDE:
            return False
        if visibility == MetadataVisibility.SHOW:
            return True
        # AUTO: show if any field has meanings
        return self._has_any_meanings(fields)

    def _should_show_descriptions(self, fields: list[FieldAnalysis]) -> bool:
        """Check if descriptions row/column should be shown based on config and data."""
        visibility = (
            self.config.metadata_visibility.description
            if self.config.metadata_visibility
            else MetadataVisibility.AUTO
        )
        if visibility == MetadataVisibility.HIDE:
            return False
        if visibility == MetadataVisibility.SHOW:
            return True
        # AUTO: show if any field has descriptions
        return self._has_any_descriptions(fields)

    def _should_show_units(self, fields: list[FieldAnalysis]) -> bool:
        """Check if units row/column should be shown based on config and data."""
        visibility = (
            self.config.metadata_visibility.unit
            if self.config.metadata_visibility
            else MetadataVisibility.AUTO
        )
        if visibility == MetadataVisibility.HIDE:
            return False
        if visibility == MetadataVisibility.SHOW:
            return True
        # AUTO: show if any field has units
        return self._has_any_units(fields)

    def _should_show_requiredness(self, fields: list[FieldAnalysis]) -> bool:
        """Check if requiredness row/column should be shown based on config.

        Unlike other metadata, requiredness is derived from field definitions
        rather than explicit XLSXMetadata. In AUTO mode, we don't show it
        to maintain backwards compatibility - users must explicitly set SHOW.
        """
        visibility = (
            self.config.metadata_visibility.requiredness
            if self.config.metadata_visibility
            else MetadataVisibility.AUTO
        )
        # HIDE or AUTO: don't show (backwards compatible)
        return visibility == MetadataVisibility.SHOW


# =============================================================================
# Table Length Adjustment
# =============================================================================


def adjust_table_length(
    worksheet: Worksheet,
    table_name: str,
    rows_pre_allocated: int = 0,
) -> None:
    """Expand a single table to include all filled rows plus pre-allocated rows.

    The table is expanded to include the last row with content plus a number of
    additional empty rows specified by rows_pre_allocated.

    Note that tables are not shrunk by this method if they contain empty rows
    at the end.

    Args:
        worksheet: The worksheet containing the table (in-memory).
        table_name: The name of the table to adjust.
        rows_pre_allocated: Number of empty rows to add after last content row.

    Note:
        Caller is responsible for saving the workbook after calling this function.
    """
    if table_name not in worksheet.tables:
        logger.warning(
            'Table "%s" not found in worksheet "%s".', table_name, worksheet.title
        )
        return

    old_range = worksheet.tables[table_name].ref
    start_col, start_row, end_col, end_row = range_boundaries(old_range)
    start = old_range.split(":")[0]

    # Find last row with content in table
    last_content_row = start_row  # At minimum, keep header row
    for row in range(start_row + 1, end_row + 1):
        row_has_content = any(
            worksheet.cell(row=row, column=col).value
            for col in range(start_col, end_col + 1)
        )
        if row_has_content:
            last_content_row = row

    new_last_row = max(last_content_row + rows_pre_allocated, worksheet.max_row)
    adjusted = f"{start}:{get_column_letter(end_col)}{new_last_row}"

    if adjusted != old_range:
        # Expanding the table is not possible with openpyxl. Instead a too
        # short table is removed, and a new adjusted table is created.
        style = copy(worksheet.tables[table_name].tableStyleInfo)
        del worksheet.tables[table_name]
        newtab = Table(displayName=table_name, ref=adjusted)
        newtab.tableStyleInfo = style
        worksheet.add_table(newtab)
        logger.debug(
            'Adjusted table "%s" in sheet "%s" from {%s} to {%s}.',
            table_name,
            worksheet.title,
            old_range,
            adjusted,
        )

    # Reset row height for all table rows including header to default
    for row in range(start_row, new_last_row + 1):
        worksheet.row_dimensions[row].height = None


def adjust_all_tables_length(
    wb_path: Path,
    rows_pre_allocated: dict[str, int] | int = 0,
    active_sheet: str | None = None,
) -> None:
    """Adjust length of all tables in a workbook file.

    This is a convenience wrapper that loads a workbook, adjusts all tables,
    and saves the workbook.

    Args:
        wb_path: Path to the Excel workbook file.
        rows_pre_allocated: Either:
            - int: same value applied to all sheets
            - dict[str, int]: per-sheet values where keys are sheet names
              (missing sheets get 0)
        active_sheet: Sheet name to set as active (visible when file opens).
            If None or sheet doesn't exist, no change is made.
    """
    wb = load_workbook(wb_path)

    # Normalize rows_pre_allocated to dict
    if isinstance(rows_pre_allocated, int):
        rows_dict = dict.fromkeys(wb.sheetnames, rows_pre_allocated)
    elif isinstance(rows_pre_allocated, dict):
        rows_dict = {sheet: rows_pre_allocated.get(sheet, 0) for sheet in wb.sheetnames}
    else:
        msg = "rows_pre_allocated must be an int or a dict"
        raise TypeError(msg)

    for ws in wb.worksheets:
        for table_name in list(ws.tables):
            adjust_table_length(ws, table_name, rows_pre_allocated=rows_dict[ws.title])

    # Set active sheet before saving
    if active_sheet and active_sheet in wb.sheetnames:
        wb.active = wb[active_sheet]

    wb.save(wb_path)
    wb.close()
