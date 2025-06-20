import os
import shutil
from pathlib import Path
from unittest import mock

import pytest
from curies import Converter
from openpyxl import load_workbook
from rdflib import SKOS, Graph, Literal, URIRef, compare

import voc4cat
from tests.test_cli import CS_SIMPLE
from voc4cat import convert
from voc4cat.utils import ConversionError


def test_empty_template():
    test_file = Path(voc4cat.__file__).parent / "blank_043.xlsx"

    assert test_file.is_file()
    with pytest.raises(ConversionError) as e:
        convert.excel_to_rdf(
            test_file,
            output_type="file",
        )
    assert "11 validation errors for ConceptScheme" in str(e)


def test_simple():
    g = convert.excel_to_rdf(
        Path(__file__).parent / "templ_versions" / "043_simple_valid.xlsx",
        output_type="graph",
    )
    assert len(g) == 147  # noqa: PLR2004
    assert (
        URIRef(
            "http://resource.geosciml.org/classifierscheme/cgi/2016.01/particletype"
        ),
        SKOS.prefLabel,
        Literal("Particle Type", lang="en"),
    ) in (g or []), "PrefLabel for vocab is not correct"
    assert (
        URIRef("http://resource.geosciml.org/classifier/cgi/particletype/bioclast"),
        SKOS.historyNote,
        Literal("NADM SLTTs 2004", lang="en"),
    ) in (g or []), "Provenance for vocab is not correct"


def test_missing_iri_concept(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_SIMPLE, tmp_path)
    monkeypatch.chdir(tmp_path)
    # change excel file: Delete vocabulary base IRI
    wb = load_workbook(filename=CS_SIMPLE)
    ws = wb["Additional Concept Features"]
    ws.cell(row=7, column=1).value = "ex:test07"
    new_filename = "bad_iri_in_AdditionalConceptFeatures.xlsx"
    wb.save(new_filename)
    wb.close()

    wb = load_workbook(filename=new_filename, read_only=True, data_only=True)
    with pytest.raises(ConversionError):
        convert.excel_to_rdf(
            Path(new_filename),
            output_type="graph",
        )


@mock.patch.dict(os.environ, clear=True)  # required to hide gh-action environment vars
def test_exhaustive_template_is_isomorphic(temp_config):
    # Load/prepare an a config with required prefix definition
    config = temp_config
    config.curies_converter = Converter.from_prefix_map({"ex": "http://example.org/"})
    config.CURIES_CONVERTER_MAP["043_exhaustive_example_perfect_output"] = (
        config.curies_converter
    )
    config.CURIES_CONVERTER_MAP["043_exhaustive_example"] = config.curies_converter

    g1 = Graph().parse(
        Path(__file__).parent
        / "templ_versions"
        / "043_exhaustive_example_perfect_output.ttl"
    )
    g2 = convert.excel_to_rdf(
        Path(__file__).parent / "templ_versions" / "043_exhaustive_example.xlsx",
        output_type="graph",
    )

    # to debug differences
    in_both, in_first, in_second = compare.graph_diff(g1, g2)
    print("\nOnly in 1st ==>\n", in_first.serialize(format="turtle"))
    print("Only in 2nd ==>\n", in_second.serialize(format="turtle"))

    assert compare.isomorphic(g1, g2), "Graphs are not Isomorphic"


@mock.patch.dict(os.environ, clear=True)  # required to hide gh-action environment vars
def test_rdf_to_excel():
    g1 = Graph().parse(
        Path(__file__).parent
        / "templ_versions"
        / "043_exhaustive_example_perfect_output.ttl"
    )
    convert.rdf_to_excel(
        Path(__file__).parent
        / "templ_versions"
        / "043_exhaustive_example_perfect_output.ttl",
    )
    g2 = convert.excel_to_rdf(
        Path(__file__).parent
        / "templ_versions"
        / "043_exhaustive_example_perfect_output.xlsx",
        output_type="graph",
    )
    # clean up file
    (
        Path(__file__).parent
        / "templ_versions"
        / "043_exhaustive_example_perfect_output.xlsx"
    ).unlink(missing_ok=True)
    # to debug differences
    in_both, in_first, in_second = compare.graph_diff(g1, g2)
    print("\nOnly in 1st ==>\n", in_first.serialize(format="turtle"))
    print("Only in 2nd ==>\n", in_second.serialize(format="turtle"))

    assert compare.isomorphic(g1, g2), "Graphs are not Isomorphic"
