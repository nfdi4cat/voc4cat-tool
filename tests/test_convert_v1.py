"""Tests for the convert_v1 module.

These tests verify that the RDF to v1.0 xlsx converter works correctly,
extracting data from RDF graphs and producing valid v1.0 template xlsx files.
"""

import logging
import re
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import ValidationError
from rdflib import (
    DCAT,
    DCTERMS,
    FOAF,
    OWL,
    PROV,
    RDF,
    RDFS,
    SDO,
    SKOS,
    XSD,
    Graph,
    Literal,
    Namespace,
    URIRef,
)

from tests.conftest import make_vocab_config_from_rdf
from voc4cat.config import Checks, Vocab
from voc4cat.convert_v1 import (
    AggregatedCollection,
    AggregatedConcept,
    aggregate_collections,
    aggregate_concepts,
    build_collection_graph,
    build_concept_graph,
    build_concept_scheme_graph,
    build_concept_to_collections_map,
    build_concept_to_ordered_collections_map,
    build_curies_converter_from_prefixes,
    build_entity_graph,
    config_to_concept_scheme_v1,
    excel_to_rdf_v1,
    extract_collections_from_rdf,
    extract_concept_scheme_from_rdf,
    extract_concepts_from_rdf,
    extract_identifier,
    extract_mappings_from_rdf,
    parse_name_url,
    parse_ordered_collection_positions,
    rdf_concept_scheme_to_v1,
    rdf_concepts_to_v1,
    rdf_mappings_to_v1,
    rdf_to_excel_v1,
    string_to_collection_obsoletion_enum,
    string_to_concept_obsoletion_enum,
    string_to_ordered_enum,
)
from voc4cat.convert_v1_helpers import (
    OBSOLETE_PREFIX,
    build_id_range_info,
    build_provenance_url,
    derive_contributors,
    extract_entity_id_from_iri,
    extract_github_repo_from_url,
    extract_used_ids,
    format_contributor_string,
    validate_deprecation,
)
from voc4cat.models_v1 import (
    TEMPLATE_VERSION,
    CollectionObsoletionReason,
    CollectionV1,
    ConceptObsoletionReason,
    ConceptSchemeV1,
    ConceptV1,
    OrderedChoice,
    PrefixV1,
)

EX = Namespace("http://example.org/")

# Test data paths
TEST_DATA_DIR = Path(__file__).parent / "data"

CS_SIMPLE_TTL = TEST_DATA_DIR / "concept-scheme-simple.ttl"
V1_COMPREHENSIVE_TTL = TEST_DATA_DIR / "v1-test-comprehensive.ttl"
VOCAB_043_TTL = TEST_DATA_DIR / "vocab-043-test.ttl"


class TestExtractConceptScheme:
    """Tests for extract_concept_scheme_from_rdf function."""

    def test_extract_basic_concept_scheme(self):
        """Test extracting basic concept scheme fields."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        data = extract_concept_scheme_from_rdf(graph)

        assert data["vocabulary_iri"] == "http://example.org/test/"
        assert data["title"] == "voc4cat-test-data"
        assert "concept scheme for unit testing" in data["description"].lower()
        assert data["created_date"] == "2022-12-01"
        assert data["modified_date"] == "2022-12-01"

    def test_extract_043_concept_scheme(self):
        """Test extracting concept scheme from 0.4.3 format vocabulary."""
        graph = Graph().parse(VOCAB_043_TTL, format="turtle")
        data = extract_concept_scheme_from_rdf(graph)

        # URI should have trailing slash (VocPub 5.2 convention)
        assert data["vocabulary_iri"] == "http://example.org/test-vocab/"
        assert "test vocabulary 043" in data["title"].lower()
        assert data["created_date"] == "2023-06-29"


class TestExtractConcepts:
    """Tests for extract_concepts_from_rdf function."""

    def test_extract_concepts_from_simple(self):
        """Test extracting concepts from simple test file."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # Should have 6 concepts (test01-test06)
        assert len(concepts) >= 6

        # Check a specific concept exists
        assert "http://example.org/test01" in concepts

    def test_concept_has_required_fields(self):
        """Test that extracted concepts have required fields."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # Get first concept's first language data
        first_iri = next(iter(concepts.keys()))
        first_lang = next(iter(concepts[first_iri].keys()))
        data = concepts[first_iri][first_lang]

        assert "preferred_label" in data
        assert "definition" in data
        assert "alternate_labels" in data
        assert "parent_iris" in data

    def test_concept_hierarchy(self):
        """Test that parent IRIs are extracted correctly."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # test02 has test01 as parent (broader)
        assert "http://example.org/test02" in concepts
        test02_data = concepts["http://example.org/test02"]
        first_lang = next(iter(test02_data.keys()))
        assert "http://example.org/test01" in test02_data[first_lang]["parent_iris"]

    def test_english_language_first(self):
        """Test that English is listed first when present."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        for iri, lang_data in concepts.items():
            if "en" in lang_data:  # pragma: no branch
                # English should be first
                first_lang = next(iter(lang_data.keys()))
                assert first_lang == "en", f"English not first for {iri}"


class TestExtractCollections:
    """Tests for extract_collections_from_rdf function."""

    def test_extract_collection_from_simple(self):
        """Test extracting collection from simple test file."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        collections = extract_collections_from_rdf(graph)

        # Should have at least 1 collection (test10)
        assert len(collections) >= 1

        # Check specific collection exists
        assert "http://example.org/test10" in collections

    def test_collection_has_required_fields(self):
        """Test that extracted collections have required fields."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        collections = extract_collections_from_rdf(graph)

        first_iri = next(iter(collections.keys()))
        first_lang = next(iter(collections[first_iri].keys()))
        data = collections[first_iri][first_lang]

        assert "preferred_label" in data
        assert "definition" in data
        assert "members" in data


class TestExtractMappings:
    """Tests for extract_mappings_from_rdf function."""

    def test_extract_mappings_from_simple(self):
        """Test extracting mappings from simple test file."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        mappings = extract_mappings_from_rdf(graph)

        # test01 has various mappings
        assert "http://example.org/test01" in mappings
        data = mappings["http://example.org/test01"]
        # Check that mapping types are present
        assert "related_matches" in data
        assert "close_matches" in data
        assert "exact_matches" in data
        assert "narrower_matches" in data
        assert "broader_matches" in data

    def test_only_concepts_with_mappings_included(self):
        """Test that only concepts with mappings are in the result."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        mappings = extract_mappings_from_rdf(graph)

        for iri, data in mappings.items():
            has_any = any(
                [
                    data.get("related_matches"),
                    data.get("close_matches"),
                    data.get("exact_matches"),
                    data.get("narrower_matches"),
                    data.get("broader_matches"),
                ]
            )
            assert has_any, f"Concept {iri} has no mappings but is in result"


class TestConceptToCollectionsMap:
    """Tests for build_concept_to_collections_map function."""

    def test_build_map_from_simple(self):
        """Test building concept-to-collections map."""
        graph = Graph().parse(CS_SIMPLE_TTL, format="turtle")
        c2c_map = build_concept_to_collections_map(graph)

        # test01-test04 are members of test10 collection
        assert c2c_map, "c2c_map should not be empty"
        # At least one concept should be in a collection
        assert any(collections for collections in c2c_map.values())


class TestModelConversion:
    """Tests for model conversion functions."""

    def test_concept_scheme_to_v1(self):
        """Test converting concept scheme data to v1.0 model."""
        data = {
            "vocabulary_iri": "http://example.org/test/",
            "title": "Test Vocabulary",
            "description": "A test vocabulary",
            "created_date": "2024-01-01",
            "modified_date": "2024-01-02",
        }

        model = rdf_concept_scheme_to_v1(data)

        assert model.template_version == TEMPLATE_VERSION
        assert model.vocabulary_iri == "http://example.org/test/"
        assert model.title == "Test Vocabulary"

    def test_concepts_to_v1_single_language(self, temp_config):
        """Test converting concepts with single language."""
        concepts_data = {
            "http://example.org/c1": {
                "en": {
                    "preferred_label": "Concept 1",
                    "definition": "Definition 1",
                    "alternate_labels": ["alt1", "alt2"],
                    "parent_iris": [],
                    "source_vocab_iri": "",
                    "change_note": "",
                }
            }
        }

        models = rdf_concepts_to_v1(concepts_data, {})

        assert len(models) == 1
        assert models[0].preferred_label == "Concept 1"
        assert models[0].language_code == "en"
        assert "alt1" in models[0].alternate_labels

    def test_concepts_to_v1_multi_language(self, temp_config):
        """Test converting concepts with multiple languages."""
        concepts_data = {
            "http://example.org/c1": {
                "en": {
                    "preferred_label": "Concept 1",
                    "definition": "Definition 1",
                    "alternate_labels": [],
                    "parent_iris": ["http://example.org/parent"],
                    "source_vocab_iri": "",
                    "change_note": "",
                },
                "de": {
                    "preferred_label": "Konzept 1",
                    "definition": "Definition 1 auf Deutsch",
                    "alternate_labels": [],
                    "parent_iris": ["http://example.org/parent"],
                    "source_vocab_iri": "",
                    "change_note": "",
                },
            }
        }

        models = rdf_concepts_to_v1(concepts_data, {})

        assert len(models) == 2

        # First row (en) should have parent_iris
        en_model = next(m for m in models if m.language_code == "en")
        assert en_model.parent_iris != ""

        # Second row (de) should NOT have parent_iris (structural data only in first row)
        de_model = next(m for m in models if m.language_code == "de")
        assert de_model.parent_iris == ""

    def test_mappings_to_v1(self, temp_config):
        """Test converting mappings to v1.0 models."""
        mappings_data = {
            "http://example.org/c1": {
                "related_matches": ["http://example.org/related1"],
                "close_matches": [],
                "exact_matches": [
                    "http://example.org/exact1",
                    "http://example.org/exact2",
                ],
                "narrower_matches": [],
                "broader_matches": [],
            }
        }

        models = rdf_mappings_to_v1(mappings_data)

        assert len(models) == 1
        assert (
            "related1" in models[0].related_matches
            or "http://example.org/related1" in models[0].related_matches
        )
        assert models[0].exact_matches != ""


class TestRdfToXlsxV1:
    """Tests for the main rdf_to_excel_v1 function."""

    def test_convert_simple_ttl(self, tmp_path, temp_config):
        """Test converting simple test TTL file."""
        output_path = tmp_path / "output.xlsx"

        result_path = rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        assert result_path.exists()
        assert result_path == output_path

        # Verify workbook structure
        wb = load_workbook(result_path)
        assert "Concept Scheme" in wb.sheetnames
        assert "Concepts" in wb.sheetnames
        assert "Collections" in wb.sheetnames
        assert "Mappings" in wb.sheetnames
        assert "Prefixes" in wb.sheetnames

    def test_convert_043_vocab(self, tmp_path, temp_config):
        """Test converting 0.4.3 format TTL file."""
        output_path = tmp_path / "vocab_043.xlsx"

        result_path = rdf_to_excel_v1(VOCAB_043_TTL, output_path)

        assert result_path.exists()

        wb = load_workbook(result_path)

        # Verify concepts sheet has data
        ws = wb["Concepts"]
        # Row 5 is first data row (after title, meaning, description, header)
        assert ws["A5"].value is not None, "Concepts sheet should have data"

    def test_default_output_path(self, tmp_path, temp_config):
        """Test that default output path uses .xlsx extension."""
        # Copy test file to tmp_path
        test_file = tmp_path / "test.ttl"
        shutil.copy(CS_SIMPLE_TTL, test_file)

        result_path = rdf_to_excel_v1(test_file)

        assert result_path == tmp_path / "test.xlsx"
        assert result_path.exists()

    def test_invalid_file_format(self, tmp_path, temp_config):
        """Test that invalid file format raises error."""
        invalid_file = tmp_path / "test.txt"
        invalid_file.write_text("not rdf")

        with pytest.raises(ValueError, match="RDF file formats"):
            rdf_to_excel_v1(invalid_file)


class TestConceptsSheetStructure:
    """Tests for the structure of the generated Concepts sheet."""

    def test_concepts_sheet_has_correct_headers(self, tmp_path, temp_config):
        """Test that Concepts sheet has correct column headers."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        expected_headers = [
            "Concept IRI*",
            "Language Code*",
            "Preferred Label*",
            "Definition*",
            "Alternate Labels",
            "Parent IRIs",
            "Member of collection(s)",
            "Member of ordered collection # position",
            "Provenance (read-only)",
            "Change Note",
            "Editorial Note",
            "Influenced by IRIs",
            "Source Vocab IRI or URL",
            "Source Vocab License",
            "Source Vocab Rights Holder",
            "Obsoletion reason",
            "dct:isReplacedBy",
        ]

        for col, expected in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual = ws[f"{col_letter}5"].value
            assert actual == expected, (
                f"Column {col} header mismatch: expected '{expected}', got '{actual}'"
            )

    def test_concepts_sheet_has_table(self, tmp_path, temp_config):
        """Test that Concepts sheet has a proper xlsx table."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        assert ws.tables, "Concepts sheet should have a table"

    def test_concepts_sheet_has_freeze_panes(self, tmp_path, temp_config):
        """Test that Concepts sheet has freeze panes."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        assert ws.freeze_panes == "A6", (
            f"Expected freeze panes at A6, got {ws.freeze_panes}"
        )


class TestCollectionsSheetStructure:
    """Tests for the structure of the generated Collections sheet."""

    def test_collections_sheet_has_correct_headers(self, tmp_path, temp_config):
        """Test that Collections sheet has correct column headers."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Collections"]

        expected_headers = [
            "Collection IRI*",
            "Language Code*",
            "Preferred Label*",
            "Definition*",
            "Parent Collection IRIs",
            "Ordered?",
            "Provenance (read-only)",
            "Change Note",
            "Editorial Note",
            "Obsoletion reason",
        ]

        for col, expected in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual = ws[f"{col_letter}5"].value
            assert actual == expected, (
                f"Column {col} header mismatch: expected '{expected}', got '{actual}'"
            )


class TestMappingsSheetStructure:
    """Tests for the structure of the generated Mappings sheet."""

    def test_mappings_sheet_has_correct_headers(self, tmp_path, temp_config):
        """Test that Mappings sheet has correct column headers."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Mappings"]

        expected_headers = [
            "Concept IRI*",
            "Related Matches",
            "Close Matches",
            "Exact Matches",
            "Narrower Matches",
            "Broader Matches",
            "Editorial Note",
        ]

        for col, expected in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual = ws[f"{col_letter}5"].value
            assert actual == expected, (
                f"Column {col} header mismatch: expected '{expected}', got '{actual}'"
            )


class TestConceptSchemeSheet:
    """Tests for the Concept Scheme sheet."""

    def test_concept_scheme_has_template_version(self, tmp_path, temp_config):
        """Test that Concept Scheme sheet has template version."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concept Scheme"]

        # Template version should be first field
        assert ws["A4"].value == "Template version"
        assert ws["B4"].value == TEMPLATE_VERSION

    def test_concept_scheme_has_vocabulary_iri(self, tmp_path, temp_config):
        """Test that Concept Scheme sheet has vocabulary IRI."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concept Scheme"]

        # Find vocabulary IRI row
        found = False
        for row in range(4, 30):  # pragma: no branch
            if ws[f"A{row}"].value == "Vocabulary IRI":  # pragma: no branch
                assert ws[f"B{row}"].value == "http://example.org/test/"
                found = True
                break

        assert found, "Vocabulary IRI field not found"


class TestPrefixesSheet:
    """Tests for the Prefixes sheet."""

    def test_prefixes_sheet_has_data(self, tmp_path, temp_config):
        """Test that Prefixes sheet has prefix data."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Prefixes"]

        # Should have at least some prefixes
        # Row 4 is header, row 5+ is data
        assert ws["A5"].value is not None, "Prefixes sheet should have data"


class TestMultiLanguageSupport:
    """Tests for multi-language handling."""

    def test_multi_language_concepts_create_multiple_rows(self, tmp_path):
        """Test that concepts with multiple languages create multiple rows."""
        # Create a simple RDF with multi-language concept
        graph = Graph()
        graph.bind("ex", EX)
        graph.bind("skos", SKOS)

        # Add concept scheme
        graph.add((EX.scheme, SKOS.prefLabel, Literal("Test Scheme", lang="en")))
        graph.add((EX.scheme, SKOS.definition, Literal("A test scheme", lang="en")))
        graph.add((EX.scheme, RDF.type, SKOS.ConceptScheme))
        graph.add(
            (EX.scheme, DCTERMS.created, Literal("2024-01-01", datatype=XSD.date))
        )
        graph.add(
            (EX.scheme, DCTERMS.modified, Literal("2024-01-01", datatype=XSD.date))
        )

        # Add multi-language concept
        graph.add((EX.concept1, RDF.type, SKOS.Concept))
        graph.add((EX.concept1, SKOS.prefLabel, Literal("Test Concept", lang="en")))
        graph.add((EX.concept1, SKOS.prefLabel, Literal("Testkonzept", lang="de")))
        graph.add((EX.concept1, SKOS.definition, Literal("A test concept", lang="en")))
        graph.add((EX.concept1, SKOS.definition, Literal("Ein Testkonzept", lang="de")))
        graph.add((EX.concept1, SKOS.inScheme, EX.scheme))

        # Save to temp file
        ttl_path = tmp_path / "multi_lang.ttl"
        graph.serialize(destination=str(ttl_path), format="turtle")

        # Convert
        output_path = tmp_path / "multi_lang.xlsx"
        rdf_to_excel_v1(ttl_path, output_path)

        # Verify
        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        # Count rows with concept IRI
        concept_rows = []
        for row in range(5, ws.max_row + 1):
            if ws[f"A{row}"].value:  # pragma: no branch
                concept_rows.append(row)

        # Should have 2 rows (en and de)
        assert len(concept_rows) >= 2, (
            "Should have at least 2 rows for multi-language concept"
        )

        # Check both languages are present
        langs = [ws[f"B{row}"].value for row in concept_rows]
        assert "en" in langs
        assert "de" in langs


class TestCollectionMembership:
    """Tests for collection membership handling."""

    def test_concept_membership_in_collection(self, tmp_path, temp_config):
        """Test that concept membership in collections is recorded."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        # Find "Member of collection(s)" column (column G)
        member_col = "G"

        # Check if any concept has collection membership
        has_membership = False
        for row in range(5, ws.max_row + 1):  # pragma: no branch
            if ws[f"{member_col}{row}"].value:  # pragma: no branch
                has_membership = True
                break

        # In concept-scheme-simple.ttl, test01-test04 are members of test10 collection
        assert has_membership, "Some concepts should have collection membership"


class TestRoundTrip:
    """Tests for lossless round-trip conversion RDF -> XLSX -> RDF."""

    def test_roundtrip_043_vocab(self, tmp_path, temp_config):
        """Test that 0.4.3 format RDF -> XLSX -> RDF preserves core data.

        Note: The 0.4.3 format uses older predicates (e.g. rdfs:isDefinedBy)
        that are now mapped to newer predicates (e.g. prov:hadPrimarySource) in v1.0.
        Therefore exact triple count match is not expected. The important thing is
        that core SKOS data is preserved.
        """
        # Load original
        original = Graph().parse(VOCAB_043_TTL, format="turtle")

        # Create vocab config from RDF for roundtrip
        vocab_config = make_vocab_config_from_rdf(original)

        # RDF -> XLSX
        xlsx_path = tmp_path / "vocab_043.xlsx"
        rdf_to_excel_v1(VOCAB_043_TTL, xlsx_path, vocab_config=vocab_config)

        # XLSX -> RDF
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        # Check that concept count matches (7 concepts in test fixture)
        original_concepts = list(original.subjects(RDF.type, SKOS.Concept))
        roundtrip_concepts = list(roundtrip.subjects(RDF.type, SKOS.Concept))
        assert len(original_concepts) == len(roundtrip_concepts), (
            f"Concept count mismatch: original={len(original_concepts)}, "
            f"roundtrip={len(roundtrip_concepts)}"
        )

        # Check that all prefLabels are preserved
        original_labels = {str(o) for o in original.objects(None, SKOS.prefLabel)}
        roundtrip_labels = {str(o) for o in roundtrip.objects(None, SKOS.prefLabel)}
        assert original_labels == roundtrip_labels, (
            f"Label mismatch.\n"
            f"Only in original: {original_labels - roundtrip_labels}\n"
            f"Only in roundtrip: {roundtrip_labels - original_labels}"
        )

    def test_roundtrip_preserves_concept_predicates(self, tmp_path, temp_config):
        """Test that concept-level predicates are preserved during round-trip.

        Note: ConceptScheme predicates may differ since scheme metadata now comes
        Uses v1.0 format test data (not 043 format which cannot be roundtripped).

        Note: Provenance predicates (dct:provenance, rdfs:seeAlso) are auto-generated
        based on repository config, not roundtripped from original data.
        """
        # Load original v1.0 format data
        original = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")

        # Create vocab config from RDF for roundtrip
        # Pass vocab_name to register ID pattern in config.ID_PATTERNS
        vocab_config = make_vocab_config_from_rdf(
            original, vocab_name="v1_comprehensive"
        )

        # RDF -> XLSX -> RDF
        xlsx_path = tmp_path / "v1_comprehensive.xlsx"
        rdf_to_excel_v1(V1_COMPREHENSIVE_TTL, xlsx_path, vocab_config=vocab_config)
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        # Get predicates used with concepts (not scheme-level)

        original_concepts = set(original.subjects(RDF.type, SKOS.Concept))
        roundtrip_concepts = set(roundtrip.subjects(RDF.type, SKOS.Concept))

        # Predicates that are auto-generated (not roundtripped from original)
        # These depend on config/environment, not on the original RDF data
        auto_generated_predicates = {DCTERMS.provenance, RDFS.seeAlso}

        # Get predicates on concepts
        original_concept_predicates = set()
        for concept in original_concepts:
            for _, p, _ in original.triples((concept, None, None)):
                original_concept_predicates.add(p)

        roundtrip_concept_predicates = set()
        for concept in roundtrip_concepts:
            for _, p, _ in roundtrip.triples((concept, None, None)):
                roundtrip_concept_predicates.add(p)

        # Exclude auto-generated predicates from comparison
        original_concept_predicates -= auto_generated_predicates
        roundtrip_concept_predicates -= auto_generated_predicates

        # Concept predicates should be the same
        assert original_concept_predicates == roundtrip_concept_predicates, (
            f"Concept predicate mismatch.\n"
            f"Only in original: {original_concept_predicates - roundtrip_concept_predicates}\n"
            f"Only in roundtrip: {roundtrip_concept_predicates - original_concept_predicates}"
        )


# =============================================================================
# V1.0 Feature Tests - New columns and predicates
# =============================================================================


class TestV1NewFeatures:
    """Tests for new v1.0 template features."""

    def test_extract_editorial_note(self):
        """Test that editorial notes are extracted from RDF."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # Concept 0000001 has editorial note
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = f"{photocat_ns}0000001"

        assert concept_iri in concepts
        en_data = concepts[concept_iri].get("en", {})
        assert (
            en_data.get("editorial_note")
            == "Review multi-language labels for accuracy."
        )

    def test_extract_obsolete_concept(self):
        """Test that obsolete concepts are extracted with reason."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # Concept 0000005 is obsolete
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = f"{photocat_ns}0000005"

        assert concept_iri in concepts
        en_data = concepts[concept_iri].get("en", {})
        assert en_data.get("obsolete_reason") == "This concept was added in error."

    def test_extract_source_vocab_attribution(self):
        """Test that source vocabulary attribution is extracted."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # Concept 0000002 has source attribution
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = f"{photocat_ns}0000002"

        assert concept_iri in concepts
        en_data = concepts[concept_iri].get("en", {})
        assert (
            en_data.get("source_vocab_iri")
            == "https://goldbook.iupac.org/terms/view/P04580"
        )
        assert (
            en_data.get("source_vocab_license")
            == "https://creativecommons.org/licenses/by/4.0/"
        )
        assert en_data.get("source_vocab_rights_holder") == "IUPAC Gold Book"

    def test_extract_influenced_by(self):
        """Test that influenced by IRIs are extracted."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        concepts = extract_concepts_from_rdf(graph)

        # Concept 0000002 has influenced_by
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = f"{photocat_ns}0000002"

        assert concept_iri in concepts
        en_data = concepts[concept_iri].get("en", {})
        influenced = en_data.get("influenced_by_iris", [])
        assert len(influenced) == 2
        assert "https://doi.org/10.1351/goldbook.P04580" in influenced

    def test_extract_ordered_collection(self):
        """Test that ordered collections are extracted."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        collections = extract_collections_from_rdf(graph)

        # coll002 is an ordered collection
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        coll_iri = f"{photocat_ns}coll002"

        assert coll_iri in collections
        en_data = collections[coll_iri].get("en", {})
        assert en_data.get("ordered") is True
        assert len(en_data.get("ordered_members", [])) == 2

    def test_extract_collection_editorial_note(self):
        """Test that collection editorial notes are extracted."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        collections = extract_collections_from_rdf(graph)

        # coll001 has editorial note
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        coll_iri = f"{photocat_ns}coll001"

        assert coll_iri in collections
        en_data = collections[coll_iri].get("en", {})
        assert "subcollections" in en_data.get("editorial_note", "")


class TestOrderedCollectionPositions:
    """Tests for ordered collection position parsing."""

    def test_parse_simple_position(self):
        """Test parsing simple position format."""
        prefixes = [PrefixV1(prefix="ex", namespace="http://example.org/")]
        converter = build_curies_converter_from_prefixes(prefixes)

        result = parse_ordered_collection_positions("ex:coll1 # 1", converter)

        assert result == {"http://example.org/coll1": 1}

    def test_parse_multiple_positions(self):
        """Test parsing multiple positions (newline-separated with optional labels)."""
        prefixes = [PrefixV1(prefix="ex", namespace="http://example.org/")]
        converter = build_curies_converter_from_prefixes(prefixes)

        result = parse_ordered_collection_positions(
            "ex:coll1 (First Collection) # 1\nex:coll2 (Second Collection) # 3",
            converter,
        )

        assert result == {
            "http://example.org/coll1": 1,
            "http://example.org/coll2": 3,
        }

    def test_parse_position_with_whitespace(self):
        """Test parsing positions with extra whitespace."""
        prefixes = [PrefixV1(prefix="ex", namespace="http://example.org/")]
        converter = build_curies_converter_from_prefixes(prefixes)

        result = parse_ordered_collection_positions("ex:coll1  #  2", converter)

        assert result == {"http://example.org/coll1": 2}

    def test_parse_empty_string(self):
        """Test parsing empty string."""
        prefixes = [PrefixV1(prefix="ex", namespace="http://example.org/")]
        converter = build_curies_converter_from_prefixes(prefixes)

        result = parse_ordered_collection_positions("", converter)

        assert result == {}


class TestBuildOrderedCollectionMaps:
    """Tests for building ordered collection maps."""

    def test_build_concept_to_ordered_collections_map(self):
        """Test building concept to ordered collections map from RDF."""
        graph = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        c2o_map = build_concept_to_ordered_collections_map(graph)

        # coll002 has ordered members: 0000003 at pos 1, 0000004 at pos 2
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"

        concept3 = f"{photocat_ns}0000003"
        assert concept3 in c2o_map
        assert f"{photocat_ns}coll002" in c2o_map[concept3]
        assert c2o_map[concept3][f"{photocat_ns}coll002"] == 1

        concept4 = f"{photocat_ns}0000004"
        assert concept4 in c2o_map
        assert f"{photocat_ns}coll002" in c2o_map[concept4]
        assert c2o_map[concept4][f"{photocat_ns}coll002"] == 2


class TestV1RoundTrip:
    """Tests for v1.0 specific round-trip conversion."""

    def test_roundtrip_comprehensive(self, tmp_path, temp_config):
        """Test round-trip with comprehensive v1.0 test data."""
        # Load original
        original = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")

        # Create vocab config from RDF for roundtrip
        vocab_config = make_vocab_config_from_rdf(original)

        # RDF -> XLSX
        xlsx_path = tmp_path / "v1_comprehensive.xlsx"
        rdf_to_excel_v1(V1_COMPREHENSIVE_TTL, xlsx_path, vocab_config=vocab_config)

        # XLSX -> RDF
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        # Check triple counts
        original_count = len(original)
        roundtrip_count = len(roundtrip)

        # Allow difference due to:
        # - BNode handling in RDF lists
        # - skos:historyNote added for first-time expressed concepts/collections
        assert abs(original_count - roundtrip_count) <= 20, (
            f"Triple count mismatch: original={original_count}, roundtrip={roundtrip_count}"
        )

    def test_roundtrip_preserves_editorial_note(self, tmp_path, temp_config):
        """Test that editorial notes survive round-trip."""
        original = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        vocab_config = make_vocab_config_from_rdf(original)

        # RDF -> XLSX -> RDF
        xlsx_path = tmp_path / "editorial_note.xlsx"
        rdf_to_excel_v1(V1_COMPREHENSIVE_TTL, xlsx_path, vocab_config=vocab_config)
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        # Check editorial note is preserved
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = URIRef(f"{photocat_ns}0000001")

        editorial_notes = list(roundtrip.objects(concept_iri, SKOS.editorialNote))
        assert len(editorial_notes) > 0
        assert any("Review multi-language" in str(n) for n in editorial_notes)

    def test_roundtrip_preserves_obsolete(self, tmp_path, temp_config):
        """Test that obsolete concepts survive round-trip."""
        original = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        vocab_config = make_vocab_config_from_rdf(original)

        # RDF -> XLSX -> RDF
        xlsx_path = tmp_path / "obsolete.xlsx"
        rdf_to_excel_v1(V1_COMPREHENSIVE_TTL, xlsx_path, vocab_config=vocab_config)
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        # Check deprecated flag and history note are preserved
        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = URIRef(f"{photocat_ns}0000005")

        # Check deprecated flag
        deprecated = list(roundtrip.objects(concept_iri, OWL.deprecated))
        assert len(deprecated) > 0

        # Check history note has obsolete reason
        history_notes = list(roundtrip.objects(concept_iri, SKOS.historyNote))
        assert len(history_notes) > 0
        assert any("added in error" in str(n) for n in history_notes)

    def test_roundtrip_preserves_source_attribution(self, tmp_path, temp_config):
        """Test that source attribution survives round-trip."""
        original = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        vocab_config = make_vocab_config_from_rdf(original)

        # RDF -> XLSX -> RDF
        xlsx_path = tmp_path / "source_attr.xlsx"
        rdf_to_excel_v1(V1_COMPREHENSIVE_TTL, xlsx_path, vocab_config=vocab_config)
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        concept_iri = URIRef(f"{photocat_ns}0000002")

        # Check hadPrimarySource
        sources = list(roundtrip.objects(concept_iri, PROV.hadPrimarySource))
        assert len(sources) > 0

        # Check license
        licenses = list(roundtrip.objects(concept_iri, DCTERMS.license))
        assert len(licenses) > 0

        # Check rightsHolder
        rights_holders = list(roundtrip.objects(concept_iri, DCTERMS.rightsHolder))
        assert len(rights_holders) > 0

    def test_roundtrip_preserves_ordered_collection(self, tmp_path, temp_config):
        """Test that ordered collections survive round-trip."""
        original = Graph().parse(V1_COMPREHENSIVE_TTL, format="turtle")
        vocab_config = make_vocab_config_from_rdf(original)

        # RDF -> XLSX -> RDF
        xlsx_path = tmp_path / "ordered_coll.xlsx"
        rdf_to_excel_v1(V1_COMPREHENSIVE_TTL, xlsx_path, vocab_config=vocab_config)
        roundtrip = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config
        )

        photocat_ns = "https://w3id.org/nfdi4cat/voc4cat-photocat/"
        coll_iri = URIRef(f"{photocat_ns}coll002")

        # Check it's an OrderedCollection
        types = list(roundtrip.objects(coll_iri, RDF.type))
        assert SKOS.OrderedCollection in types

        # Check memberList exists
        member_lists = list(roundtrip.objects(coll_iri, SKOS.memberList))
        assert len(member_lists) > 0


class TestAlternateLabelSeparator:
    """Tests for alternate label separator handling."""

    def test_alternate_labels_split_by_pipe(self, temp_config):
        """Test that alternate labels are split by pipe separator."""

        prefixes = [PrefixV1(prefix="ex", namespace="http://example.org/")]
        converter = build_curies_converter_from_prefixes(prefixes)

        concept_rows = [
            ConceptV1(
                concept_iri="ex:c1",
                language_code="en",
                preferred_label="Test",
                definition="A test",
                alternate_labels="alt1 | alt2 | alt3",
            )
        ]

        concepts = aggregate_concepts(concept_rows, converter)
        concept = concepts["http://example.org/c1"]

        assert concept.alt_labels["en"] == ["alt1", "alt2", "alt3"]

    def test_alternate_labels_trim_whitespace(self, temp_config):
        """Test that whitespace is trimmed from alternate labels."""

        prefixes = [PrefixV1(prefix="ex", namespace="http://example.org/")]
        converter = build_curies_converter_from_prefixes(prefixes)

        concept_rows = [
            ConceptV1(
                concept_iri="ex:c1",
                language_code="en",
                preferred_label="Test",
                definition="A test",
                alternate_labels="  alt1  |  alt2  ",
            )
        ]

        concepts = aggregate_concepts(concept_rows, converter)
        concept = concepts["http://example.org/c1"]

        # Whitespace should be trimmed
        assert concept.alt_labels["en"] == ["alt1", "alt2"]


# =============================================================================
# Deprecation Handling Tests
# =============================================================================


class TestValidateDeprecation:
    """Tests for validate_deprecation function."""

    def test_valid_deprecated_concept(self):
        """Test valid deprecated concept passes validation."""
        label, errors = validate_deprecation(
            pref_label="OBSOLETE Old Concept",
            is_deprecated=True,
            history_note=ConceptObsoletionReason.UNCLEAR.value,
            valid_reasons=[e.value for e in ConceptObsoletionReason],
            entity_iri="http://example.org/old",
            entity_type="concept",
        )
        assert errors == []
        assert label == "OBSOLETE Old Concept"

    def test_auto_add_obsolete_prefix(self):
        """Test that OBSOLETE prefix is auto-added when missing."""
        label, errors = validate_deprecation(
            pref_label="Old Concept",  # Missing OBSOLETE prefix
            is_deprecated=True,
            history_note=ConceptObsoletionReason.UNCLEAR.value,
            valid_reasons=[e.value for e in ConceptObsoletionReason],
            entity_iri="http://example.org/old",
            entity_type="concept",
        )
        assert errors == []
        assert label == "OBSOLETE Old Concept"

    def test_error_obsolete_prefix_without_deprecated(self):
        """Test error when OBSOLETE prefix exists but owl:deprecated is not set."""
        _label, errors = validate_deprecation(
            pref_label="OBSOLETE Old Concept",
            is_deprecated=False,  # Not deprecated!
            history_note="",
            valid_reasons=[e.value for e in ConceptObsoletionReason],
            entity_iri="http://example.org/old",
            entity_type="concept",
        )
        assert len(errors) == 1
        assert "OBSOLETE" in errors[0]
        assert "owl:deprecated" in errors[0]

    def test_error_deprecated_without_valid_reason(self):
        """Test error when owl:deprecated but historyNote is not valid."""
        _label, errors = validate_deprecation(
            pref_label="Old Concept",
            is_deprecated=True,
            history_note="Invalid reason not in enum",
            valid_reasons=[e.value for e in ConceptObsoletionReason],
            entity_iri="http://example.org/old",
            entity_type="concept",
        )
        assert len(errors) == 1
        assert "not a valid obsoletion reason" in errors[0]

    def test_valid_non_deprecated_concept(self):
        """Test valid non-deprecated concept passes validation."""
        label, errors = validate_deprecation(
            pref_label="Active Concept",
            is_deprecated=False,
            history_note="",
            valid_reasons=[e.value for e in ConceptObsoletionReason],
            entity_iri="http://example.org/active",
            entity_type="concept",
        )
        assert errors == []
        assert label == "Active Concept"

    def test_collection_uses_different_enum(self):
        """Test collection validation uses CollectionObsoletionReason."""
        # This reason is valid for collections but not concepts
        collection_reason = CollectionObsoletionReason.UNCLEAR.value

        # Should fail with concept enum
        _, errors = validate_deprecation(
            pref_label="Old",
            is_deprecated=True,
            history_note=collection_reason,
            valid_reasons=[e.value for e in ConceptObsoletionReason],
            entity_iri="http://example.org/old",
            entity_type="concept",
        )
        assert len(errors) == 1  # Invalid for concepts

        # Should pass with collection enum
        _, errors = validate_deprecation(
            pref_label="Old",
            is_deprecated=True,
            history_note=collection_reason,
            valid_reasons=[e.value for e in CollectionObsoletionReason],
            entity_iri="http://example.org/old",
            entity_type="collection",
        )
        assert errors == []


class TestDeprecationRdfExtraction:
    """Tests for extracting dct:isReplacedBy from RDF."""

    def test_extract_replaced_by_from_concept(self):
        """Test extracting dct:isReplacedBy from a concept."""
        g = Graph()
        g.bind("ex", EX)
        g.bind("skos", SKOS)

        # Create deprecated concept with replacement
        g.add((EX.oldConcept, RDF.type, SKOS.Concept))
        g.add((EX.oldConcept, SKOS.prefLabel, Literal("OBSOLETE Old", lang="en")))
        g.add((EX.oldConcept, OWL.deprecated, Literal(True)))
        g.add(
            (
                EX.oldConcept,
                SKOS.historyNote,
                Literal(ConceptObsoletionReason.UNCLEAR.value, lang="en"),
            )
        )
        g.add((EX.oldConcept, DCTERMS.isReplacedBy, EX.newConcept))

        # Create replacement concept
        g.add((EX.newConcept, RDF.type, SKOS.Concept))
        g.add((EX.newConcept, SKOS.prefLabel, Literal("New Concept", lang="en")))

        concepts = extract_concepts_from_rdf(g)

        old_data = concepts[str(EX.oldConcept)]["en"]
        assert old_data["replaced_by_iri"] == str(EX.newConcept)
        assert old_data["is_deprecated"] is True

    def test_extract_replaced_by_from_collection(self):
        """Test extracting dct:isReplacedBy from a collection."""
        g = Graph()
        g.bind("ex", EX)
        g.bind("skos", SKOS)

        # Create deprecated collection with replacement
        g.add((EX.oldColl, RDF.type, SKOS.Collection))
        g.add((EX.oldColl, SKOS.prefLabel, Literal("OBSOLETE Old Coll", lang="en")))
        g.add((EX.oldColl, OWL.deprecated, Literal(True)))
        g.add(
            (
                EX.oldColl,
                SKOS.historyNote,
                Literal(CollectionObsoletionReason.UNCLEAR.value, lang="en"),
            )
        )
        g.add((EX.oldColl, DCTERMS.isReplacedBy, EX.newColl))

        # Create replacement collection
        g.add((EX.newColl, RDF.type, SKOS.Collection))
        g.add((EX.newColl, SKOS.prefLabel, Literal("New Collection", lang="en")))

        collections = extract_collections_from_rdf(g)

        old_data = collections[str(EX.oldColl)]["en"]
        assert old_data["replaced_by_iri"] == str(EX.newColl)
        assert old_data["is_deprecated"] is True


class TestDeprecationXlsxAggregation:
    """Tests for deprecation handling in XLSX aggregation."""

    def test_aggregate_concepts_reads_replaced_by(self):
        """Test that aggregate_concepts reads replaced_by from explicit field."""
        prefixes = [
            PrefixV1(prefix="ex", namespace="http://example.org/"),
        ]
        converter = build_curies_converter_from_prefixes(prefixes)

        rows = [
            ConceptV1(
                concept_iri="ex:oldConcept",
                language_code="en",
                preferred_label="OBSOLETE Old Concept",
                definition="An old concept",
                obsolete_reason=ConceptObsoletionReason.UNCLEAR,
                replaced_by="ex:newConcept",
            ),
        ]

        concepts = aggregate_concepts(rows, converter)

        assert "http://example.org/oldConcept" in concepts
        concept = concepts["http://example.org/oldConcept"]
        assert concept.replaced_by_iri == "http://example.org/newConcept"

    def test_aggregate_concepts_adds_obsolete_prefix(self):
        """Test that aggregate_concepts auto-adds OBSOLETE prefix."""
        prefixes = [
            PrefixV1(prefix="ex", namespace="http://example.org/"),
        ]
        converter = build_curies_converter_from_prefixes(prefixes)

        rows = [
            ConceptV1(
                concept_iri="ex:oldConcept",
                language_code="en",
                preferred_label="Old Concept",  # Missing OBSOLETE prefix
                definition="An old concept",
                obsolete_reason=ConceptObsoletionReason.UNCLEAR,
            ),
        ]

        concepts = aggregate_concepts(rows, converter)

        concept = concepts["http://example.org/oldConcept"]
        assert concept.pref_labels["en"].startswith(OBSOLETE_PREFIX)

    def test_aggregate_collections_reads_replaced_by(self):
        """Test that aggregate_collections reads replaced_by from explicit field."""
        prefixes = [
            PrefixV1(prefix="ex", namespace="http://example.org/"),
        ]
        converter = build_curies_converter_from_prefixes(prefixes)

        rows = [
            CollectionV1(
                collection_iri="ex:oldColl",
                language_code="en",
                preferred_label="OBSOLETE Old Collection",
                definition="An old collection",
                obsolete_reason=CollectionObsoletionReason.UNCLEAR,
                replaced_by="ex:newColl",
            ),
        ]

        collections = aggregate_collections(rows, converter)

        assert "http://example.org/oldColl" in collections
        collection = collections["http://example.org/oldColl"]
        assert collection.replaced_by_iri == "http://example.org/newColl"


class TestDeprecationRoundTrip:
    """Tests for deprecation round-trip (RDF -> XLSX -> RDF)."""

    def test_roundtrip_deprecated_concept_with_replacement(self, tmp_path, temp_config):
        """Test round-trip of deprecated concept with dct:isReplacedBy."""
        g = Graph()
        g.bind("ex", EX)
        g.bind("skos", SKOS)

        # Create concept scheme
        g.add((EX[""], RDF.type, SKOS.ConceptScheme))
        g.add((EX[""], SKOS.prefLabel, Literal("Test Scheme", lang="en")))
        g.add(
            (EX[""], SKOS.definition, Literal("Test scheme for deprecation", lang="en"))
        )
        g.add((EX[""], DCTERMS.created, Literal("2024-01-01", datatype=XSD.date)))
        g.add(
            (EX[""], DCTERMS.creator, URIRef("https://orcid.org/0000-0001-2345-6789"))
        )

        # Create deprecated concept with replacement
        g.add((EX.oldConcept, RDF.type, SKOS.Concept))
        g.add(
            (EX.oldConcept, SKOS.prefLabel, Literal("OBSOLETE Old Concept", lang="en"))
        )
        g.add(
            (
                EX.oldConcept,
                SKOS.definition,
                Literal("An obsolete concept", lang="en"),
            )
        )
        g.add((EX.oldConcept, SKOS.inScheme, EX[""]))
        g.add((EX.oldConcept, OWL.deprecated, Literal(True)))
        g.add(
            (
                EX.oldConcept,
                SKOS.historyNote,
                Literal(ConceptObsoletionReason.UNCLEAR.value, lang="en"),
            )
        )
        g.add((EX.oldConcept, DCTERMS.isReplacedBy, EX.newConcept))

        # Create replacement concept
        g.add((EX.newConcept, RDF.type, SKOS.Concept))
        g.add((EX.newConcept, SKOS.prefLabel, Literal("New Concept", lang="en")))
        g.add((EX.newConcept, SKOS.definition, Literal("The replacement", lang="en")))
        g.add((EX.newConcept, SKOS.inScheme, EX[""]))

        # Save to TTL
        input_ttl = tmp_path / "test.ttl"
        g.serialize(destination=str(input_ttl), format="turtle")

        # Create vocab config from the graph
        vocab_config = make_vocab_config_from_rdf(g)

        # Convert RDF -> XLSX
        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(input_ttl, xlsx_path, vocab_config=vocab_config)

        # Convert XLSX -> RDF
        output_ttl = tmp_path / "output.ttl"
        excel_to_rdf_v1(xlsx_path, output_ttl, vocab_config=vocab_config)

        # Load result and verify
        result = Graph().parse(output_ttl, format="turtle")

        # Verify dct:isReplacedBy is preserved
        replaced_by = list(result.objects(EX.oldConcept, DCTERMS.isReplacedBy))
        assert len(replaced_by) == 1
        assert str(replaced_by[0]) == str(EX.newConcept)

        # Verify owl:deprecated is preserved
        deprecated = list(result.objects(EX.oldConcept, OWL.deprecated))
        assert len(deprecated) == 1
        assert str(deprecated[0]).lower() == "true"

    def test_roundtrip_obsolete_prefix_preserved(self, tmp_path, temp_config):
        """Test that OBSOLETE prefix is preserved in round-trip."""
        g = Graph()
        g.bind("ex", EX)
        g.bind("skos", SKOS)

        # Create concept scheme
        g.add((EX[""], RDF.type, SKOS.ConceptScheme))
        g.add((EX[""], SKOS.prefLabel, Literal("Test Scheme", lang="en")))
        g.add(
            (
                EX[""],
                SKOS.definition,
                Literal("Test scheme for obsolete prefix", lang="en"),
            )
        )
        g.add((EX[""], DCTERMS.created, Literal("2024-01-01", datatype=XSD.date)))
        g.add(
            (EX[""], DCTERMS.creator, URIRef("https://orcid.org/0000-0001-2345-6789"))
        )

        # Create deprecated concept with OBSOLETE prefix
        g.add((EX.old, RDF.type, SKOS.Concept))
        g.add((EX.old, SKOS.prefLabel, Literal("OBSOLETE Test Concept", lang="en")))
        g.add((EX.old, SKOS.definition, Literal("An obsolete concept", lang="en")))
        g.add((EX.old, SKOS.inScheme, EX[""]))
        g.add((EX.old, OWL.deprecated, Literal(True)))
        g.add(
            (
                EX.old,
                SKOS.historyNote,
                Literal(ConceptObsoletionReason.ADDED_IN_ERROR.value, lang="en"),
            )
        )

        # Save to TTL
        input_ttl = tmp_path / "test.ttl"
        g.serialize(destination=str(input_ttl), format="turtle")

        # Create vocab config from the graph
        vocab_config = make_vocab_config_from_rdf(g)

        # Convert RDF -> XLSX
        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(input_ttl, xlsx_path, vocab_config=vocab_config)

        # Convert XLSX -> RDF
        output_ttl = tmp_path / "output.ttl"
        excel_to_rdf_v1(xlsx_path, output_ttl, vocab_config=vocab_config)

        # Load result and verify
        result = Graph().parse(output_ttl, format="turtle")

        # Verify prefLabel still has OBSOLETE prefix
        pref_labels = list(result.objects(EX.old, SKOS.prefLabel))
        assert len(pref_labels) == 1
        assert str(pref_labels[0]).startswith("OBSOLETE ")


# =============================================================================
# Provenance URL Tests (Step 5)
# =============================================================================


class TestExtractEntityIdFromIri:
    """Tests for extract_entity_id_from_iri function."""

    def test_slash_iri(self):
        """Test extracting ID from slash-based IRI."""
        iri = "https://w3id.org/nfdi4cat/voc4cat/0000004"
        assert extract_entity_id_from_iri(iri, "voc4cat") == "0000004"

    def test_hash_iri(self):
        """Test extracting ID from hash-based IRI."""
        iri = "https://example.org/vocab#concept123"
        assert extract_entity_id_from_iri(iri, "vocab") == "concept123"

    def test_trailing_slash(self):
        """Test extracting ID from IRI with trailing slash."""
        iri = "https://example.org/vocab/term001/"
        assert extract_entity_id_from_iri(iri, "vocab") == "term001"

    def test_simple_iri(self):
        """Test extracting ID from simple IRI."""
        iri = "http://example.org/test01"
        assert extract_entity_id_from_iri(iri, "example") == "test01"

    def test_vocab_name_prefix_stripping(self):
        """Test stripping vocab prefix from IRI with underscore pattern."""
        iri = "https://w3id.org/nfdi4cat/voc4cat_0000016"
        assert extract_entity_id_from_iri(iri, "voc4cat") == "0000016"

    def test_vocab_name_no_prefix_in_iri(self):
        """Test that non-prefixed IRI is unchanged when vocab_name provided."""
        iri = "https://w3id.org/nfdi4cat/voc4cat/0000004"
        assert extract_entity_id_from_iri(iri, "voc4cat") == "0000004"

    def test_vocab_name_hash_iri(self):
        """Test hash-based IRI with vocab prefix stripping."""
        iri = "https://example.org/ns#myvocab_concept123"
        assert extract_entity_id_from_iri(iri, "myvocab") == "concept123"


class TestBuildProvenanceUrl:
    """Tests for build_provenance_url function."""

    def test_with_all_env_vars(self, monkeypatch):
        """Test URL generation with all env vars set."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/voc4cat")
        monkeypatch.setenv("VOC4CAT_VERSION", "v2025-10-14")

        url = build_provenance_url("0000004", "voc4cat")

        expected = "https://github.com/nfdi4cat/voc4cat/blame/v2025-10-14/vocabularies/voc4cat/0000004.ttl"
        assert url == expected

    def test_no_version_uses_main(self, monkeypatch):
        """Test URL generation uses 'main' when no version set."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/voc4cat")
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        url = build_provenance_url("0000004", "voc4cat")

        assert "/blame/main/" in url
        assert "v2025" not in url

    def test_empty_version_uses_main(self, monkeypatch):
        """Test URL generation uses 'main' when version is empty string."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/voc4cat")
        monkeypatch.setenv("VOC4CAT_VERSION", "")

        url = build_provenance_url("0000004", "voc4cat")

        assert "/blame/main/" in url

    def test_no_github_repo_returns_empty(self, monkeypatch):
        """Test URL generation returns empty when no GITHUB_REPOSITORY."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.setenv("VOC4CAT_VERSION", "v2025-01-01")

        url = build_provenance_url("0000004", "voc4cat")

        assert url == ""

    def test_different_vocab_names(self, monkeypatch):
        """Test URL generation with different vocabulary names."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/test-vocab")
        monkeypatch.setenv("VOC4CAT_VERSION", "v1.0.0")

        url = build_provenance_url("test123", "myvocab")

        expected = "https://github.com/nfdi4cat/test-vocab/blame/v1.0.0/vocabularies/myvocab/test123.ttl"
        assert url == expected

    def test_repository_url_fallback_when_env_not_set(self, monkeypatch):
        """Test URL generation uses repository_url when env var not set."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        url = build_provenance_url(
            "0000004",
            "voc4cat",
            repository_url="https://github.com/nfdi4cat/voc4cat",
        )

        expected = "https://github.com/nfdi4cat/voc4cat/blame/main/vocabularies/voc4cat/0000004.ttl"
        assert url == expected

    def test_env_var_takes_precedence_over_repository_url(self, monkeypatch):
        """Test that GITHUB_REPOSITORY env var takes precedence over repository_url."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/from-env")
        monkeypatch.setenv("VOC4CAT_VERSION", "v1.0.0")

        url = build_provenance_url(
            "0000004",
            "voc4cat",
            repository_url="https://github.com/nfdi4cat/from-config",
        )

        assert "nfdi4cat/from-env" in url
        assert "nfdi4cat/from-config" not in url

    def test_repository_url_with_git_suffix(self, monkeypatch):
        """Test URL generation handles repository URLs with .git suffix."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        url = build_provenance_url(
            "0000004",
            "voc4cat",
            repository_url="https://github.com/nfdi4cat/voc4cat.git",
        )

        expected = "https://github.com/nfdi4cat/voc4cat/blame/main/vocabularies/voc4cat/0000004.ttl"
        assert url == expected

    def test_repository_url_with_trailing_slash(self, monkeypatch):
        """Test URL generation handles repository URLs with trailing slash."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        url = build_provenance_url(
            "0000004",
            "voc4cat",
            repository_url="https://github.com/nfdi4cat/voc4cat/",
        )

        expected = "https://github.com/nfdi4cat/voc4cat/blame/main/vocabularies/voc4cat/0000004.ttl"
        assert url == expected

    def test_non_github_repository_url_returns_empty(self, monkeypatch):
        """Test URL generation returns empty for non-GitHub repository URLs."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        url = build_provenance_url(
            "0000004",
            "voc4cat",
            repository_url="https://gitlab.com/nfdi4cat/voc4cat",
        )

        assert url == ""

    def test_custom_template_gitlab(self, monkeypatch):
        """Test URL generation with custom GitLab template."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.setenv("VOC4CAT_VERSION", "v1.0.0")

        template = (
            "https://gitlab.com/myorg/myrepo/-/blame/{{ version }}"
            "/vocabularies/{{ vocab_name }}/{{ entity_id }}.ttl"
        )
        url = build_provenance_url("0000004", "voc4cat", provenance_template=template)

        expected = "https://gitlab.com/myorg/myrepo/-/blame/v1.0.0/vocabularies/voc4cat/0000004.ttl"
        assert url == expected

    def test_custom_template_gitea(self, monkeypatch):
        """Test URL generation with custom Gitea template."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        template = (
            "https://gitea.example.com/org/repo/blame/{{ version }}"
            "/vocabularies/{{ vocab_name }}/{{ entity_id }}.ttl"
        )
        url = build_provenance_url("test123", "myvocab", provenance_template=template)

        expected = "https://gitea.example.com/org/repo/blame/main/vocabularies/myvocab/test123.ttl"
        assert url == expected

    def test_custom_template_with_github_repo_variable(self, monkeypatch):
        """Test custom template can use github_repo variable."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "custom/repo")
        monkeypatch.setenv("VOC4CAT_VERSION", "v2.0.0")

        template = "https://example.com/{{ github_repo }}/{{ entity_id }}"
        url = build_provenance_url("0001", "vocab", provenance_template=template)

        assert url == "https://example.com/custom/repo/0001"

    def test_custom_template_takes_precedence_over_github_default(self, monkeypatch):
        """Test that custom template is used even when GitHub repo is available."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/voc4cat")
        monkeypatch.setenv("VOC4CAT_VERSION", "v1.0.0")

        template = "https://custom.example.com/{{ entity_id }}"
        url = build_provenance_url("0000004", "voc4cat", provenance_template=template)

        assert url == "https://custom.example.com/0000004"
        assert "github.com" not in url


class TestExtractGithubRepoFromUrl:
    """Tests for extract_github_repo_from_url function."""

    def test_standard_github_url(self):
        """Test extraction from standard GitHub URL."""
        result = extract_github_repo_from_url("https://github.com/nfdi4cat/voc4cat")
        assert result == "nfdi4cat/voc4cat"

    def test_github_url_with_git_suffix(self):
        """Test extraction from GitHub URL with .git suffix."""
        result = extract_github_repo_from_url("https://github.com/owner/repo.git")
        assert result == "owner/repo"

    def test_github_url_with_trailing_slash(self):
        """Test extraction from GitHub URL with trailing slash."""
        result = extract_github_repo_from_url("https://github.com/owner/repo/")
        assert result == "owner/repo"

    def test_http_url(self):
        """Test extraction from HTTP URL (not HTTPS)."""
        result = extract_github_repo_from_url("http://github.com/owner/repo")
        assert result == "owner/repo"

    def test_non_github_url_returns_empty(self):
        """Test extraction returns empty for non-GitHub URLs."""
        result = extract_github_repo_from_url("https://gitlab.com/owner/repo")
        assert result == ""

    def test_empty_string_returns_empty(self):
        """Test extraction returns empty for empty string."""
        result = extract_github_repo_from_url("")
        assert result == ""

    def test_none_like_values(self):
        """Test extraction handles None-like values."""
        result = extract_github_repo_from_url("")
        assert result == ""

    def test_github_url_with_extra_paths_returns_empty(self):
        """Test extraction returns empty for URLs with extra path components."""
        # URLs with subpaths beyond owner/repo shouldn't match
        result = extract_github_repo_from_url("https://github.com/owner/repo/tree/main")
        assert result == ""


class TestProvenanceInXlsx:
    """Tests for provenance column in generated XLSX files."""

    def test_provenance_column_has_url_when_env_set(
        self, tmp_path, temp_config, monkeypatch
    ):
        """Test that Provenance column is populated when env vars are set."""
        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/test-vocab")
        monkeypatch.setenv("VOC4CAT_VERSION", "v2025-01-01")

        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        # Find Provenance column (column I based on our header order)
        # Row 6 is first data row (row 5 is header, row 4 is requiredness)
        cell = ws["I6"]
        provenance_display = cell.value
        provenance_hyperlink = cell.hyperlink

        # Display text should be friendly format
        assert provenance_display is not None
        assert provenance_display.startswith("git blame for ")

        # Hyperlink should contain the full URL
        assert provenance_hyperlink is not None
        assert "github.com" in provenance_hyperlink.target
        assert "/blame/" in provenance_hyperlink.target
        assert "v2025-01-01" in provenance_hyperlink.target

    def test_provenance_column_empty_when_env_not_set(
        self, tmp_path, temp_config, monkeypatch
    ):
        """Test that Provenance column is empty when GITHUB_REPOSITORY not set."""
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concepts"]

        # Provenance column should be empty (row 6 is first data row)
        provenance_value = ws["I6"].value
        assert provenance_value is None or provenance_value == ""


class TestProvenanceInRdf:
    """Tests for provenance triples in generated RDF."""

    def test_xlsx_to_rdf_generates_provenance_triples(
        self, tmp_path, temp_config, test_vocab_config, monkeypatch
    ):
        """Test that XLSX->RDF generates dct:provenance triples."""

        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/voc4cat")
        monkeypatch.setenv("VOC4CAT_VERSION", "v2025-01-01")

        # First convert RDF -> XLSX
        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=test_vocab_config)

        # Then convert XLSX -> RDF
        graph = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=test_vocab_config
        )

        # Check for provenance triples on a concept
        concept_iri = URIRef("http://example.org/test01")

        provenance_urls = list(graph.objects(concept_iri, DCTERMS.provenance))
        assert len(provenance_urls) == 1
        assert "github.com" in str(provenance_urls[0])
        assert "/blame/v2025-01-01/" in str(provenance_urls[0])

        # Also check rdfs:seeAlso
        see_also_urls = list(graph.objects(concept_iri, RDFS.seeAlso))
        assert len(see_also_urls) >= 1
        # Should contain the same provenance URL
        assert str(provenance_urls[0]) in [str(u) for u in see_also_urls]

    def test_no_provenance_when_no_github_source_available(
        self, tmp_path, temp_config, monkeypatch
    ):
        """Test that no provenance triples when no GitHub source is available.

        Uses a non-GitHub repository URL (GitLab) which cannot be used for
        provenance URL generation since we only support GitHub blame URLs.
        """
        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        # Create a config with non-GitHub repository URL
        vocab_config_non_github = Vocab(
            id_length=7,
            permanent_iri_part="http://example.org/test/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "http://example.org/"},
            vocabulary_iri="http://example.org/test/",
            title="Test Vocabulary",
            description="Test vocabulary for unit tests",
            created_date="2025-01-01",
            creator="https://orcid.org/0000-0001-2345-6789",
            repository="https://gitlab.com/test/vocab",  # Non-GitHub URL
        )

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=vocab_config_non_github)

        graph = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=vocab_config_non_github
        )

        concept_iri = URIRef("http://example.org/test01")

        # Should have no dct:provenance triples (no GitHub repo available)
        provenance_urls = list(graph.objects(concept_iri, DCTERMS.provenance))
        assert len(provenance_urls) == 0

    def test_provenance_from_config_repo_when_env_not_set(
        self, tmp_path, temp_config, test_vocab_config, monkeypatch
    ):
        """Test provenance generated from config repository when env var not set."""

        monkeypatch.delenv("GITHUB_REPOSITORY", raising=False)
        monkeypatch.delenv("VOC4CAT_VERSION", raising=False)

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=test_vocab_config)

        graph = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=test_vocab_config
        )

        concept_iri = URIRef("http://example.org/test01")

        # Should have provenance from config repository
        provenance_urls = list(graph.objects(concept_iri, DCTERMS.provenance))
        assert len(provenance_urls) == 1
        # Uses 'main' as default version when VOC4CAT_VERSION not set
        assert "/blame/main/" in str(provenance_urls[0])
        # Repository is test/vocab from test_vocab_config
        assert "test/vocab" in str(provenance_urls[0])

    def test_collection_has_provenance(
        self, tmp_path, temp_config, test_vocab_config, monkeypatch
    ):
        """Test that collections also get provenance triples."""

        monkeypatch.setenv("GITHUB_REPOSITORY", "nfdi4cat/test-vocab")
        monkeypatch.setenv("VOC4CAT_VERSION", "v2025-01-01")

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=test_vocab_config)

        graph = excel_to_rdf_v1(
            xlsx_path, output_type="graph", vocab_config=test_vocab_config
        )

        # Collection IRI from concept-scheme-simple.ttl (ex:test10)
        collection_iri = URIRef("http://example.org/test10")

        provenance_urls = list(graph.objects(collection_iri, DCTERMS.provenance))
        assert len(provenance_urls) == 1
        assert "test10.ttl" in str(provenance_urls[0])


# =============================================================================
# Tests for Step 6: ConceptScheme from Config
# =============================================================================


class TestConfigToConceptSchemeV1:
    """Tests for config_to_concept_scheme_v1() function."""

    def test_config_only_creates_scheme(self, datadir, temp_config):
        """Test creating ConceptScheme from config without RDF fallback."""
        config = temp_config
        config.load_config(datadir / "idranges_with_scheme.toml")

        vocab = config.IDRANGES.vocabs["myvocab"]
        scheme = config_to_concept_scheme_v1(vocab)

        assert scheme.vocabulary_iri == "https://example.org/vocab/"
        assert scheme.prefix == "ex"
        assert scheme.title == "Test Vocabulary"
        assert scheme.description == "A test vocabulary for unit tests."
        assert scheme.created_date == "2025-01-15"
        assert "Alice Smith" in scheme.creator
        assert "Example Organization" in scheme.publisher
        assert "Bob Jones" in scheme.custodian
        assert scheme.template_version == TEMPLATE_VERSION

    def test_config_overrides_rdf(self, datadir, temp_config):
        """Test that config values override RDF values."""

        config = temp_config
        config.load_config(datadir / "idranges_with_scheme.toml")
        vocab = config.IDRANGES.vocabs["myvocab"]

        # Create an RDF scheme with different values
        rdf_scheme = ConceptSchemeV1(
            vocabulary_iri="http://rdf-value.org/",
            title="RDF Title",
            description="RDF Description",
            created_date="2020-01-01",
            modified_date="2024-06-01",  # Only in RDF
            version="1.2.3",  # Only in RDF
        )

        scheme = config_to_concept_scheme_v1(vocab, rdf_scheme)

        # Config values should override
        assert scheme.vocabulary_iri == "https://example.org/vocab/"
        assert scheme.title == "Test Vocabulary"
        assert scheme.description == "A test vocabulary for unit tests."
        assert scheme.created_date == "2025-01-15"

        # RDF-only values should be preserved
        assert scheme.modified_date == "2024-06-01"
        assert scheme.version == "1.2.3"

    def test_rdf_fills_gaps_for_optional_fields(self, temp_config):
        """Test that RDF fills gaps for optional fields when config fields are empty."""
        # Create a vocab config with mandatory fields but empty optional fields
        vocab = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            # Mandatory fields
            vocabulary_iri="https://example.org/vocab/",
            title="Config Title",
            description="Config Description",
            created_date="2025-01-01",
            creator="Test Author",
            repository="https://github.com/test/vocab",
            # Optional fields left empty
            publisher="",
            custodian="",
        )

        # Create RDF scheme with optional values that should fill gaps
        rdf_scheme = ConceptSchemeV1(
            vocabulary_iri="http://rdf-value.org/",  # Will be overridden by config
            title="RDF Title",  # Will be overridden by config
            description="RDF Description",  # Will be overridden by config
            publisher="RDF Publisher",  # Will fill gap
            custodian="RDF Custodian",  # Will fill gap
            modified_date="2024-06-01",  # Only in RDF
        )

        scheme = config_to_concept_scheme_v1(vocab, rdf_scheme)

        # Config mandatory values should override
        assert scheme.vocabulary_iri == "https://example.org/vocab/"
        assert scheme.title == "Config Title"
        assert scheme.description == "Config Description"

        # RDF optional values should fill gaps
        assert scheme.publisher == "RDF Publisher"
        assert scheme.custodian == "RDF Custodian"
        assert scheme.modified_date == "2024-06-01"

    def test_mandatory_fields_required_in_config(self, temp_config):
        """Test that mandatory fields must be in config (can't rely on RDF)."""
        # Try to create a vocab config without mandatory fields - should fail
        with pytest.raises(ValidationError) as excinfo:
            Vocab(
                id_length=7,
                permanent_iri_part="https://example.org/",
                checks=Checks(allow_delete=False),
                prefix_map={"ex": "https://example.org/"},
                # Missing: vocabulary_iri, title, description, created_date, creator, repository
            )

        error_msg = str(excinfo.value)
        assert "Mandatory ConceptScheme fields are empty" in error_msg


class TestRdfToXlsxWithConfig:
    """Tests for rdf_to_excel_v1() with vocab_config parameter."""

    def test_config_used_for_scheme_metadata(self, tmp_path, datadir, temp_config):
        """Test that config is used for ConceptScheme when provided."""
        config = temp_config
        config.load_config(datadir / "idranges_with_scheme.toml")
        vocab = config.IDRANGES.vocabs["myvocab"]

        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path, vocab_config=vocab)

        # Verify the Concept Scheme sheet has config values
        wb = load_workbook(output_path)
        ws = wb["Concept Scheme"]

        # Find vocabulary IRI row (config value should override RDF)
        found_iri = False
        found_title = False
        for row in range(4, 30):
            if ws[f"A{row}"].value == "Vocabulary IRI":  # pragma: no branch
                # Config value should be used, not RDF value
                assert ws[f"B{row}"].value == "https://example.org/vocab/"
                found_iri = True
            if ws[f"A{row}"].value == "Title":  # pragma: no branch
                assert ws[f"B{row}"].value == "Test Vocabulary"
                found_title = True

        assert found_iri, "Vocabulary IRI field not found"
        assert found_title, "Title field not found"

    def test_without_config_uses_rdf(self, tmp_path, temp_config):
        """Test that RDF is used when no config provided."""
        output_path = tmp_path / "output.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, output_path)

        wb = load_workbook(output_path)
        ws = wb["Concept Scheme"]

        # Should have RDF value (from concept-scheme-simple.ttl)
        found = False
        for row in range(4, 30):  # pragma: no branch
            if ws[f"A{row}"].value == "Vocabulary IRI":  # pragma: no branch
                assert ws[f"B{row}"].value == "http://example.org/test/"
                found = True
                break

        assert found, "Vocabulary IRI field not found"


class TestXlsxToRdfWithConfig:
    """Tests for excel_to_rdf_v1() with vocab_config parameter."""

    def test_config_used_ignores_excel_scheme(self, tmp_path, datadir, temp_config):
        """Test that config is used and xlsx scheme is ignored."""

        config = temp_config
        config.load_config(datadir / "idranges_with_scheme.toml")
        vocab = config.IDRANGES.vocabs["myvocab"]

        # First create an xlsx file with RDF data
        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path)  # Without config

        # Now convert back to RDF using config
        graph = excel_to_rdf_v1(xlsx_path, output_type="graph", vocab_config=vocab)

        # The ConceptScheme should have config values
        scheme_iri = URIRef("https://example.org/vocab/")
        titles = list(graph.objects(scheme_iri, SKOS.prefLabel))

        assert len(titles) == 1
        assert str(titles[0]) == "Test Vocabulary"


# =============================================================================
# ID Ranges Tests (Step 7)
# =============================================================================


class TestExtractUsedIds:
    """Tests for extract_used_ids() function."""

    def test_extract_ids_from_concepts(self, temp_config):
        """Test that concept IDs are extracted correctly."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            vocabulary_iri="https://example.org/vocab/",
            title="Test Vocabulary",
            description="Test description",
            created_date="2025-01-01",
            creator="Test Author",
            repository="https://github.com/test/vocab",
        )

        concepts = [
            ConceptV1(
                concept_iri="https://example.org/0000001",
                language_code="en",
                preferred_label="Concept 1",
                definition="Definition 1",
            ),
            ConceptV1(
                concept_iri="https://example.org/0000001",
                language_code="de",
                preferred_label="Konzept 1",
                definition="Definition 1",
            ),  # Duplicate IRI (different language)
            ConceptV1(
                concept_iri="https://example.org/0000005",
                language_code="en",
                preferred_label="Concept 5",
                definition="Definition 5",
            ),
            ConceptV1(
                concept_iri="https://example.org/0000010",
                language_code="en",
                preferred_label="Concept 10",
                definition="Definition 10",
            ),
        ]
        collections = []

        used_ids = extract_used_ids(concepts, collections, vocab_config)

        assert used_ids == {1, 5, 10}

    def test_extract_ids_from_collections(self, temp_config, mandatory_fields):
        """Test that collection IDs are extracted correctly."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            **mandatory_fields,
        )

        concepts = []
        collections = [
            CollectionV1(
                collection_iri="https://example.org/0000020",
                language_code="en",
                preferred_label="Collection 20",
                definition="Definition 20",
            ),
            CollectionV1(
                collection_iri="https://example.org/0000025",
                language_code="en",
                preferred_label="Collection 25",
                definition="Definition 25",
            ),
        ]

        used_ids = extract_used_ids(concepts, collections, vocab_config)

        assert used_ids == {20, 25}

    def test_extract_ids_combined(self, temp_config, mandatory_fields):
        """Test that both concept and collection IDs are extracted."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            **mandatory_fields,
        )

        concepts = [
            ConceptV1(
                concept_iri="https://example.org/0000001",
                language_code="en",
                preferred_label="Concept 1",
                definition="Definition 1",
            ),
            ConceptV1(
                concept_iri="https://example.org/0000002",
                language_code="en",
                preferred_label="Concept 2",
                definition="Definition 2",
            ),
        ]
        collections = [
            CollectionV1(
                collection_iri="https://example.org/0000003",
                language_code="en",
                preferred_label="Collection 3",
                definition="Definition 3",
            ),
        ]

        used_ids = extract_used_ids(concepts, collections, vocab_config)

        assert used_ids == {1, 2, 3}

    def test_ignores_foreign_iris(self, temp_config, mandatory_fields):
        """Test that IRIs from other vocabularies are ignored."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            **mandatory_fields,
        )

        concepts = [
            ConceptV1(
                concept_iri="https://example.org/0000001",
                language_code="en",
                preferred_label="Concept 1",
                definition="Definition 1",
            ),
            ConceptV1(
                concept_iri="https://other.org/0000002",
                language_code="en",
                preferred_label="Concept 2",
                definition="Definition 2",
            ),  # Foreign
        ]
        collections = []

        used_ids = extract_used_ids(concepts, collections, vocab_config)

        assert used_ids == {1}

    def test_handles_empty_iris(self, temp_config, mandatory_fields):
        """Test that empty IRIs are handled gracefully."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            **mandatory_fields,
        )

        concepts = [
            ConceptV1(
                concept_iri="",
                language_code="en",
                preferred_label="Empty Concept",
                definition="Empty definition",
            ),
            ConceptV1(
                concept_iri="https://example.org/0000001",
                language_code="en",
                preferred_label="Concept 1",
                definition="Definition 1",
            ),
        ]
        collections = [
            CollectionV1(
                collection_iri="",
                language_code="en",
                preferred_label="Empty Collection",
                definition="Empty definition",
            )
        ]

        used_ids = extract_used_ids(concepts, collections, vocab_config)

        assert used_ids == {1}


class TestBuildIdRangeInfo:
    """Tests for build_id_range_info() function."""

    def test_build_single_range(self, temp_config, mandatory_fields):
        """Test building info for a single ID range."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "alice"},
            ],
            **mandatory_fields,
        )

        used_ids = {1, 2, 3}
        rows = build_id_range_info(vocab_config, used_ids)

        assert len(rows) == 1
        assert rows[0].gh_name == "alice"
        assert rows[0].id_range == "0000001 - 0000010"
        assert rows[0].unused_ids == "next unused: 0000004, unused: 7"

    def test_build_multiple_ranges(self, temp_config, mandatory_fields):
        """Test building info for multiple ID ranges."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "alice"},
                {"first_id": 11, "last_id": 20, "gh_name": "bob"},
            ],
            **mandatory_fields,
        )

        used_ids = {1, 11, 12}
        rows = build_id_range_info(vocab_config, used_ids)

        assert len(rows) == 2
        assert rows[0].gh_name == "alice"
        assert rows[0].unused_ids == "next unused: 0000002, unused: 9"
        assert rows[1].gh_name == "bob"
        assert rows[1].unused_ids == "next unused: 0000013, unused: 8"

    def test_all_ids_used(self, temp_config, mandatory_fields):
        """Test when all IDs in a range are used."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            id_range=[
                {"first_id": 1, "last_id": 3, "gh_name": "alice"},
            ],
            **mandatory_fields,
        )

        used_ids = {1, 2, 3}
        rows = build_id_range_info(vocab_config, used_ids)

        assert len(rows) == 1
        assert rows[0].unused_ids == "all IDs used. Request a new range!"

    def test_orcid_fallback(self, temp_config, mandatory_fields):
        """Test that ORCID is used when gh_name is empty."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            id_range=[
                {
                    "first_id": 1,
                    "last_id": 10,
                    "gh_name": "",
                    "orcid": "https://orcid.org/0000-0001-2345-6789",
                },
            ],
            **mandatory_fields,
        )

        used_ids = set()
        rows = build_id_range_info(vocab_config, used_ids)

        assert len(rows) == 1
        assert rows[0].gh_name == "https://orcid.org/0000-0001-2345-6789"

    def test_zero_padding(self, temp_config, mandatory_fields):
        """Test that ID padding respects id_length from config."""
        vocab_config = Vocab(
            id_length=5,  # Shorter padding
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            id_range=[
                {"first_id": 1, "last_id": 100, "gh_name": "alice"},
            ],
            **mandatory_fields,
        )

        used_ids = {1}
        rows = build_id_range_info(vocab_config, used_ids)

        assert rows[0].id_range == "00001 - 00100"
        assert rows[0].unused_ids == "next unused: 00002, unused: 99"


class TestIdRangesSheetExport:
    """Tests for ID Ranges sheet in exported xlsx files."""

    def test_id_ranges_sheet_created(self, tmp_path, temp_config, datadir):
        """Test that ID Ranges sheet is created when vocab_config has id_range."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            vocabulary_iri="https://example.org/",
            title="Test Vocabulary",
            description="Test description",
            created_date="2025-01-01",
            creator="Test Author",
            repository="https://github.com/test/vocab",
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "alice"},
            ],
        )

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=vocab_config)

        wb = load_workbook(xlsx_path)
        assert "ID Ranges" in wb.sheetnames

    def test_id_ranges_sheet_position(self, tmp_path, temp_config, datadir):
        """Test that ID Ranges sheet is positioned before Prefixes."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            vocabulary_iri="https://example.org/",
            title="Test Vocabulary",
            description="Test description",
            created_date="2025-01-01",
            creator="Test Author",
            repository="https://github.com/test/vocab",
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "alice"},
            ],
        )

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=vocab_config)

        wb = load_workbook(xlsx_path)
        id_ranges_idx = wb.sheetnames.index("ID Ranges")
        prefixes_idx = wb.sheetnames.index("Prefixes")
        assert id_ranges_idx < prefixes_idx

    def test_id_ranges_sheet_content(self, tmp_path, temp_config, datadir):
        """Test that ID Ranges sheet has correct content."""
        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            vocabulary_iri="https://example.org/",
            title="Test Vocabulary",
            description="Test description",
            created_date="2025-01-01",
            creator="Test Author",
            repository="https://github.com/test/vocab",
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "alice"},
                {"first_id": 11, "last_id": 20, "gh_name": "bob"},
            ],
        )

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path, vocab_config=vocab_config)

        wb = load_workbook(xlsx_path)
        ws = wb["ID Ranges"]

        # Check title row
        assert ws["A1"].value == "ID Ranges (read-only)"

        # Check headers (row 3)
        assert ws["A3"].value == "gh-name"
        assert ws["B3"].value == "ID Range"
        assert ws["C3"].value == "Unused IDs"

        # Check data rows (row 4+)
        assert ws["A4"].value == "alice"
        assert ws["B4"].value == "0000001 - 0000010"
        assert ws["A5"].value == "bob"
        assert ws["B5"].value == "0000011 - 0000020"

    def test_id_ranges_empty_when_no_config(self, tmp_path, temp_config):
        """Test that ID Ranges sheet is empty when no config provided."""
        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(CS_SIMPLE_TTL, xlsx_path)  # No vocab_config

        wb = load_workbook(xlsx_path)
        assert "ID Ranges" in wb.sheetnames
        ws = wb["ID Ranges"]

        # Should have title and headers but no data
        assert ws["A1"].value == "ID Ranges (read-only)"
        assert ws["A3"].value == "gh-name"
        # Data row should be empty
        assert ws["A4"].value is None or ws["A4"].value == ""

    def test_unused_ids_calculation(self, tmp_path, temp_config):
        """Test that unused IDs are calculated correctly from vocabulary content."""
        # Create RDF with known concept IRIs
        graph = Graph()
        graph.bind("skos", SKOS)
        graph.bind("ex", EX)

        # ConceptScheme
        scheme = EX[""]
        graph.add((scheme, RDF.type, SKOS.ConceptScheme))
        graph.add((scheme, SKOS.prefLabel, Literal("Test", lang="en")))

        # Add concepts with specific IDs
        for concept_id in [1, 2, 5]:
            concept_iri = URIRef(f"https://example.org/{concept_id:07d}")
            graph.add((concept_iri, RDF.type, SKOS.Concept))
            graph.add(
                (
                    concept_iri,
                    SKOS.prefLabel,
                    Literal(f"Concept {concept_id}", lang="en"),
                )
            )
            graph.add(
                (
                    concept_iri,
                    SKOS.definition,
                    Literal(f"Definition {concept_id}", lang="en"),
                )
            )
            graph.add((scheme, SKOS.hasTopConcept, concept_iri))

        # Save RDF to file
        ttl_path = tmp_path / "test.ttl"
        graph.serialize(ttl_path, format="turtle")

        vocab_config = Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks=Checks(allow_delete=False),
            prefix_map={"ex": "https://example.org/"},
            vocabulary_iri="https://example.org/",
            title="Test Vocabulary",
            description="Test description",
            created_date="2025-01-01",
            creator="Test Author",
            repository="https://github.com/test/vocab",
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "alice"},
            ],
        )

        xlsx_path = tmp_path / "test.xlsx"
        rdf_to_excel_v1(ttl_path, xlsx_path, vocab_config=vocab_config)

        wb = load_workbook(xlsx_path)
        ws = wb["ID Ranges"]

        # With IDs 1, 2, 5 used in range 1-10, we have 7 unused
        # First unused should be 3
        unused_value = ws["C4"].value
        assert "next unused: 0000003" in unused_value
        assert "unused: 7" in unused_value


# =============================================================================
# Tests for Contributor Derivation
# =============================================================================


class TestFormatContributorString:
    """Tests for format_contributor_string function."""

    def test_with_orcid_and_name(self, temp_config):
        """Test formatting with ORCID and name."""
        config = temp_config
        idr = config.IdrangeItem(
            first_id=1,
            last_id=10,
            gh_name="testuser",
            name="Test User",
            orcid="0000-0001-2345-6789",
        )
        result = format_contributor_string(idr)
        assert result == "Test User https://orcid.org/0000-0001-2345-6789"

    def test_with_orcid_url_and_name(self, temp_config):
        """Test formatting with full ORCID URL and name."""
        config = temp_config
        idr = config.IdrangeItem(
            first_id=1,
            last_id=10,
            gh_name="testuser",
            name="Test User",
            orcid="https://orcid.org/0000-0001-2345-6789",
        )
        result = format_contributor_string(idr)
        assert result == "Test User https://orcid.org/0000-0001-2345-6789"

    def test_with_orcid_no_name(self, temp_config):
        """Test formatting with ORCID but no name."""
        config = temp_config
        idr = config.IdrangeItem(
            first_id=1,
            last_id=10,
            gh_name="",
            orcid="0000-0001-2345-6789",
        )
        result = format_contributor_string(idr)
        assert result == "https://orcid.org/0000-0001-2345-6789"

    def test_with_gh_name_and_name(self, temp_config):
        """Test formatting with gh_name and name (no ORCID)."""
        config = temp_config
        idr = config.IdrangeItem(
            first_id=1,
            last_id=10,
            gh_name="testuser",
            name="Test User",
        )
        result = format_contributor_string(idr)
        assert result == "Test User https://github.com/testuser"

    def test_with_gh_name_only(self, temp_config):
        """Test formatting with only gh_name (no name, no ORCID)."""
        config = temp_config
        idr = config.IdrangeItem(
            first_id=1,
            last_id=10,
            gh_name="testuser",
        )
        result = format_contributor_string(idr)
        assert result == "testuser https://github.com/testuser"

    def test_orcid_takes_priority_over_gh_name(self, temp_config):
        """Test that ORCID URL is used when both ORCID and gh_name are present."""
        config = temp_config
        idr = config.IdrangeItem(
            first_id=1,
            last_id=10,
            gh_name="testuser",
            name="Test User",
            orcid="0000-0001-2345-6789",
        )
        result = format_contributor_string(idr)
        # Should use ORCID, not GitHub URL
        assert "orcid.org" in result
        assert "github.com" not in result


class TestDeriveContributors:
    """Tests for derive_contributors function."""

    def test_derives_contributor_from_used_range(self, temp_config):
        """Test that contributor is derived when IDs from their range are used."""
        config = temp_config
        vocab = config.Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks={},
            prefix_map={},
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test",
            created_date="2025-01-01",
            creator="Creator https://orcid.org/0000-0000-0000-0001",
            repository="https://github.com/test/vocab",
            id_range=[
                {
                    "first_id": 1,
                    "last_id": 10,
                    "gh_name": "user1",
                    "name": "User One",
                    "orcid": "0000-0001-2345-6789",
                },
                {"first_id": 11, "last_id": 20, "gh_name": "user2"},
            ],
        )

        # Only user1's range has used IDs
        used_ids = {1, 2, 3}

        result = derive_contributors(vocab, used_ids)
        assert "User One" in result
        assert "0000-0001-2345-6789" in result
        assert "user2" not in result

    def test_excludes_creators_by_orcid(self, temp_config):
        """Test that creators are excluded from contributors (matching by ORCID)."""
        config = temp_config
        vocab = config.Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks={},
            prefix_map={},
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test",
            created_date="2025-01-01",
            creator="User One https://orcid.org/0000-0001-2345-6789",
            repository="https://github.com/test/vocab",
            id_range=[
                {
                    "first_id": 1,
                    "last_id": 10,
                    "gh_name": "user1",
                    "name": "User One",
                    "orcid": "0000-0001-2345-6789",
                },
            ],
        )

        used_ids = {1, 2, 3}

        result = derive_contributors(vocab, used_ids)
        # User One is excluded because they're the creator
        assert result == ""

    def test_excludes_creators_by_gh_name(self, temp_config):
        """Test that creators are excluded from contributors (matching by gh_name)."""
        config = temp_config
        vocab = config.Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks={},
            prefix_map={},
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test",
            created_date="2025-01-01",
            creator="User One https://github.com/user1",
            repository="https://github.com/test/vocab",
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "user1", "name": "User One"},
            ],
        )

        used_ids = {1, 2, 3}

        result = derive_contributors(vocab, used_ids)
        # user1 is excluded because they're the creator
        assert result == ""

    def test_no_duplicates_multiple_ranges(self, temp_config):
        """Test that same contributor from multiple ranges appears only once."""
        config = temp_config
        vocab = config.Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks={},
            prefix_map={},
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test",
            created_date="2025-01-01",
            creator="Creator https://orcid.org/0000-0000-0000-0001",
            repository="https://github.com/test/vocab",
            id_range=[
                {
                    "first_id": 1,
                    "last_id": 10,
                    "gh_name": "user1",
                    "orcid": "0000-0001-2345-6789",
                },
                {
                    "first_id": 11,
                    "last_id": 20,
                    "gh_name": "user1",
                    "orcid": "0000-0001-2345-6789",
                },
            ],
        )

        # IDs used from both ranges
        used_ids = {5, 15}

        result = derive_contributors(vocab, used_ids)
        # Should only have one entry for user1
        assert result.count("0000-0001-2345-6789") == 1

    def test_returns_empty_when_no_ids_used(self, temp_config):
        """Test that empty string is returned when no IDs are used."""
        config = temp_config
        vocab = config.Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks={},
            prefix_map={},
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test",
            created_date="2025-01-01",
            creator="Creator https://orcid.org/0000-0000-0000-0001",
            repository="https://github.com/test/vocab",
            id_range=[
                {"first_id": 1, "last_id": 10, "gh_name": "user1"},
            ],
        )

        # No IDs used
        used_ids = set()

        result = derive_contributors(vocab, used_ids)
        assert result == ""

    def test_multiple_contributors(self, temp_config):
        """Test deriving multiple contributors from different ranges."""
        config = temp_config
        vocab = config.Vocab(
            id_length=7,
            permanent_iri_part="https://example.org/",
            checks={},
            prefix_map={},
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test",
            created_date="2025-01-01",
            creator="Creator https://orcid.org/0000-0000-0000-0001",
            repository="https://github.com/test/vocab",
            id_range=[
                {
                    "first_id": 1,
                    "last_id": 10,
                    "gh_name": "alice",
                    "name": "Alice Smith",
                    "orcid": "0000-0001-2345-6789",  # Valid ORCID
                },
                {
                    "first_id": 11,
                    "last_id": 20,
                    "gh_name": "bob",
                    "name": "Bob Jones",
                    "orcid": "0000-0002-1825-0097",  # Valid ORCID
                },
                {
                    "first_id": 21,
                    "last_id": 30,
                    "gh_name": "charlie",
                },
            ],
        )

        # IDs used from alice's and charlie's ranges only
        used_ids = {5, 25}

        result = derive_contributors(vocab, used_ids)
        # Should have Alice and Charlie, but not Bob
        assert "Alice Smith" in result
        assert "0000-0001-2345-6789" in result
        assert "charlie" in result
        assert "github.com/charlie" in result
        assert "Bob Jones" not in result
        assert "0000-0002-1825-0097" not in result

        # Should be two lines
        lines = [line for line in result.split("\n") if line.strip()]
        assert len(lines) == 2


class TestHistoryNoteOnlyWithDeprecation:
    """Tests for skos:historyNote - only added when deprecated."""

    def test_concept_without_deprecation_no_history_note(self):
        """Non-deprecated concept gets no historyNote."""
        concept = AggregatedConcept(
            iri="http://example.org/concept1",
            pref_labels={"en": "Test Concept"},
            definitions={"en": "A test concept"},
        )
        scheme_iri = URIRef("http://example.org/")

        graph = build_concept_graph(
            concept, scheme_iri, narrower_map={}, id_pattern=None
        )

        history_notes = list(
            graph.objects(URIRef("http://example.org/concept1"), SKOS.historyNote)
        )
        assert len(history_notes) == 0

    def test_deprecated_concept_gets_history_note(self):
        """Deprecated concept gets historyNote with obsoletion reason."""
        concept = AggregatedConcept(
            iri="http://example.org/concept1",
            pref_labels={"en": "Test Concept"},
            definitions={"en": "A test concept"},
            obsolete_reason="Merged with concept2",
        )
        scheme_iri = URIRef("http://example.org/")

        graph = build_concept_graph(
            concept, scheme_iri, narrower_map={}, id_pattern=None
        )

        history_notes = list(
            graph.objects(URIRef("http://example.org/concept1"), SKOS.historyNote)
        )
        assert len(history_notes) == 1
        assert "Merged with concept2" in str(history_notes[0])

    def test_collection_without_deprecation_no_history_note(self):
        """Non-deprecated collection gets no historyNote."""
        collection = AggregatedCollection(
            iri="http://example.org/collection1",
            pref_labels={"en": "Test Collection"},
            definitions={"en": "A test collection"},
        )
        scheme_iri = URIRef("http://example.org/")

        graph = build_collection_graph(
            collection,
            scheme_iri,
            collection_members={},
            ordered_collection_members={},
            id_pattern=None,
        )

        history_notes = list(
            graph.objects(URIRef("http://example.org/collection1"), SKOS.historyNote)
        )
        assert len(history_notes) == 0

    def test_deprecated_collection_gets_history_note(self):
        """Deprecated collection gets historyNote with obsoletion reason."""
        collection = AggregatedCollection(
            iri="http://example.org/collection1",
            pref_labels={"en": "Test Collection"},
            definitions={"en": "A test collection"},
            obsolete_reason="No longer needed",
        )
        scheme_iri = URIRef("http://example.org/")

        graph = build_collection_graph(
            collection,
            scheme_iri,
            collection_members={},
            ordered_collection_members={},
            id_pattern=None,
        )

        history_notes = list(
            graph.objects(URIRef("http://example.org/collection1"), SKOS.historyNote)
        )
        assert len(history_notes) == 1
        assert "No longer needed" in str(history_notes[0])


# =============================================================================
# String-to-Enum Converter Tests
# =============================================================================


class TestStringToEnumConverters:
    """Tests for string_to_*_enum conversion functions."""

    # --- string_to_concept_obsoletion_enum ---

    def test_concept_obsoletion_valid_reason(self):
        """Test valid obsoletion reason returns enum."""

        result = string_to_concept_obsoletion_enum(
            ConceptObsoletionReason.UNCLEAR.value
        )
        assert result == ConceptObsoletionReason.UNCLEAR

    def test_concept_obsoletion_empty_string(self):
        """Test empty string returns None."""

        assert string_to_concept_obsoletion_enum("") is None

    def test_concept_obsoletion_nonstandard_logs_warning(self, caplog):
        """Test non-standard reason logs warning and returns None (lines 133-134)."""

        with caplog.at_level(logging.WARNING):
            result = string_to_concept_obsoletion_enum("This is not a valid reason")

        assert result is None
        assert "Non-standard obsoletion reason for concept" in caplog.text

    # --- string_to_collection_obsoletion_enum ---

    def test_collection_obsoletion_valid_reason(self):
        """Test valid collection obsoletion reason returns enum."""

        result = string_to_collection_obsoletion_enum(
            CollectionObsoletionReason.UNCLEAR.value
        )
        assert result == CollectionObsoletionReason.UNCLEAR

    def test_collection_obsoletion_empty_returns_none(self):
        """Test empty string returns None for collection."""

        assert string_to_collection_obsoletion_enum("") is None

    def test_collection_obsoletion_nonstandard_logs_warning(self, caplog):
        """Test non-standard collection reason logs warning (lines 154-155)."""

        with caplog.at_level(logging.WARNING):
            result = string_to_collection_obsoletion_enum("Invalid collection reason")

        assert result is None
        assert "Non-standard obsoletion reason for collection" in caplog.text

    # --- string_to_ordered_enum ---

    @pytest.mark.parametrize(
        ("value", "expected"),
        [
            (True, OrderedChoice.YES),
            (False, None),  # Boolean False -> None (line 168)
            ("yes", OrderedChoice.YES),
            ("YES", OrderedChoice.YES),  # Case-insensitive (line 171)
            ("Yes", OrderedChoice.YES),
            ("no", OrderedChoice.NO),  # "no" string (line 173-174)
            ("NO", OrderedChoice.NO),
            ("No", OrderedChoice.NO),
            ("", None),  # Empty string (line 169-170)
            ("  yes  ", OrderedChoice.YES),  # Whitespace trimmed
            ("maybe", None),  # Invalid string returns None (line 175)
        ],
    )
    def test_string_to_ordered_enum_parameterized(self, value, expected):
        """Parameterized test for string_to_ordered_enum (lines 167-175)."""

        result = string_to_ordered_enum(value)
        assert result == expected


# =============================================================================
# Entity Graph Building Tests
# =============================================================================


class TestBuildEntityGraph:
    """Tests for build_entity_graph function."""

    def test_orcid_url_creates_person(self):
        """Test ORCID URL always creates Person (lines 2371-2375)."""

        graph = build_entity_graph(
            "https://orcid.org/0000-0001-2345-6789",
            "John Doe",
            "publisher",  # Even though field_type is publisher, ORCID -> Person
        )

        person_uri = URIRef("https://orcid.org/0000-0001-2345-6789")
        assert (person_uri, RDF.type, SDO.Person) in graph

    def test_ror_url_creates_organization(self):
        """Test ROR URL creates Organization (lines 2375-2383)."""

        graph = build_entity_graph(
            "https://ror.org/04wx23d79",
            "Example University",
            "creator",  # Even for creator, ROR -> Organization
        )

        org_uri = URIRef("https://ror.org/04wx23d79")
        assert (org_uri, RDF.type, SDO.Organization) in graph

    def test_publisher_non_ror_creates_organization(self):
        """Test publisher field without ROR creates Organization (lines 2383-2394)."""

        graph = build_entity_graph(
            "https://example.org/publisher",
            "Example Publisher",
            "publisher",  # Publisher field -> Organization
        )

        org_uri = URIRef("https://example.org/publisher")
        assert (org_uri, RDF.type, SDO.Organization) in graph

    def test_creator_non_special_creates_person(self):
        """Test creator without ORCID/ROR creates Person (lines 2395-2403)."""

        graph = build_entity_graph(
            "https://example.org/creator",
            "Jane Smith",
            "creator",
        )

        person_uri = URIRef("https://example.org/creator")
        assert (person_uri, RDF.type, SDO.Person) in graph

    def test_contributor_creates_person_by_default(self):
        """Test contributor field defaults to Person."""

        graph = build_entity_graph(
            "https://example.org/contributor",
            "Contributor Name",
            "contributor",
        )

        person_uri = URIRef("https://example.org/contributor")
        assert (person_uri, RDF.type, SDO.Person) in graph

    def test_custodian_creates_person_by_default(self):
        """Test custodian field defaults to Person."""

        graph = build_entity_graph(
            "https://example.org/custodian",
            "Custodian Name",
            "custodian",
        )

        person_uri = URIRef("https://example.org/custodian")
        assert (person_uri, RDF.type, SDO.Person) in graph


# =============================================================================
# ConceptScheme Entity Graph Tests
# =============================================================================


class TestBuildConceptSchemeGraphEntities:
    """Tests for entity creation in build_concept_scheme_graph."""

    def test_multiple_creators_each_get_entity_graph(self):
        """Test multiple creators each generate entity graphs (lines 2414-2423)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            creator="Alice https://orcid.org/0000-0001-1111-1111\nBob https://orcid.org/0000-0002-2222-2222",
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        # Both should be Person entities
        alice_uri = URIRef("https://orcid.org/0000-0001-1111-1111")
        bob_uri = URIRef("https://orcid.org/0000-0002-2222-2222")
        assert (alice_uri, RDF.type, SDO.Person) in graph
        assert (bob_uri, RDF.type, SDO.Person) in graph

    def test_publisher_ror_creates_organization(self):
        """Test publisher with ROR URL creates Organization (lines 2430-2434)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            publisher="University https://ror.org/04wx23d79",
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        ror_uri = URIRef("https://ror.org/04wx23d79")
        assert (ror_uri, RDF.type, SDO.Organization) in graph

    def test_contributor_plain_text_creates_literal(self):
        """Test contributor without URL creates literal (line 2423)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            contributor="Plain Text Contributor",  # No URL
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        contributors = list(graph.objects(scheme_uri, DCTERMS.contributor))
        assert Literal("Plain Text Contributor") in contributors

    def test_contributor_with_url_different_from_creator(self):
        """Test contributor with URL (not in creator_urls) gets entity graph (lines 2417-2420)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            creator="Creator Person https://orcid.org/0000-0001-1111-1111",
            contributor="Contributor Person https://orcid.org/0000-0002-2222-2222",  # Different URL
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        # Contributor should have entity graph created
        contrib_uri = URIRef("https://orcid.org/0000-0002-2222-2222")
        assert (contrib_uri, RDF.type, SDO.Person) in graph

        # Both should be linked as contributor/creator
        scheme_uri = URIRef("https://example.org/vocab/")
        contributors = list(graph.objects(scheme_uri, DCTERMS.contributor))
        assert URIRef("https://orcid.org/0000-0002-2222-2222") in contributors

    def test_duplicate_creator_publisher_entity_not_duplicated(self):
        """Test same URL as creator and publisher doesn't duplicate entity (lines 2406-2409)."""

        same_url = "https://orcid.org/0000-0001-2345-6789"
        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            creator=f"Alice {same_url}",
            publisher=f"Alice {same_url}",  # Same URL as creator
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        entity_uri = URIRef(same_url)
        # Count Person type triples - should be exactly 1
        person_triples = list(graph.triples((entity_uri, RDF.type, SDO.Person)))
        assert len(person_triples) == 1

    def test_custodian_with_url_creates_entity(self):
        """Test custodian with URL creates entity graph (lines 2440-2441)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            custodian="Admin Person https://orcid.org/0000-0003-3333-3333",
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        custodian_uri = URIRef("https://orcid.org/0000-0003-3333-3333")
        assert (custodian_uri, RDF.type, SDO.Person) in graph

    def test_custodian_without_url_creates_literal(self):
        """Test custodian without URL creates literal contactPoint (lines 2442-2444)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            custodian="contact@example.org",  # No URL, email only
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        contacts = list(graph.objects(scheme_uri, DCAT.contactPoint))
        assert Literal("contact@example.org") in contacts


# =============================================================================
# Optional Metadata Fields Tests
# =============================================================================


class TestConceptSchemeOptionalMetadata:
    """Tests for optional metadata in ConceptScheme graph building."""

    def test_scheme_without_created_date(self):
        """Test ConceptScheme without created date (line 1717)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            created_date="",  # No created date
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        created_dates = list(graph.objects(scheme_uri, DCTERMS.created))
        assert len(created_dates) == 0

    def test_scheme_without_modified_date(self):
        """Test ConceptScheme without modified date (line 1856)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            modified_date="",  # No modified date
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        modified_dates = list(graph.objects(scheme_uri, DCTERMS.modified))
        assert len(modified_dates) == 0

    def test_scheme_catalogue_pid_as_uri(self):
        """Test catalogue_pid handling as URI (lines 2447-2452)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            catalogue_pid="https://doi.org/10.5281/zenodo.12345",
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        see_also = list(graph.objects(scheme_uri, RDFS.seeAlso))
        assert URIRef("https://doi.org/10.5281/zenodo.12345") in see_also

    def test_scheme_catalogue_pid_as_literal(self):
        """Test catalogue_pid handling as literal when not URL (line 2452)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            catalogue_pid="DOI:10.5281/zenodo.12345",  # Not starting with http
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        see_also = list(graph.objects(scheme_uri, RDFS.seeAlso))
        assert Literal("DOI:10.5281/zenodo.12345") in see_also

    def test_scheme_homepage_as_literal(self):
        """Test homepage handling as literal when not URL (line 2459)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            homepage="not-a-url",  # Not starting with http
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        homepages = list(graph.objects(scheme_uri, FOAF.homepage))
        assert Literal("not-a-url") in homepages

    def test_scheme_conforms_to_as_uri(self):
        """Test conforms_to handling as URI (lines 2467-2468)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            conforms_to="https://example.org/profile",
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        conforms = list(graph.objects(scheme_uri, DCTERMS.conformsTo))
        assert URIRef("https://example.org/profile") in conforms

    def test_scheme_conforms_to_as_literal(self):
        """Test conforms_to handling as literal when not URL (line 2470)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            conforms_to="SHACL Profile v1",  # Not starting with http
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        conforms = list(graph.objects(scheme_uri, DCTERMS.conformsTo))
        assert Literal("SHACL Profile v1") in conforms

    def test_scheme_conforms_to_multiline_with_empty_lines(self):
        """Test conforms_to handling with empty lines (line 2466)."""

        cs = ConceptSchemeV1(
            vocabulary_iri="https://example.org/vocab/",
            title="Test",
            description="Test vocab",
            conforms_to="https://example.org/profile1\n\nhttps://example.org/profile2",
        )

        graph = build_concept_scheme_graph(cs, {}, id_pattern=None)

        scheme_uri = URIRef("https://example.org/vocab/")
        conforms = list(graph.objects(scheme_uri, DCTERMS.conformsTo))
        assert len(conforms) == 2  # Empty line should be skipped


# =============================================================================
# IRI Parsing Edge Cases Tests
# =============================================================================


class TestIdentifierExtraction:
    """Tests for extract_identifier with ID patterns."""

    def test_extract_identifier_7_digits(self):
        """Test identifier extraction with 7-digit pattern."""

        pattern = re.compile(r"(?P<identifier>[0-9]{7})$")
        result = extract_identifier("https://example.org/vocab_0000123", pattern)
        assert result == "0000123"

    def test_extract_identifier_from_underscore_format(self):
        """Test identifier extraction from prefix_ID format."""

        pattern = re.compile(r"(?P<identifier>[0-9]{7})$")
        result = extract_identifier(
            "https://w3id.org/nfdi4cat/voc4cat_0000195", pattern
        )
        assert result == "0000195"

    def test_extract_identifier_no_match_logs_error(self):
        """Test that non-matching IRI returns empty string and logs error."""

        pattern = re.compile(r"(?P<identifier>[0-9]{7})$")
        result = extract_identifier("https://example.org/concept", pattern)
        assert result == ""

    def test_extract_identifier_different_lengths(self):
        """Test identifier extraction with different digit lengths."""

        pattern_5 = re.compile(r"(?P<identifier>[0-9]{5})$")
        result = extract_identifier("https://example.org/vocab_00123", pattern_5)
        assert result == "00123"

        pattern_4 = re.compile(r"(?P<identifier>[0-9]{4})$")
        result = extract_identifier("https://example.org/vocab_1234", pattern_4)
        assert result == "1234"

    def test_parse_name_url_name_and_url(self):
        """Test parse_name_url with name and URL."""

        name, url = parse_name_url("John Doe https://orcid.org/0000-0001-2345-6789")
        assert name == "John Doe"
        assert url == "https://orcid.org/0000-0001-2345-6789"

    def test_parse_name_url_url_only(self):
        """Test parse_name_url with URL only."""

        name, url = parse_name_url("https://orcid.org/0000-0001-2345-6789")
        assert name == ""  # Empty string, not None
        assert url == "https://orcid.org/0000-0001-2345-6789"

    def test_parse_name_url_name_only(self):
        """Test parse_name_url with name only (no URL)."""

        name, url = parse_name_url("John Doe")
        assert name == "John Doe"
        assert url == ""  # Empty string, not None

    def test_parse_name_url_empty(self):
        """Test parse_name_url with empty string."""

        name, url = parse_name_url("")
        assert name == ""  # Empty string
        assert url == ""  # Empty string
