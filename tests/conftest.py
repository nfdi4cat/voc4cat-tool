# Common pytest fixtures for all test modules
import re
import shutil
import tempfile
from datetime import date
from enum import Enum
from pathlib import Path
from typing import Annotated

import pytest
from curies import Converter
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from rdflib import DCTERMS, RDF, SKOS, Graph, URIRef
from rdflib.namespace import NamespaceManager

from voc4cat import config
from voc4cat.config import Checks, Vocab
from voc4cat.convert_v1 import rdf_to_excel_v1
from voc4cat.xlsx_common import XLSXMetadata


# Test Enums
class Priority(Enum):
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"


class Status(Enum):
    ACTIVE = "active"
    INACTIVE = "inactive"
    PENDING = "pending"


# Test Models
class Employee(BaseModel):
    """Test model for employee data."""

    employee_id: int
    first_name: str
    last_name: str
    email: str
    hire_date: date
    salary: float
    status: Status
    department: str | None = None
    is_active: bool = True


class Project(BaseModel):
    """Test model for project data."""

    project_id: int
    project_name: str
    start_date: date
    priority: Priority
    budget: float
    description: str | None = Field(None, description="Project description")
    end_date: date | None = None
    is_completed: bool = False


class SimpleModel(BaseModel):
    """Simple test model."""

    name: str
    value: int
    active: bool = True


class DemoModelWithMetadata(BaseModel):
    """Demo model with XLSX metadata for testing."""

    temp: Annotated[
        float,
        XLSXMetadata(
            unit="°C",
            meaning="Temperature measurement",
            description="Temperature reading",
        ),
    ]
    name: Annotated[
        str,
        XLSXMetadata(
            meaning="Sample name",
            description="Sample identifier",
        ),
    ]


@pytest.fixture(scope="session")
def datadir():
    """DATADIR as a LocalPath"""
    return Path(__file__).resolve().parent / "data"


@pytest.fixture(scope="session")
def cs_cycles_xlsx(datadir, tmp_path_factory) -> Path:
    """Generate concept-scheme-with-cycles.xlsx from TTL.

    Session-scoped to avoid regenerating for each test.
    Returns path to the generated xlsx file in a temp directory.
    """
    ttl_file = datadir / "concept-scheme-with-cycles.ttl"
    output_dir = tmp_path_factory.mktemp("xlsx_cycles")
    output_file = output_dir / "concept-scheme-with-cycles.xlsx"
    graph = Graph().parse(str(ttl_file), format="turtle")
    vocab_config = make_vocab_config_from_rdf(graph)
    rdf_to_excel_v1(ttl_file, output_file, vocab_config=vocab_config)
    return output_file


@pytest.fixture(scope="session")
def cs_duplicates_xlsx(cs_cycles_xlsx, tmp_path_factory) -> Path:
    """Generate xlsx with duplicate IRI for error detection testing.

    In the v1.0 template, data starts at row 6 (after title, empty, meanings,
    requiredness, headers). Creates duplicate concept row at rows 6 and 7
    for testing that the check command colors cells orange (#FFCC00).
    """
    output_dir = tmp_path_factory.mktemp("xlsx_duplicates")
    output_file = output_dir / "concept-scheme-duplicates.xlsx"
    shutil.copy(cs_cycles_xlsx, output_file)

    wb = load_workbook(output_file)
    ws = wb["Concepts"]

    # In v1.0 template: row 1=title, 2=empty, 3=meanings, 4=requiredness, 5=headers, 6+=data
    # Duplicate row 6 (first data row) to row 7
    ws.insert_rows(7)
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=7, column=col_idx).value = ws.cell(row=6, column=col_idx).value

    wb.save(output_file)
    wb.close()
    return output_file


@pytest.fixture
def temp_config():
    """
    Provides a temporary config that can be safely changed in test functions.

    After the test the config will be reset to default.
    """

    config.curies_converter.add_prefix("ex", "http://example.org/", merge=True)
    yield config

    # Reset the globally changed config to default.
    config.load_config()
    # Also reset curies_converter to default state
    config.curies_converter = Converter.from_prefix_map(
        {prefix: str(url) for prefix, url in NamespaceManager(Graph()).namespaces()}
    )
    config.curies_converter.add_prefix("ex", "http://example.org/", merge=True)


# Vocab Config Fixtures for v1.0 conversion tests
@pytest.fixture
def test_vocab_config():
    """Minimal vocab config for testing excel_to_rdf_v1."""

    return Vocab(
        id_length=7,
        permanent_iri_part="http://example.org/test/",
        checks=Checks(allow_delete=False),
        prefix_map={"ex": "http://example.org/"},
        vocabulary_iri="http://example.org/test/",
        title="Test Vocabulary",
        description="Test vocabulary for unit tests",
        created_date="2025-01-01",
        creator="https://orcid.org/0000-0001-2345-6789",
        repository="https://github.com/test/vocab",
    )


@pytest.fixture
def mandatory_fields():
    """Default mandatory fields for testing - can be merged into test Vocab configs."""
    return {
        "vocabulary_iri": "https://example.org/vocab/",
        "title": "Test Vocabulary",
        "description": "Test vocabulary for unit tests",
        "created_date": "2025-01-01",
        "creator": "https://orcid.org/0000-0001-2345-6789",
        "repository": "https://github.com/test/vocab",
    }


def make_vocab_config_from_rdf(
    graph, vocab_iri: str | None = None, vocab_name: str | None = None
):
    """Create a Vocab config from RDF graph for testing roundtrips.

    This extracts ConceptScheme metadata from the RDF and creates a Vocab config
    that can be used with excel_to_rdf_v1.

    Args:
        graph: RDF graph containing the ConceptScheme.
        vocab_iri: Optional vocabulary IRI to use.
        vocab_name: Optional vocab name to register in config.ID_PATTERNS.
    """
    # Find concept scheme
    if vocab_iri is None:  # pragma: no branch
        schemes = list(graph.subjects(RDF.type, SKOS.ConceptScheme))
        assert schemes, "No ConceptScheme found in RDF graph"
        vocab_iri = str(schemes[0])

    scheme_uri = URIRef(vocab_iri)

    # Extract metadata from the concept scheme
    title = ""
    description = ""
    created_date = ""
    creator = ""

    # Get title from concept scheme's prefLabel
    for obj in graph.objects(scheme_uri, SKOS.prefLabel):  # pragma: no branch
        title = str(obj)
        break

    # Get description from concept scheme's definition
    for obj in graph.objects(scheme_uri, SKOS.definition):  # pragma: no branch
        description = str(obj)
        break

    # Get created date
    for obj in graph.objects(scheme_uri, DCTERMS.created):  # pragma: no branch
        created_date = str(obj)
        break

    # Get creator
    for obj in graph.objects(scheme_uri, DCTERMS.creator):  # pragma: no branch
        creator = str(obj)
        break

    # Derive permanent_iri_part from vocab_iri
    permanent_iri_part = vocab_iri.rstrip("/") + "/"

    id_length = 7
    vocab_config = Vocab(
        id_length=id_length,
        permanent_iri_part=permanent_iri_part,
        checks=Checks(allow_delete=False),
        prefix_map={"ex": permanent_iri_part},
        vocabulary_iri=vocab_iri,
        title=title,
        description=description,
        created_date=created_date,
        creator=creator,
        repository="https://github.com/test/vocab",  # Always provide default
    )

    # Register ID pattern in config if vocab_name provided
    if vocab_name:
        config.ID_PATTERNS[vocab_name] = re.compile(
            rf"(?P<identifier>[0-9]{{{id_length}}})$"
        )

    return vocab_config


# XLSX Testing Fixtures
@pytest.fixture
def sample_employees():
    """Sample employee data."""
    return [
        Employee(
            employee_id=1,
            first_name="John",
            last_name="Doe",
            email="john.doe@company.com",
            hire_date=date(2023, 1, 15),
            salary=75000.0,
            status=Status.ACTIVE,
            department="Engineering",
        ),
        Employee(
            employee_id=2,
            first_name="Jane",
            last_name="Smith",
            email="jane.smith@company.com",
            hire_date=date(2023, 3, 1),
            salary=82000.0,
            status=Status.ACTIVE,
            department="Marketing",
            is_active=False,
        ),
    ]


@pytest.fixture
def sample_projects():
    """Sample project data."""
    return [
        Project(
            project_id=1,
            project_name="Website Redesign",
            description="Complete overhaul of company website",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 6, 30),
            priority=Priority.HIGH,
            budget=150000.0,
        ),
        Project(
            project_id=2,
            project_name="Mobile App",
            description="Mobile application development",
            start_date=date(2024, 2, 15),
            priority=Priority.MEDIUM,
            budget=200000.0,
        ),
    ]


@pytest.fixture
def sample_simple_model():
    """Sample simple model."""
    return SimpleModel(name="Test Item", value=42, active=True)


@pytest.fixture
def sample_model_with_metadata():
    """Sample model with XLSX metadata."""
    return DemoModelWithMetadata(temp=25.5, name="Test Sample")


@pytest.fixture
def temp_file():
    """Temporary file for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        yield Path(f.name)
    Path(f.name).unlink(missing_ok=True)


# Additional XLSX Testing Fixtures for Optimized Testing
@pytest.fixture
def sample_employee():
    """Single employee for key-value testing."""
    return Employee(
        employee_id=1,
        first_name="John",
        last_name="Doe",
        email="john@test.com",
        hire_date=date(2023, 1, 1),
        salary=50000.0,
        status=Status.ACTIVE,
        department="Engineering",
        is_active=True,
    )


@pytest.fixture
def sample_project():
    """Single project for key-value testing."""
    return Project(
        project_id=1,
        project_name="Test Project",
        start_date=date(2024, 1, 1),
        priority=Priority.HIGH,
        budget=100000.0,
        description="Test project description",
        end_date=None,
        is_completed=False,
    )


@pytest.fixture
def xlsx_metadata_map():
    """Common XLSX metadata for Employee model."""
    return {
        "salary": XLSXMetadata(unit="USD", description="Annual salary"),
        "first_name": XLSXMetadata(description="Employee first name"),
        "hire_date": XLSXMetadata(description="Date of hire"),
    }


@pytest.fixture
def large_employee_dataset():
    """Large dataset for performance testing."""
    return [
        Employee(
            employee_id=i,
            first_name=f"Employee{i}",
            last_name=f"Lastname{i}",
            email=f"emp{i}@company.com",
            hire_date=date(2023, 1, 1),
            salary=50000.0 + (i * 1000),
            status=Status.ACTIVE if i % 2 == 0 else Status.INACTIVE,
            department=f"Dept{i % 5}",
        )
        for i in range(100)  # 100 employees
    ]
