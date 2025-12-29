"""
Tests for the xlsx_common module.

This module tests the common functionality used across the XLSX processing system,
including field analysis, serialization engine, metadata handling, and converters.
"""

import logging
from datetime import date, datetime, timezone
from typing import Annotated

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
from pydantic import BaseModel, Field, HttpUrl
from pydantic_core import PydanticUndefined

from voc4cat.fields import ORCIDIdentifier, RORIdentifier
from voc4cat.xlsx_api import export_to_xlsx, import_from_xlsx
from voc4cat.xlsx_common import (
    FieldAnalysis,
    MetadataToggleConfig,
    MetadataVisibility,
    XLSXConfig,
    XLSXConverters,
    XLSXDeserializationError,
    XLSXFieldAnalyzer,
    XLSXMetadata,
    XLSXSerializationEngine,
    XLSXSerializationError,
    _validate_unit_usage,
    adjust_all_tables_length,
    adjust_table_length,
)

from .conftest import DemoModelWithMetadata, Priority, SimpleModel, Status


# ORCID/ROR Test Model (only used in this file)
class ResearcherModel(BaseModel):
    """Test model for ORCID/ROR identifiers."""

    name: str
    orcid: ORCIDIdentifier
    home_organization: RORIdentifier | None = None


@pytest.fixture
def sample_researcher():
    """Sample researcher with ORCID and ROR for testing."""
    return ResearcherModel(
        name="Dr. Jane Smith",
        orcid="https://orcid.org/0000-0002-1825-0097",
        home_organization="https://ror.org/02y72wh86",
    )


@pytest.fixture
def sample_researchers():
    """Sample researcher list for table testing."""
    return [
        ResearcherModel(
            name="Dr. Jane Smith",
            orcid="https://orcid.org/0000-0002-1825-0097",
            home_organization="https://ror.org/02y72wh86",
        ),
        ResearcherModel(
            name="Dr. John Doe",
            orcid="0000-0002-1694-233X",  # Test ID-only format
            home_organization=None,
        ),
    ]


# Tests for XLSXMetadata
class TestXLSXMetadata:
    """Tests for XLSXMetadata and Annotated field handling."""

    def test_xlsx_metadata_basic(self):
        """Test basic XLSXMetadata functionality."""
        metadata = XLSXMetadata(unit="kg")
        assert metadata.unit == "kg"
        assert metadata.meaning is None
        assert metadata.display_name is None
        assert metadata.description is None

    def test_xlsx_metadata_with_all_fields(self):
        """Test XLSXMetadata with all fields."""
        metadata = XLSXMetadata(
            unit="kg",
            meaning="http://example.org/weight",
            display_name="Weight",
            description="Weight measurement",
        )
        assert metadata.unit == "kg"
        assert metadata.meaning == "http://example.org/weight"
        assert metadata.display_name == "Weight"
        assert metadata.description == "Weight measurement"

    def test_annotated_field_extraction(self):
        """Test extraction of metadata from Annotated fields."""
        fields = XLSXFieldAnalyzer.analyze_model(DemoModelWithMetadata)
        temp_field = next(f for f in fields.values() if f.name == "temp")
        assert temp_field.xlsx_metadata is not None
        assert temp_field.xlsx_metadata.unit == "°C"
        assert temp_field.xlsx_metadata.meaning == "Temperature measurement"


# Tests for XLSXConverters
class TestXLSXConverters:
    """Tests for XLSXConverters utility class."""

    def test_comma_pattern_converters(self):
        """Test comma pattern string converters."""
        to_comma, from_comma = XLSXConverters.comma_converters()

        assert to_comma(["a", "b", "c"]) == "a, b, c"
        assert to_comma([1, 2, 3]) == "1, 2, 3"
        assert to_comma([]) == ""
        assert to_comma(None) == ""

        assert from_comma("a, b, c") == ["a", "b", "c"]
        assert from_comma("a,b,c") == ["a,b,c"]  # No spaces, so treated as one item
        assert from_comma("") == []
        assert from_comma("   ") == []
        assert from_comma("a, , c") == ["a", "", "c"]  # Empty strings are preserved

    def test_pipe_pattern_converters(self):
        """Test pipe pattern string converters."""
        to_pipe, from_pipe = XLSXConverters.pipe_converters()

        assert to_pipe(["a", "b", "c"]) == "a | b | c"
        assert to_pipe([1, 2, 3]) == "1 | 2 | 3"
        assert to_pipe([]) == ""
        assert to_pipe(None) == ""

        assert from_pipe("a | b | c") == ["a", "b", "c"]
        assert from_pipe("a|b|c") == ["a|b|c"]  # No spaces, so treated as one item
        assert from_pipe("") == []
        assert from_pipe("   ") == []
        assert from_pipe("a |  | c") == [
            "a",
            "",
            "c",
        ]  # Empty strings are preserved (two spaces)

    def test_semicolon_pattern_converters(self):
        """Test semicolon pattern string converters."""
        to_semi, from_semi = XLSXConverters.semicolon_converters()

        assert to_semi(["a", "b", "c"]) == "a; b; c"
        assert to_semi([1, 2, 3]) == "1; 2; 3"
        assert to_semi([]) == ""
        assert to_semi(None) == ""

        assert from_semi("a; b; c") == ["a", "b", "c"]
        assert from_semi("a;b;c") == ["a;b;c"]  # No spaces, so treated as one item
        assert from_semi("") == []
        assert from_semi("   ") == []
        assert from_semi("a; ; c") == ["a", "", "c"]  # Empty strings are preserved

    def test_integer_pattern_converters(self):
        """Test integer list converters."""
        to_int, from_int = XLSXConverters.create_int_separated_converters(
            XLSXConverters.COMMA
        )

        assert to_int([1, 2, 3]) == "1, 2, 3"  # COMMA pattern includes space
        assert to_int([]) == ""
        assert to_int(None) == ""

        assert from_int("1, 2, 3") == [1, 2, 3]
        assert from_int("") == []
        assert from_int("   ") == []
        # Test that empty strings in integer lists raise ValueError
        with pytest.raises(ValueError, match=r"Cannot convert.*to list of integers"):
            from_int("1, , 3")

    def test_integer_conversion_error(self):
        """Test error handling in integer list conversion."""
        _, from_int = XLSXConverters.create_int_separated_converters(",")
        with pytest.raises(ValueError, match=r"Cannot convert.*to list of integers"):
            from_int("1, abc, 3")

    def test_comma_escaped_converters(self):
        """Test comma-separated converters with backslash escaping."""
        to_escaped, from_escaped = XLSXConverters.comma_escaped_converters()

        # Basic serialization
        assert to_escaped(["a", "b", "c"]) == "a, b, c"
        assert to_escaped([]) == ""
        assert to_escaped(None) == ""

        # Basic deserialization
        assert from_escaped("a, b, c") == ["a", "b", "c"]
        assert from_escaped("") == []
        assert from_escaped("   ") == []

        # Test with embedded commas - the key feature
        data = ["item1", "item2, with comma", "item3"]
        escaped = to_escaped(data)
        assert escaped == "item1, item2\\, with comma, item3"
        assert from_escaped(escaped) == data

        # Multiple commas in one item
        data_multi = ["a, b, c", "simple", "d, e"]
        escaped_multi = to_escaped(data_multi)
        assert from_escaped(escaped_multi) == data_multi

    def test_comma_quoted_converters(self):
        """Test comma-separated converters with quote escaping."""
        to_quoted, from_quoted = XLSXConverters.comma_quoted_converters()

        # Basic serialization
        assert to_quoted(["a", "b", "c"]) == "a, b, c"
        assert to_quoted([]) == ""
        assert to_quoted(None) == ""

        # Basic deserialization
        assert from_quoted("a, b, c") == ["a", "b", "c"]
        assert from_quoted("") == []
        assert from_quoted("   ") == []

        # Test with embedded commas - items get quoted
        data = ["item1", "item2, with comma", "item3"]
        quoted = to_quoted(data)
        assert quoted == 'item1, "item2, with comma", item3'
        assert from_quoted(quoted) == data

        # Multiple commas in one item
        data_multi = ["a, b, c", "simple", "d, e"]
        quoted_multi = to_quoted(data_multi)
        assert from_quoted(quoted_multi) == data_multi

    def test_pipe_escaped_converters(self):
        """Test pipe-separated converters with backslash escaping."""
        to_pipe, from_pipe = XLSXConverters.pipe_escaped_converters()

        # Basic serialization (pipe pattern has spaces: " | ")
        assert to_pipe(["a", "b", "c"]) == "a | b | c"
        assert to_pipe([]) == ""
        assert to_pipe(None) == ""

        # Basic deserialization
        assert from_pipe("a | b | c") == ["a", "b", "c"]
        assert from_pipe("") == []
        assert from_pipe("   ") == []

        # Test with embedded pipes
        data = ["item1", "item2 | with pipe", "item3"]
        escaped = to_pipe(data)
        assert escaped == "item1 | item2 \\| with pipe | item3"
        assert from_pipe(escaped) == data

    @pytest.mark.parametrize(
        ("converter_method", "test_data", "expected_serialized"),
        [
            ("newline_converters", ["a", "b", "c"], "a\nb\nc"),
            ("tab_converters", ["a", "b", "c"], "a\tb\tc"),
            ("arrow_converters", ["a", "b", "c"], "a -> b -> c"),
            ("double_colon_converters", ["a", "b", "c"], "a::b::c"),
        ],
    )
    def test_separator_converter_variants(
        self, converter_method, test_data, expected_serialized
    ):
        """Test various separator converter methods."""
        to_sep, from_sep = getattr(XLSXConverters, converter_method)()

        # Serialization
        assert to_sep(test_data) == expected_serialized
        assert to_sep([]) == ""
        assert to_sep(None) == ""

        # Deserialization roundtrip
        assert from_sep(expected_serialized) == test_data
        assert from_sep("") == []

    def test_integer_separator_converters(self):
        """Test comma_int_converters and pipe_int_converters."""
        # Comma integer converters
        to_comma_int, from_comma_int = XLSXConverters.comma_int_converters()
        assert to_comma_int([1, 2, 3]) == "1, 2, 3"
        assert from_comma_int("1, 2, 3") == [1, 2, 3]

        # Pipe integer converters
        to_pipe_int, from_pipe_int = XLSXConverters.pipe_int_converters()
        assert to_pipe_int([1, 2, 3]) == "1 | 2 | 3"
        assert from_pipe_int("1 | 2 | 3") == [1, 2, 3]

    def test_list_to_json_string(self):
        """Test list to JSON string conversion."""
        assert XLSXConverters.list_to_json_string([1, 2, 3]) == "[1, 2, 3]"
        assert XLSXConverters.list_to_json_string([]) == "[]"
        assert XLSXConverters.list_to_json_string(None) == ""

    def test_json_string_to_list(self):
        """Test JSON string to list conversion."""
        assert XLSXConverters.json_string_to_list("[1, 2, 3]") == [1, 2, 3]
        assert XLSXConverters.json_string_to_list("[]") == []
        assert XLSXConverters.json_string_to_list("") == []
        assert XLSXConverters.json_string_to_list("   ") == []

    def test_json_string_to_list_error(self):
        """Test error handling in JSON string to list conversion."""
        with pytest.raises(ValueError, match=r"Cannot parse.*as JSON"):
            XLSXConverters.json_string_to_list("not valid json")

    def test_bool_to_yes_no(self):
        """Test boolean to Yes/No string conversion."""
        assert XLSXConverters.bool_to_yes_no(True) == "Yes"
        assert XLSXConverters.bool_to_yes_no(False) == "No"

    @pytest.mark.parametrize(
        ("input_str", "expected"),
        [
            ("yes", True),
            ("Yes", True),
            ("YES", True),
            ("y", True),
            ("true", True),
            ("1", True),
            ("no", False),
            ("No", False),
            ("NO", False),
            ("n", False),
            ("false", False),
            ("0", False),
        ],
    )
    def test_yes_no_to_bool(self, input_str, expected):
        """Test Yes/No string to boolean conversion."""
        assert XLSXConverters.yes_no_to_bool(input_str) == expected

    def test_yes_no_to_bool_error(self):
        """Test error handling for invalid Yes/No input."""
        with pytest.raises(ValueError, match=r"Cannot convert.*to boolean"):
            XLSXConverters.yes_no_to_bool("maybe")

    def test_date_converters(self):
        """Test date to/from ISO string conversion."""

        test_date = date(2025, 6, 15)
        iso_str = XLSXConverters.date_to_iso_string(test_date)
        assert iso_str == "2025-06-15"
        assert XLSXConverters.iso_string_to_date(iso_str) == test_date

    def test_iso_string_to_date_error(self):
        """Test error handling for invalid date string."""
        with pytest.raises(ValueError, match=r"Cannot parse.*as ISO date"):
            XLSXConverters.iso_string_to_date("not-a-date")

    def test_datetime_converters(self):
        """Test datetime to/from ISO string conversion."""

        test_dt = datetime(2025, 6, 15, 10, 30, 0, tzinfo=timezone.utc)
        iso_str = XLSXConverters.datetime_to_iso_string(test_dt)
        assert iso_str == "2025-06-15T10:30:00"
        assert XLSXConverters.iso_string_to_datetime(iso_str) == test_dt

    def test_iso_string_to_datetime_error(self):
        """Test error handling for invalid datetime string."""
        with pytest.raises(ValueError, match=r"Cannot parse.*as ISO datetime"):
            XLSXConverters.iso_string_to_datetime("not-a-datetime")

    def test_dict_to_json_string(self):
        """Test dictionary to JSON string conversion."""
        assert (
            XLSXConverters.dict_to_json_string({"key": "value"}) == '{"key": "value"}'
        )
        assert XLSXConverters.dict_to_json_string({}) == "{}"
        assert XLSXConverters.dict_to_json_string(None) == ""

    def test_json_string_to_dict(self):
        """Test JSON string to dictionary conversion."""
        assert XLSXConverters.json_string_to_dict('{"key": "value"}') == {
            "key": "value"
        }
        assert XLSXConverters.json_string_to_dict("{}") == {}
        assert XLSXConverters.json_string_to_dict("") == {}
        assert XLSXConverters.json_string_to_dict("   ") == {}

    def test_json_string_to_dict_error(self):
        """Test error handling in JSON string to dictionary conversion."""
        with pytest.raises(ValueError, match=r"Cannot parse.*as JSON"):
            XLSXConverters.json_string_to_dict("invalid json")


# Tests for XLSXFieldAnalyzer
class TestXLSXFieldAnalyzer:
    """Tests for XLSXFieldAnalyzer functionality."""

    def test_analyze_simple_model(self):
        """Test field analysis for simple model."""
        fields = XLSXFieldAnalyzer.analyze_model(SimpleModel)
        assert len(fields) == 3

        name_field = next(f for f in fields.values() if f.name == "name")
        assert name_field.field_type is str
        assert not name_field.is_optional
        assert name_field.enum_values == []

    def test_analyze_optional_fields(self):
        """Test field analysis for optional fields."""

        class ModelWithOptional(BaseModel):
            required_field: str
            optional_field: str | None = None

        fields = XLSXFieldAnalyzer.analyze_model(ModelWithOptional)

        required_field = next(f for f in fields.values() if f.name == "required_field")
        assert not required_field.is_optional

        optional_field = next(f for f in fields.values() if f.name == "optional_field")
        assert optional_field.is_optional
        # The actual field_type for optional fields varies by Python version
        field_type_str = str(optional_field.field_type)
        assert (
            field_type_str.startswith(("typing.Optional", "typing.Union"))
            or "| None" in field_type_str
        )

    def test_analyze_enum_fields(self):
        """Test field analysis for enum fields."""

        class ModelWithEnum(BaseModel):
            status: Status
            priority: Priority | None = None

        fields = XLSXFieldAnalyzer.analyze_model(ModelWithEnum)

        status_field = next(f for f in fields.values() if f.name == "status")
        assert status_field.field_type == Status  # Field type is the actual enum type
        assert status_field.enum_values == ["active", "inactive", "pending"]

        priority_field = next(f for f in fields.values() if f.name == "priority")
        assert priority_field.is_optional
        # The actual field_type for optional enums varies by Python version
        field_type_str = str(priority_field.field_type)
        assert (
            field_type_str.startswith(("typing.Optional", "typing.Union"))
            or "| None" in field_type_str
        )
        assert priority_field.enum_values == ["low", "medium", "high"]

    def test_analyze_fields_with_converters(self):
        """Test field analysis with pattern converters."""

        class ModelWithConverters(BaseModel):
            tags: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
            ] = Field(default=[])

        fields = XLSXFieldAnalyzer.analyze_model(ModelWithConverters)

        tags_field = next(f for f in fields.values() if f.name == "tags")
        assert tags_field.xlsx_metadata is not None
        # The serializer/deserializer are None because they're not automatically created
        # The separator_pattern is stored but converters are created when needed
        assert tags_field.xlsx_metadata.separator_pattern == XLSXConverters.COMMA

    def test_analyze_fields_with_unit_and_meaning(self):
        """Test field analysis with unit and meaning."""
        fields = XLSXFieldAnalyzer.analyze_model(DemoModelWithMetadata)

        temp_field = next(f for f in fields.values() if f.name == "temp")
        assert temp_field.xlsx_metadata is not None
        assert temp_field.xlsx_metadata.unit == "°C"
        assert temp_field.xlsx_metadata.meaning == "Temperature measurement"

    def test_validate_unit_usage_valid(self):
        """Test valid unit usage validation."""
        # Should not raise for numeric types
        _validate_unit_usage("weight", float, "kg")
        _validate_unit_usage("count", int, "units")
        _validate_unit_usage("complex_num", complex, "units")

    def test_validate_unit_usage_optional_numeric(self):
        """Test unit validation with optional numeric types."""
        # Should not raise for Optional numeric types
        _validate_unit_usage("weight", float | None, "kg")
        _validate_unit_usage("count", int | None, "units")

    def test_validate_unit_usage_invalid(self):
        """Test invalid unit usage validation."""
        with pytest.raises(ValueError, match="Units are only valid for numeric fields"):
            _validate_unit_usage("name", str, "kg")

        with pytest.raises(ValueError, match="Units are only valid for numeric fields"):
            _validate_unit_usage("active", bool, "units")

    def test_validate_unit_usage_optional_non_numeric(self):
        """Test unit validation with optional non-numeric types."""
        with pytest.raises(ValueError, match="Units are only valid for numeric fields"):
            _validate_unit_usage("name", str | None, "kg")

    def test_is_optional_type(self):
        """Test optional type detection."""
        assert XLSXFieldAnalyzer.is_optional_type(str | None) is True
        assert XLSXFieldAnalyzer.is_optional_type(str | None) is True
        assert XLSXFieldAnalyzer.is_optional_type(str | None) is True
        assert XLSXFieldAnalyzer.is_optional_type(str) is False
        assert XLSXFieldAnalyzer.is_optional_type(str | int) is False

    def test_validate_unit_usage_python310_union(self):
        """Test unit validation with Python 3.10+ union syntax (int | None)."""
        # Should not raise for int | None with valid unit
        _validate_unit_usage("count", int | None, "units")
        _validate_unit_usage("weight", float | None, "kg")

        # Should raise for str | None with unit
        with pytest.raises(ValueError, match="Units are only valid for numeric fields"):
            _validate_unit_usage("name", str | None, "kg")

    def test_get_field_display_name(self):
        """Test display name extraction."""
        # With custom display name in metadata
        metadata = XLSXMetadata(display_name="Custom Name")
        assert (
            XLSXFieldAnalyzer.get_field_display_name("field_name", metadata)
            == "Custom Name"
        )

        # Without metadata - converts underscores to spaces and title-cases
        assert (
            XLSXFieldAnalyzer.get_field_display_name("field_name", None) == "Field Name"
        )
        assert (
            XLSXFieldAnalyzer.get_field_display_name("some_long_field", None)
            == "Some Long Field"
        )

    def test_get_field_description(self):
        """Test description extraction."""
        # From metadata
        metadata = XLSXMetadata(description="Meta description")
        assert (
            XLSXFieldAnalyzer.get_field_description(None, metadata)
            == "Meta description"
        )

        # Without metadata returns None
        assert XLSXFieldAnalyzer.get_field_description(None, None) is None

    def test_get_field_unit(self):
        """Test unit extraction."""
        metadata = XLSXMetadata(unit="kg")
        assert XLSXFieldAnalyzer.get_field_unit(metadata) == "kg"
        assert XLSXFieldAnalyzer.get_field_unit(None) is None

    def test_get_field_meaning(self):
        """Test meaning extraction."""
        metadata = XLSXMetadata(meaning="http://example.org/term")
        assert (
            XLSXFieldAnalyzer.get_field_meaning(metadata) == "http://example.org/term"
        )
        assert XLSXFieldAnalyzer.get_field_meaning(None) is None

    def test_has_custom_serializers(self):
        """Test custom serializer detection."""
        # Both serializers present
        metadata = XLSXMetadata(
            xlsx_serializer=lambda x: str(x),
            xlsx_deserializer=lambda x: x,
        )
        assert XLSXFieldAnalyzer.has_custom_serializers(metadata) is True

        # Only one serializer - returns False
        metadata_partial = XLSXMetadata(xlsx_serializer=lambda x: str(x))
        assert XLSXFieldAnalyzer.has_custom_serializers(metadata_partial) is False

        # No metadata
        assert XLSXFieldAnalyzer.has_custom_serializers(None) is False

    def test_get_custom_serializer(self):
        """Test custom serializer extraction."""
        serializer = lambda x: str(x)  # noqa: E731
        metadata = XLSXMetadata(xlsx_serializer=serializer)
        assert XLSXFieldAnalyzer.get_custom_serializer(metadata) is serializer
        assert XLSXFieldAnalyzer.get_custom_serializer(None) is None

    def test_get_custom_deserializer(self):
        """Test custom deserializer extraction."""
        deserializer = lambda x: int(x)  # noqa: E731
        metadata = XLSXMetadata(xlsx_deserializer=deserializer)
        assert XLSXFieldAnalyzer.get_custom_deserializer(metadata) is deserializer
        assert XLSXFieldAnalyzer.get_custom_deserializer(None) is None

    def test_get_separator_pattern(self):
        """Test separator pattern extraction."""
        metadata = XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
        assert XLSXFieldAnalyzer.get_separator_pattern(metadata) == XLSXConverters.COMMA
        assert XLSXFieldAnalyzer.get_separator_pattern(None) is None


# Tests for FieldAnalysis
class TestFieldAnalysis:
    """Tests for FieldAnalysis dataclass."""

    def test_field_analysis_creation(self):
        """Test FieldAnalysis creation with all parameters."""
        metadata = XLSXMetadata(
            unit="kg",
            meaning="http://example.org/test",
            display_name="Test Field",
            description="A test field",
        )
        field_analysis = FieldAnalysis(
            name="test_field",
            field_type=str,
            is_optional=True,
            enum_values=["a", "b", "c"],
            xlsx_metadata=metadata,
        )

        assert field_analysis.name == "test_field"
        assert field_analysis.field_type is str
        assert field_analysis.is_optional is True
        assert field_analysis.enum_values == ["a", "b", "c"]
        assert field_analysis.xlsx_metadata == metadata
        assert field_analysis.xlsx_metadata.unit == "kg"
        assert field_analysis.xlsx_metadata.meaning == "http://example.org/test"

    def test_field_analysis_defaults(self):
        """Test FieldAnalysis with default values."""
        field_analysis = FieldAnalysis(name="test", field_type=int)

        assert field_analysis.name == "test"
        assert field_analysis.field_type is int
        assert field_analysis.is_optional is False
        assert field_analysis.enum_values == []
        assert field_analysis.xlsx_metadata is None


# Tests for XLSXSerializationEngine
class TestXLSXSerializationEngine:
    """Tests for XLSXSerializationEngine functionality."""

    def setup_method(self):
        """Set up test fixtures."""
        self.engine = XLSXSerializationEngine()

    def test_serialize_value_none(self):
        """Test serialization of None values."""
        field_analysis = FieldAnalysis(name="test", field_type=str)
        result = self.engine.serialize_value(None, field_analysis)
        assert result == ""

    def test_serialize_value_basic_types(self):
        """Test serialization of basic types."""
        field_analysis = FieldAnalysis(name="test", field_type=str)

        # Test basic types
        assert self.engine.serialize_value("hello", field_analysis) == "hello"
        assert self.engine.serialize_value(42, field_analysis) == 42
        assert self.engine.serialize_value(3.14, field_analysis) == 3.14
        assert self.engine.serialize_value(True, field_analysis) is True

        # Test date
        test_date = date(2024, 1, 1)
        assert self.engine.serialize_value(test_date, field_analysis) == test_date

    def test_serialize_value_enum(self):
        """Test serialization of enum values."""
        field_analysis = FieldAnalysis(name="status", field_type=str)
        result = self.engine.serialize_value(Status.ACTIVE, field_analysis)
        assert result == "active"

    def test_serialize_value_with_custom_serializer(self):
        """Test serialization with custom serializer."""

        def serializer(x):
            return f"custom_{x}"

        metadata = XLSXMetadata(xlsx_serializer=serializer)
        field_analysis = FieldAnalysis(
            name="test", field_type=str, xlsx_metadata=metadata
        )
        result = self.engine.serialize_value("hello", field_analysis)
        assert result == "custom_hello"

    def test_serialize_value_complex_types(self):
        """Test serialization of complex types."""
        field_analysis = FieldAnalysis(name="test", field_type=str)

        # Test list
        result = self.engine.serialize_value(["a", "b", "c"], field_analysis)
        assert result == '["a", "b", "c"]'

        # Test dict
        result = self.engine.serialize_value({"key": "value"}, field_analysis)
        assert result == '{"key": "value"}'

    def test_deserialize_value_none_empty(self):
        """Test deserialization of None/empty values."""
        field_analysis = FieldAnalysis(name="test", field_type=str)

        assert self.engine.deserialize_value(None, field_analysis, SimpleModel) is None
        assert self.engine.deserialize_value("", field_analysis, SimpleModel) is None

    def test_deserialize_value_with_custom_deserializer(self):
        """Test deserialization with custom deserializer."""

        def deserializer(x):
            return int(x) * 2

        metadata = XLSXMetadata(xlsx_deserializer=deserializer)
        field_analysis = FieldAnalysis(
            name="test", field_type=int, xlsx_metadata=metadata
        )
        result = self.engine.deserialize_value("5", field_analysis, SimpleModel)
        assert result == 10

    def test_deserialize_value_enum(self):
        """Test deserialization of enum values."""
        field_analysis = FieldAnalysis(
            name="status", field_type=str, enum_values=["active", "inactive", "pending"]
        )

        class TestModel(BaseModel):
            status: Status

        result = self.engine.deserialize_value("active", field_analysis, TestModel)
        assert result == Status.ACTIVE

    def test_deserialize_value_enum_invalid(self):
        """Test deserialization of invalid enum values."""
        field_analysis = FieldAnalysis(
            name="status", field_type=str, enum_values=["active", "inactive", "pending"]
        )

        class TestModel(BaseModel):
            status: Status

        with pytest.raises(ValueError, match="Invalid enum value"):
            self.engine.deserialize_value("invalid", field_analysis, TestModel)

    def test_convert_basic_types(self):
        """Test basic type conversion."""
        # String
        assert self.engine._convert_basic_types("hello", str) == "hello"

        # Integer
        assert self.engine._convert_basic_types("42", int) == 42
        assert self.engine._convert_basic_types(42, int) == 42

        # Float
        assert self.engine._convert_basic_types("3.14", float) == 3.14
        assert self.engine._convert_basic_types(3.14, float) == 3.14

        # Boolean
        assert self.engine._convert_basic_types("true", bool) is True
        assert self.engine._convert_basic_types("false", bool) is False
        assert self.engine._convert_basic_types("1", bool) is True
        assert self.engine._convert_basic_types("0", bool) is False

    def test_convert_bool(self):
        """Test boolean conversion."""
        assert self.engine._convert_bool(True) is True
        assert self.engine._convert_bool(False) is False
        assert self.engine._convert_bool("true") is True
        assert self.engine._convert_bool("True") is True
        assert self.engine._convert_bool("1") is True
        assert self.engine._convert_bool("yes") is True
        assert self.engine._convert_bool("on") is True
        assert self.engine._convert_bool("false") is False
        assert self.engine._convert_bool("0") is False
        assert self.engine._convert_bool("no") is False

    def test_convert_datetime(self):
        """Test date/datetime conversion."""
        # Date conversion
        date_str = "2024-01-01"
        result = self.engine._convert_datetime(date_str, date)
        assert result == date(2024, 1, 1)
        assert isinstance(result, date)

        # Datetime conversion
        datetime_str = "2024-01-01T12:00:00"
        result = self.engine._convert_datetime(datetime_str, datetime)
        assert isinstance(result, datetime)

    def test_convert_json_types(self):
        """Test JSON type conversion."""
        # When field_type is None, it returns string
        result = self.engine._convert_basic_types('["a", "b", "c"]', None)
        assert result == '["a", "b", "c"]'

        # Test that JSON parsing works when string contains valid JSON
        # The method checks if raw_value is string and tries to parse it
        result = self.engine._convert_basic_types('{"key": "value"}', str)
        # Since field_type is str, it uses the string converter which just returns str(x)
        assert result == '{"key": "value"}'

        # Test unsupported type defaults to string
        result = self.engine._convert_basic_types("test", object)
        assert result == "test"


# Tests for Custom Exceptions
class TestCustomExceptions:
    """Tests for custom exception classes."""

    def test_xlsx_serialization_error(self):
        """Test XLSXSerializationError exception."""
        original_error = ValueError("Test error")
        error = XLSXSerializationError("test_field", "test_value", original_error)

        assert error.field_name == "test_field"
        assert error.value == "test_value"
        assert error.original_error == original_error
        assert "Error serializing field 'test_field' with value 'test_value'" in str(
            error
        )

    def test_xlsx_deserialization_error(self):
        """Test XLSXDeserializationError exception."""
        original_error = ValueError("Test error")
        error = XLSXDeserializationError("test_field", "test_value", original_error)

        assert error.field_name == "test_field"
        assert error.value == "test_value"
        assert error.original_error == original_error
        assert "Error deserializing field 'test_field' with value 'test_value'" in str(
            error
        )


# Tests for Configuration
class TestConfiguration:
    """Tests for configuration system."""

    def test_base_config_inheritance(self):
        """Test configuration inheritance."""
        config = XLSXConfig(
            title="Test",
            include_fields={"field1", "field2"},
            exclude_fields={"field3"},
        )

        assert config.title == "Test"
        assert config.should_include_field("field1") is True
        assert config.should_include_field("field2") is True
        assert config.should_include_field("field3") is False
        assert config.should_include_field("field4") is False  # not in include list

    def test_config_field_filtering(self):
        """Test field filtering logic."""
        config = XLSXConfig(
            include_fields={"field1", "field2"},
            exclude_fields={"field3"},
        )

        assert config.should_include_field("field1") is True
        assert config.should_include_field("field2") is True
        assert config.should_include_field("field3") is False
        assert config.should_include_field("field4") is False  # not in include list

    def test_config_field_filtering_edge_cases(self):
        """Test edge cases in field filtering."""
        # Test with both include and exclude (include takes precedence)
        config = XLSXConfig(
            include_fields={"field1", "field2", "field3"}, exclude_fields={"field2"}
        )

        assert config.should_include_field("field1") is True
        assert config.should_include_field("field2") is True  # Include takes precedence
        assert config.should_include_field("field3") is True
        assert config.should_include_field("field4") is False  # Not in include

        # Test with no restrictions
        config_open = XLSXConfig()
        assert config_open.should_include_field("any_field") is True

    def test_config_exclude_fields_only(self):
        """Test field filtering with exclude_fields only (no include_fields)."""
        config = XLSXConfig(exclude_fields={"field2", "field3"})

        assert config.should_include_field("field1") is True
        assert config.should_include_field("field2") is False  # Excluded
        assert config.should_include_field("field3") is False  # Excluded
        assert config.should_include_field("field4") is True  # Not excluded

    def test_get_ordered_fields_custom_order(self):
        """Test field ordering with custom field_order."""
        config = XLSXConfig(field_order=["c", "a", "b"])
        available = ["a", "b", "c", "d", "e"]

        ordered = config.get_ordered_fields(available)

        # Should start with specified order, then remaining fields
        assert ordered == ["c", "a", "b", "d", "e"]

    def test_get_ordered_fields_with_filtering(self):
        """Test field ordering combined with filtering."""
        config = XLSXConfig(
            field_order=["c", "a", "b"],
            exclude_fields={"d"},
        )
        available = ["a", "b", "c", "d", "e"]

        ordered = config.get_ordered_fields(available)

        # Should order and filter
        assert ordered == ["c", "a", "b", "e"]

    def test_get_ordered_fields_no_order(self):
        """Test field ordering without custom order."""
        config = XLSXConfig()
        available = ["a", "b", "c"]

        ordered = config.get_ordered_fields(available)
        assert ordered == ["a", "b", "c"]


# Tests for Advanced Field Analysis and Edge Cases
class TestAdvancedFieldAnalysis:
    """Tests for advanced field analysis scenarios."""

    def test_analyze_model_with_all_features(self):
        """Test analysis of model with all supported features."""

        class ComplexModel(BaseModel):
            # Basic types
            name: str
            age: int
            weight: Annotated[
                float, XLSXMetadata(unit="kg", meaning="http://example.org/weight")
            ]
            active: bool = True

            # Optional types
            email: str | None = None
            birth_date: date | None = None

            # Enum types
            status: Status
            priority: Priority | None = None

            # Complex types with converters
            tags: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
            ] = Field(default=[])

        fields = XLSXFieldAnalyzer.analyze_model(ComplexModel)
        assert len(fields) == 9

        # Check specific fields
        weight_field = next(f for f in fields.values() if f.name == "weight")
        assert weight_field.xlsx_metadata is not None
        assert weight_field.xlsx_metadata.unit == "kg"
        assert weight_field.xlsx_metadata.meaning == "http://example.org/weight"

        tags_field = next(f for f in fields.values() if f.name == "tags")
        assert tags_field.xlsx_metadata is not None
        assert tags_field.xlsx_metadata.separator_pattern == XLSXConverters.COMMA

    def test_field_analysis_edge_cases(self):
        """Test FieldAnalysis edge cases."""
        # Test with None field type
        field_analysis = FieldAnalysis(name="test", field_type=None)
        assert field_analysis.field_type is None
        assert not field_analysis.is_optional

    def test_unit_validation_complex_union_types(self):
        """Test unit validation with complex Union types."""
        # Union with no numeric types should raise error
        with pytest.raises(ValueError, match="Units are only valid for numeric fields"):
            _validate_unit_usage("field", str | bool, "kg")

        # Union with more than 2 types (non-Optional) should raise error
        with pytest.raises(ValueError, match="Units are only valid for numeric fields"):
            _validate_unit_usage("field", str | int, "kg")


# Tests for ORCID and ROR Field Types in XLSX
class TestORCIDRORXLSXIntegration:
    """Tests for ORCID and ROR field types in XLSX import/export."""

    def test_orcid_ror_field_analysis(self):
        """Test field analysis for ORCID and ROR field types."""

        fields = XLSXFieldAnalyzer.analyze_model(ResearcherModel)
        assert len(fields) == 3

        # Check ORCID field - ORCIDIdentifier resolves to HttpUrl during field analysis
        orcid_field = next(f for f in fields.values() if f.name == "orcid")
        assert orcid_field.field_type == HttpUrl
        assert not orcid_field.is_optional

        # Check ROR field - RORIdentifier also resolves to HttpUrl during field analysis
        ror_field = next(f for f in fields.values() if f.name == "home_organization")
        assert ror_field.is_optional
        # The actual field_type for optional fields varies by Python version
        field_type_str = str(ror_field.field_type)
        assert (
            field_type_str.startswith(("typing.Optional", "typing.Union"))
            or "| None" in field_type_str
        )

    def test_orcid_ror_xlsx_round_trip_keyvalue(self, temp_file, sample_researcher):
        """Test ORCID and ROR fields in key-value XLSX format round-trip."""

        # Export to XLSX
        export_to_xlsx(sample_researcher, temp_file, format_type="keyvalue")

        # Import from XLSX
        imported = import_from_xlsx(temp_file, ResearcherModel, format_type="keyvalue")

        # Verify data integrity
        assert imported.name == sample_researcher.name
        assert str(imported.orcid) == str(sample_researcher.orcid)
        assert str(imported.home_organization) == str(
            sample_researcher.home_organization
        )

        # Verify types are preserved
        assert isinstance(imported.orcid, type(sample_researcher.orcid))
        assert isinstance(
            imported.home_organization, type(sample_researcher.home_organization)
        )

    def test_orcid_ror_xlsx_round_trip_table(self, temp_file, sample_researchers):
        """Test ORCID and ROR fields in table XLSX format round-trip."""

        # Export to XLSX
        export_to_xlsx(sample_researchers, temp_file, format_type="table")

        # Import from XLSX
        imported = import_from_xlsx(temp_file, ResearcherModel, format_type="table")

        # Verify data integrity
        assert len(imported) == len(sample_researchers)

        # First researcher
        assert imported[0].name == sample_researchers[0].name
        assert str(imported[0].orcid) == str(sample_researchers[0].orcid)
        assert str(imported[0].home_organization) == str(
            sample_researchers[0].home_organization
        )

        # Second researcher
        assert imported[1].name == sample_researchers[1].name
        assert str(imported[1].orcid) == str(sample_researchers[1].orcid)
        assert imported[1].home_organization is None

    def test_orcid_ror_serialization_engine(self):
        """Test ORCID and ROR field serialization/deserialization."""

        class ResearcherModel(BaseModel):
            orcid: ORCIDIdentifier
            ror: RORIdentifier

        engine = XLSXSerializationEngine()

        # Test ORCID serialization
        orcid_field = FieldAnalysis(name="orcid", field_type=ORCIDIdentifier)

        orcid_value = HttpUrl("https://orcid.org/0000-0002-1825-0097")

        serialized_orcid = engine.serialize_value(orcid_value, orcid_field)
        assert serialized_orcid == "https://orcid.org/0000-0002-1825-0097"

        # Test ORCID deserialization
        deserialized_orcid = engine.deserialize_value(
            "https://orcid.org/0000-0002-1825-0097", orcid_field, ResearcherModel
        )
        assert str(deserialized_orcid) == "https://orcid.org/0000-0002-1825-0097"

        # Test ROR serialization
        ror_field = FieldAnalysis(name="ror", field_type=RORIdentifier)
        ror_value = HttpUrl("https://ror.org/02y72wh86")

        serialized_ror = engine.serialize_value(ror_value, ror_field)
        assert serialized_ror == "https://ror.org/02y72wh86"

        # Test ROR deserialization
        deserialized_ror = engine.deserialize_value(
            "https://ror.org/02y72wh86", ror_field, ResearcherModel
        )
        assert str(deserialized_ror) == "https://ror.org/02y72wh86"

    def test_orcid_ror_validation_errors_on_import(self, temp_file, sample_researcher):
        """Test that invalid ORCID/ROR values raise validation errors during import."""

        export_to_xlsx(sample_researcher, temp_file, format_type="keyvalue")

        # Now manually corrupt the xlsx file by reading and modifying
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Find the ORCID value and corrupt it
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == str(sample_researcher.orcid):
                    cell.value = "invalid-orcid-format"
                    break

        workbook.save(temp_file)

        # Try to import corrupted data
        with pytest.raises(ValueError, match="validation error"):
            import_from_xlsx(temp_file, ResearcherModel, format_type="keyvalue")


class TestRequirednessHelpers:
    """Tests for requiredness detection and formatting."""

    def test_get_requiredness_info_required_field(self):
        """Test detection of required field (no default, not optional)."""

        class ModelWithRequired(BaseModel):
            required_field: str

        field_info = ModelWithRequired.model_fields["required_field"]
        field_analysis = XLSXFieldAnalyzer.analyze_field(
            "required_field", field_info, ModelWithRequired
        )

        is_required, _ = XLSXFieldAnalyzer.get_requiredness_info(
            field_info, field_analysis
        )

        assert is_required is True

    def test_get_requiredness_info_optional_field(self):
        """Test detection of optional field (Union with None)."""

        class ModelWithOptional(BaseModel):
            optional_field: str | None = None

        field_info = ModelWithOptional.model_fields["optional_field"]
        field_analysis = XLSXFieldAnalyzer.analyze_field(
            "optional_field", field_info, ModelWithOptional
        )

        is_required, default_value = XLSXFieldAnalyzer.get_requiredness_info(
            field_info, field_analysis
        )

        assert is_required is False
        assert default_value is None

    def test_get_requiredness_info_field_with_default(self):
        """Test detection of field with non-optional type but has default."""

        class ModelWithDefault(BaseModel):
            field_with_default: str = "default_value"

        field_info = ModelWithDefault.model_fields["field_with_default"]
        field_analysis = XLSXFieldAnalyzer.analyze_field(
            "field_with_default", field_info, ModelWithDefault
        )

        is_required, default_value = XLSXFieldAnalyzer.get_requiredness_info(
            field_info, field_analysis
        )

        assert is_required is False
        assert default_value == "default_value"

    def test_format_requiredness_text_required(self):
        """Test formatting of required field."""

        result = XLSXFieldAnalyzer.format_requiredness_text(True, PydanticUndefined)
        assert result == "Yes"

    def test_format_requiredness_text_optional_no_default(self):
        """Test formatting of optional field with no/trivial default."""
        result = XLSXFieldAnalyzer.format_requiredness_text(False, None)
        assert result == "No"

        result = XLSXFieldAnalyzer.format_requiredness_text(False, "")
        assert result == "No"

        result = XLSXFieldAnalyzer.format_requiredness_text(False, [])
        assert result == "No"

    def test_format_requiredness_text_optional_with_falsy_default(self):
        """Test formatting of optional field with falsy but meaningful defaults."""
        result = XLSXFieldAnalyzer.format_requiredness_text(False, 0)
        assert result == "No (default: 0)"

        result = XLSXFieldAnalyzer.format_requiredness_text(False, 0.0)
        assert result == "No (default: 0.0)"

        result = XLSXFieldAnalyzer.format_requiredness_text(False, False)
        assert result == "No (default: False)"

    def test_format_requiredness_text_with_default(self):
        """Test formatting of field with non-trivial default."""
        result = XLSXFieldAnalyzer.format_requiredness_text(False, "custom_default")
        assert result == "No (default: custom_default)"

        result = XLSXFieldAnalyzer.format_requiredness_text(False, 42)
        assert result == "No (default: 42)"

    def test_format_requiredness_text_enum_default(self):
        """Test formatting of field with enum default."""
        result = XLSXFieldAnalyzer.format_requiredness_text(False, Status.ACTIVE)
        assert result == "No (default: active)"


class TestMetadataVisibilityToggle:
    """Tests for metadata visibility toggle configuration."""

    def test_metadata_visibility_enum_values(self):
        """Test MetadataVisibility enum has expected values."""

        assert MetadataVisibility.AUTO.value == "auto"
        assert MetadataVisibility.SHOW.value == "show"
        assert MetadataVisibility.HIDE.value == "hide"

    def test_metadata_toggle_config_defaults(self):
        """Test MetadataToggleConfig has correct defaults."""

        config = MetadataToggleConfig()
        assert config.unit == MetadataVisibility.AUTO
        assert config.requiredness == MetadataVisibility.AUTO
        assert config.description == MetadataVisibility.AUTO
        assert config.meaning == MetadataVisibility.AUTO

    def test_xlsx_config_with_metadata_visibility(self):
        """Test XLSXConfig accepts metadata_visibility parameter."""

        toggle_config = MetadataToggleConfig(
            unit=MetadataVisibility.HIDE, requiredness=MetadataVisibility.SHOW
        )

        config = XLSXConfig(metadata_visibility=toggle_config)

        assert config.metadata_visibility is not None
        assert config.metadata_visibility.unit == MetadataVisibility.HIDE
        assert config.metadata_visibility.requiredness == MetadataVisibility.SHOW


class TestAdjustTableLength:
    """Tests for adjust_table_length and adjust_all_tables_length functions."""

    def test_adjust_table_length_single(self, tmp_path, caplog):
        """Test adjust_table_length for a single table."""
        caplog.set_level(logging.DEBUG)
        test_wb = tmp_path / "table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Letter", "value"])  # table header
        data = [
            ["A", 1],
            ["B", 2],
        ]
        for row in data:
            ws.append(row)
        tab = Table(displayName="Table1", ref="A1:B3")
        ws.add_table(tab)
        ws["A5"] = "X"  # Content outside initial table range
        wb.save(test_wb)

        # Test with no pre-allocated rows
        wb = load_workbook(test_wb)
        ws = wb.active
        adjust_table_length(ws, "Table1", rows_pre_allocated=0)
        wb.save(test_wb)
        wb.close()

        wb = load_workbook(test_wb)
        name, table_range = wb.active.tables.items()[0]
        assert name == "Table1"
        assert table_range == "A1:B5"
        assert "from {A1:B3} to {A1:B5}" in caplog.text
        wb.close()

        caplog.clear()

        # Test with pre-allocated rows
        wb = load_workbook(test_wb)
        ws = wb.active
        adjust_table_length(ws, "Table1", rows_pre_allocated=3)
        wb.save(test_wb)
        wb.close()

        wb = load_workbook(test_wb)
        name, table_range = wb.active.tables.items()[0]
        assert name == "Table1"
        # After first adjustment, table is A1:B5 with content at row 5 ("X")
        # Last content row is now 5 (includes the "X" at A5)
        # So with rows_pre_allocated=3: 5 + 3 = 8
        assert table_range == "A1:B8"
        assert "from {A1:B5} to {A1:B8}" in caplog.text
        wb.close()

    def test_adjust_table_length_missing_table(self, tmp_path, caplog):
        """Test that adjust_table_length logs warning for missing table."""
        caplog.set_level(logging.WARNING)
        test_wb = tmp_path / "table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A", 1])
        tab = Table(displayName="Table1", ref="A1:B1")
        ws.add_table(tab)
        wb.save(test_wb)
        wb.close()

        wb = load_workbook(test_wb)
        ws = wb.active
        adjust_table_length(ws, "NonExistentTable", rows_pre_allocated=5)
        wb.close()

        assert 'Table "NonExistentTable" not found' in caplog.text

    @pytest.mark.parametrize(
        ("rows_pre_allocated", "expected_range"),
        [
            (5, "A1:B8"),  # int: 3 + 5 = 8
            ({"Concepts": 10}, "A1:B13"),  # dict: 3 + 10 = 13
        ],
    )
    def test_adjust_all_tables_length(
        self, tmp_path, caplog, rows_pre_allocated, expected_range
    ):
        """Test adjust_all_tables_length with int or dict rows_pre_allocated."""
        caplog.set_level(logging.DEBUG)
        test_wb = tmp_path / "table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Concepts"
        ws.append(["Letter", "value"])  # table header
        for row in [["A", 1], ["B", 2]]:
            ws.append(row)
        tab = Table(displayName="Table1", ref="A1:B3")
        ws.add_table(tab)
        ws["A5"] = "X"
        wb.save(test_wb)

        adjust_all_tables_length(test_wb, rows_pre_allocated=rows_pre_allocated)

        wb = load_workbook(test_wb)
        name, table_range = wb.active.tables.items()[0]
        assert name == "Table1"
        assert table_range == expected_range
        wb.close()

    def test_adjust_all_tables_length_no_change_on_repeat(self, tmp_path, caplog):
        """Test that running adjustment again doesn't change if content is same."""
        caplog.set_level(logging.DEBUG)
        test_wb = tmp_path / "table.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws.append(["Letter", "value"])
        ws.append(["A", 1])
        tab = Table(displayName="Table1", ref="A1:B2")
        ws.add_table(tab)
        wb.save(test_wb)

        adjust_all_tables_length(test_wb, rows_pre_allocated={"Sheet": 5})
        wb = load_workbook(test_wb)
        assert wb.active.tables.items()[0][1] == "A1:B7"
        wb.close()

        caplog.clear()

        # Second adjustment - should not change
        adjust_all_tables_length(test_wb, rows_pre_allocated={"Sheet": 5})
        wb = load_workbook(test_wb)
        assert wb.active.tables.items()[0][1] == "A1:B7"
        assert not caplog.text  # No log because no change
        wb.close()

    @pytest.mark.parametrize(
        ("active_sheet", "expected_active"),
        [
            ("Concepts", "Concepts"),  # Existing sheet becomes active
            ("NonExistent", "Sheet1"),  # Missing sheet ignored, stays on Sheet1
            (None, "Sheet1"),  # None leaves default active
        ],
    )
    def test_adjust_all_tables_length_active_sheet(
        self, tmp_path, active_sheet, expected_active
    ):
        """Test active_sheet parameter sets the active worksheet."""
        test_wb = tmp_path / "table.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Concepts")
        ws2.append(["A", 1])
        tab = Table(displayName="Table1", ref="A1:B1")
        ws2.add_table(tab)
        wb.save(test_wb)
        wb.close()

        adjust_all_tables_length(test_wb, active_sheet=active_sheet)

        wb = load_workbook(test_wb)
        assert wb.active.title == expected_active
        wb.close()


if __name__ == "__main__":
    pytest.main([__file__])
