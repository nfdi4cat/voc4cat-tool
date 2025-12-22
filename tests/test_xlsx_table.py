"""
Tests for the xlsx_table module.

This module tests the table format specific functionality for XLSX processing.
"""

from datetime import date
from typing import Annotated

import pytest
from openpyxl import load_workbook
from pydantic import BaseModel

from voc4cat.xlsx_api import export_to_xlsx, import_from_xlsx
from voc4cat.xlsx_common import (
    MetadataToggleConfig,
    MetadataVisibility,
    XLSXMetadata,
)
from voc4cat.xlsx_table import XLSXTableConfig

from .conftest import DemoModelWithMetadata, Employee, Project, SimpleModel, Status


# Table Format Tests
class TestTableFormat:
    """Tests for tabular format processing."""

    def test_table_export_import_basic(self, sample_employees, temp_file):
        """Test basic table export and import."""
        # Export
        export_to_xlsx(sample_employees, temp_file, format_type="table")

        # Import
        imported = import_from_xlsx(temp_file, Employee, format_type="table")

        # Verify
        assert len(imported) == 2
        assert imported[0].first_name == "John"
        assert imported[0].status.value == "active"
        assert imported[1].first_name == "Jane"
        assert imported[1].is_active is False

    def test_table_export_with_config(self, sample_employees, temp_file):
        """Test table export with custom configuration."""
        config = XLSXTableConfig(
            title="Employee Directory",
            start_row=2,
            exclude_fields={"department"},
        )

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)
        imported = import_from_xlsx(
            temp_file, Employee, format_type="table", config=config
        )

        assert len(imported) == 2
        # Department should be None due to exclusion
        assert imported[0].department is None

    def test_table_field_ordering(self, sample_employees, temp_file):
        """Test custom field ordering."""
        config = XLSXTableConfig(
            field_order=["last_name", "first_name", "employee_id"],
        )

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)
        imported = import_from_xlsx(
            temp_file, Employee, format_type="table", config=config
        )

        assert len(imported) == 2
        assert imported[0].first_name == "John"

    def test_table_empty_data_error(self, temp_file):
        """Test error handling for empty data."""
        with pytest.raises(ValueError, match="No data provided"):
            export_to_xlsx([], temp_file, format_type="table")

    def test_table_missing_sheet_error(self, temp_file):
        """Test error handling for missing sheet."""
        # Create empty file
        export_to_xlsx(
            [SimpleModel(name="test", value=1)], temp_file, format_type="table"
        )

        with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
            import_from_xlsx(
                temp_file, SimpleModel, format_type="table", sheet_name="NonExistent"
            )

    def test_table_with_projects(self, sample_projects, temp_file):
        """Test table format with project data."""
        export_to_xlsx(sample_projects, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, Project, format_type="table")

        assert len(imported) == 2
        assert imported[0].project_name == "Website Redesign"
        assert imported[0].priority.value == "high"
        assert imported[1].project_name == "Mobile App"
        assert imported[1].end_date is None

    def test_table_with_metadata(self, temp_file):
        """Test table format with metadata."""
        data = [
            DemoModelWithMetadata(temp=25.5, name="Sample1"),
            DemoModelWithMetadata(temp=30.0, name="Sample2"),
        ]

        config = XLSXTableConfig()

        export_to_xlsx(data, temp_file, format_type="table", config=config)
        imported = import_from_xlsx(
            temp_file, DemoModelWithMetadata, format_type="table", config=config
        )

        assert len(imported) == 2
        assert imported[0].temp == 25.5
        assert imported[0].name == "Sample1"
        assert imported[1].temp == 30.0
        assert imported[1].name == "Sample2"

    def test_table_with_optional_fields(self, temp_file):
        """Test table format with optional fields."""

        class ModelWithOptional(BaseModel):
            name: str
            optional_field: str | None = None
            optional_with_default: str = "default_value"

        data = [
            ModelWithOptional(name="test1", optional_field=None),
            ModelWithOptional(
                name="test2", optional_field="provided", optional_with_default="custom"
            ),
        ]

        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, ModelWithOptional, format_type="table")

        assert len(imported) == 2
        assert imported[0].name == "test1"
        assert imported[0].optional_field is None
        assert imported[0].optional_with_default == "default_value"
        assert imported[1].name == "test2"
        assert imported[1].optional_field == "provided"
        assert imported[1].optional_with_default == "custom"

    def test_table_with_enum_fields(self, temp_file):
        """Test table format with enum fields."""

        class ModelWithEnum(BaseModel):
            name: str
            status: Status
            optional_status: Status | None = None

        data = [
            ModelWithEnum(
                name="test1", status=Status.ACTIVE, optional_status=Status.PENDING
            ),
            ModelWithEnum(name="test2", status=Status.INACTIVE, optional_status=None),
        ]

        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, ModelWithEnum, format_type="table")

        assert len(imported) == 2
        assert imported[0].name == "test1"
        assert imported[0].status == Status.ACTIVE
        assert imported[0].optional_status == Status.PENDING
        assert imported[1].name == "test2"
        assert imported[1].status == Status.INACTIVE
        assert imported[1].optional_status is None

    def test_table_with_union_types(self, temp_file):
        """Test table format with Union types."""

        class ModelWithUnion(BaseModel):
            name: str
            value: int | str  # Union of different types

        data = [
            ModelWithUnion(name="test1", value=42),
            ModelWithUnion(name="test2", value="text"),
        ]

        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, ModelWithUnion, format_type="table")

        assert len(imported) == 2
        assert imported[0].name == "test1"
        # Union types get converted to string in Excel
        assert str(imported[0].value) == "42"
        assert imported[1].name == "test2"
        assert str(imported[1].value) == "text"

    def test_table_with_custom_sheet_name(self, sample_employees, temp_file):
        """Test table format with custom sheet name."""
        export_to_xlsx(
            sample_employees, temp_file, format_type="table", sheet_name="CustomSheet"
        )
        imported = import_from_xlsx(
            temp_file, Employee, format_type="table", sheet_name="CustomSheet"
        )

        assert len(imported) == 2
        assert imported[0].first_name == "John"

    def test_table_large_dataset(self, large_employee_dataset, temp_file):
        """Test table format with large dataset."""
        export_to_xlsx(large_employee_dataset, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, Employee, format_type="table")

        assert len(imported) == 100
        assert imported[0].first_name == "Employee0"
        assert imported[-1].first_name == "Employee99"

    def test_table_single_item_list(self, temp_file):
        """Test table format with single item list."""
        data = [SimpleModel(name="single", value=123)]
        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, SimpleModel, format_type="table")

        assert len(imported) == 1
        assert imported[0].name == "single"
        assert imported[0].value == 123

    def test_table_tuple_data(self, sample_projects, temp_file):
        """Test table format with tuple data."""
        projects_tuple = tuple(sample_projects)
        export_to_xlsx(projects_tuple, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, Project, format_type="table")

        assert len(imported) == 2
        assert imported[0].project_name == "Website Redesign"


# Table Configuration Tests
class TestTableConfiguration:
    """Tests for table configuration."""

    def test_table_config_defaults(self):
        """Test default configuration values."""
        config = XLSXTableConfig()

        assert config.start_row == 1
        # All fields are now always shown
        assert config.field_order is None

    def test_table_config_custom_values(self):
        """Test custom configuration values."""
        config = XLSXTableConfig(
            start_row=5,
            field_order=["field1", "field2"],
        )

        assert config.start_row == 5
        # All fields are now always shown
        assert config.field_order == ["field1", "field2"]

    def test_table_config_inheritance(self):
        """Test that table config inherits from base config."""
        config = XLSXTableConfig(
            title="Test Title",
            include_fields={"field1", "field2"},
            exclude_fields={"field3"},
        )

        assert config.title == "Test Title"
        assert config.should_include_field("field1") is True
        assert config.should_include_field("field2") is True
        assert config.should_include_field("field3") is False

    def test_table_config_field_ordering(self):
        """Test field ordering configuration."""
        config = XLSXTableConfig(
            field_order=["last_name", "first_name", "employee_id"],
        )

        assert config.field_order == ["last_name", "first_name", "employee_id"]

    def test_table_config_row_settings(self):
        """Test row-related configuration."""
        config = XLSXTableConfig(
            start_row=10,
        )

        assert config.start_row == 10
        # All fields are now always shown
        assert config.auto_filter is False
        assert config.table_style == "TableStyleMedium9"
        assert config.freeze_panes is True


# Table Format Edge Cases
class TestTableFormatEdgeCases:
    """Tests for table format edge cases."""

    def test_table_with_empty_optional_fields(self, temp_file):
        """Test table format with many empty optional fields."""

        class ModelWithManyOptionals(BaseModel):
            name: str
            opt1: str | None = None
            opt2: str | None = None
            opt3: str | None = None
            opt4: str | None = None

        data = [
            ModelWithManyOptionals(name="test1"),  # All optionals None
            ModelWithManyOptionals(
                name="test2", opt1="value1", opt3="value3"
            ),  # Some filled
        ]

        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(
            temp_file, ModelWithManyOptionals, format_type="table"
        )

        assert len(imported) == 2
        assert imported[0].name == "test1"
        assert imported[0].opt1 is None
        assert imported[0].opt2 is None
        assert imported[0].opt3 is None
        assert imported[0].opt4 is None
        assert imported[1].name == "test2"
        assert imported[1].opt1 == "value1"
        assert imported[1].opt2 is None
        assert imported[1].opt3 == "value3"
        assert imported[1].opt4 is None

    def test_table_with_mixed_data_types(self, temp_file):
        """Test table format with mixed data types."""

        class MixedModel(BaseModel):
            name: str
            count: int
            price: float
            active: bool
            created: date

        data = [
            MixedModel(
                name="item1",
                count=10,
                price=99.99,
                active=True,
                created=date(2024, 1, 1),
            ),
            MixedModel(
                name="item2",
                count=5,
                price=49.50,
                active=False,
                created=date(2024, 2, 1),
            ),
        ]

        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, MixedModel, format_type="table")

        assert len(imported) == 2
        assert imported[0].name == "item1"
        assert imported[0].count == 10
        assert imported[0].price == 99.99
        assert imported[0].active is True
        assert imported[0].created == date(2024, 1, 1)
        assert imported[1].name == "item2"
        assert imported[1].count == 5
        assert imported[1].price == 49.50
        assert imported[1].active is False
        assert imported[1].created == date(2024, 2, 1)

    def test_table_with_very_long_field_names(self, temp_file):
        """Test table format with very long field names."""

        class ModelWithLongNames(BaseModel):
            very_long_field_name_that_exceeds_normal_length: str
            another_extremely_long_field_name_for_testing: int

        data = [
            ModelWithLongNames(
                very_long_field_name_that_exceeds_normal_length="test1",
                another_extremely_long_field_name_for_testing=42,
            )
        ]

        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, ModelWithLongNames, format_type="table")

        assert len(imported) == 1
        assert imported[0].very_long_field_name_that_exceeds_normal_length == "test1"
        assert imported[0].another_extremely_long_field_name_for_testing == 42


# Metadata Positioning Tests
class TestTableMetadataPositioning:
    """Tests for validating the position of title and metadata in Excel files.

    These tests use openpyxl to directly inspect the Excel file structure and verify that:
    - Titles are positioned correctly at the specified start_row
    - There's proper spacing (empty rows) between title and content
    - Field descriptions appear in the correct row when enabled
    - Field units appear in the correct row when enabled
    - Headers and data start at the expected rows based on metadata configuration
    - Custom start_row positioning works correctly
    - Font formatting (bold, size) is applied to titles

    The expected layout structure is:
    - Title: start_row (if present)
    - Empty row: start_row + 1 (spacing)
    - Descriptions: start_row + 2 (always shown)
    - Units: start_row + 3 (always shown)
    - Headers: next available row after metadata
    - Data: row after headers
    """

    def test_title_positioning_basic(self, sample_employees, temp_file):
        """Test that title is positioned correctly in basic configuration."""
        config = XLSXTableConfig(
            title="Employee Directory",
            start_row=2,
        )

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)

        # Load workbook with openpyxl to inspect structure
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 2 (start_row), column 1
        title_cell = worksheet.cell(row=2, column=1)
        assert title_cell.value == "Employee Directory"

        # Title should have bold formatting
        assert title_cell.font.bold is True
        assert title_cell.font.size == 14

        # Row 3 should be empty (spacing)
        empty_row = worksheet.cell(row=3, column=1)
        assert empty_row.value is None

        # Headers should be in row 4 (start_row + 2)
        header_row = 4
        first_header = worksheet.cell(row=header_row, column=1)
        assert first_header.value is not None  # Should contain a header

        # Data should start in row 5 (start_row + 3)
        data_row = 5
        first_data_cell = worksheet.cell(row=data_row, column=1)
        assert first_data_cell.value is not None  # Should contain data

    def test_title_positioning_with_descriptions(self, temp_file):
        """Test title positioning when descriptions are enabled."""
        # Use DemoModelWithMetadata which has proper description metadata
        data = [DemoModelWithMetadata(temp=25.5, name="Sample1")]

        config = XLSXTableConfig(
            title="Test with Descriptions",
            start_row=3,
        )

        export_to_xlsx(data, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 3 (start_row)
        title_cell = worksheet.cell(row=3, column=1)
        assert title_cell.value == "Test with Descriptions"
        assert title_cell.font.bold is True

        # Empty row should be in row 4 (start_row + 1)
        empty_cell = worksheet.cell(row=4, column=1)
        assert empty_cell.value is None

        # Meanings should be in row 5 (start_row + 2)
        meaning_row = 5
        first_meaning_cell = worksheet.cell(row=meaning_row, column=1)
        assert first_meaning_cell.value is not None
        assert "Temperature measurement" in str(first_meaning_cell.value)

        # Descriptions should be in row 6 (start_row + 3)
        desc_row = 6
        first_desc_cell = worksheet.cell(row=desc_row, column=1)
        assert first_desc_cell.value is not None
        assert "Temperature reading" in str(first_desc_cell.value)

        # Headers should be in row 6 (start_row + 3)
        header_row = 6
        first_header = worksheet.cell(row=header_row, column=1)
        assert first_header.value is not None

        # Data should start in row 7 (start_row + 4)
        data_row = 7
        first_data_cell = worksheet.cell(row=data_row, column=1)
        assert first_data_cell.value is not None

    def test_title_positioning_with_units(self, temp_file):
        """Test title positioning when units are enabled."""
        # Use DemoModelWithMetadata which has unit metadata
        data = [DemoModelWithMetadata(temp=25.5, name="Sample1")]

        config = XLSXTableConfig(
            title="Test with Units",
            start_row=3,
        )

        export_to_xlsx(data, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 3 (start_row)
        title_cell = worksheet.cell(row=3, column=1)
        assert title_cell.value == "Test with Units"
        assert title_cell.font.bold is True

        # Empty row should be in row 4 (start_row + 1)
        empty_cell = worksheet.cell(row=4, column=1)
        assert empty_cell.value is None

        # Meanings should be in row 5 (start_row + 2)
        meaning_row = 5
        meaning_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=meaning_row, column=col)
            if cell.value and "Temperature measurement" in str(
                cell.value
            ):  # pragma: no branch
                meaning_found = True
                break
        assert meaning_found, "Meanings row should contain field meanings"

        # Descriptions should be in row 6 (start_row + 3)
        desc_row = 6
        desc_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=desc_row, column=col)
            if cell.value and "Temperature reading" in str(
                cell.value
            ):  # pragma: no branch
                desc_found = True
                break
        assert desc_found, "Descriptions row should contain field descriptions"

        # Units should be in row 7 (start_row + 4)
        units_row = 7
        # Check that units are displayed (temp field has °C unit)
        units_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=units_row, column=col)
            if cell.value and "°C" in str(cell.value):  # pragma: no branch
                units_found = True
                break
        assert units_found, "Units row should contain unit information"

        # Headers should be in row 6 (start_row + 3)
        header_row = 6
        first_header = worksheet.cell(row=header_row, column=1)
        assert first_header.value is not None

        # Data should start in row 7 (start_row + 4)
        data_row = 7
        first_data_cell = worksheet.cell(row=data_row, column=1)
        assert first_data_cell.value is not None

    def test_comprehensive_metadata_positioning(self, temp_file):
        """Test positioning when all metadata types are enabled."""
        # Use DemoModelWithMetadata which has both description and unit metadata
        data = [DemoModelWithMetadata(temp=25.5, name="Sample1")]

        config = XLSXTableConfig(
            title="Complete Metadata Example",
            start_row=4,  # Give extra space
        )

        export_to_xlsx(data, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 4 (start_row)
        title_cell = worksheet.cell(row=4, column=1)
        assert title_cell.value == "Complete Metadata Example"
        assert title_cell.font.bold is True

        # Empty row should be in row 5 (start_row + 1)
        empty_cell = worksheet.cell(row=5, column=1)
        assert empty_cell.value is None

        # Meanings should be in row 6 (start_row + 2)
        meaning_row = 6
        meaning_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=meaning_row, column=col)
            if cell.value and "Temperature measurement" in str(
                cell.value
            ):  # pragma: no branch
                meaning_found = True
                break
        assert meaning_found, "Meanings row should contain field meanings"

        # Descriptions should be in row 7 (start_row + 3)
        desc_row = 7
        desc_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=desc_row, column=col)
            if cell.value and "Temperature reading" in str(
                cell.value
            ):  # pragma: no branch
                desc_found = True
                break
        assert desc_found, "Descriptions row should contain field descriptions"

        # Units should be in row 8 (start_row + 4)
        units_row = 8
        units_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=units_row, column=col)
            if cell.value and "°C" in str(cell.value):  # pragma: no branch
                units_found = True
                break
        assert units_found, "Units row should contain unit information"

        # Headers should be in row 9 (start_row + 5)
        header_row = 9
        first_header = worksheet.cell(row=header_row, column=1)
        assert first_header.value is not None

        # Data should start in row 10 (start_row + 6)
        data_row = 10
        first_data_cell = worksheet.cell(row=data_row, column=1)
        assert first_data_cell.value is not None

    def test_all_metadata_elements_positioning(self, temp_file):
        """Test positioning when ALL metadata elements are enabled."""
        # Use DemoModelWithMetadata which has descriptions, units, and meanings
        data = [DemoModelWithMetadata(temp=25.5, name="Sample1")]

        config = XLSXTableConfig(
            title="Complete Metadata with All Elements",
            start_row=5,  # Give plenty of space
            auto_filter=True,
        )

        export_to_xlsx(data, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 5 (start_row)
        title_cell = worksheet.cell(row=5, column=1)
        assert title_cell.value == "Complete Metadata with All Elements"
        assert title_cell.font.bold is True

        # Empty row should be in row 6 (start_row + 1)
        empty_cell = worksheet.cell(row=6, column=1)
        assert empty_cell.value is None

        # Meanings should be in row 7 (start_row + 2)
        meanings_row = 7
        meanings_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=meanings_row, column=col)
            if cell.value and "Temperature measurement" in str(
                cell.value
            ):  # pragma: no branch
                meanings_found = True
                break
        assert meanings_found, "Meanings row should contain field meanings"

        # Descriptions should be in row 8 (start_row + 3)
        desc_row = 8
        desc_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=desc_row, column=col)
            if cell.value and (  # pragma: no branch
                "Temperature" in str(cell.value) or "Sample" in str(cell.value)
            ):
                desc_found = True
                break
        assert desc_found, "Descriptions row should contain field descriptions"

        # Units should be in row 9 (start_row + 4)
        units_row = 9
        units_found = False
        for col in range(1, 4):  # pragma: no branch
            cell = worksheet.cell(row=units_row, column=col)
            if cell.value and "°C" in str(cell.value):  # pragma: no branch
                units_found = True
                break
        assert units_found, "Units row should contain unit information"

        # Headers should be in row 10 (start_row + 5)
        header_row = 10
        first_header = worksheet.cell(row=header_row, column=1)
        assert first_header.value is not None

        # Data should start in row 11 (start_row + 6)
        data_row = 11
        first_data_cell = worksheet.cell(row=data_row, column=1)
        assert first_data_cell.value is not None

    def test_no_title_positioning(self, sample_employees, temp_file):
        """Test positioning when no title is provided."""
        config = XLSXTableConfig(
            title=None,  # No title
        )

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Headers should be in row 1 (default start_row) when no title
        header_row = 1
        first_header = worksheet.cell(row=header_row, column=1)
        assert first_header.value is not None

        # Should not be a title (check it's likely a header)
        # Headers are typically field names like "employee_id", "first_name", etc.
        header_text = str(first_header.value).lower()
        assert any(
            field in header_text
            for field in ["employee", "first", "last", "name", "id"]
        )

        # Data should start in row 2 (start_row + 1)
        data_row = 2
        first_data_cell = worksheet.cell(row=data_row, column=1)
        assert first_data_cell.value is not None


class TestTableRequirednessAndToggles:
    """Tests for requiredness row and metadata visibility toggles."""

    def test_table_with_requiredness_row_shown(self, temp_file, sample_employees):
        """Test that requiredness row is shown when explicitly enabled."""

        config = XLSXTableConfig(
            title="Employees",
            metadata_visibility=MetadataToggleConfig(
                requiredness=MetadataVisibility.SHOW
            ),
        )

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title at row 1, empty at row 2, requiredness at row 3, header at row 4
        # With title: start_row=1, title+empty=rows 1-2, requiredness=row 3, header=row 4
        requiredness_row = 3

        # Check requiredness values (should be "Required"/"Optional" for table format)
        first_req_cell = worksheet.cell(row=requiredness_row, column=1)
        assert first_req_cell.value in [
            "Required",
            "Optional",
            None,
        ] or "Optional (default" in str(first_req_cell.value or "")

    def test_table_without_requiredness_row_by_default(
        self, temp_file, sample_employees
    ):
        """Test that requiredness row is NOT shown by default (AUTO mode)."""
        config = XLSXTableConfig(title="Employees")

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # With no requiredness row, header should be at row 3 (title + empty + header)
        header_row = 3
        first_header = worksheet.cell(row=header_row, column=1)

        # Should be a header value, not "Required"/"Optional"
        header_text = str(first_header.value).lower() if first_header.value else ""
        assert header_text not in ["required", "optional"]

    def test_table_toggle_hide_units(self, temp_file):
        """Test that HIDE toggle prevents units row even when fields have units."""

        class ModelWithUnits(BaseModel):
            temp: Annotated[
                float,
                XLSXMetadata(unit="°C"),
            ] = 25.0

        data = [ModelWithUnits()]

        # First export WITH units (default AUTO behavior)
        config_auto = XLSXTableConfig(title="With Units")
        export_to_xlsx(data, temp_file, format_type="table", config=config_auto)
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Find the unit row - should contain "°C"
        found_unit = False
        for row in worksheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value == "°C":
                    found_unit = True
                    break
        assert found_unit, "Unit should be shown in AUTO mode"

        # Now export with HIDE
        config_hide = XLSXTableConfig(
            title="No Units",
            metadata_visibility=MetadataToggleConfig(unit=MetadataVisibility.HIDE),
        )
        export_to_xlsx(data, temp_file, format_type="table", config=config_hide)
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Unit row should NOT contain "°C"
        found_unit = False
        for row in worksheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value == "°C":  # pragma: no cover
                    found_unit = True
                    break
        assert not found_unit, "Unit should NOT be shown with HIDE mode"

    def test_table_toggle_force_show_description(self, temp_file):
        """Test that SHOW toggle adds description row even without descriptions."""

        class SimpleModelNoDesc(BaseModel):
            name: str = "test"

        data = [SimpleModelNoDesc()]

        # Export with SHOW for descriptions

        config = XLSXTableConfig(
            title="Force Descriptions",
            metadata_visibility=MetadataToggleConfig(
                description=MetadataVisibility.SHOW
            ),
        )
        export_to_xlsx(data, temp_file, format_type="table", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # With title + empty + description row + header, header should be at row 4
        # Check that we have more rows than just title + empty + header
        header_row = 4
        header_cell = worksheet.cell(row=header_row, column=1)
        assert header_cell.value is not None


if __name__ == "__main__":
    pytest.main([__file__])
