"""Tests for the gen_template module.

These tests verify that the v1.0 template generator creates Excel templates
with the correct structure.
"""

import logging

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from voc4cat.checks import Voc4catError
from voc4cat.cli import main_cli
from voc4cat.config import Checks, Vocab
from voc4cat.gen_template import generate_template_v1
from voc4cat.models_v1 import (
    DEFAULT_PREFIXES,
    TEMPLATE_VERSION,
    CollectionV1,
    ConceptSchemeV1,
    ConceptV1,
    MappingV1,
    OrderedChoice,
    PrefixV1,
)


class TestTemplateGeneration:
    """Tests for template generation functionality."""

    def test_generate_template_creates_all_sheets(self, tmp_path):
        """Verify all 6 sheets are created."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)

        expected_sheets = {
            "Concept Scheme",
            "Concepts",
            "Collections",
            "Mappings",
            "ID Ranges",
            "Prefixes",
        }
        actual_sheets = set(wb.sheetnames)

        # Should have all expected sheets
        assert expected_sheets.issubset(actual_sheets), (
            f"Missing sheets: {expected_sheets - actual_sheets}"
        )

    def test_generate_template_sheet_order(self, tmp_path):
        """Verify sheets are in the correct order."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)

        expected_order = [
            "Concept Scheme",
            "Concepts",
            "Collections",
            "Mappings",
            "ID Ranges",
            "Prefixes",
        ]
        actual_order = wb.sheetnames[:6]

        assert actual_order == expected_order, (
            f"Sheet order mismatch: expected {expected_order}, got {actual_order}"
        )

    def test_concepts_sheet_has_table(self, tmp_path):
        """Verify Concepts sheet has proper table with correct columns."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Concepts"]

        # Check table exists
        assert ws.tables, "Concepts sheet should have a table"
        table_names = list(ws.tables.keys())
        assert len(table_names) == 1, "Should have exactly one table"

        # Check column headers (row 4 contains headers)
        expected_headers = [
            "Concept IRI*",
            "Language Code*",
            "Preferred Label*",
            "Definition*",
            "Alternate Labels",
            "Parent IRIs",
            "Member of collection(s)",
            "Member of ordered collection # position",
            "Provenance (read-only)",
            "Change Note",
            "Editorial Note",
            "Obsoletion reason",
            "Influenced by IRIs",
            "Source Vocab IRI or URL",
            "Source Vocab License",
            "Source Vocab Rights Holder",
        ]

        for col, expected_header in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual_header = ws[f"{col_letter}5"].value
            assert actual_header == expected_header, (
                f"Column {col} header mismatch: expected '{expected_header}', got '{actual_header}'"
            )

    def test_concepts_sheet_has_freeze_panes(self, tmp_path):
        """Verify Concepts sheet has freeze panes set."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Concepts"]

        assert ws.freeze_panes == "A6", (
            f"Expected freeze panes at A6, got {ws.freeze_panes}"
        )

    def test_collections_sheet_has_table(self, tmp_path):
        """Verify Collections sheet has proper table."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Collections"]

        assert ws.tables, "Collections sheet should have a table"

        # Check column headers
        expected_headers = [
            "Collection IRI*",
            "Language Code*",
            "Preferred Label*",
            "Definition*",
            "Parent Collection IRIs",
            "Ordered?",
            "Provenance (read-only)",
            "Change Note",
            "Editorial Note",
            "Obsoletion reason",
        ]

        for col, expected_header in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual_header = ws[f"{col_letter}5"].value
            assert actual_header == expected_header, (
                f"Column {col} header mismatch: expected '{expected_header}', got '{actual_header}'"
            )

    def test_mappings_sheet_has_table(self, tmp_path):
        """Verify Mappings sheet has proper table."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Mappings"]

        assert ws.tables, "Mappings sheet should have a table"

        # Check column headers
        expected_headers = [
            "Concept IRI*",
            "Related Matches",
            "Close Matches",
            "Exact Matches",
            "Narrower Matches",
            "Broader Matches",
            "Editorial Note",
        ]

        for col, expected_header in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual_header = ws[f"{col_letter}5"].value
            assert actual_header == expected_header, (
                f"Column {col} header mismatch: expected '{expected_header}', got '{actual_header}'"
            )

    def test_prefixes_sheet_has_default_prefixes(self, tmp_path):
        """Verify Prefixes sheet contains standard prefixes."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Prefixes"]

        # Check table exists
        assert ws.tables, "Prefixes sheet should have a table"

        # Check some key prefixes are present
        expected_prefixes = {"ex", "skos"}
        found_prefixes = set()

        # Prefixes start after header row
        for row in range(4, ws.max_row + 1):
            prefix = ws[f"A{row}"].value
            if prefix:  # pragma: no branch
                found_prefixes.add(prefix)

        assert expected_prefixes.issubset(found_prefixes), (
            f"Missing expected prefixes: {expected_prefixes - found_prefixes}"
        )

    def test_concept_scheme_has_correct_fields(self, tmp_path):
        """Verify Concept Scheme sheet has correct fields."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Concept Scheme"]

        # Check that Template version is first field
        assert ws["A4"].value == "Template version"
        assert ws["B4"].value == TEMPLATE_VERSION

        # Check other key fields exist
        expected_fields = [
            "Vocabulary IRI",
            "Prefix",
            "Title",
            "Description",
            "Created Date",
        ]

        found_fields = []
        for row in range(4, ws.max_row + 1):
            field = ws[f"A{row}"].value
            if field:  # pragma: no branch
                found_fields.append(field)

        for field in expected_fields:
            assert field in found_fields, f"Missing field: {field}"

    def test_id_ranges_sheet_has_correct_structure(self, tmp_path):
        """Verify ID Ranges sheet has title and correct headers."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["ID Ranges"]

        # Check title row
        assert ws["A1"].value == "ID Ranges (read-only)"

        # Check headers (row 3)
        expected_headers = ["gh-name", "ID Range", "Unused IDs"]
        actual_headers = [ws["A3"].value, ws["B3"].value, ws["C3"].value]
        assert actual_headers == expected_headers

        # Data row should be empty (headers only in blank template)
        assert ws["A4"].value is None or ws["A4"].value == ""


class TestModels:
    """Tests for v1.0 Pydantic models."""

    def test_concept_scheme_default_values(self):
        """Test ConceptSchemeV1 has correct defaults."""
        cs = ConceptSchemeV1()
        assert cs.template_version == TEMPLATE_VERSION
        assert cs.vocabulary_iri == ""
        assert cs.prefix == ""

    def test_concept_v1_model(self):
        """Test ConceptV1 model creation."""
        concept = ConceptV1(
            concept_iri="ex:0001",
            language_code="en",
            preferred_label="Test",
            definition="A test concept",
        )
        assert concept.concept_iri == "ex:0001"
        assert concept.language_code == "en"
        assert concept.preferred_label == "Test"
        assert concept.alternate_labels == ""  # default

    def test_collection_v1_model(self):
        """Test CollectionV1 model creation."""
        collection = CollectionV1(
            collection_iri="ex:col1",
            language_code="en",
            preferred_label="Collection",
            definition="A collection for testing",
        )
        assert collection.collection_iri == "ex:col1"
        assert collection.ordered == OrderedChoice.NO  # default

    def test_mapping_v1_model(self):
        """Test MappingV1 model creation."""
        mapping = MappingV1(
            concept_iri="ex:0001",
            exact_matches="http://example.org/ext/concept1",
        )
        assert mapping.concept_iri == "ex:0001"
        assert mapping.exact_matches == "http://example.org/ext/concept1"
        assert mapping.related_matches == ""  # default

    def test_prefix_v1_model(self):
        """Test PrefixV1 model creation."""
        prefix = PrefixV1(prefix="ex", namespace="https://example.org/")
        assert prefix.prefix == "ex"
        assert prefix.namespace == "https://example.org/"

    def test_default_prefixes_count(self):
        """Test that default prefixes list has expected entries."""
        assert len(DEFAULT_PREFIXES) == 2, (
            f"Expected 2 prefixes, got {len(DEFAULT_PREFIXES)}"
        )


class TestTemplateWithConfig:
    """Tests for template generation with vocab config."""

    @pytest.fixture
    def base_vocab_config(self):
        """Create a base vocab config with required fields."""

        return {
            "id_length": 4,
            "permanent_iri_part": "https://example.org/",
            "checks": Checks(),
            "prefix_map": {},
            # Mandatory ConceptScheme fields
            "vocabulary_iri": "https://example.org/base/",
            "title": "Base Title",
            "description": "Base description",
            "created_date": "2025-01-01",
            "creator": "Base Creator",
            "repository": "https://github.com/example/repo",
        }

    def test_generate_template_with_vocab_config(self, tmp_path, base_vocab_config):
        """Test template generation with vocabulary configuration."""

        # Override base config with specific values
        config = base_vocab_config.copy()
        config.update(
            {
                "vocabulary_iri": "https://example.org/myvocab/",
                "prefix": "myvoc",
                "title": "My Vocabulary",
                "description": "A test vocabulary",
                "created_date": "2025-01-15",
                "creator": "https://orcid.org/0000-0001-5000-0007 Test Author",
                "publisher": "Test Publisher",
                "custodian": "Test Custodian",
                "history_note": "Created for testing purposes",
                # Pass id_range as list of dicts (before validation)
                "id_range": [
                    {
                        "gh_name": "testuser",
                        "first_id": 1,
                        "last_id": 100,
                    }
                ],
            }
        )
        vocab_config = Vocab(**config)
        # Add custom prefix
        vocab_config.prefix_map["custom"] = "https://custom.example.org/"

        output_path = tmp_path / "test_vocab.xlsx"
        wb = generate_template_v1(output_path, vocab_config=vocab_config)

        # Verify Concept Scheme values from config
        ws = wb["Concept Scheme"]
        found_values = {}
        for row in range(4, ws.max_row + 1):
            field = ws[f"A{row}"].value
            value = ws[f"B{row}"].value
            if field:  # pragma: no branch
                found_values[field] = value

        assert found_values.get("Vocabulary IRI") == "https://example.org/myvocab/"
        assert found_values.get("Prefix") == "myvoc"
        assert found_values.get("Title") == "My Vocabulary"
        assert found_values.get("History Note") == "Created for testing purposes"

        # Verify ID Ranges from config
        ws_id = wb["ID Ranges"]
        assert ws_id["A4"].value == "testuser"
        assert ws_id["B4"].value == "1 - 100"

        # Verify Prefixes include custom prefix from config
        ws_pfx = wb["Prefixes"]
        found_prefixes = set()
        for row in range(4, ws_pfx.max_row + 1):
            prefix = ws_pfx[f"A{row}"].value
            if prefix:  # pragma: no branch
                found_prefixes.add(prefix)
        assert "custom" in found_prefixes

    def test_generate_template_with_history_note_auto_generation(
        self, tmp_path, base_vocab_config
    ):
        """Test history note is auto-generated when not provided in config."""

        # Override base config
        # Creator format: "<ORCID URL> <Name>"
        config = base_vocab_config.copy()
        config.update(
            {
                "vocabulary_iri": "https://example.org/autogen/",
                "prefix": "auto",
                "title": "Auto History Test",
                "description": "Testing auto-generation",
                "created_date": "2025-06-15",
                "creator": "https://orcid.org/0000-0001-5000-0007 Jane Doe",
            }
        )
        vocab_config = Vocab(**config)

        output_path = tmp_path / "autogen.xlsx"
        wb = generate_template_v1(output_path, vocab_config=vocab_config)

        # Verify history note was auto-generated
        ws = wb["Concept Scheme"]
        history_note_row = next(
            (
                row
                for row in range(4, ws.max_row + 1)
                if ws[f"A{row}"].value == "History Note"
            ),
            None,
        )
        assert history_note_row is not None, "History Note field not found"
        history_note = ws[f"B{history_note_row}"].value
        # Should contain date and creator name
        assert "2025-06-15" in history_note
        assert "Jane Doe" in history_note

    def test_generate_template_with_orcid_fallback(self, tmp_path, base_vocab_config):
        """Test ID ranges uses ORCID when gh_name not available."""

        # Override base config - use valid ORCID with correct checksum
        # id_range must be list of dicts for the validator
        config = base_vocab_config.copy()
        config.update(
            {
                "vocabulary_iri": "https://example.org/orcid/",
                "prefix": "orc",
                "title": "ORCID Test",
                "id_range": [
                    {
                        # Valid ORCID: 0000-0001-5000-0007 has correct checksum
                        "orcid": "https://orcid.org/0000-0001-5000-0007",
                        "first_id": 200,
                        "last_id": 299,
                    }
                ],
            }
        )
        vocab_config = Vocab(**config)

        output_path = tmp_path / "orcid.xlsx"
        wb = generate_template_v1(output_path, vocab_config=vocab_config)

        ws_id = wb["ID Ranges"]
        # Should use ORCID as the name since gh_name is not provided
        assert "0000-0001-5000-0007" in str(ws_id["A4"].value)

    def test_generate_template_without_output_path(self):
        """Test template generation returns workbook when no output path."""
        # When output_path is None, should still return valid workbook
        wb = generate_template_v1(output_path=None)

        # Verify workbook has expected sheets
        assert "Concept Scheme" in wb.sheetnames
        assert "Concepts" in wb.sheetnames
        assert "Prefixes" in wb.sheetnames


class TestTemplateCLI:
    """CLI integration tests for the template command."""

    def test_template_cli_default_output(self, tmp_path, monkeypatch):
        """Test template command generates file in current directory."""
        monkeypatch.chdir(tmp_path)
        main_cli(["template", "myvocab"])

        expected_file = tmp_path / "myvocab.xlsx"
        assert expected_file.exists(), f"Template file not created: {expected_file}"

        # Verify it's a valid xlsx with expected sheets
        wb = load_workbook(expected_file)
        assert "Concepts" in wb.sheetnames

    def test_template_cli_with_outdir(self, tmp_path):
        """Test template command with --outdir option."""
        outdir = tmp_path / "output"
        main_cli(["template", "--outdir", str(outdir), "testvocab"])

        expected_file = outdir / "testvocab.xlsx"
        assert expected_file.exists(), (
            f"Template file not created in outdir: {expected_file}"
        )

    def test_template_cli_with_version_option(self, tmp_path, monkeypatch):
        """Test template command with --version option."""
        monkeypatch.chdir(tmp_path)
        main_cli(["template", "--version", "v1.0", "versionedvocab"])
        assert (tmp_path / "versionedvocab.xlsx").exists()

    def test_template_cli_invalid_version(self, tmp_path):
        """Test template command with invalid --version option."""
        with pytest.raises(SystemExit) as exc_info:
            main_cli(["template", "--version", "v0.5", "somevocab"])
        assert exc_info.value.code == 2  # argparse error

    def test_template_cli_requires_vocab(self):
        """Test template command requires VOCAB argument."""
        with pytest.raises(SystemExit) as exc_info:
            main_cli(["template"])
        assert exc_info.value.code == 2  # argparse error

    def test_template_cli_no_overwrite(self, tmp_path, monkeypatch, caplog):
        """Test template command does not overwrite existing files."""
        monkeypatch.chdir(tmp_path)
        # Create first
        main_cli(["template", "existingvocab"])
        file_path = tmp_path / "existingvocab.xlsx"
        assert file_path.exists()
        original_mtime = file_path.stat().st_mtime

        # Try to create again - should log error and not overwrite
        main_cli(["template", "existingvocab"])
        assert "File already exists" in caplog.text
        assert file_path.stat().st_mtime == original_mtime  # File not modified

    def test_template_cli_strips_extension(self, tmp_path, monkeypatch):
        """Test template command strips extension from VOCAB name."""
        monkeypatch.chdir(tmp_path)
        main_cli(["template", "withext.xlsx"])

        expected_file = tmp_path / "withext.xlsx"
        assert expected_file.exists(), f"Template file not created: {expected_file}"
        # Should not create withext.xlsx.xlsx
        assert not (tmp_path / "withext.xlsx.xlsx").exists()

    def test_template_cli_with_base_template(self, tmp_path, monkeypatch):
        """Test template command with --template option preserves base sheets."""

        monkeypatch.chdir(tmp_path)

        # Create base template with custom sheets
        base_wb = Workbook()
        base_wb.active.title = "Cover"
        base_wb.create_sheet("Instructions")
        base_wb.create_sheet("Help")
        base_template = tmp_path / "base_template.xlsx"
        base_wb.save(base_template)
        base_wb.close()

        # Generate template with base
        main_cli(["template", "--template", str(base_template), "myvocab"])

        # Verify output
        output_file = tmp_path / "myvocab.xlsx"
        assert output_file.exists()

        result_wb = load_workbook(output_file)
        sheets = result_wb.sheetnames

        # Template sheets should come first
        assert sheets[0] == "Cover"
        assert sheets[1] == "Instructions"
        assert sheets[2] == "Help"

        # Auto-created sheets should follow
        assert "Concept Scheme" in sheets
        assert "Concepts" in sheets
        assert "Collections" in sheets
        assert "Mappings" in sheets
        assert "ID Ranges" in sheets
        assert "Prefixes" in sheets

        # Verify order: template sheets before auto-created
        cover_idx = sheets.index("Cover")
        cs_idx = sheets.index("Concept Scheme")
        assert cover_idx < cs_idx

        result_wb.close()

    def test_template_cli_rejects_conflicting_base_template(
        self, tmp_path, monkeypatch, caplog
    ):
        """Test template command rejects base template with reserved sheet names."""

        monkeypatch.chdir(tmp_path)

        # Create base template with a reserved sheet name
        base_wb = Workbook()
        base_wb.active.title = "Cover"
        base_wb.create_sheet("Concepts")  # Reserved name - should be rejected
        base_template = tmp_path / "bad_template.xlsx"
        base_wb.save(base_template)
        base_wb.close()

        with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
            main_cli(["template", "--template", str(base_template), "myvocab"])

        assert "reserved sheet names" in caplog.text
        assert "Concepts" in caplog.text

    def test_template_cli_rejects_nonexistent_base_template(
        self, tmp_path, monkeypatch, caplog
    ):
        """Test template command rejects non-existent base template."""

        monkeypatch.chdir(tmp_path)

        with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
            main_cli(["template", "--template", "nonexistent.xlsx", "myvocab"])

        assert "Template file not found" in caplog.text

    def test_template_cli_rejects_invalid_base_template_format(
        self, tmp_path, monkeypatch, caplog
    ):
        """Test template command rejects base template with wrong file type."""

        monkeypatch.chdir(tmp_path)

        # Create a non-xlsx file
        bad_template = tmp_path / "bad_template.txt"
        bad_template.touch()

        with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
            main_cli(["template", "--template", str(bad_template), "myvocab"])

        assert 'Template file must be of type ".xlsx"' in caplog.text

    def test_template_cli_with_config_vocab_found(self, tmp_path, monkeypatch, caplog):
        """Test template command with config file where vocab is found."""
        monkeypatch.chdir(tmp_path)

        # Create a full idranges.toml config with all required fields
        config_content = """
[vocabs.testvocab]
id_length = 4
permanent_iri_part = "https://example.org/testvocab/"
prefix_map = {}
vocabulary_iri = "https://example.org/testvocab/"
prefix = "test"
title = "Test Vocabulary"
description = "A test vocabulary from config"
created_date = "2025-01-01"
creator = "Test Creator"
repository = "https://github.com/example/testvocab"

[vocabs.testvocab.checks]
"""
        config_file = tmp_path / "idranges.toml"
        config_file.write_text(config_content)

        with caplog.at_level(logging.INFO):
            main_cli(["template", "--config", str(config_file), "testvocab"])

        # Verify template was created
        assert (tmp_path / "testvocab.xlsx").exists()

        # Verify config was used
        assert "Using config for vocabulary" in caplog.text

    def test_template_cli_with_config_vocab_not_found(
        self, tmp_path, monkeypatch, caplog
    ):
        """Test template command with config file where vocab is not found."""
        monkeypatch.chdir(tmp_path)

        # Create a config without the requested vocabulary
        config_content = """
[vocabs.othervocab]
id_length = 4
permanent_iri_part = "https://example.org/other/"
prefix_map = {}
vocabulary_iri = "https://example.org/other/"
prefix = "other"
title = "Other Vocabulary"
description = "Other description"
created_date = "2025-01-01"
creator = "Other Creator"
repository = "https://github.com/example/other"

[vocabs.othervocab.checks]
"""
        config_file = tmp_path / "idranges.toml"
        config_file.write_text(config_content)

        with caplog.at_level(logging.WARNING):
            main_cli(["template", "--config", str(config_file), "myvocab"])

        # Template should still be created (with example data)
        assert (tmp_path / "myvocab.xlsx").exists()

        # Warning should be logged
        assert "not found in config file" in caplog.text

    def test_template_cli_with_missing_config_file(self, tmp_path, monkeypatch, caplog):
        """Test template command with non-existent config file raises error."""

        monkeypatch.chdir(tmp_path)

        with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
            main_cli(["template", "--config", "nonexistent.toml", "myvocab"])

        # Error should be logged
        assert "Config file not found" in caplog.text
