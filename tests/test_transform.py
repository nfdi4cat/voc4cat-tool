import logging
import os
import shutil
from unittest import mock

import pytest
from openpyxl import load_workbook
from rdflib import DCTERMS, OWL, SKOS, XSD, Graph, Literal

from tests.test_cli import (
    CS_CYCLES_MULTI_LANG,
    CS_SIMPLE,
    CS_SIMPLE_TURTLE,
)
from voc4cat.checks import Voc4catError
from voc4cat.cli import main_cli

# ===== Tests for no option set =====


def test_transform_no_option(monkeypatch, datadir, tmp_path, caplog):
    shutil.copy(datadir / CS_SIMPLE, tmp_path / CS_SIMPLE)
    monkeypatch.chdir(tmp_path)

    with caplog.at_level(logging.DEBUG):
        main_cli(["transform", str(tmp_path)])
    assert "nothing to do for xlsx files" in caplog.text

    shutil.copy(datadir / CS_SIMPLE_TURTLE, tmp_path / CS_SIMPLE_TURTLE)
    with caplog.at_level(logging.DEBUG):
        main_cli(["transform", str(tmp_path)])
    assert "nothing to do for rdf files" in caplog.text


def test_transform_unsupported_filetype(monkeypatch, datadir, tmp_path, caplog):
    shutil.copy(datadir / "README.md", tmp_path)
    monkeypatch.chdir(tmp_path)
    # Try to run a command that would overwrite the input file with the output file
    with caplog.at_level(logging.WARNING):
        main_cli(["transform", str(tmp_path / "README.md")])
    assert "Unsupported filetype: " in caplog.text


# ===== Tests for option --make-ids =====


def test_make_ids_missing_file(caplog):
    with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
        main_cli(["transform", "--make-ids", "ex", "1", "missing.xyz"])
    assert "File/dir not found: missing.xyz" in caplog.text


def test_make_ids_overwrite_warning(monkeypatch, datadir, tmp_path, caplog):
    shutil.copy(datadir / CS_SIMPLE, tmp_path / CS_SIMPLE)
    monkeypatch.chdir(tmp_path)
    # Try to run a command that would overwrite the input file with the output file
    # a) dir as input:
    with caplog.at_level(logging.WARNING):
        main_cli(["transform", "--make-ids", "ex", "1", str(tmp_path)])
    assert "This command will overwrite the existing file" in caplog.text
    # b) file as input
    with caplog.at_level(logging.WARNING):
        main_cli(["transform", "--make-ids", "ex", "1", str(tmp_path / CS_SIMPLE)])
    assert "This command will overwrite the existing file" in caplog.text


def test_make_ids_no_voc_base_iri(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_SIMPLE, tmp_path)
    monkeypatch.chdir(tmp_path)
    # change excel file: Delete vocabulary base IRI
    wb = load_workbook(filename=CS_SIMPLE)
    ws = wb["Concept Scheme"]
    ws.cell(row=2, column=2).value = None
    new_filename = "no_voc_base_iri.xlsx"
    wb.save(new_filename)
    wb.close()

    main_cli(
        [
            "transform",
            "--make-ids",
            "na",
            "1001",
            "--inplace",
            str(tmp_path / new_filename),
        ]
    )
    wb = load_workbook(filename=new_filename, read_only=True, data_only=True)
    ws = wb["Concept Scheme"]
    assert ws.cell(row=2, column=2).value == "https://example.org/"


def test_make_ids_invalid_id(datadir):
    with pytest.raises(
        Voc4catError,
        match="Start ID must be an integer number.",
    ):
        main_cli(
            [
                "transform",
                "--make-ids",
                "ex",
                "###",
                "--inplace",
                str(datadir / CS_SIMPLE),
            ]
        )


def test_make_ids_invalid_negative_id(datadir):
    with pytest.raises(
        Voc4catError,
        match="Start ID must be greater than zero.",
    ):
        main_cli(
            [
                "transform",
                "--make-ids",
                "ex",
                "-1",
                "--inplace",
                str(datadir / CS_SIMPLE),
            ]
        )


def test_make_ids_invalid_base_iri(datadir):
    with pytest.raises(
        Voc4catError,
        match='The base_iri must be in IRI-form and start with "http".',
    ):
        main_cli(
            [
                "transform",
                "--make-ids",
                "ex:ftp://example.org",
                "1",
                "--inplace",
                str(datadir / CS_SIMPLE),
            ]
        )


@pytest.mark.parametrize(
    ("indir", "outdir"),
    [(True, ""), (True, "out"), (False, ""), (False, "out")],
    ids=[
        "in:dir, out:default",
        "in:dir, out:dir",
        "in:file, out:default",
        "in:file, out:dir",
    ],
)
def test_make_ids_variants(monkeypatch, datadir, tmp_path, indir, outdir):
    # fmt: off
    expected_concepts = [
        ("ex:test/0001001", "term1", "en", "def for term1", "en", "AltLbl for term1", None,),
        ("ex:test/0001002", "term2", "en", "def for term2", "en", "AltLbl for term2", "ex:test/0001001",),
        ("ex:test/0001003", "term3", "en", "def for term3", "en", "AltLbl for term3", "ex:test/0001001",),
        ("ex:test/0001004", "term4", "en", "def for term4", "en", "AltLbl for term4", "ex:test/0001003",),
        ("ex:test/0001005", "term5", "en", "def for term5", "en", "AltLbl for term5", None, ),
        ("ex:test/0001006", "term6", "en", "def for term6", "en", "AltLbl for term6", None,),
    ]
    expected_collections = [
        ("ex:test/0001007", "con", "def for con", "ex:test/0001001, ex:test/0001002, ex:test/0001003, ex:test/0001004",),
    ]
    # fmt: on
    expected_additional = (
        "ex:test/0001001",
        "ex:test/0001002",
        "ex:test/0001003",
        "ex:test/0001004",
        "ex:test/0001005",
        "ex:test/0001006",
    )

    shutil.copy(datadir / CS_SIMPLE, tmp_path)
    monkeypatch.chdir(tmp_path)
    main_cli(
        ["transform", "--make-ids", "ex", "1001", "--inplace"]
        + (["--outdir", outdir] if outdir else [])
        + ([str(tmp_path)] if indir else [str(tmp_path / CS_SIMPLE)])
    )
    xlsxfile = tmp_path / outdir / CS_SIMPLE if outdir else tmp_path / CS_SIMPLE
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=7, values_only=True), expected_concepts
    ):
        assert row == expected_row
    ws = wb["Collections"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=4, values_only=True), expected_collections
    ):
        assert row == expected_row

    ws = wb["Additional Concept Features"]
    count_present = 0
    for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
        if row[0] == expected_additional[0]:  # same concept IRI
            assert row == expected_additional
        elif row[0]:
            count_present += 1
    # Only 4 of 6 concepts are in the Additional Concept Features
    assert count_present == 3  # noqa: PLR2004


def test_make_ids_base_iri(monkeypatch, datadir, tmp_path):
    # fmt: off
    expected_concepts = [
        ("https://example.com/new_0001001", "term1", "en", "def for term1", "en", "AltLbl for term1", None,),
        ("https://example.com/new_0001002", "term2", "en", "def for term2", "en", "AltLbl for term2", "https://example.com/new_0001001",),
        ("https://example.com/new_0001003", "term3", "en", "def for term3", "en", "AltLbl for term3", "https://example.com/new_0001001",),
        ("https://example.com/new_0001004", "term4", "en", "def for term4", "en", "AltLbl for term4", "https://example.com/new_0001003", ),
        ("https://example.com/new_0001005", "term5", "en", "def for term5", "en", "AltLbl for term5", None, ),
        ("https://example.com/new_0001006", "term6", "en", "def for term6", "en", "AltLbl for term6", None,),
    ]
    expected_collections = [
        ("https://example.com/new_0001007", "con", "def for con", "https://example.com/new_0001001, https://example.com/new_0001002, https://example.com/new_0001003, https://example.com/new_0001004",),
    ]
    # fmt: on
    expected_additional = (
        "https://example.com/new_0001001",
        "https://example.com/new_0001002",
        "https://example.com/new_0001003",
        "https://example.com/new_0001004",
        "https://example.com/new_0001005",
        "https://example.com/new_0001006",
    )

    shutil.copy(datadir / CS_SIMPLE, tmp_path)
    monkeypatch.chdir(tmp_path)
    main_cli(
        [
            "transform",
            "--make-ids",
            "ex:https://example.com/new_",
            "1001",
            "--inplace",
            str(tmp_path / CS_SIMPLE),
        ]
    )
    xlsxfile = tmp_path / CS_SIMPLE
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=7, values_only=True), expected_concepts
    ):
        assert row == expected_row
    ws = wb["Collections"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=4, values_only=True), expected_collections
    ):
        assert row == expected_row
    ws = wb["Additional Concept Features"]
    match = False
    for row in ws.iter_rows(min_row=3, max_col=6, values_only=True):
        if row[0] == expected_additional[0]:
            match = True
            assert row == expected_additional
    assert match


def test_make_ids_multilang(tmp_path, datadir):
    # fmt: off
    expected_concepts = [
        ("ex:test/0001001", "term1", "en", "def for term1", "en", "AltLbl for term1", None,),
        ("ex:test/0001002", "term2", "en", "def for term2", "en", "AltLbl for term2", "ex:test/0001001",),
        ("ex:test/0001003", "term3", "en", "def for term3", "en", "AltLbl for term3", "ex:test/0001001",),
        ("ex:test/0001004", "term4", "en", "def for term4", "en", "AltLbl for term4", "ex:test/0001003",),
        ("ex:test/0001004", "Begr4", "de", "Def für Begr4", "de", "AltLbl für Begr4", None, ),
        ("ex:test/0001001", "Begr1", "de", "Def für Begr1", "de", "AltLbl für Begr1", None, ),
    ]
    # fmt: on
    main_cli(
        [
            "transform",
            "--make-ids",
            "ex",
            "1001",
            "--outdir",
            str(tmp_path),
            str(datadir / CS_CYCLES_MULTI_LANG),
        ]
    )
    xlsxfile = tmp_path / CS_CYCLES_MULTI_LANG
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=7, values_only=True), expected_concepts
    ):
        assert row == expected_row


# ===== Tests for options --split / --join =====


@pytest.mark.parametrize(
    "opt",
    [None, "--inplace"],
    ids=["default", "inplace"],
)
def test_split(monkeypatch, datadir, tmp_path, opt, caplog):
    shutil.copy(datadir / CS_SIMPLE_TURTLE, tmp_path)
    cmd = ["transform", "-v", "--split"]
    if opt:
        cmd.append("--inplace")
    cmd.append(str(tmp_path))
    monkeypatch.chdir(tmp_path)

    with caplog.at_level(logging.DEBUG):
        main_cli(cmd)
    assert "-> wrote split vocabulary to" in caplog.text

    vocdir = (tmp_path / CS_SIMPLE_TURTLE).with_suffix("")
    assert vocdir.is_dir()
    assert (vocdir / "concept_scheme.ttl").exists()
    assert len([*vocdir.glob("*.ttl")]) == 8  # noqa: PLR2004
    if opt:
        assert not (tmp_path / CS_SIMPLE_TURTLE).exists()


@pytest.mark.parametrize(
    "opt",
    [None, "inplace", "outdir"],
)
def test_join(monkeypatch, datadir, tmp_path, opt, caplog):
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_SIMPLE_TURTLE, tmp_path)
    # create dir with split files
    main_cli(["transform", "-v", "--split", "--inplace", str(tmp_path)])
    # join files again as test
    cmd = ["transform", "-v", "--join"]
    if opt == "inplace":
        cmd.append("--inplace")
    elif opt == "outdir":
        cmd.append("--outdir")
        cmd.append(str(tmp_path / "outdir"))
    cmd.append(str(tmp_path))

    with caplog.at_level(logging.DEBUG):
        main_cli(cmd)
    assert "-> joined vocabulary into" in caplog.text

    vocdir = (tmp_path / CS_SIMPLE_TURTLE).with_suffix("")
    if opt == "inplace":
        assert not vocdir.exists()
        assert (vocdir.parent / CS_SIMPLE_TURTLE).exists()
    elif opt == "outdir":
        assert (tmp_path / "outdir" / CS_SIMPLE_TURTLE).exists()
        assert not (vocdir.parent / CS_SIMPLE_TURTLE).exists()
    else:
        assert (vocdir.parent / CS_SIMPLE_TURTLE).exists()


@mock.patch.dict(
    os.environ, {"CI": "", "VOC4CAT_VERSION": "v2.0", "VOC4CAT_MODIFIED": "2023-08-15"}
)
def test_join_with_envvars(monkeypatch, datadir, tmp_path, caplog):
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_SIMPLE_TURTLE, tmp_path)
    # create dir with split files
    main_cli(["transform", "-v", "--split", "--inplace", str(tmp_path)])
    # join files again as test
    cmd = ["transform", "-v", "--join"]
    cmd.append(str(tmp_path))
    with caplog.at_level(logging.DEBUG):
        main_cli(cmd)
    assert "-> joined vocabulary into" in caplog.text
    # Were version and modified date correctly set?
    graph = Graph().parse(CS_SIMPLE_TURTLE, format="turtle")
    cs_query = "SELECT ?iri WHERE {?iri a skos:ConceptScheme.}"
    qresults = [*graph.query(cs_query, initNs={"skos": SKOS})]
    assert len(qresults) == 1

    cs_iri = qresults[0][0]
    assert (
        len(
            [
                *graph.triples(
                    (cs_iri, OWL.versionInfo, Literal("v2.0")),
                )
            ]
        )
        == 1
    )

    assert (
        len(
            [
                *graph.triples(
                    (
                        cs_iri,
                        DCTERMS.modified,
                        Literal("2023-08-15", datatype=XSD.date),
                    ),
                )
            ]
        )
        == 1
    )


@mock.patch.dict(
    os.environ, {"CI": "", "VOC4CAT_VERSION": "2.0", "VOC4CAT_MODIFIED": "2023-08-15"}
)
def test_join_with_invalid_envvar(monkeypatch, datadir, tmp_path, caplog):
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_SIMPLE_TURTLE, tmp_path)
    # create dir with split files
    main_cli(["transform", "-v", "--split", "--inplace", str(tmp_path)])
    # join files again as test
    cmd = ["transform", "-v", "--join"]
    cmd.append(str(tmp_path))
    with (
        caplog.at_level(logging.ERROR),
        pytest.raises(
            Voc4catError, match="Invalid environment variable VOC4CAT_VERSION"
        ),
    ):
        main_cli(cmd)
    assert 'Invalid environment variable VOC4CAT_VERSION "2.0"' in caplog.text
