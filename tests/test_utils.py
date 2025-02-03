import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table

from voc4cat.utils import adjust_length_of_tables, split_and_tidy


def test_default_action():
    assert split_and_tidy(None) == []
    assert split_and_tidy("") == []
    assert split_and_tidy("a,b") == ["a", "b"]
    assert split_and_tidy(" a , b ") == ["a", "b"]


def test_trailing_comma():
    assert split_and_tidy("a,") == ["a"]
    assert split_and_tidy("a,b,") == ["a", "b"]


def test_expand_tables(tmp_path, caplog):
    caplog.set_level(logging.DEBUG)
    test_wb = tmp_path / "table.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Concepts"
    ws.append(["Letter", "value"])  # table header
    data = [
        ["A", 1],
        ["B", 2],
    ]
    for row in data:
        ws.append(row)
    tab = Table(displayName="Table1", ref="A1:B3")
    ws.add_table(tab)
    # style A2 with bold font
    ws["A2"].font = Font(bold=True)
    ws["B3"].alignment = Alignment(indent=2.0)
    ws["A5"] = "X"
    wb.save(test_wb)

    print(f"\nFile: {test_wb}")

    # Test with no expansion of table beyond last row used.
    adjust_length_of_tables(test_wb, rows_pre_allocated=0, copy_style=False)
    wb = load_workbook(test_wb)
    name, table_range = wb.active.tables.items()[0]
    assert name == "Table1"
    assert table_range == "A1:B5"
    assert "from {A1:B3} to {A1:B5}" in caplog.text
    assert not ws["A3"].font.bold


    caplog.clear()
    # Test with expansion of table given rows beyond last data block row.
    adjust_length_of_tables(test_wb, rows_pre_allocated=5) # adds 5 rows (default of rows_pre_allocated)
    wb = load_workbook(test_wb)
    name, table_range = wb.active.tables.items()[0]
    assert name == "Table1"
    assert table_range == "A1:B8"
    assert "from {A1:B5} to {A1:B8}" in caplog.text
    # alignment should be kept in each row col B in sheet "Concepts"
    assert wb["Concepts"]["B3"].alignment.indent == 2.0

    caplog.clear()
    # Run once again: The size should not change because the contents are the same
    adjust_length_of_tables(test_wb, rows_pre_allocated={"Sheet": 5})  # add 5 rows
    wb = load_workbook(test_wb)
    name, table_range = wb.active.tables.items()[0]
    assert name == "Table1"
    assert table_range == "A1:B8"
    assert not caplog.text
