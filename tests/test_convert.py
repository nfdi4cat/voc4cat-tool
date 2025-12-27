import contextlib
import logging
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from rdflib import SH

from tests.test_cli import (
    CS_CYCLES,
    CS_CYCLES_TURTLE,
)
from voc4cat.checks import Voc4catError
from voc4cat.cli import main_cli
from voc4cat.convert import format_log_msg, resolve_profile, validate_with_profile
from voc4cat.utils import ConversionError


@pytest.mark.parametrize(
    ("outputdir", "testfile"),
    [
        ("out", CS_CYCLES_TURTLE),
        ("", CS_CYCLES_TURTLE),
    ],
    ids=["out:dir & ttl", "out:default & ttl"],
)
def test_run_voc4cat_outputdir(
    monkeypatch, datadir, tmp_path, outputdir, testfile, temp_config
):
    """Check that an xlsx file is converted to ttl by voc4cat."""
    shutil.copy(datadir / testfile, tmp_path)
    monkeypatch.chdir(tmp_path)

    # Load/prepare an a config with required prefix definition
    config = temp_config
    vocab_name = testfile.split(".")[0]
    config.CURIES_CONVERTER_MAP[vocab_name] = config.curies_converter

    # Check if log is placed in out folder.
    log = Path(outputdir) / "test-run.log"
    main_cli(
        ["convert", "--logfile", str(log)]
        + (["--outdir", str(outputdir)] if outputdir else [])
        + [str(tmp_path)]
    )
    outdir = tmp_path / outputdir
    assert (outdir / testfile).with_suffix(".xlsx").exists()
    assert (outdir / log.name).exists()


def test_duplicates(datadir, tmp_path, caplog, cs_cycles_xlsx):
    """Check that files do not have the same stem."""
    shutil.copy(cs_cycles_xlsx, tmp_path / CS_CYCLES)
    shutil.copy(datadir / CS_CYCLES_TURTLE, tmp_path)
    with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
        main_cli(["convert", str(tmp_path)])
    assert "Files may only be present in one format." in caplog.text


def test_template_missing_warns_and_continues(monkeypatch, datadir, tmp_path, caplog):
    """Check that missing template logs warning and continues without template."""
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_CYCLES_TURTLE, tmp_path)
    with caplog.at_level(logging.WARNING):
        main_cli(["convert", "--template", "missing.xlsx", str(tmp_path)])
    assert "Template file not found:" in caplog.text
    assert "Continuing without template" in caplog.text
    # Conversion should succeed
    assert (tmp_path / CS_CYCLES_TURTLE).with_suffix(".xlsx").exists()


def test_template_wrong_extension_raises_error(monkeypatch, datadir, tmp_path, caplog):
    """Check that template with wrong extension raises error."""
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_CYCLES_TURTLE, tmp_path)
    Path(tmp_path / "template.txt").touch()
    with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
        main_cli(["convert", "--template", "template.txt", str(tmp_path)])
    assert 'Template file must be of type ".xlsx".' in caplog.text


def test_template_with_conflicting_sheets(monkeypatch, datadir, tmp_path, caplog):
    """Check that templates with reserved sheet names are rejected."""
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_CYCLES_TURTLE, tmp_path)

    # Create template with a reserved sheet name
    wb = Workbook()
    wb.active.title = "Branding"  # OK - not reserved
    wb.create_sheet("Concepts")  # CONFLICT - reserved name
    template_path = tmp_path / "bad_template.xlsx"
    wb.save(template_path)
    wb.close()

    with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
        main_cli(["convert", "--template", str(template_path), str(tmp_path)])

    assert "reserved sheet names" in caplog.text
    assert "Concepts" in caplog.text


def test_template_sheets_preserved_and_ordered(
    monkeypatch, datadir, tmp_path, temp_config
):
    """Check that template sheets are preserved and placed first."""
    monkeypatch.chdir(tmp_path)
    vocab_dir = tmp_path / "vocab"
    vocab_dir.mkdir()
    shutil.copy(datadir / CS_CYCLES_TURTLE, vocab_dir)

    # Create valid template with branding sheets (outside vocab_dir)
    wb = Workbook()
    wb.active.title = "Cover"
    wb.create_sheet("Instructions")
    wb.create_sheet("Help")
    template_path = tmp_path / "branding_template.xlsx"
    wb.save(template_path)
    wb.close()

    # Set up config for the vocabulary (name matches input file stem)
    config = temp_config
    vocab_name = "concept-scheme-with-cycles"
    config.CURIES_CONVERTER_MAP[vocab_name] = config.curies_converter

    main_cli(["convert", "--template", str(template_path), str(vocab_dir)])

    # Check output (name matches input file stem)
    output_path = vocab_dir / "concept-scheme-with-cycles.xlsx"
    assert output_path.exists()

    result_wb = load_workbook(output_path)
    sheets = result_wb.sheetnames

    # Template sheets should come first
    assert sheets[0] == "Cover"
    assert sheets[1] == "Instructions"
    assert sheets[2] == "Help"

    # Auto-created sheets should follow
    assert "Concept Scheme" in sheets
    assert "Concepts" in sheets
    assert "Collections" in sheets
    assert "Mappings" in sheets
    assert "ID Ranges" in sheets
    assert "Prefixes" in sheets

    # Verify order: template sheets before auto-created
    cover_idx = sheets.index("Cover")
    cs_idx = sheets.index("Concept Scheme")
    assert cover_idx < cs_idx

    result_wb.close()


def test_template_warning_for_xlsx_input(
    monkeypatch, datadir, tmp_path, caplog, cs_cycles_xlsx
):
    """Check that a warning is logged when template is used with xlsx input."""
    monkeypatch.chdir(tmp_path)
    vocab_dir = tmp_path / "vocab"
    vocab_dir.mkdir()
    shutil.copy(cs_cycles_xlsx, vocab_dir / CS_CYCLES)

    # Create a valid template (outside vocab_dir)
    wb = Workbook()
    wb.active.title = "Branding"
    template_path = tmp_path / "template.xlsx"
    wb.save(template_path)
    wb.close()

    # The conversion may fail for other reasons (missing config), but the warning
    # should be logged before that happens
    with caplog.at_level(logging.WARNING), contextlib.suppress(Voc4catError):
        main_cli(["convert", "--template", str(template_path), str(vocab_dir)])

    assert "Template option ignored for xlsx->RDF conversion" in caplog.text


class TestResolveProfile:
    """Tests for resolve_profile function."""

    def test_resolve_bundled_profile(self):
        """Test resolving a bundled profile token."""
        # Use an actual bundled profile
        path, name = resolve_profile("vocpub-4.7")
        assert path.exists()
        assert name == "vocpub-4.7"

    def test_resolve_custom_profile_file(self, tmp_path):
        """Test resolving a custom SHACL profile file."""
        # Create a dummy profile file
        custom_profile = tmp_path / "my_profile.ttl"
        custom_profile.write_text("# Custom SHACL profile")

        path, name = resolve_profile(str(custom_profile))
        assert path == custom_profile
        assert name == "my_profile"

    def test_resolve_profile_file_not_found(self, tmp_path):
        """Test error when profile file doesn't exist."""
        with pytest.raises(Voc4catError, match="SHACL profile file not found"):
            resolve_profile(str(tmp_path / "nonexistent.ttl"))

    def test_resolve_profile_unknown_token(self):
        """Test error when unknown profile token is given."""
        with pytest.raises(Voc4catError, match="Unknown profile"):
            resolve_profile("unknown_profile_xyz")


class TestConvertFrom043:
    """Tests for --from 043 conversion path."""

    def test_from_043_requires_config(self, tmp_path, monkeypatch):
        """Test that --from 043 requires idranges.toml config."""
        # Create a dummy RDF file
        rdf_file = tmp_path / "test.ttl"
        rdf_file.write_text("# Dummy RDF")

        monkeypatch.chdir(tmp_path)

        with pytest.raises(Voc4catError, match=r"--from 043 requires an idranges.toml"):
            main_cli(["convert", "--from", "043", str(tmp_path)])

    def test_from_043_requires_v1_config(self, tmp_path, monkeypatch):
        """Test that --from 043 requires v1.0 config format."""
        # Create a dummy RDF file
        rdf_file = tmp_path / "test.ttl"
        rdf_file.write_text("# Dummy RDF")

        # Create an old-style config (no config_version)
        config_content = """
[vocabs.test]
id_length = 4
permanent_iri_part = "https://example.org/test/"
prefix_map = {}
vocabulary_iri = "https://example.org/test/"
title = "Test"
description = "Test"
created_date = "2025-01-01"
creator = "https://orcid.org/0000-0001-5000-0007 Test"
repository = "https://github.com/example/test"

[vocabs.test.checks]
"""
        config_file = tmp_path / "idranges.toml"
        config_file.write_text(config_content)

        monkeypatch.chdir(tmp_path)

        with pytest.raises(Voc4catError, match=r"Pre-v1.0 idranges.toml detected"):
            main_cli(
                [
                    "convert",
                    "--config",
                    str(config_file),
                    "--from",
                    "043",
                    str(tmp_path),
                ]
            )

    def test_from_043_converts_rdf(self, datadir, tmp_path, monkeypatch, temp_config):
        """Test that --from 043 converts RDF files successfully."""
        # Copy test file
        shutil.copy(datadir / "vocab-043-test.ttl", tmp_path)

        # Create v1.0 config (creator is just the ORCID URL)
        config_content = """
config_version = "v1.0"
single_vocab = true

[vocabs.vocab-043-test]
id_length = 7
permanent_iri_part = "http://example.org/test-vocab/"
vocabulary_iri = "http://example.org/test-vocab/"
title = "Test Vocabulary 043"
description = "Test vocabulary"
created_date = "2023-06-29"
creator = "Test Author https://orcid.org/0000-0001-5000-0007"
repository = "https://github.com/example/test"

[vocabs.vocab-043-test.checks]

[vocabs.vocab-043-test.prefix_map]
ex = "http://example.org/"
"""
        config_file = tmp_path / "idranges.toml"
        config_file.write_text(config_content)

        monkeypatch.chdir(tmp_path)

        main_cli(
            ["convert", "--config", str(config_file), "--from", "043", str(tmp_path)]
        )

        # Check output was created
        assert (tmp_path / "vocab-043-test.ttl").exists()


class TestFormatLogMsg:
    """Tests for format_log_msg function."""

    def test_format_info_message(self):
        """Test formatting INFO severity message."""
        result_dict = {
            "sourceConstraintComponent": f"{SH}MinCountConstraintComponent",
            "resultSeverity": str(SH.Info),
            "sourceShape": "http://example.org/shape",
            "focusNode": "http://example.org/node",
            "resultMessage": "Test info message",
        }
        msg = format_log_msg(result_dict, colored=True)
        assert "INFO:" in msg
        assert "Test info message" in msg

    def test_format_warning_message(self):
        """Test formatting WARNING severity message."""
        result_dict = {
            "sourceConstraintComponent": f"{SH}MinCountConstraintComponent",
            "resultSeverity": str(SH.Warning),
            "sourceShape": "http://example.org/shape",
            "focusNode": "http://example.org/node",
            "resultMessage": "Test warning message",
        }
        msg = format_log_msg(result_dict, colored=True)
        assert "WARNING:" in msg
        assert "Test warning message" in msg

    def test_format_violation_message(self):
        """Test formatting VIOLATION severity message."""
        result_dict = {
            "sourceConstraintComponent": f"{SH}MinCountConstraintComponent",
            "resultSeverity": str(SH.Violation),
            "sourceShape": "http://example.org/shape",
            "focusNode": "http://example.org/node",
            "resultMessage": "Test violation message",
        }
        msg = format_log_msg(result_dict, colored=True)
        assert "VIOLATION:" in msg
        assert "Test violation message" in msg

    def test_format_message_with_value(self):
        """Test formatting message that includes a value node."""
        result_dict = {
            "sourceConstraintComponent": f"{SH}MinCountConstraintComponent",
            "resultSeverity": str(SH.Info),
            "sourceShape": "http://example.org/shape",
            "focusNode": "http://example.org/node",
            "value": "test-value",
            "resultMessage": "Test message with value",
        }
        msg = format_log_msg(result_dict)
        assert "test-value" in msg


class TestValidateWithProfile:
    """Tests for validate_with_profile function."""

    def test_validate_valid_vocabulary(self, datadir):
        """Test validation of a valid vocabulary passes."""
        # concept-scheme-simple.ttl should be valid
        ttl_file = datadir / "concept-scheme-simple.ttl"
        # Should not raise
        validate_with_profile(str(ttl_file), profile="vp4cat-5.2")

    def test_validate_invalid_vocabulary_raises(self, datadir):
        """Test validation of invalid vocabulary raises ConversionError."""
        ttl_file = datadir / "concept-scheme-badfile.ttl"
        with pytest.raises(ConversionError, match="not valid according to"):
            validate_with_profile(str(ttl_file), profile="vp4cat-5.2")

    def test_validate_with_error_level_3(self, datadir, caplog):
        """Test validation with error_level=3 only fails on violations."""
        # badfile should have violations
        ttl_file = datadir / "concept-scheme-badfile.ttl"
        with caplog.at_level(logging.ERROR), pytest.raises(ConversionError):
            validate_with_profile(str(ttl_file), profile="vp4cat-5.2", error_level=3)


class TestXlsxToRdfConversion:
    """Tests for xlsx to RDF conversion path."""

    def test_xlsx_to_rdf_requires_config(self, tmp_path, cs_cycles_xlsx, caplog):
        """Test that xlsx to RDF conversion requires vocab config."""
        vocab_dir = tmp_path / "vocab"
        vocab_dir.mkdir()
        shutil.copy(cs_cycles_xlsx, vocab_dir / CS_CYCLES)

        with pytest.raises(Voc4catError, match=r"No idranges.toml config found"):
            main_cli(["convert", str(vocab_dir)])

    def test_xlsx_to_rdf_with_config(
        self, datadir, tmp_path, monkeypatch, cs_cycles_xlsx, temp_config
    ):
        """Test xlsx to RDF conversion with proper config."""
        vocab_dir = tmp_path / "vocab"
        vocab_dir.mkdir()
        shutil.copy(cs_cycles_xlsx, vocab_dir / CS_CYCLES)

        # Load config for the vocabulary
        config = temp_config
        vocab_name = "concept-scheme-with-cycles"
        config.CURIES_CONVERTER_MAP[vocab_name] = config.curies_converter

        # Create a v1.0 config file (creator is just the ORCID URL)
        config_content = """
config_version = "v1.0"
single_vocab = true

[vocabs.concept-scheme-with-cycles]
id_length = 7
permanent_iri_part = "http://example.org/test/"
vocabulary_iri = "http://example.org/test/"
title = "Test Vocabulary"
description = "Test vocabulary"
created_date = "2022-12-01"
creator = "Test Author https://orcid.org/0000-0001-5000-0007"
repository = "https://github.com/example/test"

[vocabs.concept-scheme-with-cycles.checks]

[vocabs.concept-scheme-with-cycles.prefix_map]
ex = "http://example.org/"
"""
        config_file = tmp_path / "idranges.toml"
        config_file.write_text(config_content)

        monkeypatch.chdir(tmp_path)

        main_cli(["convert", "--config", str(config_file), str(vocab_dir)])

        # Check output was created
        assert (vocab_dir / "concept-scheme-with-cycles.ttl").exists()
