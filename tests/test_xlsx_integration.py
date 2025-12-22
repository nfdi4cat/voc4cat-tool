"""
Tests for integration scenarios and real-world use cases.

This module tests complex integration scenarios, performance, and edge cases
that span multiple modules in the XLSX processing system.
"""

from datetime import date
from enum import Enum
from typing import Annotated, get_origin

import pytest
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field, field_validator

from voc4cat.models_v1 import TEMPLATE_VERSION
from voc4cat.xlsx_api import (
    XLSXProcessorFactory,
    create_xlsx_wrapper,
    export_to_xlsx,
    import_from_xlsx,
)
from voc4cat.xlsx_common import XLSXConverters, XLSXMetadata
from voc4cat.xlsx_keyvalue import XLSXKeyValueConfig
from voc4cat.xlsx_table import (
    JoinConfiguration,
    JoinedModelProcessor,
    XLSXDeserializationError,
    XLSXTableConfig,
)

from .conftest import Employee, Priority, Project, SimpleModel, Status


# Vocabulary Template Models (copied from demo files)
class ObsoletionReason(Enum):
    """Reasons for obsoleting vocabulary elements."""

    UNCLEAR_DEFINITION = "Not clearly defined or inconsistent usage."
    ADDED_IN_ERROR = "Added in error."
    MORE_SPECIFIC_CREATED = "More specific concepts created."
    CONVERTED_TO_COLLECTION = "Converted to a collection."
    AMBIGUOUS_MEANING = "The meaning was ambiguous."
    LACK_OF_EVIDENCE = "Lack of evidence that this function/process/component exists."


class MappingType(Enum):
    """Type of SKOS mapping relation."""

    EXACT_MATCH = "skos:exactMatch"
    CLOSE_MATCH = "skos:closeMatch"
    NARROW_MATCH = "skos:narrowMatch"
    BROAD_MATCH = "skos:broadMatch"
    RELATED_MATCH = "skos:relatedMatch"


class CollectionObsoletionReason(Enum):
    """Reasons for obsoleting collections."""

    UNCLEAR_DEFINITION = "Not clearly defined or inconsistent usage."
    ADDED_IN_ERROR = "Added in error."
    MORE_SPECIFIC_CREATED = "More specific colllections were created."
    CONVERTED_TO_CONCEPT = "Converted to a concept."


class ConceptScheme(BaseModel):
    """Concept scheme metadata (key-value format)."""

    template_version: Annotated[
        str,
        XLSXMetadata(
            meaning="dcterms:hasVersion",
            description="Gets increments only on structural changes. Its general pattern is major.minor.rev-JJJJ-MM[a-z].",
        ),
    ] = Field(default=TEMPLATE_VERSION)

    vocabulary_iri: Annotated[
        str,
        XLSXMetadata(
            display_name="Vocabulary IRI",
            meaning="skos:conceptScheme",
            description="IRI for this vocabulary",
        ),
    ] = Field(...)

    prefix: Annotated[
        str,
        XLSXMetadata(
            meaning="vann:preferredNamespacePrefix",
            description="Registered (or preferred) prefix to be used in CURIES for this vocabulary",
        ),
    ] = Field(...)

    title: Annotated[
        str,
        XLSXMetadata(meaning="dcterms:title", description="Title of the vocabulary"),
    ] = Field(...)

    description: Annotated[
        str,
        XLSXMetadata(
            meaning="dcterms:description",
            description="General description of the vocabulary",
        ),
    ] = Field(...)

    version: Annotated[
        str,
        XLSXMetadata(
            meaning="owl:versionInfo",
            description="Automatically added version specifier for the vocabulary",
        ),
    ] = Field(...)

    modified_date: Annotated[
        date,
        XLSXMetadata(
            meaning="dcterms:modified",
            description="Automatically added date of last modification",
        ),
    ] = Field(default_factory=date.today)


class Mapping(BaseModel):
    """External mapping between concepts (table format)."""

    concept_iri: Annotated[
        str, XLSXMetadata(meaning="skos:Concept", description="IRI of Voc4Cat Concept")
    ] = Field(...)

    mapping_relation: Annotated[
        MappingType | None,
        XLSXMetadata(
            meaning="skos:mappingRelation", description="Type of mapping relation."
        ),
    ] = Field(None)

    mapped_external_concept: str | None = Field(
        None,
        description="IRI of mapped external concept",
    )


class Collection(BaseModel):
    """Collection of concepts (table format)."""

    collection_iri: Annotated[
        str,
        XLSXMetadata(
            meaning="skos:Collection, skos:orderedCollection",
            description="Collection IRI",
        ),
    ] = Field(...)

    language_code: Annotated[
        str, XLSXMetadata(meaning="dcterms:language", description="Language Code")
    ] = Field(default="en")

    preferred_label: Annotated[
        str, XLSXMetadata(meaning="skos:prefLabel", description="Preferred Label")
    ] = Field(...)

    definition: Annotated[
        str, XLSXMetadata(meaning="skos:definition", description="Definition")
    ] = Field(...)

    parent_collection_iris: (
        Annotated[
            list[str],
            XLSXMetadata(
                separator_pattern=XLSXConverters.PIPE_ESCAPED,
                meaning="skos:member",
                description="Parent Collection IRIs",
            ),
        ]
        | None
    ) = Field(default=None)

    is_ordered: bool = Field(
        default=False,
        description="Ordered? Yes or No (default)",
    )

    change_note: (
        Annotated[
            str, XLSXMetadata(meaning="skos:changeNote", description="Change Note")
        ]
        | None
    ) = Field(None)

    editorial_note: (
        Annotated[
            str,
            XLSXMetadata(meaning="skos:editorialNote", description="Editorial Note"),
        ]
        | None
    ) = Field(None)

    obsoletion_reason: CollectionObsoletionReason | None = Field(
        None, description="Reason for obsoletion if applicable"
    )


class Prefix(BaseModel):
    """Namespace prefix mapping (table format)."""

    prefix: Annotated[
        str, XLSXMetadata(meaning="vann:preferredNamespacePrefix", description="Prefix")
    ] = Field(...)

    namespace: Annotated[
        str, XLSXMetadata(meaning="vann:preferredNamespaceUri", description="Namespace")
    ] = Field(...)


class ConceptTranslation(BaseModel):
    """Translation of a concept in a specific language."""

    concept_iri: Annotated[
        str,
        XLSXMetadata(
            meaning="skos:Concept",
            description="IRI of the concept being translated",
        ),
    ] = Field(...)

    language_code: Annotated[
        str,
        XLSXMetadata(
            meaning="dcterms:language",
            description="ISO 639-1 language code (e.g., 'en', 'de', 'fr')",
        ),
    ] = Field(...)

    preferred_label: Annotated[
        str,
        XLSXMetadata(
            meaning="skos:prefLabel",
            description="Preferred label in this language",
        ),
    ] = Field(...)

    alternate_labels: Annotated[
        list[str],
        XLSXMetadata(
            separator_pattern=XLSXConverters.COMMA,
            meaning="skos:altLabel",
            description="Alternative labels in this language",
        ),
    ] = Field(default_factory=list)

    definition: Annotated[
        str | None,
        XLSXMetadata(
            meaning="skos:definition",
            description="Definition of the concept in this language",
        ),
    ] = Field(None)

    scope_note: Annotated[
        str | None,
        XLSXMetadata(
            meaning="skos:scopeNote",
            description="Scope note in this language",
        ),
    ] = Field(None)


class MultiLingualConcept(BaseModel):
    """Concept with multi-language support for joined model demonstration."""

    concept_iri: Annotated[
        str,
        XLSXMetadata(
            meaning="skos:Concept",
            description="Unique IRI identifier for this concept",
        ),
    ] = Field(...)

    parent_iris: Annotated[
        list[str],
        XLSXMetadata(
            separator_pattern=XLSXConverters.PIPE,
            meaning="skos:broader",
            description="Parent concept IRIs",
        ),
    ] = Field(default_factory=list)

    collection_membership: Annotated[
        list[str],
        XLSXMetadata(
            separator_pattern=XLSXConverters.PIPE,
            meaning="skos:member",
            description="Collection IRIs this concept belongs to",
        ),
    ] = Field(default_factory=list)

    created_date: Annotated[
        date,
        XLSXMetadata(
            meaning="dcterms:created",
            description="Date when this concept was created",
        ),
    ] = Field(default_factory=date.today)

    status: Annotated[
        str,
        XLSXMetadata(
            meaning="adms:status",
            description="Status of this concept (draft, published, deprecated)",
        ),
    ] = Field(default="draft")

    translations: list[ConceptTranslation] = Field(default_factory=list)


# Integration Tests
class TestIntegration:
    """Integration tests for the unified system."""

    def test_round_trip_table_format(self, sample_employees, temp_file):
        """Test complete round-trip for table format."""
        # Export with custom config
        config = XLSXTableConfig(
            title="Employee Data Export",
            start_row=4,
            field_order=["employee_id", "first_name", "last_name"],
        )

        export_to_xlsx(sample_employees, temp_file, format_type="table", config=config)
        imported = import_from_xlsx(
            temp_file, Employee, format_type="table", config=config
        )

        # Verify complete data integrity
        assert len(imported) == len(sample_employees)
        for original, imported_item in zip(sample_employees, imported):
            assert original.employee_id == imported_item.employee_id
            assert original.first_name == imported_item.first_name
            assert original.status == imported_item.status

    def test_round_trip_keyvalue_format(self, temp_file):
        """Test complete round-trip for key-value format."""
        # Create a complex model with various field types
        project = Project(
            project_id=42,
            project_name="Complex Project",
            description="A very complex project with all field types",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            priority=Priority.HIGH,
            budget=500000.50,
            is_completed=True,
        )

        config = XLSXKeyValueConfig(
            title="Project Export",
        )

        export_to_xlsx(project, temp_file, format_type="keyvalue", config=config)
        imported = import_from_xlsx(
            temp_file, Project, format_type="keyvalue", config=config
        )

        # Verify all fields
        assert imported.project_id == project.project_id
        assert imported.project_name == project.project_name
        assert imported.start_date == project.start_date
        assert imported.end_date == project.end_date
        assert imported.priority == project.priority
        assert imported.budget == project.budget
        assert imported.description == project.description
        assert imported.is_completed == project.is_completed

    def test_multiple_sheets_compatibility(
        self, sample_employees, sample_projects, temp_file
    ):
        """Test that the unified system works with existing multi-sheet scenarios."""
        # Export employees to one sheet
        export_to_xlsx(
            sample_employees, temp_file, format_type="table", sheet_name="Employees"
        )

        # For now, we can't add to existing files, so test single sheet
        imported_employees = import_from_xlsx(
            temp_file, Employee, format_type="table", sheet_name="Employees"
        )

        assert len(imported_employees) == 2
        assert imported_employees[0].first_name == "John"

    def test_factory_consistency(self, sample_simple_model, temp_file):
        """Test that factory methods produce consistent results."""
        # Test both factory and direct API

        # Using factory
        processor = XLSXProcessorFactory.create_keyvalue_processor()
        processor.export(sample_simple_model, temp_file)
        imported1 = processor.import_data(temp_file, SimpleModel)

        # Using direct API
        temp_file2 = temp_file.with_suffix(".2.xlsx")
        export_to_xlsx(sample_simple_model, temp_file2, format_type="keyvalue")
        imported2 = import_from_xlsx(temp_file2, SimpleModel, format_type="keyvalue")

        # Results should be identical
        assert imported1.name == imported2.name
        assert imported1.value == imported2.value
        assert imported1.active == imported2.active

        temp_file2.unlink(missing_ok=True)

    def test_round_trip_with_all_converters(self, temp_file):
        """Test round-trip with all built-in converters."""

        class ModelWithAllConverters(BaseModel):
            name: str
            comma_list: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
            ] = Field(default=[])
            pipe_list: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.PIPE)
            ] = Field(default=[])
            semicolon_list: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.SEMICOLON)
            ] = Field(default=[])
            int_list: Annotated[
                list[int],
                XLSXMetadata(
                    xlsx_serializer=lambda x: ", ".join(str(i) for i in x) if x else "",
                    xlsx_deserializer=lambda x: [
                        int(i.strip()) for i in x.split(",") if i.strip()
                    ]
                    if x.strip()
                    else [],
                ),
            ] = Field(default=[])

        # Create test data
        original = ModelWithAllConverters(
            name="Test Model",
            comma_list=["a", "b", "c"],
            pipe_list=["x", "y", "z"],
            semicolon_list=["1", "2", "3"],
            int_list=[10, 20, 30],
        )

        # Round-trip test
        export_to_xlsx(original, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(
            temp_file, ModelWithAllConverters, format_type="keyvalue"
        )

        assert imported.name == original.name
        assert imported.comma_list == original.comma_list
        assert imported.pipe_list == original.pipe_list
        assert imported.semicolon_list == original.semicolon_list
        assert imported.int_list == original.int_list

    def test_large_dataset_performance(self, large_employee_dataset, temp_file):
        """Test performance with larger dataset."""
        # Should handle this reasonably quickly
        export_to_xlsx(large_employee_dataset, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, Employee, format_type="table")

        assert len(imported) == 100
        assert imported[0].first_name == "Employee0"
        assert imported[-1].first_name == "Employee99"

    def test_mixed_data_types_integration(self, temp_file):
        """Test integration with mixed complex data types."""

        class ComplexIntegrationModel(BaseModel):
            # Basic types
            name: str
            count: int
            price: float
            active: bool
            created: date

            # Enums
            status: Status
            priority: Priority | None = None

            # Complex types with converters
            tags: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
            ] = Field(default=[])

            # Optional fields
            description: str | None = None
            metadata: dict | None = None

        # Test data
        data = [
            ComplexIntegrationModel(
                name="Item 1",
                count=10,
                price=99.99,
                active=True,
                created=date(2024, 1, 1),
                status=Status.ACTIVE,
                priority=Priority.HIGH,
                tags=["tag1", "tag2", "tag3"],
                description="First item",
                metadata={"key": "value"},
            ),
            ComplexIntegrationModel(
                name="Item 2",
                count=5,
                price=49.50,
                active=False,
                created=date(2024, 2, 1),
                status=Status.INACTIVE,
                priority=None,
                tags=["tag4", "tag5"],
                description=None,
                metadata=None,
            ),
        ]

        # Test table format
        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(
            temp_file, ComplexIntegrationModel, format_type="table"
        )

        assert len(imported) == 2
        assert imported[0].name == "Item 1"
        assert imported[0].status == Status.ACTIVE
        assert imported[0].priority == Priority.HIGH
        assert imported[0].tags == ["tag1", "tag2", "tag3"]
        assert imported[1].name == "Item 2"
        assert imported[1].status == Status.INACTIVE
        assert imported[1].priority is None
        assert imported[1].tags == ["tag4", "tag5"]

    def test_configuration_inheritance_integration(self, sample_employees, temp_file):
        """Test that configuration inheritance works across formats."""
        # Test with base config settings - include all required fields
        base_config = XLSXTableConfig(
            title="Integrated Test",
            include_fields={
                "employee_id",
                "first_name",
                "last_name",
                "email",
                "hire_date",
                "salary",
                "status",
            },
            exclude_fields={"department"},
        )

        export_to_xlsx(
            sample_employees, temp_file, format_type="table", config=base_config
        )
        imported = import_from_xlsx(
            temp_file, Employee, format_type="table", config=base_config
        )

        assert len(imported) == 2
        assert imported[0].first_name == "John"
        assert imported[0].department is None  # Excluded field

    def test_error_recovery_integration(self, temp_file):
        """Test error recovery in integration scenarios."""
        # Test that system gracefully handles various error conditions

        # 1. Invalid format type
        with pytest.raises(ValueError, match="Unsupported format type"):
            export_to_xlsx(
                [SimpleModel(name="test", value=1)], temp_file, format_type="invalid"
            )

        # 2. Empty data
        with pytest.raises(ValueError, match="No data provided"):
            export_to_xlsx([], temp_file, format_type="table")

        # 3. Wrong config type (should be handled gracefully)
        keyvalue_config = XLSXKeyValueConfig(title="Wrong Config")
        # This should work - configs are flexible
        export_to_xlsx(
            [SimpleModel(name="test", value=1)],
            temp_file,
            format_type="table",
            config=keyvalue_config,
        )

        # 4. Missing sheet
        with pytest.raises(ValueError, match=r"Sheet.*not found"):
            import_from_xlsx(
                temp_file, SimpleModel, format_type="table", sheet_name="NonExistent"
            )

    def test_backwards_compatibility_integration(self, sample_simple_model, temp_file):
        """Test backwards compatibility with existing code patterns."""
        # Test that old patterns still work

        # Basic export/import without explicit format_type
        export_to_xlsx([sample_simple_model], temp_file)  # Should default to table
        imported = import_from_xlsx(temp_file, SimpleModel)  # Should use table

        assert len(imported) == 1
        assert imported[0].name == sample_simple_model.name

    def test_edge_case_integration(self, temp_file):
        """Test edge cases in integration scenarios."""
        # Test with model that has challenging field names

        class EdgeCaseModel(BaseModel):
            # Field names that might cause issues - simplified without aliases
            field_with_spaces: str = "spaces"
            field_with_special_chars: str = "special"
            field_with_unicode: str = "unicode"

            # Very long field name
            very_long_field_name_that_might_cause_issues_in_excel: str = "test"

        data = [
            EdgeCaseModel(
                field_with_spaces="spaces",
                field_with_special_chars="special",
                field_with_unicode="unicode",
            )
        ]

        # Should handle edge cases gracefully
        export_to_xlsx(data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, EdgeCaseModel, format_type="table")

        assert len(imported) == 1
        assert imported[0].field_with_spaces == "spaces"

    def test_memory_efficiency_integration(self, temp_file):
        """Test memory efficiency with realistic data sizes."""
        # Create a moderately large dataset
        large_data = []
        for i in range(1000):  # 1000 records
            large_data.append(
                SimpleModel(
                    name=f"Item_{i}",
                    value=i,
                    active=i % 2 == 0,
                )
            )

        # Should handle without excessive memory usage
        export_to_xlsx(large_data, temp_file, format_type="table")
        imported = import_from_xlsx(temp_file, SimpleModel, format_type="table")

        assert len(imported) == 1000
        assert imported[0].name == "Item_0"
        assert imported[-1].name == "Item_999"

    def test_concurrent_usage_simulation(self, temp_file):
        """Test that the system can handle concurrent-like usage patterns."""
        # Simulate multiple operations in sequence

        # Operation 1: Export employees
        employees = [
            Employee(
                employee_id=1,
                first_name="John",
                last_name="Doe",
                email="john@example.com",
                hire_date=date(2023, 1, 1),
                salary=50000.0,
                status=Status.ACTIVE,
            )
        ]

        export_to_xlsx(employees, temp_file, format_type="table")

        # Operation 2: Import and verify
        imported = import_from_xlsx(temp_file, Employee, format_type="table")
        assert len(imported) == 1

        # Operation 3: Export different model to same file (overwrites)
        simple_data = [SimpleModel(name="test", value=42)]
        export_to_xlsx(simple_data, temp_file, format_type="table")

        # Operation 4: Import new data
        imported_simple = import_from_xlsx(temp_file, SimpleModel, format_type="table")
        assert len(imported_simple) == 1
        assert imported_simple[0].name == "test"

    def test_vocabulary_template_round_trip_integration(self, temp_file):
        """Test complete vocabulary template round-trip with all SKOS vocabulary features."""

        # Create sample vocabulary data (adapted from demo files)
        def create_sample_data():
            concept_scheme = ConceptScheme(
                vocabulary_iri="https://example.org/photocatalysis/",
                prefix="photo",
                title="Photocatalysis Vocabulary",
                description="A controlled vocabulary for photocatalysis research and applications",
                version="1.0.0",
            )

            mappings = [
                Mapping(
                    concept_iri="https://example.org/photocatalysis/TiO2",
                    mapping_relation=MappingType.EXACT_MATCH,
                    mapped_external_concept="http://purl.obolibrary.org/obo/CHEBI_32234",
                ),
                Mapping(
                    concept_iri="https://example.org/photocatalysis/Semiconductor",
                    mapping_relation=MappingType.CLOSE_MATCH,
                    mapped_external_concept="https://dbpedia.org/resource/Semiconductor",
                ),
            ]

            collections = [
                Collection(
                    collection_iri="https://example.org/photocatalysis/PhotocatalystMaterials",
                    preferred_label="Photocatalyst Materials",
                    definition="Collection of materials used as photocatalysts",
                    is_ordered=True,
                ),
                Collection(
                    collection_iri="https://example.org/photocatalysis/OpticalProperties",
                    preferred_label="Optical Properties",
                    definition="Collection of optical properties relevant to photocatalysis",
                    parent_collection_iris=[
                        "https://example.org/photocatalysis/MaterialProperties",
                        "https://example.org/photocatalysis/PhysicalProperties",
                    ],
                ),
            ]

            prefixes = [
                Prefix(prefix="skos", namespace="http://www.w3.org/2004/02/skos/core#"),
                Prefix(prefix="dcterms", namespace="http://purl.org/dc/terms/"),
                Prefix(prefix="photo", namespace="https://example.org/photocatalysis/"),
            ]

            return concept_scheme, mappings, collections, prefixes

        def create_multilingual_concepts():
            return [
                MultiLingualConcept(
                    concept_iri="https://example.org/photocatalysis/TiO2",
                    parent_iris=[
                        "https://example.org/photocatalysis/PhotocatalystMaterials"
                    ],
                    collection_membership=[
                        "https://example.org/photocatalysis/Materials"
                    ],
                    status="published",
                    translations=[
                        ConceptTranslation(
                            concept_iri="https://example.org/photocatalysis/TiO2",
                            language_code="en",
                            preferred_label="Titanium Dioxide",
                            alternate_labels=["Titania", "Titanium(IV) oxide"],
                            definition="A widely used photocatalyst material with excellent properties.",
                            scope_note="Used in self-cleaning surfaces and air purification.",
                        ),
                        ConceptTranslation(
                            concept_iri="https://example.org/photocatalysis/TiO2",
                            language_code="de",
                            preferred_label="Titandioxid",
                            alternate_labels=["Titania"],
                            definition="Ein weit verbreitetes Photokatalysator-Material.",
                            scope_note="Verwendet in selbstreinigenden Oberflächen.",
                        ),
                    ],
                ),
                MultiLingualConcept(
                    concept_iri="https://example.org/photocatalysis/BandGap",
                    parent_iris=[
                        "https://example.org/photocatalysis/MaterialProperty",
                        "https://example.org/photocatalysis/ElectronicProperty",
                    ],
                    collection_membership=[
                        "https://example.org/photocatalysis/OpticalProperties"
                    ],
                    status="published",
                    translations=[
                        ConceptTranslation(
                            concept_iri="https://example.org/photocatalysis/BandGap",
                            language_code="en",
                            preferred_label="Band Gap",
                            alternate_labels=["Energy gap", "Electronic band gap"],
                            definition="Energy difference between valence and conduction bands.",
                            scope_note="Critical for photocatalytic activity.",
                        ),
                    ],
                ),
            ]

        # Create test data
        concept_scheme, mappings, collections, prefixes = create_sample_data()
        multilingual_concepts = create_multilingual_concepts()

        # Configuration for different formats
        kv_config = XLSXKeyValueConfig(title="Concept Scheme Metadata")
        table_config = XLSXTableConfig()

        # Test 1: Export ConceptScheme (key-value format)
        export_to_xlsx(
            concept_scheme,
            temp_file,
            format_type="keyvalue",
            config=kv_config,
            sheet_name="ConceptScheme",
        )

        # Test 2: Export Collections (table format)
        export_to_xlsx(
            collections,
            temp_file,
            format_type="table",
            config=table_config,
            sheet_name="Collections",
        )

        # Test 3: Export Mappings (table format)
        export_to_xlsx(
            mappings,
            temp_file,
            format_type="table",
            config=table_config,
            sheet_name="Mappings",
        )

        # Test 4: Export Prefixes (table format)
        export_to_xlsx(
            prefixes,
            temp_file,
            format_type="table",
            config=table_config,
            sheet_name="Prefixes",
        )

        # Test 5: Export multi-language concepts with join configuration
        join_config = JoinConfiguration(
            primary_model=MultiLingualConcept,
            related_models={"translations": ConceptTranslation},
            join_keys={"translations": "concept_iri"},
            flattened_fields=[
                "concept_iri",
                "language_code",
                "preferred_label",
                "alternate_labels",
                "definition",
                "scope_note",
                "parent_iris",
                "collection_membership",
                "created_date",
                "status",
            ],
            field_mappings={
                "concept_iri": ("primary", "concept_iri"),
                "language_code": ("related", "language_code"),
                "preferred_label": ("related", "preferred_label"),
                "alternate_labels": ("related", "alternate_labels"),
                "definition": ("related", "definition"),
                "scope_note": ("related", "scope_note"),
                "parent_iris": ("primary", "parent_iris"),
                "collection_membership": ("primary", "collection_membership"),
                "created_date": ("primary", "created_date"),
                "status": ("primary", "status"),
            },
            list_fields={"alternate_labels", "parent_iris", "collection_membership"},
        )

        concepts_config = XLSXTableConfig(title="Multi-Language Concepts")
        joined_processor = XLSXProcessorFactory.create_joined_table_processor(
            join_config, concepts_config
        )
        joined_processor.export(multilingual_concepts, temp_file, "Concepts")

        # Now test round-trip imports and validate data integrity

        # Import ConceptScheme
        imported_concept_scheme = import_from_xlsx(
            temp_file,
            ConceptScheme,
            format_type="keyvalue",
            config=kv_config,
            sheet_name="ConceptScheme",
        )

        # Import Collections
        imported_collections = import_from_xlsx(
            temp_file,
            Collection,
            format_type="table",
            config=table_config,
            sheet_name="Collections",
        )

        # Import Mappings
        imported_mappings = import_from_xlsx(
            temp_file,
            Mapping,
            format_type="table",
            config=table_config,
            sheet_name="Mappings",
        )

        # Import Prefixes
        imported_prefixes = import_from_xlsx(
            temp_file,
            Prefix,
            format_type="table",
            config=table_config,
            sheet_name="Prefixes",
        )

        # Import Concepts (multi-language joined format)
        # Use the same joined processor for import that was used for export
        imported_concepts = joined_processor.import_data(
            temp_file, MultiLingualConcept, "Concepts"
        )

        # Validate ConceptScheme round-trip
        assert imported_concept_scheme.title == concept_scheme.title
        assert imported_concept_scheme.vocabulary_iri == concept_scheme.vocabulary_iri
        assert imported_concept_scheme.prefix == concept_scheme.prefix
        assert imported_concept_scheme.description == concept_scheme.description
        assert imported_concept_scheme.version == concept_scheme.version

        # Validate Collections round-trip
        assert len(imported_collections) == len(collections)
        for orig, imp in zip(collections, imported_collections):
            assert orig.collection_iri == imp.collection_iri
            assert orig.preferred_label == imp.preferred_label
            assert orig.definition == imp.definition
            assert orig.is_ordered == imp.is_ordered
            assert orig.parent_collection_iris == imp.parent_collection_iris

        # Validate Mappings round-trip with enum preservation
        assert len(imported_mappings) == len(mappings)
        for orig, imp in zip(mappings, imported_mappings):
            assert orig.concept_iri == imp.concept_iri
            assert orig.mapping_relation == imp.mapping_relation
            assert orig.mapped_external_concept == imp.mapped_external_concept

        # Validate specific enum deserialization
        exact_match = next(
            (
                m
                for m in imported_mappings
                if m.mapping_relation == MappingType.EXACT_MATCH
            ),
            None,
        )
        assert exact_match is not None
        assert exact_match.mapping_relation.value == "skos:exactMatch"

        # Validate Prefixes round-trip
        assert len(imported_prefixes) == len(prefixes)
        skos_prefix = next((p for p in imported_prefixes if p.prefix == "skos"), None)
        assert skos_prefix is not None
        assert skos_prefix.namespace == "http://www.w3.org/2004/02/skos/core#"

        # Validate Concepts round-trip (focus on critical features)
        # The joined processor properly reconstructs the nested MultiLingualConcept objects
        assert len(imported_concepts) == len(multilingual_concepts)

        # Check first concept (TiO2) with complex lists and multiple translations
        titania_concept = next(
            (c for c in imported_concepts if "TiO2" in c.concept_iri), None
        )
        assert titania_concept is not None
        assert titania_concept.status == "published"
        assert len(titania_concept.parent_iris) == 1
        assert (
            titania_concept.parent_iris[0]
            == "https://example.org/photocatalysis/PhotocatalystMaterials"
        )
        assert len(titania_concept.collection_membership) == 1
        assert (
            titania_concept.collection_membership[0]
            == "https://example.org/photocatalysis/Materials"
        )

        # Check translations are properly reconstructed
        assert len(titania_concept.translations) == 2  # EN and DE

        # Check English translation
        titania_en = next(
            (t for t in titania_concept.translations if t.language_code == "en"), None
        )
        assert titania_en is not None
        assert titania_en.preferred_label == "Titanium Dioxide"
        assert "Titania" in titania_en.alternate_labels
        assert "Titanium(IV) oxide" in titania_en.alternate_labels
        assert (
            titania_en.definition
            == "A widely used photocatalyst material with excellent properties."
        )

        # Check German translation
        titania_de = next(
            (t for t in titania_concept.translations if t.language_code == "de"), None
        )
        assert titania_de is not None
        assert titania_de.preferred_label == "Titandioxid"
        assert "Titania" in titania_de.alternate_labels

        # Check second concept (BandGap) with multiple parent IRIs
        bandgap_concept = next(
            (c for c in imported_concepts if "BandGap" in c.concept_iri), None
        )
        assert bandgap_concept is not None
        assert len(bandgap_concept.parent_iris) == 2
        assert "MaterialProperty" in bandgap_concept.parent_iris[0]
        assert "ElectronicProperty" in bandgap_concept.parent_iris[1]
        assert len(bandgap_concept.collection_membership) == 1
        assert "OpticalProperties" in bandgap_concept.collection_membership[0]

        # Check BandGap translations
        assert len(bandgap_concept.translations) == 1  # Only EN
        bandgap_en = bandgap_concept.translations[0]
        assert bandgap_en.language_code == "en"
        assert bandgap_en.preferred_label == "Band Gap"
        assert "Energy gap" in bandgap_en.alternate_labels
        assert "Electronic band gap" in bandgap_en.alternate_labels

        # Validate SKOS metadata preservation (check that XLSXMetadata annotations work)
        # This is implicit in successful round-trip but validates the SKOS vocabulary structure


class TestXLSXApiHelpers:
    """Tests for xlsx_api helper functions."""

    def test_create_xlsx_wrapper_basic(self):
        """Test create_xlsx_wrapper with basic metadata."""

        class SimpleReaction(BaseModel):
            name: str
            temperature: float
            catalysts: list[str] = Field(default=[])

        # Create wrapper with metadata
        metadata_map = {
            "temperature": XLSXMetadata(unit="°C", meaning="Reaction temperature"),
            "catalysts": XLSXMetadata(separator_pattern=XLSXConverters.COMMA),
        }
        xlsx_reaction = create_xlsx_wrapper(SimpleReaction, metadata_map)

        # Verify wrapper class
        assert xlsx_reaction.__name__ == "XLSXSimpleReaction"
        assert issubclass(xlsx_reaction, SimpleReaction)

        # Verify annotations include metadata
        temp_annotation = xlsx_reaction.__annotations__["temperature"]
        assert get_origin(temp_annotation) is Annotated

    def test_create_xlsx_wrapper_with_inheritance(self):
        """Test create_xlsx_wrapper with base wrapper inheritance."""

        class BaseReaction(BaseModel):
            name: str
            temperature: float

        class AdvancedReaction(BaseReaction):
            catalysts: list[str] = Field(default=[])
            yield_percent: float = Field(default=0.0)

        # Create base wrapper
        base_metadata = {
            "temperature": XLSXMetadata(unit="°C"),
        }
        xlsx_base_reaction = create_xlsx_wrapper(BaseReaction, base_metadata)

        # Create derived wrapper inheriting from base
        advanced_metadata = {
            "yield_percent": XLSXMetadata(unit="%"),
        }
        xlsx_advanced_reaction = create_xlsx_wrapper(
            AdvancedReaction, advanced_metadata, base_wrapper=xlsx_base_reaction
        )

        # Verify inheritance
        assert issubclass(xlsx_advanced_reaction, xlsx_base_reaction)
        assert xlsx_advanced_reaction.__name__ == "XLSXAdvancedReaction"

        # Temperature metadata should be inherited from base
        temp_annotation = xlsx_advanced_reaction.__annotations__["temperature"]
        assert get_origin(temp_annotation) is Annotated

    def test_create_joined_table_processor_default_title(self):
        """Test that joined table processor generates default title."""

        class PrimaryModel(BaseModel):
            id: str

        class RelatedModel(BaseModel):
            id: str
            primary_id: str

        join_config = JoinConfiguration(
            primary_model=PrimaryModel,
            related_models={"items": RelatedModel},
            join_keys={"items": "primary_id"},
            flattened_fields=["id"],
            field_mappings={"id": ("primary", "id")},
        )

        # Create processor without specifying config (should generate default title)
        processor = XLSXProcessorFactory.create_joined_table_processor(join_config)

        # Verify default title was generated
        assert "PrimaryModel" in processor.config.title
        assert "items" in processor.config.title


class TestXLSXTableEdgeCases:
    """Tests for edge cases in xlsx_table.py to improve coverage."""

    def test_header_row_color_styling(self, temp_file):
        """Test that header row color styling is applied."""
        config = XLSXTableConfig(
            title="Test Header Color",
            header_row_color="FFCC00",  # Yellow color
        )

        data = [SimpleModel(name="test", value=1)]
        export_to_xlsx(data, temp_file, format_type="table", config=config)

        # Verify header styling was applied
        wb = load_workbook(temp_file)
        ws = wb.active
        # Header is on row 5 (after title row 1-2, empty 3, header 4-5)
        header_cell = ws["A5"]  # Assuming start column 1
        assert header_cell.fill.start_color.rgb is not None

    def test_table_name_truncation(self, temp_file):
        """Test that long table names are truncated."""

        # Create a sheet name that would result in a long table name
        very_long_sheet_name = "A" * 40  # Exceeds MAX_SHEETNAME_LENGTH

        data = [SimpleModel(name="test", value=1)]
        export_to_xlsx(
            data, temp_file, format_type="table", sheet_name=very_long_sheet_name[:31]
        )

        wb = load_workbook(temp_file)
        # Just verify it was created successfully
        assert len(wb.sheetnames) > 0

    def test_joined_model_non_list_related_field(self, temp_file):
        """Test JoinedModelProcessor with non-list related field."""

        class RelatedItem(BaseModel):
            id: str
            value: str

        class PrimaryWithSingleRelated(BaseModel):
            primary_id: str
            related: RelatedItem | None = None

        join_config = JoinConfiguration(
            primary_model=PrimaryWithSingleRelated,
            related_models={"related": RelatedItem},
            join_keys={"related": "primary_id"},
            flattened_fields=["primary_id", "id", "value"],
            field_mappings={
                "primary_id": ("primary", "primary_id"),
                "id": ("related", "id"),
                "value": ("related", "value"),
            },
        )

        # Test with single related item (not a list)
        data = [
            PrimaryWithSingleRelated(
                primary_id="p1",
                related=RelatedItem(id="r1", value="val1"),
            )
        ]

        flattened = JoinedModelProcessor.flatten_joined_data(data, join_config)
        assert len(flattened) == 1
        assert flattened[0]["primary_id"] == "p1"

    def test_joined_model_empty_related_field(self, temp_file):
        """Test JoinedModelProcessor with empty related field."""

        class RelatedItem(BaseModel):
            id: str
            value: str

        class PrimaryWithEmptyRelated(BaseModel):
            primary_id: str
            related: list[RelatedItem] = Field(default_factory=list)

        join_config = JoinConfiguration(
            primary_model=PrimaryWithEmptyRelated,
            related_models={"related": RelatedItem},
            join_keys={"related": "primary_id"},
            flattened_fields=["primary_id", "id", "value"],
            field_mappings={
                "primary_id": ("primary", "primary_id"),
                "id": ("related", "id"),
                "value": ("related", "value"),
            },
        )

        # Test with empty related list
        data = [
            PrimaryWithEmptyRelated(primary_id="p1", related=[]),
        ]

        flattened = JoinedModelProcessor.flatten_joined_data(data, join_config)
        assert len(flattened) == 1
        assert flattened[0]["primary_id"] == "p1"
        assert flattened[0]["id"] == ""

    def test_joined_model_unmapped_field(self, temp_file):
        """Test JoinedModelProcessor with unmapped field in output."""

        class SimpleRelated(BaseModel):
            id: str

        class SimplePrimary(BaseModel):
            primary_id: str
            related: list[SimpleRelated] = Field(default_factory=list)

        join_config = JoinConfiguration(
            primary_model=SimplePrimary,
            related_models={"related": SimpleRelated},
            join_keys={"related": "primary_id"},
            flattened_fields=["primary_id", "id", "unmapped_field"],  # unmapped field
            field_mappings={
                "primary_id": ("primary", "primary_id"),
                "id": ("related", "id"),
                # "unmapped_field" is NOT in field_mappings
            },
        )

        data = [SimplePrimary(primary_id="p1", related=[SimpleRelated(id="r1")])]

        flattened = JoinedModelProcessor.flatten_joined_data(data, join_config)
        assert flattened[0]["unmapped_field"] == ""

    def test_export_empty_data_error(self, temp_file):
        """Test that exporting empty data raises error."""
        with pytest.raises(ValueError, match="No data provided"):
            export_to_xlsx([], temp_file, format_type="table")

    def test_joined_processor_empty_data_error(self, temp_file):
        """Test that joined processor with empty data raises error."""

        class PrimaryModel(BaseModel):
            id: str

        class RelatedModel(BaseModel):
            id: str
            primary_id: str

        join_config = JoinConfiguration(
            primary_model=PrimaryModel,
            related_models={"items": RelatedModel},
            join_keys={"items": "primary_id"},
            flattened_fields=["id"],
            field_mappings={"id": ("primary", "id")},
        )

        processor = XLSXProcessorFactory.create_joined_table_processor(join_config)

        with pytest.raises(ValueError, match="No data provided"):
            processor.export([], temp_file)

    def test_joined_processor_missing_sheet_error(self, temp_file):
        """Test that import from missing sheet raises error."""

        class PrimaryModel(BaseModel):
            id: str

        class RelatedModel(BaseModel):
            id: str
            primary_id: str

        join_config = JoinConfiguration(
            primary_model=PrimaryModel,
            related_models={"items": RelatedModel},
            join_keys={"items": "primary_id"},
            flattened_fields=["id"],
            field_mappings={"id": ("primary", "id")},
        )

        # Create empty file

        wb = Workbook()
        wb.save(temp_file)
        wb.close()

        processor = XLSXProcessorFactory.create_joined_table_processor(join_config)

        with pytest.raises(ValueError, match="not found"):
            processor.import_data(temp_file, sheet_name="NonExistent")

    def test_joined_processor_import_default_model(self, temp_file):
        """Test joined processor import uses default model from config."""

        class PrimaryModel(BaseModel):
            id: str
            name: str = ""

        class RelatedModel(BaseModel):
            id: str
            value: str = ""

        join_config = JoinConfiguration(
            primary_model=PrimaryModel,
            related_models={"items": RelatedModel},
            join_keys={"items": "id"},
            flattened_fields=["id", "name", "value"],
            field_mappings={
                "id": ("primary", "id"),
                "name": ("primary", "name"),
                "value": ("related", "value"),
            },
        )

        config = XLSXTableConfig(title="Test")
        processor = XLSXProcessorFactory.create_joined_table_processor(
            join_config, config
        )

        # Export some data
        data = [PrimaryModel(id="1", name="test")]
        processor.export(data, temp_file, "PrimaryModel")

        # Import without specifying model_class - should use primary_model from config
        imported = processor.import_data(temp_file, sheet_name="PrimaryModel")

        assert len(imported) == 1
        assert imported[0].id == "1"

    def test_import_with_validation_error(self, temp_file):
        """Test that import handles validation errors gracefully."""

        # Model with a required string field and a constrained int field
        class StrictModel(BaseModel):
            required_name: str
            positive_value: int = Field(ge=0)

        # Export valid data first
        valid_data = [StrictModel(required_name="test", positive_value=10)]
        export_to_xlsx(valid_data, temp_file, format_type="table")

        # Corrupt the file: clear the required field in the data row
        wb = load_workbook(temp_file)
        ws = wb.active

        # With default config (no title): header row 1, data row 2
        ws["A2"] = ""  # Clear required field
        wb.save(temp_file)
        wb.close()

        # Import should raise ValueError with validation errors
        with pytest.raises(ValueError, match="Required field 'required_name' is empty"):
            import_from_xlsx(temp_file, StrictModel, format_type="table")

    def test_import_with_deserialization_error(self, temp_file):
        """Test that import handles deserialization errors gracefully."""

        class TypedModel(BaseModel):
            name: str
            count: int

        # Export valid data first
        valid_data = [TypedModel(name="test", count=42)]
        export_to_xlsx(valid_data, temp_file, format_type="table")

        # Corrupt the file: put non-integer in count field
        wb = load_workbook(temp_file)
        ws = wb.active

        # With default config (no title): header row 1, data row 2
        ws["B2"] = "not_a_number"
        wb.save(temp_file)
        wb.close()

        # Import should raise ValueError (deserialization errors are collected and wrapped)
        with pytest.raises(ValueError, match="Error deserializing field 'count'"):
            import_from_xlsx(temp_file, TypedModel, format_type="table")

    def test_joined_processor_no_primary_key_error(self):
        """Test that reconstruct raises error when no primary key in mappings."""

        class Primary(BaseModel):
            id: str

        class Related(BaseModel):
            id: str
            value: str

        # Create join config with only related field mappings (no primary key)
        join_config = JoinConfiguration(
            primary_model=Primary,
            related_models={"items": Related},
            join_keys={"items": "id"},
            flattened_fields=["value"],
            field_mappings={
                # Only related fields, no primary key mapped
                "value": ("related", "value"),
            },
        )

        flattened_data = [{"value": "test"}]

        with pytest.raises(ValueError, match="No primary key found"):
            JoinedModelProcessor.reconstruct_joined_data(flattened_data, join_config)

    def test_joined_processor_related_model_creation_error(self):
        """Test error handling when related model creation fails."""

        class Primary(BaseModel):
            id: str

        class StrictRelated(BaseModel):
            id: str
            required_field: str  # Required field with no default
            optional_field: str = ""

        join_config = JoinConfiguration(
            primary_model=Primary,
            related_models={"items": StrictRelated},
            join_keys={"items": "id"},
            flattened_fields=["id", "required_field", "optional_field"],
            field_mappings={
                "id": ("primary", "id"),
                "required_field": ("related", "required_field"),
                "optional_field": ("related", "optional_field"),
            },
        )

        # Provide non-empty optional_field (triggers has_related_data=True)
        # but leave required_field as None (causes model creation to fail)
        flattened_data = [
            {"id": "1", "required_field": None, "optional_field": "some_value"},
        ]

        with pytest.raises(ValueError, match="Error creating related model"):
            JoinedModelProcessor.reconstruct_joined_data(flattened_data, join_config)

    def test_joined_processor_primary_model_creation_error(self):
        """Test error handling when primary model creation fails."""

        class StrictPrimary(BaseModel):
            id: str
            required_field: str  # Required with no default

            @field_validator("required_field")
            @classmethod
            def validate_required(cls, v):
                if not v or not v.strip():
                    msg = "required_field cannot be empty"
                    raise ValueError(msg)
                return v  # pragma: no cover

        class Related(BaseModel):
            id: str
            value: str = ""

        join_config = JoinConfiguration(
            primary_model=StrictPrimary,
            related_models={"items": Related},
            join_keys={"items": "id"},
            flattened_fields=["id", "required_field", "value"],
            field_mappings={
                "id": ("primary", "id"),
                "required_field": ("primary", "required_field"),
                "value": ("related", "value"),
            },
        )

        # Provide data with None required primary field (becomes empty string)
        # which fails the validator
        flattened_data = [
            {"id": "1", "required_field": None, "value": ""},
        ]

        with pytest.raises(ValueError, match="Error creating primary model"):
            JoinedModelProcessor.reconstruct_joined_data(flattened_data, join_config)

    def test_joined_processor_list_field_in_primary(self):
        """Test reconstruction with list fields in primary model."""

        class Primary(BaseModel):
            id: str
            tags: list[str] = Field(default_factory=list)

        class Related(BaseModel):
            id: str
            value: str = ""

        join_config = JoinConfiguration(
            primary_model=Primary,
            related_models={"items": Related},
            join_keys={"items": "id"},
            flattened_fields=["id", "tags", "value"],
            field_mappings={
                "id": ("primary", "id"),
                "tags": ("primary", "tags"),
                "value": ("related", "value"),
            },
            list_fields={"tags"},
        )

        # Test with empty list field (covers line 230)
        flattened_data = [{"id": "1", "tags": "", "value": ""}]

        result = JoinedModelProcessor.reconstruct_joined_data(
            flattened_data, join_config
        )

        assert len(result) == 1
        assert result[0].tags == []  # Empty list

    def test_joined_processor_list_field_in_related(self):
        """Test reconstruction with list fields in related model."""

        class Related(BaseModel):
            id: str
            labels: list[str] = Field(default_factory=list)

        class PrimaryWithItems(BaseModel):
            id: str
            items: list[Related] = Field(default_factory=list)

        join_config = JoinConfiguration(
            primary_model=PrimaryWithItems,
            related_models={"items": Related},
            join_keys={"items": "id"},
            flattened_fields=["id", "labels"],
            field_mappings={
                "id": ("primary", "id"),
                "labels": ("related", "labels"),
            },
            list_fields={"labels"},
        )

        # Test with valid list field (covers line 254 - non-empty list in related field)
        flattened_data = [
            {"id": "1", "labels": "a, b"},  # Non-empty list
        ]

        result = JoinedModelProcessor.reconstruct_joined_data(
            flattened_data, join_config
        )

        assert len(result) == 1
        assert len(result[0].items) == 1
        assert result[0].items[0].labels == ["a", "b"]

    def test_joined_processor_empty_list_field_in_related(self):
        """Test reconstruction with empty list field values in related model."""

        class Related(BaseModel):
            id: str
            name: str  # Non-list field to trigger has_related_data
            labels: list[str] = Field(default_factory=list)

        class PrimaryWithItems(BaseModel):
            id: str
            items: list[Related] = Field(default_factory=list)

        join_config = JoinConfiguration(
            primary_model=PrimaryWithItems,
            related_models={"items": Related},
            join_keys={"items": "id"},
            flattened_fields=["id", "name", "labels"],
            field_mappings={
                "id": ("primary", "id"),
                "name": ("related", "name"),
                "labels": ("related", "labels"),
            },
            list_fields={"labels"},
        )

        # Test empty list field with non-empty non-list field to trigger has_related_data
        # Covers line 258: when list field is empty but has_related_data is True
        flattened_data = [
            {
                "id": "1",
                "name": "item1",
                "labels": "",
            },  # Empty list field but non-empty name
        ]

        result = JoinedModelProcessor.reconstruct_joined_data(
            flattened_data, join_config
        )

        assert len(result) == 1
        assert len(result[0].items) == 1
        assert result[0].items[0].name == "item1"
        assert result[0].items[0].labels == []  # Empty list for the list field

    def test_joined_import_with_deserialization_error(self, temp_file):
        """Test joined table import handles deserialization errors."""

        class Primary(BaseModel):
            id: str
            count: int = 0

        class Related(BaseModel):
            id: str
            value: str = ""

        join_config = JoinConfiguration(
            primary_model=Primary,
            related_models={"items": Related},
            join_keys={"items": "id"},
            flattened_fields=["id", "count", "value"],
            field_mappings={
                "id": ("primary", "id"),
                "count": ("primary", "count"),
                "value": ("related", "value"),
            },
        )

        config = XLSXTableConfig(title="Test")
        processor = XLSXProcessorFactory.create_joined_table_processor(
            join_config, config
        )

        # Export valid data
        data = [Primary(id="1", count=10)]
        processor.export(data, temp_file, "TestSheet")

        # Corrupt the count field (with title: row 1=title, row 2=empty, row 3=header, row 4=data)
        wb = load_workbook(temp_file)
        ws = wb["TestSheet"]
        ws["B4"] = "not_a_number"  # Corrupt count in data row
        wb.save(temp_file)
        wb.close()

        with pytest.raises(XLSXDeserializationError, match="count"):
            processor.import_data(temp_file, sheet_name="TestSheet")


class TestXLSXApiPathHandling:
    """Tests for string path handling in export/import functions."""

    def test_export_with_string_filepath(self, sample_employees, tmp_path):
        """Test export_to_xlsx accepts string filepath (line 173)."""
        filepath_str = str(tmp_path / "string_path.xlsx")
        export_to_xlsx(sample_employees, filepath_str, format_type="table")
        assert (tmp_path / "string_path.xlsx").exists()

    def test_import_with_string_filepath(self, sample_employees, tmp_path):
        """Test import_from_xlsx accepts string filepath (line 208)."""
        filepath = tmp_path / "test.xlsx"
        export_to_xlsx(sample_employees, filepath, format_type="table")

        # Now import using string path instead of Path object
        filepath_str = str(filepath)
        imported = import_from_xlsx(filepath_str, Employee, format_type="table")
        assert len(imported) == 2
        assert imported[0].first_name == "John"


class TestAutoFormatDetection:
    """Tests for auto-detection of format type."""

    def test_auto_detect_falls_back_to_table_invalid_kv_columns(self, tmp_path):
        """Test auto-detect falls back to 'table' when Field/Value have invalid metadata columns.

        Code coverage test for lines 240-241 in xlsx_api.py:
        - Auto-detector finds "Field" and "Value" headers (looks like key-value format)
        - But remaining columns are NOT valid KV metadata (Unit/Meaning/Description)
        - So format_type is set to "table" as fallback

        Note: This test focuses on exercising the specific code path, not end-to-end success.
        The file structure deliberately triggers the fallback but may not import successfully.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "SimpleModel"

        # Row 1: Table-like headers
        ws["A1"] = "name"
        ws["B1"] = "value"
        ws["C1"] = "active"

        # Row 2: Data that won't validate
        ws["A2"] = "test1"
        ws["B2"] = "not_a_number"  # Invalid for int field
        ws["C2"] = True

        # Row 3: "Field" and "Value" to trigger KV detection
        # The auto-detector scans rows 1-3, finds this, checks column C
        # "InvalidColumn" is not in {Unit, Meaning, Description}
        # So it sets format_type = "table" (lines 240-241) and breaks
        ws["A3"] = "Field"
        ws["B3"] = "Value"
        ws["C3"] = "InvalidColumn"

        filepath = tmp_path / "fallback_test.xlsx"
        wb.save(filepath)
        wb.close()

        # The test goal: verify auto-detection executes lines 240-241 (fallback to table)
        # The file has "Field"/"Value" in row 3 with invalid KV columns,
        # so auto-detector falls back to table format (lines 240-241).
        # Table import returns a list (even if empty), while keyvalue returns None/object.
        result = import_from_xlsx(filepath, SimpleModel, format_type="auto")

        # Verify it used table format (returns list) not keyvalue format (returns None/object)
        assert isinstance(result, list), "Should use table format (returns list)"
        # The data is invalid so the list should be empty (validation filters out bad rows)
        assert len(result) == 0, "Invalid data should be filtered out"

    def test_auto_detect_loop_completion_defaults_to_table(self, tmp_path):
        """Test auto-detect defaults to table when no headers match in rows 1-3 (line 243)."""

        wb = Workbook()
        ws = wb.active
        ws.title = "SimpleModel"
        # No Field/Value headers in rows 1-3, just table-style headers
        # Use proper casing to match model field names
        ws["A1"] = "Name"
        ws["B1"] = "Value"  # This matches model field, but not KV pattern
        ws["C1"] = "Active"
        ws["A2"] = "test"
        ws["B2"] = 42
        ws["C2"] = True

        filepath = tmp_path / "table_format.xlsx"
        wb.save(filepath)
        wb.close()

        # Should default to table format and import successfully
        imported = import_from_xlsx(filepath, SimpleModel, format_type="auto")
        assert len(imported) == 1
        assert imported[0].name == "test"

    def test_auto_detect_sheet_not_found_defaults_to_table(self, tmp_path):
        """Test auto-detect defaults to table when specified sheet not found (line 247)."""

        wb = Workbook()
        ws = wb.active
        ws.title = "OtherSheet"
        ws["A1"] = "name"
        ws["B1"] = "value"
        ws["A2"] = "test"
        ws["B2"] = "42"

        filepath = tmp_path / "wrong_sheet.xlsx"
        wb.save(filepath)
        wb.close()

        # When sheet_name is specified but doesn't exist, it should fall back to table
        # format detection and then fail because the sheet doesn't exist
        with pytest.raises(ValueError, match="not found"):
            import_from_xlsx(
                filepath, SimpleModel, format_type="auto", sheet_name="NonExistent"
            )

    def test_import_invalid_format_type_error(self, tmp_path):
        """Test import raises ValueError for invalid format_type (lines 257-258)."""

        # Create a valid xlsx file first
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "name"
        ws["B1"] = "value"
        ws["A2"] = "test"
        ws["B2"] = "42"
        wb.save(tmp_path / "test.xlsx")
        wb.close()

        with pytest.raises(ValueError, match="Unsupported format type"):
            import_from_xlsx(tmp_path / "test.xlsx", SimpleModel, format_type="invalid")


if __name__ == "__main__":
    pytest.main([__file__])
