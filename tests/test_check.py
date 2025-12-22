import logging
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tests.test_cli import (
    CS_CYCLES,
)
from voc4cat.checks import Voc4catError
from voc4cat.cli import main_cli
from voc4cat.utils import ConversionError

CS_SIMPLE_TURTLE = "concept-scheme-simple.ttl"


@pytest.mark.parametrize(
    ("test_file", "msg"),
    [
        (CS_CYCLES, "xlsx check passed"),
    ],
    ids=["no error"],
)
def test_check_xlsx(tmp_path, caplog, test_file, msg, cs_cycles_xlsx):
    dst = tmp_path / test_file
    shutil.copy(cs_cycles_xlsx, dst)
    with caplog.at_level(logging.INFO):
        main_cli(["check", "--inplace", str(dst)])
    assert msg in caplog.text


def test_check_xlsx_with_cell_coloring(tmp_path, caplog, cs_duplicates_xlsx):
    """Test that erroneous cells get colored when duplicates are found."""
    test_file = "concept-scheme-duplicates.xlsx"
    dst = tmp_path / test_file
    shutil.copy(cs_duplicates_xlsx, dst)

    with caplog.at_level(logging.ERROR):
        main_cli(["check", "--inplace", str(dst)])

    # Check that the error was logged (IRI may vary depending on the source TTL)
    assert "Same Concept IRI" in caplog.text
    assert 'used more than once for language "en"' in caplog.text

    # Check that cells were colored orange
    wb = load_workbook(dst)
    ws = wb["Concepts"]

    # Expected orange color from check.py: PatternFill("solid", start_color="00FFCC00")
    expected_color = "00FFCC00"

    # In v1.0 template, data starts at row 6. Duplicate rows are 6 and 7.
    # Column A = Concept IRI, Column B = Language Code
    assert ws["A6"].fill.start_color.rgb == expected_color, (
        "First duplicate IRI cell should be colored orange"
    )
    assert ws["B6"].fill.start_color.rgb == expected_color, (
        "First duplicate language cell should be colored orange"
    )
    assert ws["A7"].fill.start_color.rgb == expected_color, (
        "Second duplicate IRI cell should be colored orange"
    )
    assert ws["B7"].fill.start_color.rgb == expected_color, (
        "Second duplicate language cell should be colored orange"
    )

    wb.close()


def test_check_skos_rdf(datadir, tmp_path, caplog):
    shutil.copy(datadir / CS_SIMPLE_TURTLE, tmp_path)
    with caplog.at_level(logging.INFO):
        main_cli(["check", str(tmp_path / CS_SIMPLE_TURTLE)])
    assert "The file is valid according to the vp4cat-5.2 profile." in caplog.text


def test_check_skos_badfile(monkeypatch, datadir, tmp_path, temp_config, caplog):
    """Check failing profile validation."""
    # Load/prepare an a config with required prefix definition
    config = temp_config
    vocab_name = "concept-scheme-badfile"
    config.CURIES_CONVERTER_MAP[vocab_name] = config.curies_converter
    with caplog.at_level(logging.ERROR), pytest.raises(ConversionError):
        main_cli(["check", str(Path(datadir / vocab_name).with_suffix(".ttl"))])
    assert "VIOLATION: Validation Result in MinCountConstraintComponent" in caplog.text


def test_check_list_profiles(capsys):
    main_cli(["check", "--listprofiles"])
    captured = capsys.readouterr()
    assert "Known profiles:" in captured.out
    assert "bundled\tvocpub-4.7" in captured.out


def test_check_list_profiles_with_config(datadir, capsys, temp_config):
    """Test --listprofiles shows custom profiles when config is provided."""
    main_cli(
        [
            "check",
            "--config",
            str(datadir / "idranges_with_scheme.toml"),
            "--listprofiles",
        ]
    )
    captured = capsys.readouterr()
    assert "Known profiles:" in captured.out
    assert "bundled\tvocpub-4.7" in captured.out
    assert "config\tmyvocab" in captured.out
    assert "custom_profile.ttl" in captured.out


def test_check_missing_vocab(capsys):
    with pytest.raises(SystemExit) as exc_info:
        main_cli(["check"])
    assert exc_info.value.code == 2
    captured = capsys.readouterr()
    assert "usage: voc4cat check" in captured.out
    assert "--listprofiles" in captured.out


def test_check_ci_pre_one_folder(tmp_path, caplog):
    with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
        main_cli(["check", "-v", "--ci-pre", str(tmp_path)])
    assert "Need two dirs for ci_pre!" in caplog.text


def test_check_ci_post_one_folder(tmp_path, caplog):
    with caplog.at_level(logging.ERROR), pytest.raises(Voc4catError):
        main_cli(["check", "-v", "--ci-post", str(tmp_path)])
    assert "Need two dirs for ci_post!" in caplog.text


def test_check_xlsx_no_inplace_warning(tmp_path, caplog, cs_cycles_xlsx):
    """Test warning when xlsx check would overwrite file without --inplace."""
    dst = tmp_path / CS_CYCLES
    shutil.copy(cs_cycles_xlsx, dst)

    with caplog.at_level(logging.WARNING):
        main_cli(["check", str(dst)])  # No --inplace, no --outdir

    assert "This command will overwrite the existing file" in caplog.text
    assert '--inplace"' in caplog.text or "--inplace" in caplog.text


def test_check_ci_pre_valid_directories(datadir, tmp_path, caplog):
    """Test ci_pre check with valid inbox and vocab directories."""
    inbox_dir = tmp_path / "inbox"
    vocab_dir = tmp_path / "vocab"
    inbox_dir.mkdir()
    vocab_dir.mkdir()

    # Copy a TTL file to vocab_dir (ci_pre checks vocab files exist)
    shutil.copy(datadir / "concept-scheme-simple.ttl", vocab_dir)

    with caplog.at_level(logging.INFO):
        main_cli(["check", "--ci-pre", str(inbox_dir), str(vocab_dir)])

    assert "ci-pre passed" in caplog.text


def test_check_ci_post_valid_directories(datadir, tmp_path, caplog):
    """Test ci_post check with valid previous and new vocab directories."""
    prev_dir = tmp_path / "prev"
    new_dir = tmp_path / "new"
    prev_dir.mkdir()
    new_dir.mkdir()

    # Copy same TTL to both dirs (no removed IRIs)
    shutil.copy(datadir / "concept-scheme-simple.ttl", prev_dir)
    shutil.copy(datadir / "concept-scheme-simple.ttl", new_dir)

    with caplog.at_level(logging.INFO):
        main_cli(["check", "--ci-post", str(prev_dir), str(new_dir)])

    assert "ci-post passed" in caplog.text


def test_check_ci_post_prev_vocab_not_exists(datadir, tmp_path, caplog):
    """Test ci_post when previous version of vocab doesn't exist."""
    prev_dir = tmp_path / "prev"
    new_dir = tmp_path / "new"
    prev_dir.mkdir()
    new_dir.mkdir()

    # Only new dir has the file (prev doesn't have it)
    shutil.copy(datadir / "concept-scheme-simple.ttl", new_dir)

    with caplog.at_level(logging.DEBUG):
        main_cli(["check", "-v", "--ci-post", str(prev_dir), str(new_dir)])

    assert "does not exist" in caplog.text


def test_check_ci_post_with_split_vocab(datadir, tmp_path, caplog):
    """Test ci_post handles split vocabulary directories."""
    prev_dir = tmp_path / "prev"
    new_dir = tmp_path / "new"
    prev_dir.mkdir()
    new_dir.mkdir()

    # Create a split vocabulary directory in prev (multiple TTL files)
    split_vocab_dir = prev_dir / "concept-scheme-simple"
    split_vocab_dir.mkdir()
    # Add TTL files to simulate a split vocabulary
    (split_vocab_dir / "part1.ttl").write_text(
        Path(datadir / "concept-scheme-simple.ttl").read_text()
    )

    # New dir has the joined vocab
    shutil.copy(datadir / "concept-scheme-simple.ttl", new_dir)

    with caplog.at_level(logging.DEBUG):
        main_cli(["check", "-v", "--ci-post", str(prev_dir), str(new_dir)])

    assert "split vocabulary, joining" in caplog.text


def test_check_with_explicit_profile(datadir, tmp_path, caplog):
    """Test check command with explicit --profile option."""
    shutil.copy(datadir / "concept-scheme-simple.ttl", tmp_path)

    with caplog.at_level(logging.INFO):
        main_cli(
            [
                "check",
                "--profile",
                "vocpub-4.7",
                str(tmp_path / "concept-scheme-simple.ttl"),
            ]
        )

    assert "valid according to the vocpub-4.7 profile" in caplog.text


def test_check_with_config_profile(datadir, tmp_path, caplog, monkeypatch, temp_config):
    """Test check uses profile from vocab config when not explicitly set."""
    # Copy TTL file with matching name from config
    ttl_file = tmp_path / "myvocab.ttl"
    shutil.copy(datadir / "concept-scheme-simple.ttl", ttl_file)

    # Create custom profile file (referenced in idranges_with_scheme.toml)
    # The profile is at config_dir/custom_profile.ttl relative to config file
    config_dir = tmp_path / "config"
    config_dir.mkdir()
    shutil.copy(datadir / "idranges_with_scheme.toml", config_dir)

    # Create custom profile (just a valid SHACL stub - use bundled one as source)
    from voc4cat.convert import get_bundled_profiles

    bundled = get_bundled_profiles()
    bundled_profile = bundled.get("vocpub-4.7")
    shutil.copy(bundled_profile, config_dir / "custom_profile.ttl")

    with caplog.at_level(logging.INFO):
        main_cli(
            [
                "check",
                "--config",
                str(config_dir / "idranges_with_scheme.toml"),
                str(ttl_file),
            ]
        )

    # The config's profile_local_path should be used
    assert "valid according to the custom_profile profile" in caplog.text
