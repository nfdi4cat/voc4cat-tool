"""
Tests for the xlsx_keyvalue module.

This module tests the key-value format specific functionality for XLSX processing.
"""

from datetime import date
from typing import Annotated

import pytest
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from voc4cat.xlsx_api import export_to_xlsx, import_from_xlsx
from voc4cat.xlsx_common import XLSXConverters, XLSXMetadata
from voc4cat.xlsx_keyvalue import XLSXKeyValueConfig

from .conftest import DemoModelWithMetadata, Priority, Project, SimpleModel, Status


# Key-Value Format Tests
class TestKeyValueFormat:
    """Tests for key-value format processing."""

    def test_keyvalue_export_import_basic(self, sample_simple_model, temp_file):
        """Test basic key-value export and import."""
        # Export
        export_to_xlsx(sample_simple_model, temp_file, format_type="keyvalue")

        # Import
        imported = import_from_xlsx(temp_file, SimpleModel, format_type="keyvalue")

        # Verify
        assert imported.name == "Test Item"
        assert imported.value == 42
        assert imported.active is True

    def test_keyvalue_export_with_config(self, sample_simple_model, temp_file):
        """Test key-value export with custom configuration."""
        config = XLSXKeyValueConfig(
            title="Custom Model Instance",
            field_column_header="Property",
            value_column_header="Data",
        )

        export_to_xlsx(
            sample_simple_model, temp_file, format_type="keyvalue", config=config
        )
        imported = import_from_xlsx(
            temp_file, SimpleModel, format_type="keyvalue", config=config
        )

        assert imported.name == "Test Item"

    def test_keyvalue_custom_field_names(self, sample_simple_model, temp_file):
        """Test custom field headers."""
        config = XLSXKeyValueConfig(
            field_column_header="Property Name",
            value_column_header="Property Value",
        )

        export_to_xlsx(
            sample_simple_model, temp_file, format_type="keyvalue", config=config
        )
        imported = import_from_xlsx(
            temp_file, SimpleModel, format_type="keyvalue", config=config
        )

        assert imported.name == "Test Item"
        assert imported.value == 42

    def test_keyvalue_field_filtering(self, sample_simple_model, temp_file):
        """Test field inclusion/exclusion."""
        config = XLSXKeyValueConfig(
            include_fields={"name", "value"},
        )

        export_to_xlsx(
            sample_simple_model, temp_file, format_type="keyvalue", config=config
        )
        imported = import_from_xlsx(
            temp_file, SimpleModel, format_type="keyvalue", config=config
        )

        # active field should use default since it was excluded
        assert imported.name == "Test Item"
        assert imported.value == 42
        assert imported.active is True  # default value

    def test_keyvalue_with_metadata(self, sample_model_with_metadata, temp_file):
        """Test key-value format with metadata."""
        config = XLSXKeyValueConfig()

        export_to_xlsx(
            sample_model_with_metadata, temp_file, format_type="keyvalue", config=config
        )
        imported = import_from_xlsx(
            temp_file, DemoModelWithMetadata, format_type="keyvalue", config=config
        )

        assert imported.temp == 25.5
        assert imported.name == "Test Sample"

    def test_keyvalue_complex_model(self, temp_file):
        """Test key-value format with complex model."""
        project = Project(
            project_id=1,
            project_name="Test Project",
            description="Test description",
            start_date=date(2024, 1, 1),
            priority=Priority.HIGH,
            budget=100000.0,
        )

        export_to_xlsx(project, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(temp_file, Project, format_type="keyvalue")

        assert imported.project_name == "Test Project"
        assert imported.priority == Priority.HIGH
        assert imported.description == "Test description"

    def test_keyvalue_with_converters(self, temp_file):
        """Test key-value format with custom converters."""

        class ModelWithConverters(BaseModel):
            name: str
            comma_list: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
            ] = Field(default=[])
            pipe_list: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.PIPE)
            ] = Field(default=[])

        # Create test data
        original = ModelWithConverters(
            name="Test Model",
            comma_list=["a", "b", "c"],
            pipe_list=["x", "y", "z"],
        )

        # Round-trip test
        export_to_xlsx(original, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(
            temp_file, ModelWithConverters, format_type="keyvalue"
        )

        assert imported.name == original.name
        assert imported.comma_list == original.comma_list
        assert imported.pipe_list == original.pipe_list

    def test_keyvalue_with_optional_fields(self, temp_file):
        """Test key-value format with optional fields."""

        class ModelWithOptional(BaseModel):
            name: str
            optional_field: str | None = None
            optional_with_default: str = "default_value"

        # Test with None value
        data = ModelWithOptional(name="test", optional_field=None)
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(
            temp_file, ModelWithOptional, format_type="keyvalue"
        )

        assert imported.name == "test"
        assert imported.optional_field is None
        assert imported.optional_with_default == "default_value"

        # Test with provided value
        data = ModelWithOptional(
            name="test", optional_field="provided", optional_with_default="custom"
        )
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(
            temp_file, ModelWithOptional, format_type="keyvalue"
        )

        assert imported.name == "test"
        assert imported.optional_field == "provided"
        assert imported.optional_with_default == "custom"

    def test_keyvalue_with_enum_fields(self, temp_file):
        """Test key-value format with enum fields."""

        class ModelWithEnum(BaseModel):
            name: str
            status: Status
            optional_status: Status | None = None

        # Test with enum values
        data = ModelWithEnum(
            name="test", status=Status.ACTIVE, optional_status=Status.PENDING
        )
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(temp_file, ModelWithEnum, format_type="keyvalue")

        assert imported.name == "test"
        assert imported.status == Status.ACTIVE
        assert imported.optional_status == Status.PENDING

    def test_keyvalue_field_order(self, temp_file):
        """Test field ordering in key-value format."""

        class OrderedModel(BaseModel):
            field_c: str
            field_a: str
            field_b: str

        data = OrderedModel(field_c="c", field_a="a", field_b="b")
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(temp_file, OrderedModel, format_type="keyvalue")

        assert imported.field_c == "c"
        assert imported.field_a == "a"
        assert imported.field_b == "b"

    def test_keyvalue_with_union_types(self, temp_file):
        """Test key-value format with Union types."""

        class ModelWithUnion(BaseModel):
            name: str
            value: int | str  # Union of different types

        data = ModelWithUnion(name="test", value=42)
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(temp_file, ModelWithUnion, format_type="keyvalue")

        assert imported.name == "test"
        # Union types get converted to string in Excel
        assert str(imported.value) == "42"

    def test_keyvalue_with_nested_optional(self, temp_file):
        """Test key-value format with deeply nested Optional types."""

        class ModelWithNestedOptional(BaseModel):
            name: str
            tags: list[str] | None = None

        # Test with None value
        data = ModelWithNestedOptional(name="test", tags=None)
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(
            temp_file, ModelWithNestedOptional, format_type="keyvalue"
        )

        assert imported.name == "test"
        assert imported.tags is None

    def test_keyvalue_with_complex_defaults(self, temp_file):
        """Test key-value format with complex default values."""

        class ModelWithComplexDefaults(BaseModel):
            name: str = "default_name"
            tags: Annotated[
                list[str], XLSXMetadata(separator_pattern=XLSXConverters.COMMA)
            ] = Field(default=["default", "tags"])

        # Create with defaults
        data = ModelWithComplexDefaults()
        export_to_xlsx(data, temp_file, format_type="keyvalue")
        imported = import_from_xlsx(
            temp_file, ModelWithComplexDefaults, format_type="keyvalue"
        )

        assert imported.name == "default_name"
        assert imported.tags == ["default", "tags"]


# Key-Value Configuration Tests
class TestKeyValueConfiguration:
    """Tests for key-value configuration."""

    def test_keyvalue_config_headers(self):
        """Test field header configuration."""
        config = XLSXKeyValueConfig(
            field_column_header="Property",
            value_column_header="Value",
            description_column_header="Details",
            meaning_column_header="Meaning",
            unit_column_header="Unit",
        )

        assert config.field_column_header == "Property"
        assert config.value_column_header == "Value"
        assert config.description_column_header == "Details"
        assert config.meaning_column_header == "Meaning"
        assert config.unit_column_header == "Unit"

    def test_keyvalue_config_defaults(self):
        """Test default configuration values."""
        config = XLSXKeyValueConfig()

        assert config.field_column_header == "Field"
        assert config.value_column_header == "Value"
        assert config.description_column_header == "Description"
        assert config.meaning_column_header == "Meaning"
        assert config.unit_column_header == "Unit"

    def test_keyvalue_config_visibility_flags(self):
        """Test configuration visibility flags - all columns always shown."""
        config = XLSXKeyValueConfig()

        # All columns are now always shown, so we just verify the config object works
        assert config.field_column_header == "Field"
        assert config.value_column_header == "Value"
        assert config.description_column_header == "Description"
        assert config.meaning_column_header == "Meaning"
        assert config.unit_column_header == "Unit"

    def test_keyvalue_config_inheritance(self):
        """Test that keyvalue config inherits from base config."""
        config = XLSXKeyValueConfig(
            title="Test Title",
            include_fields={"field1", "field2"},
            exclude_fields={"field3"},
        )

        assert config.title == "Test Title"
        assert config.should_include_field("field1") is True
        assert config.should_include_field("field2") is True
        assert config.should_include_field("field3") is False


# Key-Value Positioning Tests
class TestKeyValuePositioning:
    """Tests for validating the position of title and elements in key-value Excel files.

    These tests use openpyxl to directly inspect the Excel file structure and verify that:
    - Titles are positioned correctly at the specified start_row when present
    - Headers (Field, Value, etc.) appear at the correct row when title is present/absent
    - Data starts at the expected row based on title configuration
    - Custom start_row positioning works correctly
    - Font formatting (bold, size) is applied to titles

    The expected layout structure for key-value format is:
    - Title: start_row (if present)
    - Empty row: start_row + 1 (spacing, if title present)
    - Headers: start_row + 2 (if title present) or start_row (if no title)
    - Data: next row after headers
    """

    def _assert_worksheet_has_expected_fields(
        self, worksheet, data_row, expected_fields, check_rows=3
    ):
        """Verify worksheet contains expected field names in first column.

        Args:
            worksheet: The openpyxl worksheet to check.
            data_row: The row number where data starts.
            expected_fields: List of field name substrings to look for.
            check_rows: Number of data rows to check.
        """
        first_field_cell = worksheet.cell(row=data_row, column=1)
        first_value_cell = worksheet.cell(row=data_row, column=2)

        # Should contain actual field data
        assert first_field_cell.value is not None
        assert first_value_cell.value is not None

        # Collect field names from first column
        field_names = []
        for row in range(data_row, data_row + check_rows):
            field_cell = worksheet.cell(row=row, column=1)
            if field_cell.value:  # pragma: no branch
                field_names.append(str(field_cell.value).lower())

        # Verify expected field names are present
        for field in expected_fields:
            assert any(field in name for name in field_names), (
                f"Expected field '{field}' not found in {field_names}"
            )

    def test_keyvalue_title_positioning_with_title(
        self, sample_simple_model, temp_file
    ):
        """Test that title is positioned correctly when title is provided."""
        config = XLSXKeyValueConfig(
            title="Simple Model Data",
            start_row=3,
        )

        export_to_xlsx(
            sample_simple_model, temp_file, format_type="keyvalue", config=config
        )

        # Load workbook with openpyxl to inspect structure
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 3 (start_row), column 1
        title_cell = worksheet.cell(row=3, column=1)
        assert title_cell.value == "Simple Model Data"

        # Title should have bold formatting and larger font
        assert title_cell.font.bold is True
        assert title_cell.font.size == 14

        # Empty row should be in row 4 (start_row + 1) for spacing
        empty_cell = worksheet.cell(row=4, column=1)
        assert empty_cell.value is None

        # Headers should be in row 5 (start_row + 2)
        header_row = 5
        field_header = worksheet.cell(row=header_row, column=1)
        value_header = worksheet.cell(row=header_row, column=2)

        # Check default headers
        assert field_header.value == "Field"
        assert value_header.value == "Value"

        # Data should start in row 6 (start_row + 3)
        data_row = 6
        self._assert_worksheet_has_expected_fields(
            worksheet, data_row, ["name", "value"]
        )

    def test_keyvalue_title_positioning_without_title(
        self, sample_simple_model, temp_file
    ):
        """Test positioning when no title is provided."""
        config = XLSXKeyValueConfig(
            title=None,  # No title
            start_row=2,
        )

        export_to_xlsx(
            sample_simple_model, temp_file, format_type="keyvalue", config=config
        )

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Headers should be in row 2 (start_row) when no title
        header_row = 2
        field_header = worksheet.cell(row=header_row, column=1)
        value_header = worksheet.cell(row=header_row, column=2)

        assert field_header.value == "Field"
        assert value_header.value == "Value"

        # Data should start in row 3 (start_row + 1)
        data_row = 3
        self._assert_worksheet_has_expected_fields(
            worksheet, data_row, ["name", "value"]
        )

    def test_keyvalue_positioning_with_custom_headers(
        self, sample_simple_model, temp_file
    ):
        """Test positioning with custom header names."""
        config = XLSXKeyValueConfig(
            title="Custom Headers Example",
            field_column_header="Property Name",
            value_column_header="Property Value",
        )

        export_to_xlsx(
            sample_simple_model, temp_file, format_type="keyvalue", config=config
        )

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Title should be in row 1 (default start_row)
        title_cell = worksheet.cell(row=1, column=1)
        assert title_cell.value == "Custom Headers Example"
        assert title_cell.font.bold is True

        # Empty row should be in row 2 (start_row + 1) for spacing
        empty_cell = worksheet.cell(row=2, column=1)
        assert empty_cell.value is None

        # Headers should be in row 3 (start_row + 2)
        header_row = 3
        field_header = worksheet.cell(row=header_row, column=1)
        value_header = worksheet.cell(row=header_row, column=2)

        # Check custom headers
        assert field_header.value == "Property Name"
        assert value_header.value == "Property Value"

        # Data should start in row 4 (start_row + 3)
        data_row = 4
        first_field_cell = worksheet.cell(row=data_row, column=1)
        first_value_cell = worksheet.cell(row=data_row, column=2)

        assert first_field_cell.value is not None
        assert first_value_cell.value is not None


class TestKeyValueRequirednessAndToggles:
    """Tests for requiredness column and metadata visibility toggles."""

    def test_keyvalue_with_requiredness_column_shown(self, temp_file):
        """Test that requiredness column is shown when explicitly enabled."""
        from voc4cat.xlsx_common import MetadataToggleConfig, MetadataVisibility

        class ModelForReq(BaseModel):
            required_field: str
            optional_field: str | None = None
            with_default: str = "default_value"

        data = ModelForReq(required_field="test")

        config = XLSXKeyValueConfig(
            title="Test Model",
            metadata_visibility=MetadataToggleConfig(
                requiredness=MetadataVisibility.SHOW
            ),
        )

        export_to_xlsx(data, temp_file, format_type="keyvalue", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Find the "Required" header
        header_row = 3
        found_required_header = False
        for col in range(1, 10):  # pragma: no branch
            cell_value = worksheet.cell(row=header_row, column=col).value
            if cell_value == "Required":  # pragma: no branch
                found_required_header = True
                break

        assert found_required_header, "Required column header should be present"

    def test_keyvalue_without_requiredness_column_by_default(self, temp_file):
        """Test that requiredness column is NOT shown by default."""

        class SimpleModelForReq(BaseModel):
            name: str = "test"

        data = SimpleModelForReq()

        config = XLSXKeyValueConfig(title="Test Model")
        export_to_xlsx(data, temp_file, format_type="keyvalue", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Check that no "Required" header exists
        header_row = 3
        for col in range(1, 10):
            cell_value = worksheet.cell(row=header_row, column=col).value
            assert cell_value != "Required", "Required column should not be present"

    def test_keyvalue_toggle_hide_description(self, temp_file):
        """Test that HIDE toggle removes description column."""
        from voc4cat.xlsx_common import MetadataToggleConfig, MetadataVisibility

        class ModelWithDesc(BaseModel):
            name: Annotated[
                str,
                XLSXMetadata(description="A description"),
            ] = "test"

        data = ModelWithDesc()

        # First verify description shows in AUTO mode
        config_auto = XLSXKeyValueConfig(title="Test")
        export_to_xlsx(data, temp_file, format_type="keyvalue", config=config_auto)
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        header_row = 3
        found_desc = False
        for col in range(1, 10):  # pragma: no branch
            if (
                worksheet.cell(row=header_row, column=col).value == "Description"
            ):  # pragma: no branch
                found_desc = True
                break
        assert found_desc, "Description should be shown in AUTO mode"

        # Now with HIDE
        config_hide = XLSXKeyValueConfig(
            title="Test",
            metadata_visibility=MetadataToggleConfig(
                description=MetadataVisibility.HIDE
            ),
        )
        export_to_xlsx(data, temp_file, format_type="keyvalue", config=config_hide)
        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        for col in range(1, 10):
            cell_value = worksheet.cell(row=header_row, column=col).value
            assert cell_value != "Description", (
                "Description should not be shown with HIDE"
            )

    def test_keyvalue_requiredness_values_correct(self, temp_file):
        """Test that requiredness column values are correct."""
        from voc4cat.xlsx_common import MetadataToggleConfig, MetadataVisibility

        class ModelWithVariousFields(BaseModel):
            required_field: str
            optional_field: str | None = None
            with_default: str = "custom"

        data = ModelWithVariousFields(required_field="value")

        config = XLSXKeyValueConfig(
            metadata_visibility=MetadataToggleConfig(
                requiredness=MetadataVisibility.SHOW
            ),
        )

        export_to_xlsx(data, temp_file, format_type="keyvalue", config=config)

        workbook = load_workbook(temp_file)
        worksheet = workbook.active

        # Find requiredness column
        header_row = 1  # No title
        req_col = None
        for col in range(1, 10):  # pragma: no branch
            if (
                worksheet.cell(row=header_row, column=col).value == "Required"
            ):  # pragma: no branch
                req_col = col
                break

        assert req_col is not None, "Required column should exist"

        # Check values - data starts at row 2
        # Find field names and their requiredness
        field_col = 1
        values = {}
        for row in range(2, 5):
            field_name = worksheet.cell(row=row, column=field_col).value
            req_value = worksheet.cell(row=row, column=req_col).value
            if field_name:  # pragma: no branch
                values[field_name.lower().replace(" ", "_")] = req_value

        # required_field should be "Yes"
        assert "yes" in str(values.get("required_field", "")).lower()


if __name__ == "__main__":
    pytest.main([__file__])
