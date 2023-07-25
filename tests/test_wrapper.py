import logging
import os
import shutil
from pathlib import Path
from unittest import mock

import pytest
from openpyxl.reader.excel import load_workbook
from voc4cat.wrapper import build_docs, main_cli

CS_SIMPLE = "concept-scheme-simple.xlsx"
CS_SIMPLE_TURTLE = "concept-scheme-simple.ttl"
CS_CYCLES = "concept-scheme-with-cycles.xlsx"
CS_CYCLES_TURTLE = "concept-scheme-with-cycles.ttl"
CS_CYCLES_INDENT = "concept-scheme-with-cycles_indent.xlsx"
CS_CYCLES_INDENT_IRI = "concept-scheme-with-cycles_indent_iri.xlsx"
CS_CYCLES_INDENT_DOT = "concept-scheme-with-cycles_indent-by-dot.xlsx"
CS_CYCLES_MULTI_LANG = "concept-scheme-with-cycles_multilang.xlsx"
CS_CYCLES_MULTI_LANG_IND = "concept-scheme-with-cycles_multilang_indent_iri.xlsx"


def test_main_no_args_entrypoint(monkeypatch, capsys):
    monkeypatch.setattr("sys.argv", ["voc4at"])
    exit_code = main_cli()
    captured = capsys.readouterr()
    assert "usage: voc4cat" in captured.out
    assert exit_code == 0


def test_main_no_args(capsys):
    exit_code = main_cli([])
    captured = capsys.readouterr()
    assert "usage: voc4cat" in captured.out
    assert exit_code == 0


def test_main_unknown_arg(caplog):
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli(["--unknown-arg"])
    assert "Unknown voc4cat option: ['--unknown-arg']" in caplog.text
    assert exit_code == 1


def test_main_version(capsys):
    exit_code = main_cli(["--version"])
    captured = capsys.readouterr()
    assert captured.out.startswith("voc4cat")
    assert exit_code == 0


def test_make_ids_missing_file(caplog):
    # Try to run a command that would overwrite the input file with the output file
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli(["--make-ids", "ex", "1", "missing.xyz"])
    assert "Expected xlsx-file or directory but got:" in caplog.text
    assert exit_code == 1


def test_make_ids_overwrite_warning(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_SIMPLE, tmp_path / CS_SIMPLE)
    monkeypatch.chdir(tmp_path)
    # Try to run a command that would overwrite the input file with the output file
    # a) dir as input:
    with pytest.warns(
        UserWarning, match='Option "--make-ids" will overwrite the existing file'
    ):
        main_cli(["--make-ids", "ex", "1", str(tmp_path)])
    # b) file as input
    with pytest.warns(
        UserWarning, match='Option "--make-ids" will overwrite the existing file'
    ):
        main_cli(["--make-ids", "ex", "1", str(tmp_path / CS_SIMPLE)])


def test_make_ids_no_voc_base_iri(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_SIMPLE, tmp_path)
    monkeypatch.chdir(tmp_path)
    # change excel file: Delete vocabulary base IRI
    wb = load_workbook(filename=CS_SIMPLE)
    ws = wb["Concept Scheme"]
    ws.cell(row=2, column=2).value = None
    new_filename = "no_voc_base_iri.xlsx"
    wb.save(new_filename)
    wb.close()

    main_cli(["--make-ids", "na", "1001", "--no-warn", str(tmp_path / new_filename)])
    wb = load_workbook(filename=new_filename, read_only=True, data_only=True)
    ws = wb["Concept Scheme"]
    assert ws.cell(row=2, column=2).value == "https://example.org/"


def test_make_ids_invalid_id(datadir):
    with pytest.raises(
        ValueError,
        match='For option --make-ids the "start_id" must be an integer greater than 0.',
    ):
        main_cli(["--make-ids", "ex", "###", "--no-warn", str(datadir / CS_SIMPLE)])


@pytest.mark.parametrize(
    ("indir", "outdir"),
    [(True, ""), (True, "out"), (False, ""), (False, "out")],
    ids=[
        "in:dir, out:default",
        "in:dir, out:dir",
        "in:file, out:default",
        "in:file, out:dir",
    ],
)
def test_make_ids_variants(monkeypatch, datadir, tmp_path, indir, outdir):
    # fmt: off
    expected_concepts = [
        ("ex:test/0001001", "term1", "en", "def for term1", "en", "AltLbl for term1", "ex:test/0001002, ex:test/0001003",),
        ("ex:test/0001002", "term2", "en", "def for term2", "en", "AltLbl for term2", None,),
        ("ex:test/0001003", "term3", "en", "def for term3", "en", "AltLbl for term3", "ex:test/0001004",),
        ("ex:test/0001004", "term4", "en", "def for term4", "en", "AltLbl for term4", None, ),
        ("ex:test/0001005", "term5", "en", "def for term5", "en", "AltLbl for term5", None, ),
        ("ex:test/0001006", "term6", "en", "def for term6", "en", "AltLbl for term6", None,),
    ]
    expected_collections = [
        ("ex:test/0001007", "con", "def for con", "ex:test/0001001, ex:test/0001002, ex:test/0001003, ex:test/0001004",),
    ]
    expected_additional = [
        ("ex:test/0001001", "ex:test/0001002", "ex:test/0001003", "ex:test/0001004", "ex:test/0001005", "ex:test/0001006", ),
    ]
    # fmt: on
    shutil.copy(datadir / CS_SIMPLE, tmp_path)
    monkeypatch.chdir(tmp_path)
    main_cli(
        ["--make-ids", "ex", "1001", "--no-warn"]
        + (["--output-directory", outdir] if outdir else [])
        + ([str(tmp_path)] if indir else [str(tmp_path / CS_SIMPLE)])
    )
    xlsxfile = tmp_path / outdir / CS_SIMPLE if outdir else tmp_path / CS_SIMPLE
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=7, values_only=True), expected_concepts
    ):
        assert row == expected_row
    ws = wb["Collections"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=4, values_only=True), expected_collections
    ):
        assert row == expected_row
    ws = wb["Additional Concept Features"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=6, values_only=True), expected_additional
    ):
        assert row == expected_row


def test_make_ids_multilang(tmp_path, datadir):
    # fmt: off
    expected_concepts = [
        ("ex:test/0001001", "term1", "en", "def for term1", "en", "AltLbl for term1", "ex:test/0001002, ex:test/0001003",),
        ("ex:test/0001002", "term2", "en", "def for term2", "en", "AltLbl for term2", "ex:test/0001004",),
        ("ex:test/0001003", "term3", "en", "def for term3", "en", "AltLbl for term3", "ex:test/0001004",),
        ("ex:test/0001004", "term4", "en", "def for term4", "en", "AltLbl for term4", None, ),
        ("ex:test/0001004", "Begr4", "de", "Def für Begr4", "de", "AltLbl für Begr4", None, ),
        ("ex:test/0001001", "Begr1", "de", "Def für Begr1", "de", "AltLbl für Begr1", None, ),
    ]
    # fmt: on
    main_cli(
        [
            "--make-ids",
            "ex",
            "1001",
            "--output-directory",
            str(tmp_path),
            str(datadir / CS_CYCLES_MULTI_LANG),
        ]
    )
    xlsxfile = tmp_path / CS_CYCLES_MULTI_LANG
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=7, values_only=True), expected_concepts
    ):
        assert row == expected_row


def test_hierarchy_from_indent_on_dir(tmp_path, caplog):
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli(
            ["--hierarchy-from-indent", str(tmp_path / "missing.xlsx")]
        )
    assert exit_code == 1
    assert "File not found:" in caplog.text

    # exit_code = main_cli(["--hierarchy-from-indent", "tmp_path"])


@pytest.mark.parametrize(
    ("xlsxfile", "indent"),
    [(CS_CYCLES_INDENT_IRI, None), (CS_CYCLES_INDENT_DOT, "..")],
    ids=["indent:Excel", "indent:dots"],
)
def test_hierarchy_from_indent(monkeypatch, datadir, tmp_path, xlsxfile, indent):
    # fmt: off
    expected = [  # data in children-IRI-representation
        ("ex:test/term1", "term1", "en", "def for term1", "en", "AltLbl for term1", "ex:test/term2, ex:test/term3", "Prov for term1", "ex:XYZ/term1"),
        ("ex:test/term2", "term2", "en", "def for term2", "en", "AltLbl for term2", "ex:test/term4",                "Prov for term2", "ex:XYZ/term2"),
        ("ex:test/term3", "term3", "en", "def for term3", "en", "AltLbl for term3", "ex:test/term4",                "Prov for term3", "ex:XYZ/term3"),
        ("ex:test/term4", "term4", "en", "def for term4", "en", "AltLbl for term4", None,                           "Prov for term4", "ex:XYZ/term4"),
        (None, None, None, None, None, None, None, None, None)
    ]
    # fmt: on
    expected_len = len(expected[0])
    monkeypatch.chdir(datadir)
    main_cli(
        ["--hierarchy-from-indent"]
        + (
            [
                "--indent-separator",
                indent,
            ]
            if indent
            else []
        )
        + [
            "--output-directory",
            str(tmp_path),
            xlsxfile,
        ]
    )
    monkeypatch.chdir(tmp_path)
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, _expected_row in zip(ws.iter_rows(min_row=3, values_only=True), expected):
        assert len(row) == expected_len
        assert row in expected  # We intentionally don't check the row position here!


def test_hierarchy_from_indent_multilang(monkeypatch, datadir, tmp_path):
    # fmt: off
    expected = [  # data in children-IRI-representation
        ("ex:test/term1", "term1", "en", "def for term1", "en", "AltLbl for term1", "ex:test/term2, ex:test/term3", "Prov for term1", "ex:XYZ/term1"),
        ("ex:test/term1", "Begr1", "de", "Def für Begr1", "de", "AltLbl für Begr1", "ex:test/term2, ex:test/term3", "Prov for term1", "ex:XYZ/term1"),
        ("ex:test/term2", "term2", "en", "def for term2", "en", "AltLbl for term2", "ex:test/term4", "Prov for term2", "ex:XYZ/term2"),
        ("ex:test/term3", "term3", "en", "def for term3", "en", "AltLbl for term3", "ex:test/term4", "Prov for term3", "ex:XYZ/term3"),
        ("ex:test/term4", "term4", "en", "def for term4", "en", "AltLbl for term4", None,            "Prov for term4", "ex:XYZ/term4"),
        ("ex:test/term4", "Begr4", "de", "Def für Begr4", "de", "AltLbl für Begr4", None,            "Prov for term4", "ex:XYZ/term4"),
        (None, None, None, None, None, None, None, None, None)
    ]
    # fmt: on
    expected_len = len(expected[0])
    monkeypatch.chdir(datadir)
    main_cli(
        [
            "--hierarchy-from-indent",
            "--output-directory",
            str(tmp_path),
            CS_CYCLES_MULTI_LANG_IND,
        ]
    )
    monkeypatch.chdir(tmp_path)
    wb = load_workbook(
        filename=CS_CYCLES_MULTI_LANG_IND, read_only=True, data_only=True
    )
    ws = wb["Concepts"]
    for row, _expected_row in zip(ws.iter_rows(min_row=3, values_only=True), expected):
        assert len(row) == expected_len
        assert row in expected  # We intentionally don't check the row position here!


def test_hierarchy_from_indent_merge(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_CYCLES_INDENT_IRI, tmp_path)
    monkeypatch.chdir(tmp_path)
    # change excel file: Delete vocabulary base IRI
    wb = load_workbook(filename=CS_CYCLES_INDENT_IRI)
    ws = wb["Concepts"]
    ws.cell(row=8, column=4).value = "Contradicting def."
    iri = ws.cell(row=8, column=1).value
    new_filename = "indent_merge_problem.xlsx"
    wb.save(new_filename)
    wb.close()
    with pytest.raises(
        ValueError, match=f"Cannot merge rows for {iri}. Resolve differences manually."
    ):
        main_cli(["--hierarchy-from-indent", "--no-warn", str(tmp_path / new_filename)])


@pytest.mark.parametrize(
    "indent",
    ["..", None],
    ids=["indent:dots", "indent:Excel"],
)
def test_hierarchy_to_indent(monkeypatch, datadir, tmp_path, indent):
    # fmt: off
    expected_rows = [  # data in children-IRI-representation
        ("ex:test/term1", "term1",     "en", "def for term1", "en", "AltLbl for term1", None, "Prov for term1", "ex:XYZ/term1"),
        ("ex:test/term3", "..term3",   "en", "def for term3", "en", "AltLbl for term3", None, "Prov for term3", "ex:XYZ/term3"),
        ("ex:test/term4", "....term4", "en", "def for term4", "en", "AltLbl for term4", None, "Prov for term4", "ex:XYZ/term4"),
        ("ex:test/term2", "term2",     "en", "def for term2", "en", "AltLbl for term2", None, "Prov for term2", "ex:XYZ/term2"),
        ("ex:test/term4", "..term4",   "en", None, None, None, None, None, None),
        ("ex:test/term1", "term1",     "en", None, None, None, None, None, None),
        ("ex:test/term2", "..term2",   "en", None, None, None, None, None, None),
        (None, None, None, None, None, None, None, None, None),
    ]
    # fmt: on
    expected_levels = [0, 1, 2, 0, 1, 0, 1, 0]
    assert len(expected_rows) == len(expected_levels)
    expected_len = len(expected_rows[0])

    monkeypatch.chdir(datadir)
    main_cli(
        ["--hierarchy-to-indent"]
        + (
            [
                "--indent-separator",
                indent,
            ]
            if indent
            else []
        )
        + [
            "--output-directory",
            str(tmp_path),
            CS_CYCLES,
        ]
    )
    monkeypatch.chdir(tmp_path)
    wb = load_workbook(filename=CS_CYCLES, read_only=True)
    ws = wb["Concepts"]
    for row, expected_row, expected_level in zip(
        ws.iter_rows(min_row=3), expected_rows, expected_levels
    ):
        assert len(row) == expected_len
        if indent is None:  # Excel-indent
            assert int(row[1].alignment.indent) == expected_level

        for col in range(len(expected_rows)):
            if indent is None and col == 1:  # Excel-indent
                continue
            assert row[col].value == expected_row[col]


def test_hierarchy_to_indent_multilanguage(monkeypatch, datadir, tmp_path):
    # fmt: off
    expected_rows = [  # data in children-IRI-representation
        ("ex:test/term1", "term1",     "en", "def for term1", "en", "AltLbl for term1", None, "Prov for term1", "ex:XYZ/term1"),
        ("ex:test/term1", "Begr1",     "de", "Def für Begr1", "de", "AltLbl für Begr1", None, "Prov for term1", "ex:XYZ/term1"),
        ("ex:test/term3", "..term3",   "en", "def for term3", "en", "AltLbl for term3", None, "Prov for term3", "ex:XYZ/term3"),
        ("ex:test/term4", "....term4", "en", "def for term4", "en", "AltLbl for term4", None, "Prov for term4", "ex:XYZ/term4"),
        ("ex:test/term4", "....Begr4", "de", "Def für Begr4", "de", "AltLbl für Begr4", None, "Prov for term4", "ex:XYZ/term4"),
        ("ex:test/term2", "term2",     "en", "def for term2", "en", "AltLbl for term2", None, "Prov for term2", "ex:XYZ/term2"),
        ("ex:test/term4", "..term4",   "en", None, None, None, None, None, None),
        ("ex:test/term1", "term1",     "en", None, None, None, None, None, None),
        ("ex:test/term2", "..term2",   "en", None, None, None, None, None, None),
        (None, None, None, None, None, None, None, None, None),
    ]
    # fmt: on
    expected_levels = [0, 0, 1, 2, 2, 0, 1, 0, 1, 0]
    assert len(expected_rows) == len(expected_levels)
    monkeypatch.chdir(datadir)
    main_cli(
        [
            "--hierarchy-to-indent",
            "--output-directory",
            str(tmp_path),
            CS_CYCLES_MULTI_LANG,
        ]
    )
    monkeypatch.chdir(tmp_path)
    wb = load_workbook(filename=CS_CYCLES_MULTI_LANG, read_only=True)
    ws = wb["Concepts"]
    for row, expected_row, expected_level in zip(
        ws.iter_rows(min_row=3), expected_rows, expected_levels
    ):
        # Excel-indent
        assert int(row[1].alignment.indent) == expected_level

        for col in range(len(expected_rows[0])):
            if col == 1:  # Excel-indent
                assert row[col].value == expected_row[col].strip(".")
                continue
            assert row[col].value == expected_row[col]


def test_hierarchy_to_indent_merge(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_CYCLES_MULTI_LANG, tmp_path)
    monkeypatch.chdir(tmp_path)
    # change excel file: Delete vocabulary base IRI
    wb = load_workbook(filename=CS_CYCLES_MULTI_LANG)
    ws = wb["Concepts"]
    ws.cell(row=8, column=8).value = "Contradicting def."
    iri = ws.cell(row=8, column=1).value
    new_filename = "indent_merge_problem.xlsx"
    wb.save(new_filename)
    wb.close()
    with pytest.raises(ValueError, match=f"Merge conflict for concept {iri}"):
        main_cli(["--hierarchy-to-indent", "--no-warn", str(tmp_path / new_filename)])


@pytest.mark.parametrize(
    "outdir",
    [None, "out"],
    ids=["no outdir", "with outdir"],
)
def test_outdir_variants(monkeypatch, datadir, tmp_path, outdir):
    shutil.copy(datadir / CS_CYCLES_INDENT_IRI, tmp_path)
    cmd = ["--hierarchy-from-indent"]
    if outdir:
        cmd.extend(["--output-directory", str(tmp_path / outdir)])
    cmd.append(str(tmp_path / CS_CYCLES_INDENT_IRI))
    # print(f"\n>>> cmd {cmd}")
    monkeypatch.chdir(tmp_path)
    main_cli(cmd)

    expected = [
        ("ex:test/term1", "term1"),
        ("ex:test/term3", "term3"),
        ("ex:test/term4", "term4"),
        ("ex:test/term2", "term2"),
        (None, None),
    ]
    if outdir:
        xlsxfile = tmp_path / outdir / CS_CYCLES_INDENT_IRI
    else:
        xlsxfile = tmp_path / CS_CYCLES_INDENT_IRI
    wb = load_workbook(filename=xlsxfile, read_only=True, data_only=True)
    ws = wb["Concepts"]
    for row, expected_row in zip(
        ws.iter_rows(min_row=3, max_col=2, values_only=True), expected
    ):
        assert row == expected_row


@pytest.mark.parametrize(
    "test_file",
    [CS_CYCLES_TURTLE, ""],
    ids=["single file", "dir of files"],
)
def test_build_docs_ontospy(datadir, tmp_path, test_file):
    """Check that ontospy generates the expected output."""
    dst = tmp_path / test_file
    shutil.copy(datadir / CS_CYCLES_TURTLE, tmp_path)
    outdir = tmp_path / "ontospy"
    # To test the code-path, outdir is created automatically here.
    main_cli(["--docs", "ontospy", "--output-directory", str(outdir), str(dst)])
    assert (outdir / Path(CS_CYCLES_TURTLE).stem / "dendro" / "index.html").exists()
    assert (outdir / Path(CS_CYCLES_TURTLE).stem / "docs" / "index.html").exists()


@pytest.mark.parametrize(
    "doc_builder",
    ["pylode", "ontospy"],
)
def test_build_docs(tmp_path, caplog, doc_builder):
    """Check handling of missing dir/file on documentation build."""
    with caplog.at_level(logging.INFO):
        exit_code = build_docs(tmp_path, tmp_path, doc_builder)
    assert exit_code == 1
    assert f"No turtle file(s) found to document in {tmp_path}" in caplog.text

    with caplog.at_level(logging.INFO):
        exit_code = build_docs(tmp_path / CS_CYCLES_TURTLE, tmp_path, doc_builder)
    assert exit_code == 1
    assert f"File/dir not found (for docs): {tmp_path/CS_CYCLES_TURTLE}" in caplog.text

    with caplog.at_level(logging.INFO):
        exit_code = build_docs(tmp_path / CS_CYCLES_TURTLE, tmp_path, "ontospy")
    assert exit_code == 1
    assert f"File/dir not found (for docs): {tmp_path/CS_CYCLES_TURTLE}" in caplog.text


def test_build_docs_unknown_builder(tmp_path, caplog):
    """Check handling of unknown documentation builder."""
    unknown_doc_builder = "123doc"
    with caplog.at_level(logging.INFO):
        exit_code = build_docs(tmp_path, tmp_path, unknown_doc_builder)
    assert exit_code == 1
    assert f"Unsupported document builder '{unknown_doc_builder}'." in caplog.text


@pytest.mark.parametrize(
    ("test_file", "err", "msg"),
    [
        (CS_CYCLES, 0, "All checks passed successfully."),
        (
            CS_CYCLES_INDENT_IRI,
            1,
            'Same Concept IRI "ex:test/term1" used more than once for language "en"',
        ),
    ],
    ids=["no error", "with error"],
)
def test_check(datadir, tmp_path, caplog, test_file, err, msg):  # noqa: PLR0913
    dst = tmp_path / test_file
    shutil.copy(datadir / test_file, dst)
    with caplog.at_level(logging.INFO):
        exit_code = main_cli(["--check", "--no-warn", str(dst)])
    assert exit_code == err
    assert msg in caplog.text
    # TODO check that erroneous cells get colored.


def test_check_overwrite_warning(monkeypatch, datadir, tmp_path):
    shutil.copy(datadir / CS_SIMPLE, tmp_path / CS_SIMPLE)
    monkeypatch.chdir(tmp_path)
    # Try to run a command that would overwrite the input file with the output file
    # a) dir as input:
    with pytest.warns(
        UserWarning, match='Option "--check" will overwrite the existing file'
    ):
        main_cli(["--check", str(tmp_path)])
    # b) file as input
    with pytest.warns(
        UserWarning, match='Option "--check" will overwrite the existing file'
    ):
        main_cli(["--check", str(tmp_path / CS_SIMPLE)])


def test_ci_check_skipped(datadir, tmp_path, caplog):
    # Test if skipping works if no output-directory is given.
    dst = tmp_path
    shutil.copy(datadir / CS_SIMPLE, dst)
    exit_code = main_cli(["--ci-check", str(dst)])
    assert exit_code == 0


@mock.patch.dict(os.environ, {"CI_RUN": "", "LOGLEVEL": "DEBUG"})
def test_ci_check_local(datadir, tmp_path, caplog):
    dst = tmp_path
    shutil.copy(datadir / CS_SIMPLE, dst)
    outdir = tmp_path / "out"

    exit_code = main_cli(["--ci-check", "--output-directory", str(outdir), str(dst)])
    assert exit_code == 0


@pytest.fixture()
def main_branch():
    """
    Provides a temporary dir "_main_branch" in the repo as in CI.
    """
    main_branch = Path(__file__).resolve().parents[1] / "_main_branch"
    main_branch.mkdir(exist_ok=True)
    yield main_branch
    # Remove the directory tree after the test
    shutil.rmtree(main_branch, ignore_errors=True)


@mock.patch.dict(os.environ, {"CI_RUN": "true"})
def test_ci_check_in_ci(datadir, tmp_path, temp_config, main_branch, caplog):
    dst = tmp_path / "inbox"
    dst.mkdir()
    outdir = tmp_path / "vocabularies"
    # Could be probably solved better. We put the dir/files in real repo dir here
    # not in tmp_dir to make it findable
    (main_branch / "vocabularies").mkdir(parents=True, exist_ok=True)

    shutil.copy(datadir / CS_SIMPLE, dst / "myvocab.xlsx")
    shutil.copy(datadir / "valid_idranges.toml", main_branch / "idranges.toml")

    # Load/prepare a valid strict config
    config = temp_config
    config.load_config(main_branch / "idranges.toml")
    config.IDRANGES.vocabs["myvocab"].id_length = 2
    config.IDRANGES.vocabs["myvocab"].permanent_iri_part = "http://example.org/test"
    config.load_config(config=config.IDRANGES)

    # Test without a previous vocabulary version in main_branch
    with caplog.at_level(logging.DEBUG):
        exit_code = main_cli(
            ["-v", "--ci-check", "--output-directory", str(outdir), str(dst)]
        )
    assert exit_code == 0
    assert 'previous version of vocabulary "myvocab.ttl" does not exist.' in caplog.text

    # Test with a previous vocabulary version in main_branch
    shutil.copy(
        datadir / CS_SIMPLE_TURTLE, main_branch / "vocabularies" / "myvocab.ttl"
    )
    with caplog.at_level(logging.DEBUG):
        exit_code = main_cli(
            ["-v", "--ci-check", "--output-directory", str(outdir), str(dst)]
        )
    assert exit_code == 0
    assert "-> Checking changes between" in caplog.text

    # Test loading of a config (which makes the command fail)
    exit_code = main_cli(
        [
            "-v",
            "--config",
            str(main_branch / "idranges.toml"),
            "--ci-check",
            "--output-directory",
            str(outdir),
            str(dst),
        ]
    )
    assert exit_code == 1
    assert config.IDRANGES.vocabs["myvocab"].id_length == 7  # noqa: PLR2004


def test_unsupported_filetype(monkeypatch, datadir, caplog):
    monkeypatch.chdir(datadir)
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli(["README.md"])
    assert exit_code == 1
    assert "Cannot convert file" in caplog.text
    assert "Files for processing must end with" in caplog.text


def test_nonexisting_file(monkeypatch, datadir, caplog):
    monkeypatch.chdir(datadir)
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli(["missing.txt"])
    assert exit_code == 1
    assert "File not found:" in caplog.text


def test_nonexisting_config(monkeypatch, datadir, caplog):
    monkeypatch.chdir(datadir)
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli(["--config", "missing.toml", CS_SIMPLE])
    assert exit_code == 1
    assert "Config file not found at" in caplog.text


def test_no_separator(monkeypatch, datadir):
    monkeypatch.chdir(datadir)
    with pytest.raises(
        ValueError, match="Setting the indent separator to zero length is not allowed."
    ):
        main_cli(["--indent-separator", "", CS_CYCLES])


def test_duplicates(datadir, tmp_path, caplog):
    """Check that files do not have the same stem."""
    shutil.copy(datadir / CS_CYCLES, tmp_path)
    shutil.copy(datadir / CS_CYCLES_TURTLE, tmp_path)
    with caplog.at_level(logging.ERROR):
        exit_code = main_cli([str(tmp_path)])
    assert exit_code == 1
    assert "Files may only be present in one format." in caplog.text


def test_run_vocexcel_badfile(monkeypatch, datadir, tmp_path, caplog):
    """Check handling of failing run of vocexcel."""
    shutil.copy(datadir / CS_CYCLES_INDENT, tmp_path)
    monkeypatch.chdir(tmp_path)
    exit_code = main_cli([CS_CYCLES_INDENT])
    assert exit_code > 0
    # The next message is logged by vocexcel so it may change.
    assert "VIOLATION: Validation Result in MinCountConstraintComponent" in caplog.text


@pytest.mark.parametrize(
    "test_file",
    [CS_CYCLES, ""],
    ids=["single file", "dir of files"],
)
def test_run_vocexcel(datadir, tmp_path, test_file):
    """Check that an xlsx file is converted to ttl by vocexcel."""
    dst = tmp_path / test_file
    shutil.copy(datadir / CS_CYCLES, dst)
    # We also test if the logging option is passed on to vocexcel.
    log = tmp_path / "logs" / "test-run.log"
    main_cli(["--logfile", str(log), str(dst)])
    expected = (tmp_path / CS_CYCLES).with_suffix(".ttl")
    assert expected.exists()
    assert log.exists()


@pytest.mark.parametrize(
    ("outputdir", "testfile"),
    [
        ("out", CS_CYCLES),
        ("", CS_CYCLES),
        ("out", CS_CYCLES_TURTLE),
        ("", CS_CYCLES_TURTLE),
    ],
    ids=["out:dir & xlsx", "out:default & xlsx", "out:dir & ttl", "out:default & ttl"],
)
def test_run_vocexcel_outputdir(monkeypatch, datadir, tmp_path, outputdir, testfile):
    """Check that an xlsx file is converted to ttl by vocexcel."""
    shutil.copy(datadir / testfile, tmp_path)
    monkeypatch.chdir(tmp_path)
    # Check if log is placed in out folder.
    log = "test-run.log"
    main_cli(
        ["--logfile", str(log)]
        + (["--output-directory", str(outputdir)] if outputdir else [])
        + [str(tmp_path)]
    )
    outdir = tmp_path / outputdir
    if testfile.endswith("xlsx"):
        assert (outdir / testfile).with_suffix(".ttl").exists()
    else:
        assert (outdir / testfile).with_suffix(".xlsx").exists()
    assert (outdir / log).exists()


@pytest.mark.parametrize(
    "test_file",
    [CS_CYCLES, ""],
    ids=["single file", "dir of files"],
)
def test_forwarding_3stages(monkeypatch, datadir, tmp_path, test_file):
    """Check a file by voc4cat then forward it to vocexcel then to pyLODE."""
    dst = tmp_path / test_file
    shutil.copy(datadir / CS_CYCLES, tmp_path)
    monkeypatch.chdir(tmp_path)
    main_cli(
        [
            "--check",
            "--forward",
            "--logfile",
            "test.log",
            "--no-warn",
            "--docs",
            "pylode",
            str(dst),
        ]
    )
    assert (tmp_path / CS_CYCLES).with_suffix(".ttl").exists()
    assert (tmp_path / Path(CS_CYCLES).stem / "index.html").exists()
    assert (tmp_path / "test.log").exists()


@pytest.mark.parametrize(
    "test_file",
    [CS_CYCLES, ""],
    ids=["single file", "dir of files"],
)
def test_forwarding_3stages_outdir(monkeypatch, datadir, tmp_path, test_file):
    """Check file by voc4cat, write it to output folder, forward to vocexcel & pyLODE.

    Related: #issue106
    """
    monkeypatch.chdir(tmp_path)
    shutil.copy(datadir / CS_CYCLES, tmp_path)
    main_cli(
        [
            "--check",
            "--forward",
            "--output-directory",
            str(tmp_path / "out"),
            "--logfile",
            "test.log",
            "--docs",
            "pylode",
            str(tmp_path / test_file),
        ]
    )
    assert (tmp_path / "out" / CS_CYCLES).with_suffix(".ttl").exists()
    assert (tmp_path / "out" / Path(CS_CYCLES).stem / "index.html").exists()
    assert (tmp_path / "out" / "test.log").exists()


@pytest.mark.parametrize(
    "test_file",
    [CS_CYCLES, ""],
    ids=["single file", "dir of files"],
)
def test_forwarding_2stages(monkeypatch, datadir, tmp_path, test_file):
    """Use voc4cat to run vocexcel then forward result to ontospy."""
    dst = tmp_path / test_file
    shutil.copy(datadir / CS_CYCLES, tmp_path)
    monkeypatch.chdir(tmp_path)
    main_cli(
        [
            "--forward",
            "--logfile",
            "test.log",
            "--no-warn",
            "--docs",
            "ontospy",
            str(dst),
        ]
    )
    assert (tmp_path / CS_CYCLES).with_suffix(".ttl").exists()
    assert (tmp_path / Path(CS_CYCLES).stem / "dendro" / "index.html").exists()
    assert (tmp_path / Path(CS_CYCLES).stem / "docs" / "index.html").exists()
    assert (tmp_path / "test.log").exists()
