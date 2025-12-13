"""Tests for the template_v1 module.

These tests verify that the v1.0 template generator creates Excel templates
with the correct structure matching the reference template.
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from voc4cat.cli import main_cli
from voc4cat.models_v1 import (
    DEFAULT_PREFIXES,
    TEMPLATE_VERSION,
    CollectionV1,
    ConceptSchemeV1,
    ConceptV1,
    MappingV1,
    PrefixV1,
)
from voc4cat.template_v1 import generate_template_v1

# Path to reference template for comparison
REFERENCE_TEMPLATE = (
    Path(__file__).parent.parent
    / "src"
    / "voc4cat"
    / "templates"
    / "vocab"
    / "blank_1.0_min.xlsx"
)


class TestTemplateGeneration:
    """Tests for template generation functionality."""

    def test_generate_template_creates_all_sheets(self, tmp_path):
        """Verify all 6 sheets are created."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)

        expected_sheets = {
            "Concept Scheme",
            "Concepts",
            "Collections",
            "Mappings",
            "ID Ranges",
            "Prefixes",
        }
        actual_sheets = set(wb.sheetnames)

        # Should have all expected sheets
        assert expected_sheets.issubset(actual_sheets), (
            f"Missing sheets: {expected_sheets - actual_sheets}"
        )

    def test_generate_template_sheet_order(self, tmp_path):
        """Verify sheets are in the correct order."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)

        expected_order = [
            "Concept Scheme",
            "Concepts",
            "Collections",
            "Mappings",
            "ID Ranges",
            "Prefixes",
        ]
        actual_order = wb.sheetnames[:6]

        assert actual_order == expected_order, (
            f"Sheet order mismatch: expected {expected_order}, got {actual_order}"
        )

    def test_concepts_sheet_has_table(self, tmp_path):
        """Verify Concepts sheet has proper table with correct columns."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Concepts"]

        # Check table exists
        assert ws.tables, "Concepts sheet should have a table"
        table_names = list(ws.tables.keys())
        assert len(table_names) == 1, "Should have exactly one table"

        # Check column headers (row 4 contains headers)
        expected_headers = [
            "Concept IRI*",
            "Language Code*",
            "Preferred Label*",
            "Definition*",
            "Alternate Labels",
            "Parent IRIs",
            "Member of collection(s)",
            "Member of ordered collection # position",
            "Provenance (read-only)",
            "Change Note",
            "Editorial Note",
            "Obsoletion reason",
            "Influenced by IRIs",
            "Source Vocab IRI or URL",
            "Source Vocab License",
            "Source Vocab Rights Holder",
        ]

        for col, expected_header in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual_header = ws[f"{col_letter}4"].value
            assert actual_header == expected_header, (
                f"Column {col} header mismatch: expected '{expected_header}', got '{actual_header}'"
            )

    def test_concepts_sheet_has_freeze_panes(self, tmp_path):
        """Verify Concepts sheet has freeze panes set."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Concepts"]

        assert ws.freeze_panes == "A5", (
            f"Expected freeze panes at A5, got {ws.freeze_panes}"
        )

    def test_collections_sheet_has_table(self, tmp_path):
        """Verify Collections sheet has proper table."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Collections"]

        assert ws.tables, "Collections sheet should have a table"

        # Check column headers
        expected_headers = [
            "Collection IRI",
            "Language Code*",
            "Preferred Label",
            "Definition",
            "Parent Collection IRIs",
            "Ordered? Yes or No (default)",
            "Provenance (read-only)",
            "Change Note*",
            "Editorial Note",
            "Obsoletion reason",
        ]

        for col, expected_header in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual_header = ws[f"{col_letter}4"].value
            assert actual_header == expected_header, (
                f"Column {col} header mismatch: expected '{expected_header}', got '{actual_header}'"
            )

    def test_mappings_sheet_has_table(self, tmp_path):
        """Verify Mappings sheet has proper table."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Mappings"]

        assert ws.tables, "Mappings sheet should have a table"

        # Check column headers
        expected_headers = [
            "Concept IRI*",
            "Related Matches",
            "Close Matches",
            "Exact Matches",
            "Narrower Matches",
            "Broader Matches",
            "Editorial Note",
        ]

        for col, expected_header in enumerate(expected_headers, 1):
            col_letter = get_column_letter(col)
            actual_header = ws[f"{col_letter}4"].value
            assert actual_header == expected_header, (
                f"Column {col} header mismatch: expected '{expected_header}', got '{actual_header}'"
            )

    def test_prefixes_sheet_has_default_prefixes(self, tmp_path):
        """Verify Prefixes sheet contains standard prefixes."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Prefixes"]

        # Check table exists
        assert ws.tables, "Prefixes sheet should have a table"

        # Check some key prefixes are present
        expected_prefixes = {"ex", "skos"}
        found_prefixes = set()

        # Prefixes start after header row
        for row in range(4, ws.max_row + 1):
            prefix = ws[f"A{row}"].value
            if prefix:
                found_prefixes.add(prefix)

        assert expected_prefixes.issubset(found_prefixes), (
            f"Missing expected prefixes: {expected_prefixes - found_prefixes}"
        )

    def test_concept_scheme_has_correct_fields(self, tmp_path):
        """Verify Concept Scheme sheet has correct fields."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["Concept Scheme"]

        # Check that Template version is first field
        assert ws["A4"].value == "Template version"
        assert ws["B4"].value == TEMPLATE_VERSION

        # Check other key fields exist
        expected_fields = [
            "Vocabulary IRI",
            "Prefix",
            "Title",
            "Description",
            "Created Date",
        ]

        found_fields = []
        for row in range(4, ws.max_row + 1):
            field = ws[f"A{row}"].value
            if field:
                found_fields.append(field)

        for field in expected_fields:
            assert field in found_fields, f"Missing field: {field}"

    def test_id_ranges_sheet_has_correct_structure(self, tmp_path):
        """Verify ID Ranges sheet has title and correct headers."""
        output_path = tmp_path / "test_template.xlsx"
        wb = generate_template_v1(output_path)
        ws = wb["ID Ranges"]

        # Check title row
        assert ws["A1"].value == "ID Ranges (read-only)"

        # Check headers (row 3)
        expected_headers = ["gh-name", "ID Range", "Unused IDs"]
        actual_headers = [ws["A3"].value, ws["B3"].value, ws["C3"].value]
        assert actual_headers == expected_headers

        # Data row should be empty (headers only in blank template)
        assert ws["A4"].value is None or ws["A4"].value == ""


class TestTemplateComparisonWithReference:
    """Tests comparing generated template against reference template."""

    @pytest.fixture
    def reference_workbook(self):
        """Load reference template for comparison."""
        if not REFERENCE_TEMPLATE.exists():
            pytest.skip(f"Reference template not found: {REFERENCE_TEMPLATE}")
        return load_workbook(REFERENCE_TEMPLATE)

    @pytest.fixture
    def generated_workbook(self, tmp_path):
        """Generate a template for comparison."""
        output_path = tmp_path / "generated_template.xlsx"
        return generate_template_v1(output_path)

    def test_concepts_columns_match_reference(
        self, reference_workbook, generated_workbook
    ):
        """Compare Concepts column headers against reference."""
        ref_ws = reference_workbook["Concepts"]
        gen_ws = generated_workbook["Concepts"]

        # Compare headers in row 4
        for col in range(1, 16):
            col_letter = get_column_letter(col)
            ref_header = ref_ws[f"{col_letter}4"].value
            gen_header = gen_ws[f"{col_letter}4"].value
            assert ref_header == gen_header, (
                f"Column {col} header mismatch: reference='{ref_header}', generated='{gen_header}'"
            )

    def test_concepts_meanings_match_reference(
        self, reference_workbook, generated_workbook
    ):
        """Compare Concepts SKOS meanings against reference."""
        ref_ws = reference_workbook["Concepts"]
        gen_ws = generated_workbook["Concepts"]

        # Compare meanings in row 3
        for col in range(1, 16):
            col_letter = get_column_letter(col)
            ref_meaning = ref_ws[f"{col_letter}3"].value
            gen_meaning = gen_ws[f"{col_letter}3"].value
            assert ref_meaning == gen_meaning, (
                f"Column {col} meaning mismatch: reference='{ref_meaning}', generated='{gen_meaning}'"
            )

    def test_collections_columns_match_reference(
        self, reference_workbook, generated_workbook
    ):
        """Compare Collections column headers against reference."""
        ref_ws = reference_workbook["Collections"]
        gen_ws = generated_workbook["Collections"]

        # Compare headers in row 4
        for col in range(1, 10):
            col_letter = get_column_letter(col)
            ref_header = ref_ws[f"{col_letter}4"].value
            gen_header = gen_ws[f"{col_letter}4"].value
            assert ref_header == gen_header, (
                f"Column {col} header mismatch: reference='{ref_header}', generated='{gen_header}'"
            )

    def test_mappings_columns_match_reference(
        self, reference_workbook, generated_workbook
    ):
        """Compare Mappings column headers against reference."""
        ref_ws = reference_workbook["Mappings"]
        gen_ws = generated_workbook["Mappings"]

        # Compare headers in row 4
        for col in range(1, 8):
            col_letter = get_column_letter(col)
            ref_header = ref_ws[f"{col_letter}4"].value
            gen_header = gen_ws[f"{col_letter}4"].value
            assert ref_header == gen_header, (
                f"Column {col} header mismatch: reference='{ref_header}', generated='{gen_header}'"
            )


class TestModels:
    """Tests for v1.0 Pydantic models."""

    def test_concept_scheme_default_values(self):
        """Test ConceptSchemeV1 has correct defaults."""
        cs = ConceptSchemeV1()
        assert cs.template_version == TEMPLATE_VERSION
        assert cs.vocabulary_iri == ""
        assert cs.prefix == ""

    def test_concept_v1_model(self):
        """Test ConceptV1 model creation."""
        concept = ConceptV1(
            concept_iri="ex:0001",
            language_code="en",
            preferred_label="Test",
            definition="A test concept",
        )
        assert concept.concept_iri == "ex:0001"
        assert concept.language_code == "en"
        assert concept.preferred_label == "Test"
        assert concept.alternate_labels == ""  # default

    def test_collection_v1_model(self):
        """Test CollectionV1 model creation."""
        collection = CollectionV1(
            collection_iri="ex:col1",
            language_code="en",
            preferred_label="Collection",
        )
        assert collection.collection_iri == "ex:col1"
        assert collection.ordered == ""  # default

    def test_mapping_v1_model(self):
        """Test MappingV1 model creation."""
        mapping = MappingV1(
            concept_iri="ex:0001",
            exact_matches="http://example.org/ext/concept1",
        )
        assert mapping.concept_iri == "ex:0001"
        assert mapping.exact_matches == "http://example.org/ext/concept1"
        assert mapping.related_matches == ""  # default

    def test_prefix_v1_model(self):
        """Test PrefixV1 model creation."""
        prefix = PrefixV1(prefix="ex", namespace="https://example.org/")
        assert prefix.prefix == "ex"
        assert prefix.namespace == "https://example.org/"

    def test_default_prefixes_count(self):
        """Test that default prefixes list has expected entries."""
        assert len(DEFAULT_PREFIXES) == 2, (
            f"Expected 2 prefixes, got {len(DEFAULT_PREFIXES)}"
        )


class TestTemplateCLI:
    """CLI integration tests for the template command."""

    def test_template_cli_default_output(self, tmp_path, monkeypatch):
        """Test template command generates file in current directory."""
        monkeypatch.chdir(tmp_path)
        main_cli(["template"])

        expected_file = tmp_path / "blank_v1.0.xlsx"
        assert expected_file.exists(), f"Template file not created: {expected_file}"

        # Verify it's a valid xlsx with expected sheets
        wb = load_workbook(expected_file)
        assert "Concepts" in wb.sheetnames

    def test_template_cli_with_outdir(self, tmp_path):
        """Test template command with --outdir option."""
        outdir = tmp_path / "output"
        main_cli(["template", "--outdir", str(outdir)])

        expected_file = outdir / "blank_v1.0.xlsx"
        assert expected_file.exists(), (
            f"Template file not created in outdir: {expected_file}"
        )

    def test_template_cli_with_version_option(self, tmp_path, monkeypatch):
        """Test template command with --version option."""
        monkeypatch.chdir(tmp_path)
        main_cli(["template", "--version", "v1.0"])
        assert (tmp_path / "blank_v1.0.xlsx").exists()

    def test_template_cli_invalid_version(self):
        """Test template command with invalid --version option."""
        with pytest.raises(SystemExit) as exc_info:
            main_cli(["template", "--version", "v0.5"])
        assert exc_info.value.code == 2  # argparse error
