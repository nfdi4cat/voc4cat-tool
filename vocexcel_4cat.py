"""
Add mechanism to extend VocExcel with more commands

The new commands be may be run either before or after VocExcel.

Copyright (c) 2022 David Linke (ORCID: 0000-0002-5898-1820)
"""
import argparse
from bdb import set_trace
import os
import sys
from pathlib import Path

import openpyxl

from rdflib import URIRef

from vocexcel.convert import main
from vocexcel.utils import EXCEL_FILE_ENDINGS
from vocexcel.models import ORGANISATIONS

from django.utils.text import slugify

__version__ = "0.1.0"

ORGANISATIONS["NFDI4Cat"] = URIRef("http://example.org/nfdi4cat/")
ORGANISATIONS["LIKAT"] = URIRef("https://www.catalsis.de/")


def is_excel_file_available(fname):
    if fname is None or not fname.suffix.lower().endswith(tuple(EXCEL_FILE_ENDINGS)):
        print("Files for preprocessing must end with .xlsx (Excel).")
        return False
    if not os.path.exists(fname):
        print(f"Excel file not found: {fname}s")
        return False
    return True


def is_supported_template(wb):
    #TODO write check
    return True


def add_IRI(fpath):
    """
    Add IRIs from preferred label in col A of sheets Concepts & Collections

    Safe and valid IRIs are created using django's slugify function.
    Column A is only updated for the rows where the cell is empty.
    """
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)
    VOC_BASE_IRI = wb["Concept Scheme"].cell(row=2, column=2).value
    if not VOC_BASE_IRI.endswith("/"):
        VOC_BASE_IRI += "/"

    # process both worksheets
    subsequent_empty_rows = 0
    for sheet in ["Concepts", "Collections"]:
        ws = wb[sheet]
        # iterate over first two columns; skip header and start from row 3
        for row in ws.iter_rows(min_row=3, max_col=2):
            if not row[0].value and row[1].value:
                concept_iri = VOC_BASE_IRI + slugify(row[1].value)
                ws.cell(row[0].row, column=1).value = concept_iri
                subsequent_empty_rows = 0
            elif row[0].value is None and row[1].value is None:
                # stop processing a sheet after 3 empty rows
                if subsequent_empty_rows < 2:
                    subsequent_empty_rows += 1
                else:
                    break

    output_fnanme = "%s_i.%s" % tuple(str(fpath).rsplit(".", 1))
    wb.save(output_fnanme)
    print(f"Saved updated file as {output_fnanme}")


def add_related(fpath):
    """
    Add related Children URI and Members URI by preferred label if none is present.

    In detail the folowing sheets are modified:
    "Concepts": update col. G "Children URI" from col. J "Children by Pref. Label"
    "Collections": update col. D "Members URI" from col. F "Members by Pref. Label"
    """
    wb = openpyxl.load_workbook(fpath)
    is_supported_template(wb)

    # process both worksheets
    subsequent_empty_rows = 0
    for sheet in ["Concepts", "Collections"]:
        print(f"Processing sheet '{sheet}'")
        ws = wb[sheet]
        pref_label_lookup = {}
        # read all peferred labels from the sheet
        for row in ws.iter_rows(min_row=3, max_col=2):
            if row[0].value and row[1].value:
                pref_label_lookup[slugify(row[1].value)] = row[0].value
                subsequent_empty_rows = 0
            elif row[0].value is None and row[1].value is None:
                # stop processing a sheet after 3 empty rows
                if subsequent_empty_rows < 2:
                    subsequent_empty_rows += 1
                else:
                    subsequent_empty_rows = 0
                    break

        # update read all peferred labels from the sheet
        col_to_fill = 6 if sheet == "Concepts" else 3
        col_preflabel = 9 if sheet == "Concepts" else 5
        for row in ws.iter_rows(min_row=3, max_col=col_preflabel + 1):
            if not row[col_to_fill].value and row[col_preflabel].value:
                # update cell in col_to_fill
                pref_label_slugs = [
                    slugify(pl.strip()) for pl in row[col_preflabel].value.split(",")
                ]
                found = [
                    pref_label_lookup[pl]
                    for pl in pref_label_slugs
                    if pl in pref_label_lookup
                ]
                ws.cell(row[0].row, column=col_to_fill + 1).value = ", ".join(found)
                subsequent_empty_rows = 0
            elif row[0].value is None and row[1].value is None:
                # stop processing a sheet after 3 empty rows
                if subsequent_empty_rows < 2:
                    subsequent_empty_rows += 1
                else:
                    break

    output_fnanme = "%s_r.%s" % tuple(str(fpath).rsplit(".", 1))
    wb.save(output_fnanme)
    print(f"Saved updated file as {output_fnanme}")


def run_vocexcel(args=None):
    main(args)


def wrapper(args=None):
    print("Running VocExcel wrapper")
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    parser.add_argument(
        "-v",
        "--version",
        help="The version of this wrapper of VocExel.",
        action="store_true",
    )

    parser.add_argument(
        "-i",
        "--addIRI",
        help="Add IRI (http://example.org/...) for concepts and collections if none is present.",
        action="store_true",
    )

    parser.add_argument(
        "-r",
        "--add_related",
        help="Add related Children URI and Members URI by preferred label if none is present.",
        action="store_true",
    )

    parser.add_argument(
        "file_to_preprocess",
        nargs="?",  # allow 0 or 1 file name as argument
        type=Path,
        help="The Excel file to preprocess and convert to a SKOS vocabulary.",
    )

    args_wrapper = parser.parse_args()

    if len(sys.argv) == 1:
        # show help if no args are given
        parser.print_help()
        parser.exit()

    if args_wrapper.version:
        print(f"{__version__}")
    elif args_wrapper.addIRI:
        if is_excel_file_available(args_wrapper.file_to_preprocess):
            print(f"Processing file {args_wrapper.file_to_preprocess}")
            add_IRI(args_wrapper.file_to_preprocess)
        else:
            parser.exit()
    elif args_wrapper.add_related:
        if is_excel_file_available(args_wrapper.file_to_preprocess):
            print(f"Processing file {args_wrapper.file_to_preprocess}")
            add_related(args_wrapper.file_to_preprocess)
        else:
            parser.exit()
    elif args_wrapper and args_wrapper.file_to_preprocess:
        if is_excel_file_available(args_wrapper.file_to_preprocess):
            print(f"Processing file {args_wrapper.file_to_preprocess}")
            print("\nCalling VocExcel")
            main(args)
        else:
            print("Files for preprocessing must end with .xlsx (Excel).")
            parser.exit()
    else:
        print("\nCalling VocExcel")
        # We add a dummy since vocexcel must be called with a filename.
        main(args.insert(1, "."))

    return args_wrapper


if __name__ == "__main__":
    w = wrapper(sys.argv[1:])
